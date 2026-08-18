# app/blueprints/transport/services.py

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable

from sqlalchemy import exists, func, or_, select
from sqlalchemy.exc import IntegrityError
from io import BytesIO
from datetime import date, datetime, timedelta
from sqlalchemy.orm import joinedload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from app.extensions import db
from app.models.transport import (
    Driver,
    DriverApmRecord,
    DriverDocument,
    DriverExitPermission,
    DriverTruckAssignment,
    TransportDocumentChange,
    TransportIncident,
    TransportIncidentFollowUp,
    Truck,
    TruckDocument,
    TruckOwner,
)


# =========================================================
# EXCEPCIONES DEL MÓDULO
# =========================================================
class TransportServiceError(Exception):
    """
    Error base controlado del módulo de transportistas.
    """


class TransportValidationError(TransportServiceError):
    """
    Error de validación de datos o reglas de negocio.
    """


class TransportNotFoundError(TransportServiceError):
    """
    El registro solicitado no existe.
    """


class TransportConflictError(TransportServiceError):
    """
    La operación entra en conflicto con otro registro existente.
    """


# =========================================================
# CONSTANTES
# =========================================================
DRIVER_STATUSES = {
    "ACTIVE",
    "INACTIVE",
    "SUSPENDED",
}

DRIVER_DOCUMENT_TYPES = {
    "DOCK_PERMIT",
    "GENERAL_CARD",
    "CHEMICAL_PERMIT",
    "LICENSE",
    "CRIMINAL_RECORD",
}

DOCUMENT_STATUSES = {
    "PENDING",
    "VALID",
    "EXPIRED",
    "NOT_APPLICABLE",
}

APM_TRAINING_STATUSES = {
    "PENDING",
    "YES",
}

APM_CARD_STATUSES = {
    "PENDING",
    "EXPIRED",
    "YES",
}

APM_EXPIRY_MODES = {
    "PENDING",
    "EXPIRED",
    "NO_EXPIRY",
    "DATE",
}

TRUCK_STATUSES = {
    "ACTIVE",
    "INACTIVE",
    "DAMAGED",
    "STRANDED",
}

TRUCK_BONDED_STATUSES = {
    "BONDED",
    "PENDING",
}

ASSIGNMENT_STATUSES = {
    "ACTIVE",
    "ENDED",
}

EXIT_PERMISSION_STATUSES = {
    "AUTHORIZED",
    "RETURNED",
    "CANCELLED",
}

INCIDENT_TYPES = {
    "DAMAGED",
    "STRANDED",
}

INCIDENT_STATUSES = {
    "OPEN",
    "FOLLOW_UP",
    "RESOLVED",
    "CANCELLED",
}


# =========================================================
# HELPERS DE NORMALIZACIÓN
# =========================================================
def _clean_text(
    value: Any,
    *,
    upper: bool = False,
    max_length: int | None = None,
) -> str | None:
    """
    Limpia texto recibido desde formularios o archivos.

    Retorna None cuando el valor viene vacío.
    """
    if value is None:
        return None

    cleaned = str(value).strip()

    if not cleaned:
        return None

    if upper:
        cleaned = cleaned.upper()

    if max_length is not None:
        cleaned = cleaned[:max_length]

    return cleaned


def _required_text(
    value: Any,
    field_name: str,
    *,
    upper: bool = False,
    max_length: int | None = None,
) -> str:
    cleaned = _clean_text(
        value,
        upper=upper,
        max_length=max_length,
    )

    if not cleaned:
        raise TransportValidationError(
            f"El campo {field_name} es obligatorio."
        )

    return cleaned


def _normalize_identification(value: Any) -> str:
    """
    Conserva letras y números, pero elimina espacios y separadores
    comunes para evitar duplicados por formato.
    """
    cleaned = _required_text(
        value,
        "identificación",
        upper=True,
        max_length=40,
    )

    return (
        cleaned
        .replace(" ", "")
        .replace("-", "")
        .replace(".", "")
    )


def _normalize_plate(value: Any) -> str:
    """
    Normaliza la placa para que abc-123 y ABC123 no se dupliquen.
    """
    cleaned = _required_text(
        value,
        "placa",
        upper=True,
        max_length=40,
    )

    return (
        cleaned
        .replace(" ", "")
        .replace("-", "")
        .replace(".", "")
    )


def _parse_int(
    value: Any,
    field_name: str,
    *,
    required: bool = False,
) -> int | None:
    if value in (None, ""):
        if required:
            raise TransportValidationError(
                f"El campo {field_name} es obligatorio."
            )

        return None

    try:
        parsed = int(value)
    except (TypeError, ValueError):
        raise TransportValidationError(
            f"El campo {field_name} debe ser numérico."
        )

    return parsed


def _parse_decimal(
    value: Any,
    field_name: str,
) -> Decimal | None:
    if value in (None, ""):
        return None

    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise TransportValidationError(
            f"El campo {field_name} debe ser numérico."
        )

    return parsed


def _parse_bool(
    value: Any,
    *,
    default: bool = False,
) -> bool:
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    normalized = str(value).strip().lower()

    return normalized in {
        "1",
        "true",
        "yes",
        "si",
        "sí",
        "on",
    }


def _parse_date(
    value: Any,
    field_name: str,
    *,
    required: bool = False,
) -> date | None:
    if value in (None, ""):
        if required:
            raise TransportValidationError(
                f"El campo {field_name} es obligatorio."
            )

        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    try:
        return datetime.strptime(
            str(value).strip(),
            "%Y-%m-%d",
        ).date()
    except (TypeError, ValueError):
        raise TransportValidationError(
            f"El campo {field_name} debe tener formato YYYY-MM-DD."
        )


def _parse_datetime(
    value: Any,
    field_name: str,
    *,
    required: bool = False,
) -> datetime | None:
    if value in (None, ""):
        if required:
            raise TransportValidationError(
                f"El campo {field_name} es obligatorio."
            )

        return None

    if isinstance(value, datetime):
        return value

    raw = str(value).strip()

    formats = (
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    )

    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue

    raise TransportValidationError(
        f"El campo {field_name} contiene una fecha u hora inválida."
    )


def _validate_choice(
    value: Any,
    field_name: str,
    allowed_values: set[str],
    *,
    default: str | None = None,
) -> str:
    cleaned = _clean_text(
        value,
        upper=True,
    )

    if not cleaned:
        if default is not None:
            return default

        raise TransportValidationError(
            f"El campo {field_name} es obligatorio."
        )

    if cleaned not in allowed_values:
        allowed = ", ".join(sorted(allowed_values))

        raise TransportValidationError(
            f"El valor de {field_name} no es válido. "
            f"Valores permitidos: {allowed}."
        )

    return cleaned


def _different(
    old_value: Any,
    new_value: Any,
) -> bool:
    return str(old_value or "") != str(new_value or "")


def _display_value(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, bool):
        return "SI" if value else "NO"

    return str(value)


# =========================================================
# HELPERS DE BASE DE DATOS
# =========================================================
def _commit_or_raise(
    conflict_message: str = "No fue posible completar la operación.",
) -> None:
    try:
        db.session.commit()
    except IntegrityError as exc:
        db.session.rollback()

        raise TransportConflictError(
            conflict_message
        ) from exc


def _driver_exists(driver_id: int) -> bool:
    return bool(
        db.session.scalar(
            select(
                exists().where(
                    Driver.id == driver_id
                )
            )
        )
    )


def _truck_exists(truck_id: int) -> bool:
    return bool(
        db.session.scalar(
            select(
                exists().where(
                    Truck.id == truck_id
                )
            )
        )
    )


def get_driver_or_404(driver_id: int) -> Driver:
    driver = db.session.get(
        Driver,
        driver_id,
    )

    if driver is None:
        raise TransportNotFoundError(
            "El chofer solicitado no existe."
        )

    return driver


def get_truck_or_404(truck_id: int) -> Truck:
    truck = db.session.get(
        Truck,
        truck_id,
    )

    if truck is None:
        raise TransportNotFoundError(
            "El cabezal solicitado no existe."
        )

    return truck


def get_assignment_or_404(
    assignment_id: int,
) -> DriverTruckAssignment:
    assignment = db.session.get(
        DriverTruckAssignment,
        assignment_id,
    )

    if assignment is None:
        raise TransportNotFoundError(
            "La asignación solicitada no existe."
        )

    return assignment


def get_exit_permission_or_404(
    permission_id: int,
) -> DriverExitPermission:
    permission = db.session.get(
        DriverExitPermission,
        permission_id,
    )

    if permission is None:
        raise TransportNotFoundError(
            "El permiso de salida solicitado no existe."
        )

    return permission


def get_incident_or_404(
    incident_id: int,
) -> TransportIncident:
    incident = db.session.get(
        TransportIncident,
        incident_id,
    )

    if incident is None:
        raise TransportNotFoundError(
            "El incidente solicitado no existe."
        )

    return incident


# =========================================================
# HISTORIAL DOCUMENTAL
# =========================================================
def register_document_change(
    *,
    entity_type: str,
    entity_id: int,
    document_type: str,
    field_name: str,
    old_value: Any,
    new_value: Any,
    changed_by_user_id: int,
    notes: str | None = None,
) -> TransportDocumentChange | None:
    """
    Registra únicamente cambios reales.

    No ejecuta commit por sí solo para permitir que forme parte de
    la misma transacción del registro principal.
    """
    entity_type = _validate_choice(
        entity_type,
        "tipo de entidad",
        {
            "DRIVER",
            "TRUCK",
            "APM",
        },
    )

    if not _different(
        old_value,
        new_value,
    ):
        return None

    change = TransportDocumentChange(
        entity_type=entity_type,
        entity_id=entity_id,
        document_type=_required_text(
            document_type,
            "tipo de documento",
            upper=True,
            max_length=50,
        ),
        field_name=_required_text(
            field_name,
            "campo modificado",
            max_length=80,
        ),
        old_value=_display_value(old_value),
        new_value=_display_value(new_value),
        changed_by_user_id=changed_by_user_id,
        notes=_clean_text(notes),
    )

    db.session.add(change)

    return change


# =========================================================
# PROPIETARIOS
# =========================================================
def get_or_create_truck_owner(
    *,
    name: Any,
    phone: Any = None,
    email: Any = None,
    notes: Any = None,
) -> TruckOwner | None:
    """
    Busca un propietario por nombre y teléfono.

    Usa una consulta puntual y no carga toda la tabla.
    No ejecuta commit independiente.
    """
    owner_name = _clean_text(
        name,
        upper=True,
        max_length=160,
    )

    if not owner_name:
        return None

    owner_phone = _clean_text(
        phone,
        max_length=40,
    )

    owner = db.session.scalar(
        select(TruckOwner)
        .where(
            func.upper(TruckOwner.name) == owner_name,
            TruckOwner.phone == owner_phone,
        )
        .limit(1)
    )

    if owner is not None:
        if email is not None:
            owner.email = _clean_text(
                email,
                max_length=160,
            )

        if notes is not None:
            owner.notes = _clean_text(notes)

        if not owner.is_active:
            owner.is_active = True

        return owner

    owner = TruckOwner(
        name=owner_name,
        phone=owner_phone,
        email=_clean_text(
            email,
            max_length=160,
        ),
        notes=_clean_text(notes),
        is_active=True,
    )

    db.session.add(owner)
    db.session.flush()

    return owner


# =========================================================
# CHOFERES
# =========================================================
def create_driver(
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> Driver:
    identification = _normalize_identification(
        data.get("identification")
    )

    duplicate = db.session.scalar(
        select(
            exists().where(
                Driver.identification == identification
            )
        )
    )

    if duplicate:
        raise TransportConflictError(
            "Ya existe un chofer con esa identificación."
        )

    habitual_site_id = _parse_int(
        data.get("habitual_site_id"),
        "predio habitual",
    )

    driver = Driver(
        name=_required_text(
            data.get("name"),
            "nombre",
            upper=True,
            max_length=180,
        ),
        residence=_clean_text(
            data.get("residence"),
            upper=True,
            max_length=240,
        ),
        identification=identification,
        phone_1=_clean_text(
            data.get("phone_1"),
            max_length=40,
        ),
        phone_2=_clean_text(
            data.get("phone_2"),
            max_length=40,
        ),
        habitual_site_id=habitual_site_id,
        status=_validate_choice(
            data.get("status"),
            "estado",
            DRIVER_STATUSES,
            default="ACTIVE",
        ),
        notes=_clean_text(
            data.get("notes")
        ),
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )

    db.session.add(driver)
    db.session.flush()

    if commit:
        _commit_or_raise(
            "No fue posible crear el chofer. "
            "Verifique que la identificación no esté repetida."
        )

    return driver


def update_driver(
    driver: Driver,
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> Driver:
    new_identification = _normalize_identification(
        data.get(
            "identification",
            driver.identification,
        )
    )

    duplicate = db.session.scalar(
        select(
            exists().where(
                Driver.identification == new_identification,
                Driver.id != driver.id,
            )
        )
    )

    if duplicate:
        raise TransportConflictError(
            "Ya existe otro chofer con esa identificación."
        )

    driver.name = _required_text(
        data.get(
            "name",
            driver.name,
        ),
        "nombre",
        upper=True,
        max_length=180,
    )

    driver.residence = _clean_text(
        data.get(
            "residence",
            driver.residence,
        ),
        upper=True,
        max_length=240,
    )

    driver.identification = new_identification

    driver.phone_1 = _clean_text(
        data.get(
            "phone_1",
            driver.phone_1,
        ),
        max_length=40,
    )

    driver.phone_2 = _clean_text(
        data.get(
            "phone_2",
            driver.phone_2,
        ),
        max_length=40,
    )

    driver.habitual_site_id = _parse_int(
        data.get(
            "habitual_site_id",
            driver.habitual_site_id,
        ),
        "predio habitual",
    )

    driver.status = _validate_choice(
        data.get(
            "status",
            driver.status,
        ),
        "estado",
        DRIVER_STATUSES,
        default=driver.status,
    )

    driver.notes = _clean_text(
        data.get(
            "notes",
            driver.notes,
        )
    )

    driver.updated_by_user_id = user_id
    driver.updated_at = datetime.utcnow()

    if commit:
        _commit_or_raise(
            "No fue posible actualizar el chofer."
        )

    return driver


def set_driver_status(
    driver: Driver,
    status: str,
    *,
    user_id: int,
    commit: bool = True,
) -> Driver:
    new_status = _validate_choice(
        status,
        "estado",
        DRIVER_STATUSES,
    )

    if (
        new_status != "ACTIVE"
        and get_active_assignment_for_driver(
            driver.id
        ) is not None
    ):
        raise TransportValidationError(
            "No se puede desactivar o suspender al chofer "
            "mientras tenga un cabezal asignado."
        )

    driver.status = new_status
    driver.updated_by_user_id = user_id
    driver.updated_at = datetime.utcnow()

    if commit:
        _commit_or_raise(
            "No fue posible actualizar el estado del chofer."
        )

    return driver


# =========================================================
# DOCUMENTOS DE CHOFER
# =========================================================
def calculate_document_status(
    *,
    expiry_date: date | None,
    no_expiry: bool,
    requested_status: str | None = None,
) -> str:
    if requested_status:
        normalized = _validate_choice(
            requested_status,
            "estado del documento",
            DOCUMENT_STATUSES,
        )

        if normalized == "NOT_APPLICABLE":
            return normalized

    if no_expiry:
        return "VALID"

    if expiry_date is None:
        return "PENDING"

    if expiry_date < date.today():
        return "EXPIRED"

    return "VALID"


def upsert_driver_document(
    driver: Driver,
    document_type: str,
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> DriverDocument:
    normalized_type = _validate_choice(
        document_type,
        "tipo de documento",
        DRIVER_DOCUMENT_TYPES,
    )

    document = db.session.scalar(
        select(DriverDocument)
        .where(
            DriverDocument.driver_id == driver.id,
            DriverDocument.document_type == normalized_type,
        )
        .limit(1)
    )

    if document is None:
        document = DriverDocument(
            driver_id=driver.id,
            document_type=normalized_type,
            updated_by_user_id=user_id,
        )

        db.session.add(document)
        db.session.flush()

    old_values = {
        "document_number": document.document_number,
        "status": document.status,
        "issue_date": document.issue_date,
        "expiry_date": document.expiry_date,
        "no_expiry": document.no_expiry,
        "notes": document.notes,
    }

    issue_date = _parse_date(
        data.get("issue_date"),
        "fecha de emisión",
    )

    expiry_date = _parse_date(
        data.get("expiry_date"),
        "fecha de vencimiento",
    )

    no_expiry = _parse_bool(
        data.get("no_expiry"),
        default=False,
    )

    if no_expiry:
        expiry_date = None

    status = calculate_document_status(
        expiry_date=expiry_date,
        no_expiry=no_expiry,
        requested_status=data.get("status"),
    )

    document.document_number = _clean_text(
        data.get("document_number"),
        upper=True,
        max_length=100,
    )

    document.issue_date = issue_date
    document.expiry_date = expiry_date
    document.no_expiry = no_expiry
    document.status = status
    document.notes = _clean_text(
        data.get("notes")
    )
    document.updated_by_user_id = user_id
    document.updated_at = datetime.utcnow()

    new_values = {
        "document_number": document.document_number,
        "status": document.status,
        "issue_date": document.issue_date,
        "expiry_date": document.expiry_date,
        "no_expiry": document.no_expiry,
        "notes": document.notes,
    }

    for field_name, old_value in old_values.items():
        register_document_change(
            entity_type="DRIVER",
            entity_id=driver.id,
            document_type=normalized_type,
            field_name=field_name,
            old_value=old_value,
            new_value=new_values[field_name],
            changed_by_user_id=user_id,
        )

    if commit:
        _commit_or_raise(
            "No fue posible guardar el documento del chofer."
        )

    return document


def refresh_driver_document_statuses(
    *,
    driver_id: int | None = None,
    commit: bool = True,
) -> int:
    """
    Actualiza únicamente documentos con fecha de vencimiento.

    No carga relaciones ni objetos de choferes.
    """
    query = select(DriverDocument).where(
        DriverDocument.no_expiry.is_(False),
        DriverDocument.expiry_date.is_not(None),
        DriverDocument.status != "NOT_APPLICABLE",
    )

    if driver_id is not None:
        query = query.where(
            DriverDocument.driver_id == driver_id
        )

    documents = db.session.scalars(query).all()

    changed = 0
    today = date.today()

    for document in documents:
        expected_status = (
            "EXPIRED"
            if document.expiry_date < today
            else "VALID"
        )

        if document.status != expected_status:
            document.status = expected_status
            document.updated_at = datetime.utcnow()
            changed += 1

    if commit and changed:
        _commit_or_raise(
            "No fue posible actualizar los vencimientos."
        )

    return changed


# =========================================================
# REGISTRO APM
# =========================================================
def upsert_driver_apm_record(
    driver: Driver,
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> DriverApmRecord:
    record = db.session.scalar(
        select(DriverApmRecord)
        .where(
            DriverApmRecord.driver_id == driver.id
        )
        .limit(1)
    )

    if record is None:
        record = DriverApmRecord(
            driver_id=driver.id,
            updated_by_user_id=user_id,
        )

        db.session.add(record)
        db.session.flush()

    old_values = {
        "training_status": record.training_status,
        "card_status": record.card_status,
        "card_number": record.card_number,
        "expiry_mode": record.expiry_mode,
        "expiry_date": record.expiry_date,
        "notes": record.notes,
    }

    training_status = _validate_choice(
        data.get("training_status"),
        "estado de capacitación",
        APM_TRAINING_STATUSES,
        default="PENDING",
    )

    card_status = _validate_choice(
        data.get("card_status"),
        "estado del carné",
        APM_CARD_STATUSES,
        default="PENDING",
    )

    expiry_mode = _validate_choice(
        data.get("expiry_mode"),
        "modo de vencimiento",
        APM_EXPIRY_MODES,
        default="PENDING",
    )

    expiry_date = _parse_date(
        data.get("expiry_date"),
        "vencimiento APM",
    )

    if expiry_mode == "DATE" and expiry_date is None:
        raise TransportValidationError(
            "Debe indicar la fecha de vencimiento APM."
        )

    if expiry_mode != "DATE":
        expiry_date = None

    if (
        expiry_mode == "DATE"
        and expiry_date is not None
        and expiry_date < date.today()
    ):
        expiry_mode = "EXPIRED"
        card_status = "EXPIRED"

    if expiry_mode == "EXPIRED":
        card_status = "EXPIRED"

    record.training_status = training_status
    record.card_status = card_status
    record.card_number = _clean_text(
        data.get("card_number"),
        upper=True,
        max_length=100,
    )
    record.expiry_mode = expiry_mode
    record.expiry_date = expiry_date
    record.notes = _clean_text(
        data.get("notes")
    )
    record.updated_by_user_id = user_id
    record.updated_at = datetime.utcnow()

    new_values = {
        "training_status": record.training_status,
        "card_status": record.card_status,
        "card_number": record.card_number,
        "expiry_mode": record.expiry_mode,
        "expiry_date": record.expiry_date,
        "notes": record.notes,
    }

    for field_name, old_value in old_values.items():
        register_document_change(
            entity_type="APM",
            entity_id=driver.id,
            document_type="APM",
            field_name=field_name,
            old_value=old_value,
            new_value=new_values[field_name],
            changed_by_user_id=user_id,
        )

    if commit:
        _commit_or_raise(
            "No fue posible guardar la información APM."
        )

    return record


# =========================================================
# CABEZALES
# =========================================================
def create_truck(
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> Truck:
    plate = _normalize_plate(
        data.get("plate")
    )

    duplicate = db.session.scalar(
        select(
            exists().where(
                Truck.plate == plate
            )
        )
    )

    if duplicate:
        raise TransportConflictError(
            "Ya existe un cabezal con esa placa."
        )

    owner = get_or_create_truck_owner(
        name=data.get("owner_name"),
        phone=data.get("owner_phone"),
        email=data.get("owner_email"),
        notes=data.get("owner_notes"),
    )

    dekra_month = _parse_int(
        data.get("dekra_month"),
        "mes DEKRA",
    )

    dekra_year = _parse_int(
        data.get("dekra_year"),
        "año DEKRA",
    )

    if (
        dekra_month is not None
        and not 1 <= dekra_month <= 12
    ):
        raise TransportValidationError(
            "El mes DEKRA debe estar entre 1 y 12."
        )

    truck = Truck(
        registration_date=_parse_date(
            data.get("registration_date"),
            "fecha de registro",
            required=True,
        ),
        registered_site_id=_parse_int(
            data.get("registered_site_id"),
            "predio de registro",
            required=True,
        ),
        plate=plate,
        owner_id=owner.id if owner else None,
        status=_validate_choice(
            data.get("status"),
            "estado",
            TRUCK_STATUSES,
            default="ACTIVE",
        ),
        dock_permit_number=_clean_text(
            data.get("dock_permit_number"),
            upper=True,
            max_length=100,
        ),
        dock_permit_expiry_date=_parse_date(
            data.get("dock_permit_expiry_date"),
            "vencimiento del permiso de muelle",
        ),
        circulation_card=_clean_text(
            data.get("circulation_card"),
            upper=True,
            max_length=180,
        ),
        dekra_month=dekra_month,
        dekra_year=dekra_year,
        insurance_name=_clean_text(
            data.get("insurance_name"),
            upper=True,
            max_length=160,
        ),
        insurance_expiry_date=_parse_date(
            data.get("insurance_expiry_date"),
            "vencimiento del seguro",
        ),
        is_payroll=_parse_bool(
            data.get("is_payroll")
        ),
        rt_name=_clean_text(
            data.get("rt_name"),
            upper=True,
            max_length=180,
        ),
        rt_expiry_date=_parse_date(
            data.get("rt_expiry_date"),
            "vencimiento RT",
        ),
        weights_dimensions=_parse_decimal(
            data.get("weights_dimensions"),
            "pesos y dimensiones",
        ),
        policy_number=_clean_text(
            data.get("policy_number"),
            upper=True,
            max_length=120,
        ),
        bonded_status=_validate_choice(
            data.get("bonded_status"),
            "estado de caución",
            TRUCK_BONDED_STATUSES,
            default="PENDING",
        ),
        notes=_clean_text(
            data.get("notes")
        ),
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )

    db.session.add(truck)
    db.session.flush()

    if commit:
        _commit_or_raise(
            "No fue posible crear el cabezal. "
            "Verifique que la placa no esté repetida."
        )

    return truck


def update_truck(
    truck: Truck,
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> Truck:
    plate = _normalize_plate(
        data.get(
            "plate",
            truck.plate,
        )
    )

    duplicate = db.session.scalar(
        select(
            exists().where(
                Truck.plate == plate,
                Truck.id != truck.id,
            )
        )
    )

    if duplicate:
        raise TransportConflictError(
            "Ya existe otro cabezal con esa placa."
        )

    owner = get_or_create_truck_owner(
        name=data.get("owner_name"),
        phone=data.get("owner_phone"),
        email=data.get("owner_email"),
        notes=data.get("owner_notes"),
    )

    old_values = {
        "dock_permit_number": truck.dock_permit_number,
        "dock_permit_expiry_date": truck.dock_permit_expiry_date,
        "circulation_card": truck.circulation_card,
        "dekra_month": truck.dekra_month,
        "dekra_year": truck.dekra_year,
        "insurance_name": truck.insurance_name,
        "insurance_expiry_date": truck.insurance_expiry_date,
        "rt_name": truck.rt_name,
        "rt_expiry_date": truck.rt_expiry_date,
        "policy_number": truck.policy_number,
        "bonded_status": truck.bonded_status,
    }

    dekra_month = _parse_int(
        data.get(
            "dekra_month",
            truck.dekra_month,
        ),
        "mes DEKRA",
    )

    dekra_year = _parse_int(
        data.get(
            "dekra_year",
            truck.dekra_year,
        ),
        "año DEKRA",
    )

    if (
        dekra_month is not None
        and not 1 <= dekra_month <= 12
    ):
        raise TransportValidationError(
            "El mes DEKRA debe estar entre 1 y 12."
        )

    truck.registration_date = _parse_date(
        data.get(
            "registration_date",
            truck.registration_date,
        ),
        "fecha de registro",
        required=True,
    )

    truck.registered_site_id = _parse_int(
        data.get(
            "registered_site_id",
            truck.registered_site_id,
        ),
        "predio de registro",
        required=True,
    )

    truck.plate = plate
    truck.owner_id = owner.id if owner else None

    truck.status = _validate_choice(
        data.get(
            "status",
            truck.status,
        ),
        "estado",
        TRUCK_STATUSES,
        default=truck.status,
    )

    truck.dock_permit_number = _clean_text(
        data.get(
            "dock_permit_number",
            truck.dock_permit_number,
        ),
        upper=True,
        max_length=100,
    )

    truck.dock_permit_expiry_date = _parse_date(
        data.get(
            "dock_permit_expiry_date",
            truck.dock_permit_expiry_date,
        ),
        "vencimiento del permiso de muelle",
    )

    truck.circulation_card = _clean_text(
        data.get(
            "circulation_card",
            truck.circulation_card,
        ),
        upper=True,
        max_length=180,
    )

    truck.dekra_month = dekra_month
    truck.dekra_year = dekra_year

    truck.insurance_name = _clean_text(
        data.get(
            "insurance_name",
            truck.insurance_name,
        ),
        upper=True,
        max_length=160,
    )

    truck.insurance_expiry_date = _parse_date(
        data.get(
            "insurance_expiry_date",
            truck.insurance_expiry_date,
        ),
        "vencimiento del seguro",
    )

    truck.is_payroll = _parse_bool(
        data.get(
            "is_payroll",
            truck.is_payroll,
        ),
        default=truck.is_payroll,
    )

    truck.rt_name = _clean_text(
        data.get(
            "rt_name",
            truck.rt_name,
        ),
        upper=True,
        max_length=180,
    )

    truck.rt_expiry_date = _parse_date(
        data.get(
            "rt_expiry_date",
            truck.rt_expiry_date,
        ),
        "vencimiento RT",
    )

    truck.weights_dimensions = _parse_decimal(
        data.get(
            "weights_dimensions",
            truck.weights_dimensions,
        ),
        "pesos y dimensiones",
    )

    truck.policy_number = _clean_text(
        data.get(
            "policy_number",
            truck.policy_number,
        ),
        upper=True,
        max_length=120,
    )

    truck.bonded_status = _validate_choice(
        data.get(
            "bonded_status",
            truck.bonded_status,
        ),
        "estado de caución",
        TRUCK_BONDED_STATUSES,
        default=truck.bonded_status,
    )

    truck.notes = _clean_text(
        data.get(
            "notes",
            truck.notes,
        )
    )

    truck.updated_by_user_id = user_id
    truck.updated_at = datetime.utcnow()

    new_values = {
        "dock_permit_number": truck.dock_permit_number,
        "dock_permit_expiry_date": truck.dock_permit_expiry_date,
        "circulation_card": truck.circulation_card,
        "dekra_month": truck.dekra_month,
        "dekra_year": truck.dekra_year,
        "insurance_name": truck.insurance_name,
        "insurance_expiry_date": truck.insurance_expiry_date,
        "rt_name": truck.rt_name,
        "rt_expiry_date": truck.rt_expiry_date,
        "policy_number": truck.policy_number,
        "bonded_status": truck.bonded_status,
    }

    for field_name, old_value in old_values.items():
        register_document_change(
            entity_type="TRUCK",
            entity_id=truck.id,
            document_type="TRUCK",
            field_name=field_name,
            old_value=old_value,
            new_value=new_values[field_name],
            changed_by_user_id=user_id,
        )

    if commit:
        _commit_or_raise(
            "No fue posible actualizar el cabezal."
        )

    return truck


def set_truck_status(
    truck: Truck,
    status: str,
    *,
    user_id: int,
    commit: bool = True,
) -> Truck:
    new_status = _validate_choice(
        status,
        "estado",
        TRUCK_STATUSES,
    )

    if (
        new_status != "ACTIVE"
        and get_active_assignment_for_truck(
            truck.id
        ) is not None
    ):
        raise TransportValidationError(
            "No se puede cambiar el cabezal a ese estado "
            "mientras tenga un chofer asignado."
        )

    truck.status = new_status
    truck.updated_by_user_id = user_id
    truck.updated_at = datetime.utcnow()

    if commit:
        _commit_or_raise(
            "No fue posible actualizar el estado del cabezal."
        )

    return truck


# =========================================================
# DOCUMENTOS DE CABEZAL
# =========================================================
def upsert_truck_document(
    truck: Truck,
    document_type: str,
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> TruckDocument:
    normalized_type = _required_text(
        document_type,
        "tipo de documento",
        upper=True,
        max_length=50,
    )

    document = db.session.scalar(
        select(TruckDocument)
        .where(
            TruckDocument.truck_id == truck.id,
            TruckDocument.document_type == normalized_type,
        )
        .limit(1)
    )

    if document is None:
        document = TruckDocument(
            truck_id=truck.id,
            document_type=normalized_type,
            updated_by_user_id=user_id,
        )

        db.session.add(document)
        db.session.flush()

    old_values = {
        "document_number": document.document_number,
        "status": document.status,
        "issue_date": document.issue_date,
        "expiry_date": document.expiry_date,
        "no_expiry": document.no_expiry,
        "notes": document.notes,
    }

    issue_date = _parse_date(
        data.get("issue_date"),
        "fecha de emisión",
    )

    expiry_date = _parse_date(
        data.get("expiry_date"),
        "fecha de vencimiento",
    )

    no_expiry = _parse_bool(
        data.get("no_expiry"),
        default=False,
    )

    if no_expiry:
        expiry_date = None

    document.document_number = _clean_text(
        data.get("document_number"),
        upper=True,
        max_length=120,
    )

    document.status = calculate_document_status(
        expiry_date=expiry_date,
        no_expiry=no_expiry,
        requested_status=data.get("status"),
    )

    document.issue_date = issue_date
    document.expiry_date = expiry_date
    document.no_expiry = no_expiry
    document.notes = _clean_text(
        data.get("notes")
    )
    document.extra_data = data.get("extra_data")
    document.updated_by_user_id = user_id
    document.updated_at = datetime.utcnow()

    new_values = {
        "document_number": document.document_number,
        "status": document.status,
        "issue_date": document.issue_date,
        "expiry_date": document.expiry_date,
        "no_expiry": document.no_expiry,
        "notes": document.notes,
    }

    for field_name, old_value in old_values.items():
        register_document_change(
            entity_type="TRUCK",
            entity_id=truck.id,
            document_type=normalized_type,
            field_name=field_name,
            old_value=old_value,
            new_value=new_values[field_name],
            changed_by_user_id=user_id,
        )

    if commit:
        _commit_or_raise(
            "No fue posible guardar el documento del cabezal."
        )

    return document


# =========================================================
# ASIGNACIONES
# =========================================================
def get_active_assignment_for_driver(
    driver_id: int,
) -> DriverTruckAssignment | None:
    return db.session.scalar(
        select(DriverTruckAssignment)
        .where(
            DriverTruckAssignment.driver_id == driver_id,
            DriverTruckAssignment.status == "ACTIVE",
        )
        .limit(1)
    )

def update_driver_complete_row(
    driver: Driver,
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> Driver:
    """
    Actualiza una fila completa de la matriz de choferes.

    En una sola transacción puede modificar:

    - Datos principales del chofer.
    - Documentos del chofer.
    - Información APM.
    - Asignación activa chofer-cabezal.
    - Datos del cabezal asignado.
    - Propietario del cabezal.

    IMPORTANTE:
    - Solo procesa el chofer recibido.
    - No guarda la tabla completa.
    - Todo queda dentro de una única transacción.
    """

    # =====================================================
    # 1. DATOS GENERALES DEL CHOFER
    # =====================================================
    driver_data = {
        "name": data.get(
            "name",
            driver.name,
        ),
        "residence": data.get(
            "residence",
            driver.residence,
        ),
        "identification": data.get(
            "identification",
            driver.identification,
        ),
        "phone_1": data.get(
            "phone_1",
            driver.phone_1,
        ),
        "phone_2": data.get(
            "phone_2",
            driver.phone_2,
        ),
        "habitual_site_id": data.get(
            "habitual_site_id",
            driver.habitual_site_id,
        ),
        "status": data.get(
            "status",
            driver.status,
        ),
        "notes": data.get(
            "notes",
            driver.notes,
        ),
    }

    update_driver(
        driver,
        driver_data,
        user_id=user_id,
        commit=False,
    )

    # =====================================================
    # 2. DOCUMENTOS DEL CHOFER
    # =====================================================
    document_config = {
        "DOCK_PERMIT": "dock_permit",
        "GENERAL_CARD": "general_card",
        "CHEMICAL_PERMIT": "chemical_permit",
        "LICENSE": "license",
        "CRIMINAL_RECORD": "criminal_record",
    }

    for document_type, prefix in document_config.items():
        document_keys = {
            f"{prefix}_document_number",
            f"{prefix}_status",
            f"{prefix}_issue_date",
            f"{prefix}_expiry_date",
            f"{prefix}_no_expiry",
            f"{prefix}_notes",
        }

        # Solo modifica el documento si la fila envió
        # al menos uno de sus campos.
        if not any(
            key in data
            for key in document_keys
        ):
            continue

        document_data = {
            "document_number": data.get(
                f"{prefix}_document_number"
            ),
            "status": data.get(
                f"{prefix}_status"
            ),
            "issue_date": data.get(
                f"{prefix}_issue_date"
            ),
            "expiry_date": data.get(
                f"{prefix}_expiry_date"
            ),
            "no_expiry": data.get(
                f"{prefix}_no_expiry"
            ),
            "notes": data.get(
                f"{prefix}_notes"
            ),
        }

        upsert_driver_document(
            driver,
            document_type,
            document_data,
            user_id=user_id,
            commit=False,
        )

    # =====================================================
    # 3. INFORMACIÓN APM
    # =====================================================
    apm_keys = {
        "apm_training_status",
        "apm_card_status",
        "apm_card_number",
        "apm_expiry_mode",
        "apm_expiry_date",
        "apm_notes",
    }

    if any(
        key in data
        for key in apm_keys
    ):
        apm_data = {
            "training_status": data.get(
                "apm_training_status"
            ),
            "card_status": data.get(
                "apm_card_status"
            ),
            "card_number": data.get(
                "apm_card_number"
            ),
            "expiry_mode": data.get(
                "apm_expiry_mode"
            ),
            "expiry_date": data.get(
                "apm_expiry_date"
            ),
            "notes": data.get(
                "apm_notes"
            ),
        }

        upsert_driver_apm_record(
            driver,
            apm_data,
            user_id=user_id,
            commit=False,
        )

    # =====================================================
    # 4. ASIGNACIÓN CHOFER-CABEZAL
    # =====================================================
    if "truck_id" in data:
        requested_truck_id = _parse_int(
            data.get("truck_id"),
            "cabezal",
        )

        active_assignment = (
            get_active_assignment_for_driver(
                driver.id
            )
        )

        current_truck_id = (
            active_assignment.truck_id
            if active_assignment is not None
            else None
        )

        # Solo actúa cuando realmente cambió el cabezal.
        if requested_truck_id != current_truck_id:

            # ---------------------------------------------
            # Retirar cabezal
            # ---------------------------------------------
            if requested_truck_id is None:
                if active_assignment is not None:
                    _end_assignment_without_commit(
                        active_assignment,
                        user_id=user_id,
                        ended_at=datetime.utcnow(),
                        end_reason=(
                            _clean_text(
                                data.get(
                                    "assignment_change_reason"
                                )
                            )
                            or "RETIRO DESDE MATRIZ DE CHOFERES"
                        ),
                    )

                    db.session.flush()

            # ---------------------------------------------
            # Asignar / cambiar cabezal
            # ---------------------------------------------
            else:
                target_truck = get_truck_or_404(
                    requested_truck_id
                )

                if target_truck.status != "ACTIVE":
                    raise TransportValidationError(
                        "El cabezal seleccionado no está activo."
                    )

                target_assignment = (
                    get_active_assignment_for_truck(
                        requested_truck_id
                    )
                )

                if (
                    target_assignment is not None
                    and target_assignment.driver_id
                    != driver.id
                ):
                    raise TransportConflictError(
                        "El cabezal seleccionado ya está asignado "
                        "a otro chofer."
                    )

                if active_assignment is not None:
                    _end_assignment_without_commit(
                        active_assignment,
                        user_id=user_id,
                        ended_at=datetime.utcnow(),
                        end_reason=(
                            _clean_text(
                                data.get(
                                    "assignment_change_reason"
                                )
                            )
                            or "CAMBIO DESDE MATRIZ DE CHOFERES"
                        ),
                    )

                    db.session.flush()

                assign_driver_to_truck(
                    driver_id=driver.id,
                    truck_id=requested_truck_id,
                    user_id=user_id,
                    started_at=data.get(
                        "assignment_started_at"
                    ),
                    notes=data.get(
                        "assignment_notes"
                    ),
                    replace_existing=False,
                    commit=False,
                )

                db.session.flush()

    # =====================================================
    # 5. CABEZAL ACTUAL
    #
    # Se vuelve a consultar la asignación porque pudo haber
    # cambiado en el bloque anterior.
    # =====================================================
    active_assignment = (
        get_active_assignment_for_driver(
            driver.id
        )
    )

    current_truck = None

    if active_assignment is not None:
        current_truck = get_truck_or_404(
            active_assignment.truck_id
        )

    # =====================================================
    # 6. DATOS DEL CABEZAL Y PROPIETARIO
    #
    # Todos los nombres enviados por la matriz llevan
    # prefijo truck_ para no colisionar con datos del chofer.
    # =====================================================
    truck_keys = {
        "truck_registration_date",
        "truck_registered_site_id",
        "truck_plate",
        "truck_status",

        "truck_owner_name",
        "truck_owner_phone",
        "truck_owner_email",
        "truck_owner_notes",

        "truck_dock_permit_number",
        "truck_dock_permit_expiry_date",

        "truck_circulation_card",

        "truck_dekra_month",
        "truck_dekra_year",

        "truck_insurance_name",
        "truck_insurance_expiry_date",

        "truck_is_payroll",

        "truck_rt_name",
        "truck_rt_expiry_date",

        "truck_weights_dimensions",
        "truck_policy_number",
        "truck_bonded_status",

        "truck_notes",
    }

    wants_truck_update = any(
        key in data
        for key in truck_keys
    )

    if wants_truck_update:
        if current_truck is None:
            raise TransportValidationError(
                "No se pueden guardar datos del cabezal porque "
                "el chofer no tiene un cabezal asignado."
            )

        # ---------------------------------------------
        # Datos actuales del propietario
        # ---------------------------------------------
        current_owner = current_truck.owner

        current_owner_name = (
            current_owner.name
            if current_owner is not None
            else None
        )

        current_owner_phone = (
            current_owner.phone
            if current_owner is not None
            else None
        )

        current_owner_email = (
            current_owner.email
            if current_owner is not None
            else None
        )

        current_owner_notes = (
            current_owner.notes
            if current_owner is not None
            else None
        )

        # ---------------------------------------------
        # Adaptamos nombres truck_* a los nombres que
        # update_truck() ya utiliza.
        # ---------------------------------------------
        truck_data = {
            "registration_date": data.get(
                "truck_registration_date",
                current_truck.registration_date,
            ),

            "registered_site_id": data.get(
                "truck_registered_site_id",
                current_truck.registered_site_id,
            ),

            "plate": data.get(
                "truck_plate",
                current_truck.plate,
            ),

            "status": data.get(
                "truck_status",
                current_truck.status,
            ),

            "owner_name": data.get(
                "truck_owner_name",
                current_owner_name,
            ),

            "owner_phone": data.get(
                "truck_owner_phone",
                current_owner_phone,
            ),

            "owner_email": data.get(
                "truck_owner_email",
                current_owner_email,
            ),

            "owner_notes": data.get(
                "truck_owner_notes",
                current_owner_notes,
            ),

            "dock_permit_number": data.get(
                "truck_dock_permit_number",
                current_truck.dock_permit_number,
            ),

            "dock_permit_expiry_date": data.get(
                "truck_dock_permit_expiry_date",
                current_truck.dock_permit_expiry_date,
            ),

            "circulation_card": data.get(
                "truck_circulation_card",
                current_truck.circulation_card,
            ),

            "dekra_month": data.get(
                "truck_dekra_month",
                current_truck.dekra_month,
            ),

            "dekra_year": data.get(
                "truck_dekra_year",
                current_truck.dekra_year,
            ),

            "insurance_name": data.get(
                "truck_insurance_name",
                current_truck.insurance_name,
            ),

            "insurance_expiry_date": data.get(
                "truck_insurance_expiry_date",
                current_truck.insurance_expiry_date,
            ),

            "is_payroll": data.get(
                "truck_is_payroll",
                current_truck.is_payroll,
            ),

            "rt_name": data.get(
                "truck_rt_name",
                current_truck.rt_name,
            ),

            "rt_expiry_date": data.get(
                "truck_rt_expiry_date",
                current_truck.rt_expiry_date,
            ),

            "weights_dimensions": data.get(
                "truck_weights_dimensions",
                current_truck.weights_dimensions,
            ),

            "policy_number": data.get(
                "truck_policy_number",
                current_truck.policy_number,
            ),

            "bonded_status": data.get(
                "truck_bonded_status",
                current_truck.bonded_status,
            ),

            "notes": data.get(
                "truck_notes",
                current_truck.notes,
            ),
        }

        update_truck(
            current_truck,
            truck_data,
            user_id=user_id,
            commit=False,
        )

    # =====================================================
    # 7. GUARDADO ÚNICO
    # =====================================================
    if commit:
        _commit_or_raise(
            "No fue posible guardar la fila del chofer. "
            "Verifique los datos del chofer, documentos, "
            "APM, asignación y cabezal."
        )

    return driver

def bulk_import_transport_excel(
    file_stream,
    *,
    user_id: int,
    registered_site_id: int,
    habitual_site_id: int | None = None,
) -> dict[str, Any]:
    """
    Importa choferes/cabezales desde la plantilla Excel.

    Diseño optimizado para Render + PostgreSQL:

    - Excel en read_only=True.
    - No precarga tablas completas.
    - Procesa en bloques pequeños.
    - No usa SAVEPOINT por fila.
    - No hace flush por cada objeto.
    - Consulta únicamente cédulas, placas y propietarios
      presentes en cada bloque.
    - 1 flush principal por bloque.
    - 1 commit por bloque.
    - Documentos/APM/asignaciones se crean después de
      obtener los IDs del bloque.

    Conceptos:

    registered_site_id
        Dato técnico obligatorio para cabezales nuevos.

    habitual_site_id
        Únicamente distintivo de patiero.
        None = chofer normal.
    """

    # =====================================================
    # CONFIGURACIÓN
    # =====================================================
    BATCH_SIZE = 20

    if not registered_site_id:
        raise TransportValidationError(
            "No fue posible determinar el predio técnico "
            "para registrar los cabezales."
        )

    # =====================================================
    # HELPERS LOCALES
    # =====================================================
    def clean(value) -> str:
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            value = int(value)

        return str(value).strip()

    def clean_upper(value) -> str:
        return clean(value).upper()

    def normalize_identification(value) -> str:
        value = clean(value)

        if not value:
            return ""

        return (
            value
            .replace(" ", "")
            .replace("-", "")
            .replace(".", "")
        )

    def normalize_plate(value) -> str:
        value = clean_upper(value)

        if not value:
            return ""

        return (
            value
            .replace(" ", "")
            .replace("-", "")
            .replace(".", "")
        )

    def excel_date(value):
        if value in (None, ""):
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)):
            try:
                return (
                    datetime(1899, 12, 30)
                    + timedelta(
                        days=float(value)
                    )
                ).date()

            except Exception:
                return None

        raw = clean(value)

        if not raw:
            return None

        for fmt in (
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
        ):
            try:
                return datetime.strptime(
                    raw,
                    fmt,
                ).date()

            except ValueError:
                continue

        return None

    def yes_pending(value) -> str:
        raw = clean_upper(value)

        if raw in {
            "SI",
            "SÍ",
            "YES",
            "OK",
            "VALIDO",
            "VÁLIDO",
            "VALID",
        }:
            return "YES"

        return "PENDING"

    def document_payload(
        *,
        number=None,
        expiry=None,
        no_expiry=False,
    ) -> dict[str, Any]:
        expiry_date = excel_date(
            expiry
        )

        if no_expiry:
            status = "VALID"

        elif expiry_date is None:
            status = "PENDING"

        elif expiry_date < date.today():
            status = "EXPIRED"

        else:
            status = "VALID"

        return {
            "document_number": (
                clean_upper(number)
                or None
            ),
            "issue_date": None,
            "expiry_date": expiry_date,
            "no_expiry": bool(
                no_expiry
            ),
            "status": status,
            "notes": None,
        }

    # =====================================================
    # RESULTADO GLOBAL
    # =====================================================
    result = {
        "processed": 0,
        "created_drivers": 0,
        "updated_drivers": 0,
        "created_trucks": 0,
        "updated_trucks": 0,
        "created_assignments": 0,
        "updated_documents": 0,
        "updated_apm": 0,
        "skipped": 0,
        "errors": [],
    }

    # =====================================================
    # PROCESAR UN BLOQUE
    # =====================================================
    def process_batch(
        batch: list[dict[str, Any]],
    ) -> None:
        """
        Procesa un bloque completo dentro de una transacción.

        Si PostgreSQL rechaza el bloque, se revierte únicamente
        este bloque. Los bloques previamente confirmados permanecen.
        """

        if not batch:
            return

        # =================================================
        # ESTADÍSTICAS LOCALES
        #
        # Solo se suman al resultado global después del COMMIT.
        # =================================================
        batch_stats = {
            "processed": len(batch),
            "created_drivers": 0,
            "updated_drivers": 0,
            "created_trucks": 0,
            "updated_trucks": 0,
            "created_assignments": 0,
            "updated_documents": 0,
            "updated_apm": 0,
        }

        batch_observations: list[
            dict[str, Any]
        ] = []

        try:
            # =================================================
            # 1. VALORES PRESENTES EN EL BLOQUE
            # =================================================
            identifications = {
                item["identification"]
                for item in batch
                if item["identification"]
            }

            plates = {
                item["plate"]
                for item in batch
                if item["plate"]
            }

            owner_names = {
                item["owner_name"]
                for item in batch
                if item["owner_name"]
            }

            # =================================================
            # 2. CHOFERES EXISTENTES
            # =================================================
            drivers_by_identification: dict[
                str,
                Driver,
            ] = {}

            if identifications:
                existing_drivers = list(
                    db.session.scalars(
                        select(Driver)
                        .where(
                            Driver.identification.in_(
                                identifications
                            )
                        )
                    ).all()
                )

                drivers_by_identification = {
                    driver.identification: driver
                    for driver in existing_drivers
                }

            # =================================================
            # 3. CABEZALES EXISTENTES
            # =================================================
            trucks_by_plate: dict[
                str,
                Truck,
            ] = {}

            if plates:
                existing_trucks = list(
                    db.session.scalars(
                        select(Truck)
                        .where(
                            Truck.plate.in_(
                                plates
                            )
                        )
                    )
                    .unique()
                    .all()
                )

                trucks_by_plate = {
                    normalize_plate(
                        truck.plate
                    ): truck
                    for truck in existing_trucks
                }

            # =================================================
            # 4. PROPIETARIOS EXISTENTES
            #
            # El modelo realmente usa UNIQUE(name, phone),
            # por lo que la clave correcta incluye ambos.
            # =================================================
            owners_by_key: dict[
                tuple[str, str],
                TruckOwner,
            ] = {}

            if owner_names:
                existing_owners = list(
                    db.session.scalars(
                        select(TruckOwner)
                        .where(
                            func.upper(
                                TruckOwner.name
                            ).in_(
                                owner_names
                            )
                        )
                    ).all()
                )

                owners_by_key = {
                    (
                        clean_upper(
                            owner.name
                        ),
                        clean(
                            owner.phone
                        ),
                    ): owner
                    for owner in existing_owners
                }

            # =================================================
            # 5. IDS EXISTENTES PARA CONSULTAS DEPENDIENTES
            # =================================================
            existing_driver_ids = {
                driver.id
                for driver
                in drivers_by_identification.values()
            }

            existing_truck_ids = {
                truck.id
                for truck
                in trucks_by_plate.values()
            }

            # =================================================
            # 6. DOCUMENTOS EXISTENTES
            # =================================================
            documents_by_key: dict[
                tuple[int, str],
                DriverDocument,
            ] = {}

            if existing_driver_ids:
                existing_documents = list(
                    db.session.scalars(
                        select(DriverDocument)
                        .where(
                            DriverDocument.driver_id.in_(
                                existing_driver_ids
                            )
                        )
                    ).all()
                )

                documents_by_key = {
                    (
                        document.driver_id,
                        document.document_type,
                    ): document
                    for document
                    in existing_documents
                }

            # =================================================
            # 7. APM EXISTENTE
            # =================================================
            apm_by_driver: dict[
                int,
                DriverApmRecord,
            ] = {}

            if existing_driver_ids:
                existing_apm = list(
                    db.session.scalars(
                        select(DriverApmRecord)
                        .where(
                            DriverApmRecord.driver_id.in_(
                                existing_driver_ids
                            )
                        )
                    ).all()
                )

                apm_by_driver = {
                    record.driver_id: record
                    for record in existing_apm
                }

            # =================================================
            # 8. ASIGNACIONES ACTIVAS EXISTENTES
            # =================================================
            assignment_by_driver: dict[
                int,
                DriverTruckAssignment,
            ] = {}

            assignment_by_truck: dict[
                int,
                DriverTruckAssignment,
            ] = {}

            assignment_conditions = []

            if existing_driver_ids:
                assignment_conditions.append(
                    DriverTruckAssignment.driver_id.in_(
                        existing_driver_ids
                    )
                )

            if existing_truck_ids:
                assignment_conditions.append(
                    DriverTruckAssignment.truck_id.in_(
                        existing_truck_ids
                    )
                )

            if assignment_conditions:
                active_assignments = list(
                    db.session.scalars(
                        select(
                            DriverTruckAssignment
                        )
                        .where(
                            DriverTruckAssignment.status
                            == "ACTIVE",
                            or_(
                                *assignment_conditions
                            ),
                        )
                    ).all()
                )

                assignment_by_driver = {
                    assignment.driver_id: assignment
                    for assignment
                    in active_assignments
                }

                assignment_by_truck = {
                    assignment.truck_id: assignment
                    for assignment
                    in active_assignments
                }

            # =================================================
            # 9. PRIMERA FASE
            #
            # Driver + Truck + Owner.
            #
            # No hacemos flush dentro del loop.
            # =================================================
            row_objects: list[
                dict[str, Any]
            ] = []

            now = datetime.utcnow()

            for item in batch:
                # =============================================
                # DRIVER
                # =============================================
                driver = (
                    drivers_by_identification.get(
                        item[
                            "identification"
                        ]
                    )
                )

                new_driver = False

                if driver is None:
                    driver = Driver(
                        name=(
                            item["name"]
                        ),

                        residence=(
                            item["residence"]
                            or None
                        ),

                        identification=(
                            item[
                                "identification"
                            ]
                        ),

                        phone_1=(
                            item["phone"]
                            or None
                        ),

                        habitual_site_id=(
                            habitual_site_id
                        ),

                        status="ACTIVE",

                        created_by_user_id=(
                            user_id
                        ),

                        updated_by_user_id=(
                            user_id
                        ),
                    )

                    db.session.add(
                        driver
                    )

                    # Cache local para duplicados
                    # dentro del mismo Excel.
                    drivers_by_identification[
                        item["identification"]
                    ] = driver

                    new_driver = True

                    batch_stats[
                        "created_drivers"
                    ] += 1

                else:
                    driver.name = (
                        item["name"]
                    )

                    if item["residence"]:
                        driver.residence = (
                            item["residence"]
                        )

                    if item["phone"]:
                        driver.phone_1 = (
                            item["phone"]
                        )

                    # None significa:
                    # no cambiar condición existente.
                    if habitual_site_id:
                        driver.habitual_site_id = (
                            habitual_site_id
                        )

                    driver.updated_by_user_id = (
                        user_id
                    )

                    driver.updated_at = (
                        now
                    )

                    batch_stats[
                        "updated_drivers"
                    ] += 1

                # =============================================
                # TRUCK
                # =============================================
                truck = None
                new_truck = False

                if item["plate"]:
                    truck = (
                        trucks_by_plate.get(
                            item["plate"]
                        )
                    )

                    if truck is None:
                        truck = Truck(
                            plate=(
                                item["plate"]
                            ),

                            registration_date=(
                                item[
                                    "registration_date"
                                ]
                                or date.today()
                            ),

                            # Dato técnico.
                            registered_site_id=(
                                registered_site_id
                            ),

                            status="ACTIVE",

                            bonded_status=(
                                "PENDING"
                            ),

                            created_by_user_id=(
                                user_id
                            ),

                            updated_by_user_id=(
                                user_id
                            ),
                        )

                        db.session.add(
                            truck
                        )

                        trucks_by_plate[
                            item["plate"]
                        ] = truck

                        new_truck = True

                        batch_stats[
                            "created_trucks"
                        ] += 1

                    else:
                        if item[
                            "registration_date"
                        ]:
                            truck.registration_date = (
                                item[
                                    "registration_date"
                                ]
                            )

                        truck.updated_by_user_id = (
                            user_id
                        )

                        truck.updated_at = (
                            now
                        )

                        batch_stats[
                            "updated_trucks"
                        ] += 1

                    # =========================================
                    # OWNER
                    # =========================================
                    owner = None

                    if item["owner_name"]:
                        owner_key = (
                            item["owner_name"],
                            item["owner_phone"],
                        )

                        owner = (
                            owners_by_key.get(
                                owner_key
                            )
                        )

                        if owner is None:
                            owner = TruckOwner(
                                name=(
                                    item[
                                        "owner_name"
                                    ]
                                ),

                                phone=(
                                    item[
                                        "owner_phone"
                                    ]
                                    or None
                                ),

                                is_active=True,
                            )

                            db.session.add(
                                owner
                            )

                            owners_by_key[
                                owner_key
                            ] = owner

                        else:
                            if (
                                item["owner_phone"]
                                and owner.phone
                                != item[
                                    "owner_phone"
                                ]
                            ):
                                owner.phone = (
                                    item[
                                        "owner_phone"
                                    ]
                                )

                            if not owner.is_active:
                                owner.is_active = (
                                    True
                                )

                            owner.updated_at = (
                                now
                            )

                        # Usar relación ORM.
                        # SQLAlchemy resuelve el owner_id
                        # en el mismo flush.
                        truck.owner = owner

                    # =========================================
                    # DATOS DEL CABEZAL
                    # =========================================

                    if item[
                        "truck_dock_expiry"
                    ]:
                        truck.dock_permit_expiry_date = (
                            item[
                                "truck_dock_expiry"
                            ]
                        )

                    if item[
                        "circulation_card"
                    ]:
                        truck.circulation_card = (
                            item[
                                "circulation_card"
                            ]
                        )

                    if item["rtv_date"]:
                        truck.dekra_month = (
                            item[
                                "rtv_date"
                            ].month
                        )

                        truck.dekra_year = (
                            item[
                                "rtv_date"
                            ].year
                        )

                    if item[
                        "insurance_name"
                    ]:
                        truck.insurance_name = (
                            item[
                                "insurance_name"
                            ]
                        )

                    if item[
                        "rt_expiry"
                    ]:
                        truck.rt_expiry_date = (
                            item[
                                "rt_expiry"
                            ]
                        )

                    if item["rt_name"]:
                        truck.rt_name = (
                            item["rt_name"]
                        )

                    if item[
                        "weights_dimensions"
                    ] is not None:
                        truck.weights_dimensions = (
                            item[
                                "weights_dimensions"
                            ]
                        )

                    if item[
                        "policy_number"
                    ]:
                        truck.policy_number = (
                            item[
                                "policy_number"
                            ]
                        )

                    if item[
                        "bonded_status"
                    ]:
                        truck.bonded_status = (
                            item[
                                "bonded_status"
                            ]
                        )

                row_objects.append({
                    "row": item["row"],
                    "raw": item,
                    "driver": driver,
                    "truck": truck,
                    "new_driver": (
                        new_driver
                    ),
                    "new_truck": (
                        new_truck
                    ),
                })

            # =================================================
            # 10. ÚNICO FLUSH EXPLÍCITO DEL BLOQUE
            #
            # Obtiene:
            # - driver.id
            # - truck.id
            # - owner.id
            #
            # y envía juntos los cambios principales.
            # =================================================
            db.session.flush()

            # =================================================
            # 11. SEGUNDA FASE
            #
            # Documentos + APM + asignaciones.
            #
            # Ahora todos los IDs están disponibles.
            # =================================================
            for row_data in row_objects:
                item = (
                    row_data["raw"]
                )

                driver = (
                    row_data["driver"]
                )

                truck = (
                    row_data["truck"]
                )

                # =============================================
                # DOCUMENTOS
                # =============================================
                document_values = {
                    "DOCK_PERMIT": (
                        document_payload(
                            expiry=(
                                item[
                                    "driver_dock_expiry"
                                ]
                            )
                        )
                    ),

                    "GENERAL_CARD": (
                        document_payload(
                            number=(
                                item[
                                    "general_card_number"
                                ]
                            ),

                            expiry=(
                                item[
                                    "general_card_expiry"
                                ]
                            ),
                        )
                    ),

                    "CHEMICAL_PERMIT": (
                        document_payload(
                            expiry=(
                                item[
                                    "chemical_expiry"
                                ]
                            )
                        )
                    ),

                    "LICENSE": (
                        document_payload(
                            expiry=(
                                item[
                                    "license_expiry"
                                ]
                            )
                        )
                    ),

                    "CRIMINAL_RECORD": (
                        document_payload(
                            expiry=(
                                item[
                                    "criminal_expiry"
                                ]
                            )
                        )
                    ),
                }

                for (
                    document_type,
                    payload,
                ) in document_values.items():

                    key = (
                        driver.id,
                        document_type,
                    )

                    document = (
                        documents_by_key.get(
                            key
                        )
                    )

                    # Igual que el importador anterior:
                    # no crear documento nuevo completamente vacío.
                    if (
                        document is None
                        and not payload[
                            "document_number"
                        ]
                        and not payload[
                            "expiry_date"
                        ]
                    ):
                        continue

                    if document is None:
                        document = (
                            DriverDocument(
                                driver_id=(
                                    driver.id
                                ),

                                document_type=(
                                    document_type
                                ),

                                # DriverDocument NO tiene
                                # created_by_user_id.
                                updated_by_user_id=(
                                    user_id
                                ),
                            )
                        )

                        db.session.add(
                            document
                        )

                        documents_by_key[
                            key
                        ] = document

                    document.document_number = (
                        payload[
                            "document_number"
                        ]
                    )

                    document.issue_date = (
                        payload[
                            "issue_date"
                        ]
                    )

                    document.expiry_date = (
                        payload[
                            "expiry_date"
                        ]
                    )

                    document.no_expiry = (
                        payload[
                            "no_expiry"
                        ]
                    )

                    document.status = (
                        payload[
                            "status"
                        ]
                    )

                    document.notes = (
                        payload[
                            "notes"
                        ]
                    )

                    document.updated_by_user_id = (
                        user_id
                    )

                    document.updated_at = (
                        now
                    )

                    if (
                        payload[
                            "document_number"
                        ]
                        or payload[
                            "expiry_date"
                        ]
                    ):
                        batch_stats[
                            "updated_documents"
                        ] += 1

                # =============================================
                # APM
                # =============================================
                if (
                    item["apm_training"]
                    or item["apm_card"]
                    or item["apm_expiry_raw"]
                ):
                    record = (
                        apm_by_driver.get(
                            driver.id
                        )
                    )

                    if record is None:
                        record = (
                            DriverApmRecord(
                                driver_id=(
                                    driver.id
                                ),

                                # DriverApmRecord NO tiene
                                # created_by_user_id.
                                updated_by_user_id=(
                                    user_id
                                ),
                            )
                        )

                        db.session.add(
                            record
                        )

                        apm_by_driver[
                            driver.id
                        ] = record

                    record.training_status = (
                        yes_pending(
                            item[
                                "apm_training"
                            ]
                        )
                    )

                    card_upper = (
                        item[
                            "apm_card"
                        ].upper()
                    )

                    if card_upper in {
                        "",
                        "PENDIENTE",
                    }:
                        record.card_status = (
                            "PENDING"
                        )

                        record.card_number = (
                            None
                        )

                    elif "VENC" in card_upper:
                        record.card_status = (
                            "EXPIRED"
                        )

                        record.card_number = (
                            None
                        )

                    else:
                        record.card_status = (
                            "YES"
                        )

                        record.card_number = (
                            item["apm_card"]
                        )

                    apm_expiry_text = (
                        clean_upper(
                            item[
                                "apm_expiry_raw"
                            ]
                        )
                    )

                    if (
                        "SIN VENCIMIENTO"
                        in apm_expiry_text
                    ):
                        record.expiry_mode = (
                            "NO_EXPIRY"
                        )

                        record.expiry_date = (
                            None
                        )

                    elif (
                        "PENDIENTE"
                        in apm_expiry_text
                    ):
                        record.expiry_mode = (
                            "PENDING"
                        )

                        record.expiry_date = (
                            None
                        )

                    elif (
                        "VENCIDO"
                        in apm_expiry_text
                    ):
                        record.expiry_mode = (
                            "EXPIRED"
                        )

                        record.card_status = (
                            "EXPIRED"
                        )

                        record.expiry_date = (
                            None
                        )

                    elif item[
                        "apm_expiry_date"
                    ]:
                        record.expiry_date = (
                            item[
                                "apm_expiry_date"
                            ]
                        )

                        if (
                            item[
                                "apm_expiry_date"
                            ]
                            < date.today()
                        ):
                            record.expiry_mode = (
                                "EXPIRED"
                            )

                            record.card_status = (
                                "EXPIRED"
                            )

                        else:
                            record.expiry_mode = (
                                "DATE"
                            )

                    else:
                        record.expiry_mode = (
                            "PENDING"
                        )

                        record.expiry_date = (
                            None
                        )

                    record.updated_by_user_id = (
                        user_id
                    )

                    record.updated_at = (
                        now
                    )

                    batch_stats[
                        "updated_apm"
                    ] += 1

                # =============================================
                # ASIGNACIÓN CHOFER - CABEZAL
                # =============================================
                if truck is not None:
                    current_driver_assignment = (
                        assignment_by_driver.get(
                            driver.id
                        )
                    )

                    current_truck_assignment = (
                        assignment_by_truck.get(
                            truck.id
                        )
                    )

                    if (
                        current_driver_assignment
                        is None
                        and current_truck_assignment
                        is None
                    ):
                        assignment = (
                            DriverTruckAssignment(
                                driver_id=(
                                    driver.id
                                ),

                                truck_id=(
                                    truck.id
                                ),

                                status="ACTIVE",

                                started_at=(
                                    datetime.combine(
                                        item[
                                            "registration_date"
                                        ]
                                        or date.today(),

                                        datetime.min.time(),
                                    )
                                ),

                                notes=(
                                    "Carga masiva inicial"
                                ),

                                created_by_user_id=(
                                    user_id
                                ),
                            )
                        )

                        db.session.add(
                            assignment
                        )

                        # Actualizar cache local inmediatamente
                        # para detectar duplicados dentro del bloque.
                        assignment_by_driver[
                            driver.id
                        ] = assignment

                        assignment_by_truck[
                            truck.id
                        ] = assignment

                        batch_stats[
                            "created_assignments"
                        ] += 1

                    elif (
                        current_driver_assignment
                        is not None
                        and
                        current_driver_assignment.truck_id
                        == truck.id
                    ):
                        # Ya está correctamente ligado.
                        pass

                    else:
                        batch_observations.append({
                            "row": (
                                item["row"]
                            ),
                            "message": (
                                f"No se cambió la asignación "
                                f"de {item['name']}: "
                                f"el chofer o la placa "
                                f"{item['plate']} ya tienen "
                                "una asignación activa distinta."
                            ),
                        })

            # =================================================
            # 12. COMMIT DEL BLOQUE
            #
            # SQLAlchemy realiza aquí el flush final de:
            # - documentos
            # - APM
            # - asignaciones
            #
            # No necesitamos llamar db.session.flush()
            # nuevamente.
            # =================================================
            db.session.commit()

            # =================================================
            # 13. SOLO AHORA SUMAR CONTADORES
            # =================================================
            for key, value in (
                batch_stats.items()
            ):
                result[key] += value

            result["errors"].extend(
                batch_observations
            )

        except Exception as exc:
            # =================================================
            # Falla solamente este bloque.
            #
            # Los bloques anteriores ya hicieron COMMIT.
            # =================================================
            db.session.rollback()

            result["skipped"] += (
                len(batch)
            )

            # No contamos como procesadas las filas
            # de un bloque que fue revertido.
            for item in batch:
                result["errors"].append({
                    "row": (
                        item["row"]
                    ),
                    "message": (
                        "El bloque de importación fue revertido. "
                        f"Motivo: {exc}"
                    ),
                })

    # =====================================================
    # ABRIR EXCEL
    # =====================================================
    try:
        workbook = load_workbook(
            file_stream,
            read_only=True,
            data_only=True,
        )

    except Exception as exc:
        raise TransportValidationError(
            "No fue posible leer el archivo Excel."
        ) from exc

    # =====================================================
    # HOJA
    # =====================================================
    allowed_sheet_names = (
        "VIGENCIA PERMISOS CALDERA",
        "TRANSPORTES",
    )

    worksheet = None

    for sheet_name in (
        allowed_sheet_names
    ):
        if (
            sheet_name
            in workbook.sheetnames
        ):
            worksheet = (
                workbook[
                    sheet_name
                ]
            )

            break

    if worksheet is None:
        workbook.close()

        raise TransportValidationError(
            "No se encontró una hoja válida. "
            "La hoja debe llamarse "
            "'VIGENCIA PERMISOS CALDERA' "
            "o 'TRANSPORTES'."
        )

    # =====================================================
    # LEER SECUENCIALMENTE
    #
    # No guardamos todo el Excel en memoria.
    # Solo mantenemos máximo BATCH_SIZE registros.
    # =====================================================
    batch: list[
        dict[str, Any]
    ] = []

    try:
        for excel_row, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            values = tuple(
                values
            )

            # Necesitamos como mínimo hasta AE (índice 30).
            if len(values) < 31:
                values = (
                    values
                    + (
                        None,
                    ) * (
                        31
                        - len(values)
                    )
                )

            # =================================================
            # DATOS PRINCIPALES
            # =================================================
            name = clean_upper(
                values[2]
            )

            identification = (
                normalize_identification(
                    values[4]
                )
            )

            plate = normalize_plate(
                values[5]
            )

            # Fila completamente vacía.
            if (
                not name
                and not identification
                and not plate
            ):
                continue

            # =================================================
            # VALIDACIONES QUE NO NECESITAN BD
            # =================================================
            if (
                not name
                or not identification
            ):
                result["skipped"] += 1

                result["errors"].append({
                    "row": excel_row,
                    "message": (
                        "Fila omitida: falta nombre "
                        "o cédula del chofer."
                    ),
                })

                continue

            # =================================================
            # PESOS Y DIMENSIONES
            # =================================================
            weights_dimensions = None

            raw_weights = (
                values[28]
            )

            if raw_weights not in (
                None,
                "",
            ):
                try:
                    weights_dimensions = (
                        Decimal(
                            str(
                                raw_weights
                            )
                        )
                    )

                except (
                    InvalidOperation,
                    TypeError,
                    ValueError,
                ):
                    result["errors"].append({
                        "row": excel_row,
                        "message": (
                            "Pesos y dimensiones contiene "
                            "un valor no numérico. "
                            "Ese campo se ignorará."
                        ),
                    })

                    weights_dimensions = (
                        None
                    )

            # =================================================
            # CAUCIONADO
            # =================================================
            bonded_raw = clean_upper(
                values[30]
            )

            bonded_status = None

            if bonded_raw:
                bonded_status = (
                    "BONDED"
                    if "CAU" in bonded_raw
                    else "PENDING"
                )

            # =================================================
            # REGISTRO NORMALIZADO
            # =================================================
            item = {
                "row": excel_row,

                # ---------------------------------------------
                # Driver
                # ---------------------------------------------
                "name": name,

                "residence": (
                    clean_upper(
                        values[3]
                    )
                ),

                "identification": (
                    identification
                ),

                "phone": clean(
                    values[6]
                ),

                # ---------------------------------------------
                # Truck
                # ---------------------------------------------
                "plate": plate,

                "registration_date": (
                    excel_date(
                        values[1]
                    )
                ),

                # ---------------------------------------------
                # Documentos driver
                # ---------------------------------------------
                "driver_dock_expiry": (
                    excel_date(
                        values[10]
                    )
                ),

                "general_card_number": (
                    clean_upper(
                        values[12]
                    )
                ),

                "general_card_expiry": (
                    excel_date(
                        values[13]
                    )
                ),

                "chemical_expiry": (
                    excel_date(
                        values[14]
                    )
                ),

                "license_expiry": (
                    excel_date(
                        values[18]
                    )
                ),

                "criminal_expiry": (
                    excel_date(
                        values[19]
                    )
                ),

                # ---------------------------------------------
                # APM
                # ---------------------------------------------
                "apm_training": (
                    clean_upper(
                        values[15]
                    )
                ),

                "apm_card": (
                    clean(
                        values[16]
                    )
                ),

                "apm_expiry_raw": (
                    values[17]
                ),

                "apm_expiry_date": (
                    excel_date(
                        values[17]
                    )
                ),

                # ---------------------------------------------
                # Owner
                # ---------------------------------------------
                "owner_name": (
                    clean_upper(
                        values[20]
                    )
                ),

                "owner_phone": (
                    clean(
                        values[21]
                    )
                ),

                # ---------------------------------------------
                # Truck documentos/datos
                # ---------------------------------------------
                "truck_dock_expiry": (
                    excel_date(
                        values[11]
                    )
                ),

                "circulation_card": (
                    clean_upper(
                        values[22]
                    )
                ),

                "rtv_date": (
                    excel_date(
                        values[23]
                    )
                ),

                "insurance_name": (
                    clean_upper(
                        values[25]
                    )
                ),

                "rt_expiry": (
                    excel_date(
                        values[26]
                    )
                ),

                "rt_name": (
                    clean_upper(
                        values[27]
                    )
                ),

                "weights_dimensions": (
                    weights_dimensions
                ),

                "policy_number": (
                    clean_upper(
                        values[29]
                    )
                ),

                "bonded_status": (
                    bonded_status
                ),
            }

            batch.append(
                item
            )

            # =================================================
            # BLOQUE LLENO
            # =================================================
            if (
                len(batch)
                >= BATCH_SIZE
            ):
                process_batch(
                    batch
                )

                batch = []

        # =====================================================
        # ÚLTIMO BLOQUE
        # =====================================================
        if batch:
            process_batch(
                batch
            )

    except Exception as exc:
        db.session.rollback()

        raise TransportServiceError(
            "No fue posible completar la carga masiva."
        ) from exc

    finally:
        workbook.close()

    # =====================================================
    # RESULTADO FINAL
    # =====================================================
    result["error_count"] = (
        len(
            result["errors"]
        )
    )

    # Evitar mandar una cantidad enorme de texto
    # al navegador.
    result["errors"] = (
        result["errors"][:100]
    )

    return result

def build_transport_bulk_template() -> BytesIO:
    """
    Genera la plantilla oficial para la carga masiva
    de choferes y cabezales.

    IMPORTANTE:
    El importador actual procesa las columnas por posición,
    por lo que NO se deben eliminar ni reordenar columnas.

    La hoja principal conserva el nombre:
        VIGENCIA PERMISOS CALDERA

    Esto se mantiene por compatibilidad con
    bulk_import_transport_excel().
    """

    # =====================================================
    # 1. CREAR LIBRO
    # =====================================================
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "VIGENCIA PERMISOS CALDERA"

    # =====================================================
    # 2. ENCABEZADOS
    #
    # Deben mantenerse EXACTAMENTE en este orden.
    #
    # A  = 0
    # B  = 1
    # ...
    # AF = 31
    #
    # bulk_import_transport_excel() trabaja por índice.
    # =====================================================
    headers = [
        "ORDINAL",                    # A  - ignorado
        "FECHA INGRESO",              # B
        "CHOFERES",                   # C
        "RESIDENCIA",                 # D
        "CEDULA",                     # E
        "PLACA",                      # F
        "TELEFONO",                   # G

        "VACUNA COVID 1 DOSIS",       # H  - ignorado
        "VACUNA COVID 2 DOSIS",       # I  - ignorado
        "CARTA INS",                  # J  - ignorado

        "PERMISO MUELLE CHOFER",      # K
        "PERMISO MUELLE CAMION",      # L

        "CARNET",                     # M
        "VENCIMIENTO CARNET",         # N

        "PERMISO QUIMICO",            # O

        "CAPACITACION APM",           # P
        "CARNET APM",                 # Q
        "VENCIMIENTO CARNET APM",     # R

        "LICENCIA",                   # S
        "HD",                         # T

        "PROPIETARIO",                # U
        "TELEFONO PROPIETARIO",       # V

        "TC",                         # W
        "RTV",                        # X

        "RESERVADO",                  # Y - no procesado actualmente

        "NOMBRE AUT",                 # Z
        "RT ALAMO",                   # AA
        "NOMBRE RT",                  # AB

        "PESOS Y DIMENSIONES",        # AC
        "NUMERO DE POLIZA",           # AD
        "CAUCIONADO",                 # AE

        "PENDIENTES",                 # AF - ignorado
    ]

    worksheet.append(headers)

    # =====================================================
    # 3. ESTILO DE ENCABEZADOS
    # =====================================================
    header_fill = PatternFill(
        "solid",
        fgColor="0F3B63",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    worksheet.row_dimensions[1].height = 42

    # =====================================================
    # 4. ANCHOS DE COLUMNAS
    # =====================================================
    widths = {
        "A": 12,
        "B": 16,
        "C": 32,
        "D": 24,
        "E": 18,
        "F": 16,
        "G": 18,

        "H": 18,
        "I": 18,
        "J": 18,

        "K": 22,
        "L": 22,

        "M": 18,
        "N": 20,

        "O": 20,

        "P": 22,
        "Q": 18,
        "R": 22,

        "S": 18,
        "T": 18,

        "U": 28,
        "V": 22,

        "W": 20,
        "X": 16,
        "Y": 16,

        "Z": 22,
        "AA": 18,
        "AB": 22,

        "AC": 24,
        "AD": 22,
        "AE": 18,

        "AF": 28,
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width

    # =====================================================
    # 5. CONGELAR ENCABEZADO
    # =====================================================
    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        f"A1:AF1"
    )

    # =====================================================
    # 6. FORMATOS DE FECHA
    #
    # Columnas que bulk_import_transport_excel()
    # interpreta como fecha.
    # =====================================================
    date_columns = [
        "B",   # fecha ingreso
        "K",   # permiso muelle chofer
        "L",   # permiso muelle camión
        "N",   # vence carnet
        "O",   # permiso químico
        "R",   # vence APM
        "S",   # licencia
        "T",   # HD
        "X",   # RTV
        "AA",  # RT Álamo
    ]

    for column_letter in date_columns:
        for row_number in range(
            2,
            5001,
        ):
            worksheet[
                f"{column_letter}{row_number}"
            ].number_format = "yyyy-mm-dd"

    # =====================================================
    # 7. VALIDACIONES
    # =====================================================

    # Capacitación APM
    apm_training_validation = DataValidation(
        type="list",
        formula1='"SI,PENDIENTE"',
        allow_blank=True,
    )

    # Caucionado
    bonded_validation = DataValidation(
        type="list",
        formula1='"CAUCIONADO,PENDIENTE"',
        allow_blank=True,
    )

    worksheet.add_data_validation(
        apm_training_validation
    )

    worksheet.add_data_validation(
        bonded_validation
    )

    apm_training_validation.add(
        "P2:P5000"
    )

    bonded_validation.add(
        "AE2:AE5000"
    )

    # =====================================================
    # 8. COMENTARIO VISUAL PARA COLUMNAS NO PROCESADAS
    # =====================================================
    ignored_fill = PatternFill(
        "solid",
        fgColor="E2E8F0",
    )

    ignored_font = Font(
        color="64748B",
        bold=True,
    )

    ignored_columns = {
        "A",
        "H",
        "I",
        "J",
        "Y",
        "AF",
    }

    for column_letter in ignored_columns:
        cell = worksheet[
            f"{column_letter}1"
        ]

        cell.fill = ignored_fill
        cell.font = ignored_font

    # =====================================================
    # 9. HOJA DE INSTRUCCIONES
    # =====================================================
    instructions_sheet = workbook.create_sheet(
        "INSTRUCCIONES"
    )

    instructions = [
        [
            "PLANTILLA DE CARGA MASIVA - CHOFERES Y CABEZALES",
            "",
        ],
        [
            "",
            "",
        ],
        [
            "REGLA IMPORTANTE",
            (
                "No elimine, agregue ni cambie de posición las columnas. "
                "El importador actual procesa los datos por posición."
            ),
        ],
        [
            "Hoja que se importa",
            "VIGENCIA PERMISOS CALDERA",
        ],
        [
            "Fila de encabezados",
            "Fila 1. Los datos deben comenzar en la fila 2.",
        ],
        [
            "",
            "",
        ],
        [
            "CAMPOS OBLIGATORIOS",
            "CHOFERES y CEDULA.",
        ],
        [
            "PLACA",
            (
                "Opcional. Si se indica, el sistema crea o localiza "
                "el cabezal y trata de ligarlo al chofer."
            ),
        ],
        [
            "FECHAS",
            (
                "Use preferiblemente formato YYYY-MM-DD. "
                "Ejemplo: 2026-08-17."
            ),
        ],
        [
            "CEDULA",
            (
                "Puede contener guiones o espacios. "
                "El sistema normaliza el valor al importar."
            ),
        ],
        [
            "PLACA",
            (
                "Puede contener guiones o espacios. "
                "El sistema normaliza y convierte a mayúsculas."
            ),
        ],
        [
            "",
            "",
        ],
        [
            "CAPACITACION APM",
            "Valores reconocidos como aprobado: SI, SÍ, YES, OK, VALIDO, VALID.",
        ],
        [
            "CARNET APM",
            (
                "PENDIENTE = pendiente. "
                "Un valor que contenga VENC se considera vencido. "
                "Cualquier otro valor se toma como número de carnet."
            ),
        ],
        [
            "VENCIMIENTO CARNET APM",
            (
                "Puede ser una fecha, PENDIENTE, VENCIDO "
                "o SIN VENCIMIENTO."
            ),
        ],
        [
            "CAUCIONADO",
            (
                "Un valor que contenga CAU se registra como CAUCIONADO. "
                "Los demás valores quedan PENDIENTE."
            ),
        ],
        [
            "",
            "",
        ],
        [
            "COLUMNAS DE COMPATIBILIDAD",
            (
                "ORDINAL, VACUNA COVID 1 DOSIS, VACUNA COVID 2 DOSIS, "
                "CARTA INS, RESERVADO y PENDIENTES no se guardan actualmente."
            ),
        ],
        [
            "IMPORTANTE",
            (
                "Estas columnas deben permanecer en la plantilla porque "
                "mantienen las posiciones que espera el importador."
            ),
        ],
        [
            "",
            "",
        ],
        [
            "PENDIENTES",
            (
                "No se importa. La aplicación calcula los pendientes "
                "a partir de documentos y vencimientos."
            ),
        ],
        [
            "PATIERO",
            (
                "La condición de patiero NO se define dentro del Excel. "
                "Se selecciona desde la pantalla de carga masiva."
            ),
        ],
    ]

    for instruction_row in instructions:
        instructions_sheet.append(
            instruction_row
        )

    instructions_sheet.column_dimensions[
        "A"
    ].width = 34

    instructions_sheet.column_dimensions[
        "B"
    ].width = 95

    instructions_sheet[
        "A1"
    ].font = Font(
        bold=True,
        size=14,
    )

    for row in instructions_sheet.iter_rows(
        min_row=1,
        max_col=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # =====================================================
    # 10. GUARDAR EN MEMORIA
    # =====================================================
    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output

def get_active_assignment_for_truck(
    truck_id: int,
) -> DriverTruckAssignment | None:
    return db.session.scalar(
        select(DriverTruckAssignment)
        .where(
            DriverTruckAssignment.truck_id == truck_id,
            DriverTruckAssignment.status == "ACTIVE",
        )
        .limit(1)
    )


def assign_driver_to_truck(
    *,
    driver_id: int,
    truck_id: int,
    user_id: int,
    started_at: Any = None,
    notes: Any = None,
    replace_existing: bool = False,
    replacement_reason: Any = None,
    commit: bool = True,
) -> DriverTruckAssignment:
    if not driver_id:
        raise TransportValidationError(
            "Debe seleccionar un chofer."
        )

    if not truck_id:
        raise TransportValidationError(
            "Debe seleccionar un cabezal."
        )
    driver = get_driver_or_404(driver_id)
    truck = get_truck_or_404(truck_id)

    if driver.status != "ACTIVE":
        raise TransportValidationError(
            "Solo se pueden asignar choferes activos."
        )

    if truck.status != "ACTIVE":
        raise TransportValidationError(
            "Solo se pueden asignar cabezales activos."
        )

    active_driver_assignment = (
        get_active_assignment_for_driver(driver.id)
    )

    active_truck_assignment = (
        get_active_assignment_for_truck(truck.id)
    )

    if (
        active_driver_assignment is not None
        and active_driver_assignment.truck_id == truck.id
    ):
        raise TransportConflictError(
            "El chofer ya se encuentra asignado a ese cabezal."
        )

    if not replace_existing:
        if active_driver_assignment is not None:
            raise TransportConflictError(
                "El chofer ya tiene un cabezal asignado."
            )

        if active_truck_assignment is not None:
            raise TransportConflictError(
                "El cabezal ya tiene un chofer asignado."
            )

    reason = _clean_text(
        replacement_reason,
        max_length=240,
    )

    if replace_existing:
        if active_driver_assignment is not None:
            _end_assignment_without_commit(
                active_driver_assignment,
                user_id=user_id,
                ended_at=datetime.utcnow(),
                end_reason=(
                    reason
                    or "REEMPLAZO DE ASIGNACIÓN"
                ),
            )

        if (
            active_truck_assignment is not None
            and (
                active_driver_assignment is None
                or active_truck_assignment.id
                != active_driver_assignment.id
            )
        ):
            _end_assignment_without_commit(
                active_truck_assignment,
                user_id=user_id,
                ended_at=datetime.utcnow(),
                end_reason=(
                    reason
                    or "REEMPLAZO DE ASIGNACIÓN"
                ),
            )

    assignment = DriverTruckAssignment(
        driver_id=driver.id,
        truck_id=truck.id,
        status="ACTIVE",
        started_at=(
            _parse_datetime(
                started_at,
                "fecha de inicio",
            )
            or datetime.utcnow()
        ),
        notes=_clean_text(notes),
        created_by_user_id=user_id,
    )

    db.session.add(assignment)
    db.session.flush()

    if commit:
        _commit_or_raise(
            "No fue posible realizar la asignación. "
            "El chofer o el cabezal podrían tener otra "
            "asignación activa."
        )

    return assignment


def _end_assignment_without_commit(
    assignment: DriverTruckAssignment,
    *,
    user_id: int,
    ended_at: datetime,
    end_reason: str,
    notes: str | None = None,
) -> DriverTruckAssignment:
    if assignment.status != "ACTIVE":
        raise TransportValidationError(
            "La asignación ya se encuentra finalizada."
        )

    assignment.status = "ENDED"
    assignment.ended_at = ended_at
    assignment.end_reason = _required_text(
        end_reason,
        "motivo de finalización",
        max_length=240,
    )
    assignment.ended_by_user_id = user_id

    if notes is not None:
        assignment.notes = _clean_text(notes)

    return assignment


def end_assignment(
    assignment: DriverTruckAssignment,
    *,
    user_id: int,
    ended_at: Any = None,
    end_reason: Any,
    notes: Any = None,
    commit: bool = True,
) -> DriverTruckAssignment:
    result = _end_assignment_without_commit(
        assignment,
        user_id=user_id,
        ended_at=(
            _parse_datetime(
                ended_at,
                "fecha de finalización",
            )
            or datetime.utcnow()
        ),
        end_reason=_required_text(
            end_reason,
            "motivo de finalización",
            max_length=240,
        ),
        notes=_clean_text(notes),
    )

    if commit:
        _commit_or_raise(
            "No fue posible finalizar la asignación."
        )

    return result


# =========================================================
# PERMISOS DE SALIDA
# =========================================================
def create_exit_permission(
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> DriverExitPermission:
    driver_id = _parse_int(
        data.get("driver_id"),
        "chofer",
        required=True,
    )

    truck_id = _parse_int(
        data.get("truck_id"),
        "cabezal",
    )

    driver = get_driver_or_404(driver_id)

    truck = (
        get_truck_or_404(truck_id)
        if truck_id is not None
        else None
    )

    if driver.status != "ACTIVE":
        raise TransportValidationError(
            "El chofer debe estar activo para autorizar una salida."
        )

    if truck is not None:
        active_assignment = get_active_assignment_for_driver(
            driver.id
        )

        if (
            active_assignment is None
            or active_assignment.truck_id != truck.id
        ):
            raise TransportValidationError(
                "El cabezal seleccionado no está asignado "
                "actualmente a ese chofer."
            )

    open_permission = db.session.scalar(
        select(
            exists().where(
                DriverExitPermission.driver_id == driver.id,
                DriverExitPermission.status == "AUTHORIZED",
            )
        )
    )

    if open_permission:
        raise TransportConflictError(
            "El chofer ya tiene un permiso de salida activo."
        )

    departure_at = (
        _parse_datetime(
            data.get("departure_at"),
            "fecha y hora de salida",
        )
        or datetime.utcnow()
    )

    expected_return_at = _parse_datetime(
        data.get("expected_return_at"),
        "fecha y hora esperada de regreso",
    )

    if (
        expected_return_at is not None
        and expected_return_at < departure_at
    ):
        raise TransportValidationError(
            "La fecha esperada de regreso no puede ser "
            "anterior a la salida."
        )

    permission = DriverExitPermission(
        driver_id=driver.id,
        truck_id=truck.id if truck else None,
        departure_at=departure_at,
        reason=_required_text(
            data.get("reason"),
            "motivo",
            max_length=240,
        ),
        destination=_clean_text(
            data.get("destination"),
            upper=True,
            max_length=240,
        ),
        expected_return_at=expected_return_at,
        status="AUTHORIZED",
        notes=_clean_text(
            data.get("notes")
        ),
        authorized_by_user_id=user_id,
    )

    db.session.add(permission)
    db.session.flush()

    if commit:
        _commit_or_raise(
            "No fue posible registrar el permiso de salida."
        )

    return permission


def register_exit_permission_return(
    permission: DriverExitPermission,
    *,
    user_id: int,
    actual_return_at: Any = None,
    notes: Any = None,
    commit: bool = True,
) -> DriverExitPermission:
    if permission.status != "AUTHORIZED":
        raise TransportValidationError(
            "Solo se puede registrar el regreso de un permiso activo."
        )

    return_at = (
        _parse_datetime(
            actual_return_at,
            "fecha y hora de regreso",
        )
        or datetime.utcnow()
    )

    if return_at < permission.departure_at:
        raise TransportValidationError(
            "La fecha de regreso no puede ser anterior a la salida."
        )

    permission.actual_return_at = return_at
    permission.returned_by_user_id = user_id
    permission.status = "RETURNED"

    additional_notes = _clean_text(notes)

    if additional_notes:
        permission.notes = (
            f"{permission.notes}\n{additional_notes}"
            if permission.notes
            else additional_notes
        )

    if commit:
        _commit_or_raise(
            "No fue posible registrar el regreso."
        )

    return permission


def cancel_exit_permission(
    permission: DriverExitPermission,
    *,
    user_id: int,
    notes: Any = None,
    commit: bool = True,
) -> DriverExitPermission:
    if permission.status != "AUTHORIZED":
        raise TransportValidationError(
            "Solo se puede cancelar un permiso activo."
        )

    permission.status = "CANCELLED"
    permission.returned_by_user_id = user_id

    cancellation_note = _clean_text(notes)

    if cancellation_note:
        permission.notes = (
            f"{permission.notes}\nCANCELACIÓN: {cancellation_note}"
            if permission.notes
            else f"CANCELACIÓN: {cancellation_note}"
        )

    if commit:
        _commit_or_raise(
            "No fue posible cancelar el permiso."
        )

    return permission


# =========================================================
# INCIDENTES
# =========================================================
def create_incident(
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> TransportIncident:
    truck_id = _parse_int(
        data.get("truck_id"),
        "cabezal",
        required=True,
    )

    truck = get_truck_or_404(truck_id)

    driver_id = _parse_int(
        data.get("driver_id"),
        "chofer",
    )

    if driver_id is not None:
        get_driver_or_404(driver_id)
    else:
        active_assignment = get_active_assignment_for_truck(
            truck.id
        )

        if active_assignment is not None:
            driver_id = active_assignment.driver_id

    incident_type = _validate_choice(
        data.get("incident_type"),
        "tipo de incidente",
        INCIDENT_TYPES,
    )

    incident = TransportIncident(
        driver_id=driver_id,
        truck_id=truck.id,
        incident_type=incident_type,
        occurred_at=(
            _parse_datetime(
                data.get("occurred_at"),
                "fecha del incidente",
            )
            or datetime.utcnow()
        ),
        location=_clean_text(
            data.get("location"),
            upper=True,
            max_length=240,
        ),
        description=_required_text(
            data.get("description"),
            "descripción",
        ),
        status="OPEN",
        next_follow_up_at=_parse_datetime(
            data.get("next_follow_up_at"),
            "próximo seguimiento",
        ),
        reported_by_user_id=user_id,
    )

    truck.status = (
        "DAMAGED"
        if incident_type == "DAMAGED"
        else "STRANDED"
    )
    truck.updated_by_user_id = user_id
    truck.updated_at = datetime.utcnow()

    db.session.add(incident)
    db.session.flush()

    if commit:
        _commit_or_raise(
            "No fue posible registrar el incidente."
        )

    return incident


def add_incident_follow_up(
    incident: TransportIncident,
    data: dict[str, Any],
    *,
    user_id: int,
    commit: bool = True,
) -> TransportIncidentFollowUp:
    if incident.status in {
        "RESOLVED",
        "CANCELLED",
    }:
        raise TransportValidationError(
            "No se puede agregar seguimiento a un incidente cerrado."
        )

    contacted_at = (
        _parse_datetime(
            data.get("contacted_at"),
            "fecha de contacto",
        )
        or datetime.utcnow()
    )

    next_follow_up_at = _parse_datetime(
        data.get("next_follow_up_at"),
        "próximo seguimiento",
    )

    resolved = _parse_bool(
        data.get("resolved")
    )

    follow_up = TransportIncidentFollowUp(
        incident_id=incident.id,
        contacted_at=contacted_at,
        contact_name=_clean_text(
            data.get("contact_name"),
            upper=True,
            max_length=180,
        ),
        current_situation=_required_text(
            data.get("current_situation"),
            "situación actual",
        ),
        repair_estimate=_clean_text(
            data.get("repair_estimate"),
            max_length=240,
        ),
        notes=_clean_text(
            data.get("notes")
        ),
        next_follow_up_at=next_follow_up_at,
        resolved=resolved,
        created_by_user_id=user_id,
    )

    incident.last_follow_up_at = contacted_at
    incident.next_follow_up_at = next_follow_up_at
    incident.updated_at = datetime.utcnow()

    if resolved:
        incident.status = "RESOLVED"
        incident.resolved_at = contacted_at
        incident.resolved_by_user_id = user_id
        incident.resolution = follow_up.current_situation

        truck = db.session.get(
            Truck,
            incident.truck_id,
        )

        if truck is not None:
            truck.status = "ACTIVE"
            truck.updated_by_user_id = user_id
            truck.updated_at = datetime.utcnow()
    else:
        incident.status = "FOLLOW_UP"

    db.session.add(follow_up)
    db.session.flush()

    if commit:
        _commit_or_raise(
            "No fue posible registrar el seguimiento."
        )

    return follow_up


def resolve_incident(
    incident: TransportIncident,
    *,
    user_id: int,
    resolution: Any,
    resolved_at: Any = None,
    truck_status: str = "ACTIVE",
    commit: bool = True,
) -> TransportIncident:
    if incident.status in {
        "RESOLVED",
        "CANCELLED",
    }:
        raise TransportValidationError(
            "El incidente ya se encuentra cerrado."
        )

    final_truck_status = _validate_choice(
        truck_status,
        "estado final del cabezal",
        TRUCK_STATUSES,
        default="ACTIVE",
    )

    incident.status = "RESOLVED"
    incident.resolution = _required_text(
        resolution,
        "resolución",
    )
    incident.resolved_at = (
        _parse_datetime(
            resolved_at,
            "fecha de resolución",
        )
        or datetime.utcnow()
    )
    incident.resolved_by_user_id = user_id
    incident.next_follow_up_at = None
    incident.updated_at = datetime.utcnow()

    truck = db.session.get(
        Truck,
        incident.truck_id,
    )

    if truck is not None:
        truck.status = final_truck_status
        truck.updated_by_user_id = user_id
        truck.updated_at = datetime.utcnow()

    if commit:
        _commit_or_raise(
            "No fue posible resolver el incidente."
        )

    return incident


def cancel_incident(
    incident: TransportIncident,
    *,
    user_id: int,
    reason: Any,
    restore_truck_status: bool = True,
    commit: bool = True,
) -> TransportIncident:
    if incident.status in {
        "RESOLVED",
        "CANCELLED",
    }:
        raise TransportValidationError(
            "El incidente ya se encuentra cerrado."
        )

    incident.status = "CANCELLED"
    incident.resolution = _required_text(
        reason,
        "motivo de cancelación",
    )
    incident.resolved_at = datetime.utcnow()
    incident.resolved_by_user_id = user_id
    incident.next_follow_up_at = None
    incident.updated_at = datetime.utcnow()

    if restore_truck_status:
        truck = db.session.get(
            Truck,
            incident.truck_id,
        )

        if truck is not None:
            truck.status = "ACTIVE"
            truck.updated_by_user_id = user_id
            truck.updated_at = datetime.utcnow()

    if commit:
        _commit_or_raise(
            "No fue posible cancelar el incidente."
        )

    return incident


# =========================================================
# CONSULTAS LIVIANAS PARA SELECTORES
# =========================================================
def get_available_drivers(
    *,
    search: str | None = None,
    limit: int = 50,
) -> list[Driver]:
    """
    Retorna únicamente choferes activos sin asignación activa.

    Se usa para selectores y nunca trae relaciones adicionales.
    """
    safe_limit = max(
        1,
        min(limit, 100),
    )

    active_assignment_exists = (
        select(DriverTruckAssignment.id)
        .where(
            DriverTruckAssignment.driver_id == Driver.id,
            DriverTruckAssignment.status == "ACTIVE",
        )
        .exists()
    )

    query = (
        select(Driver)
        .where(
            Driver.status == "ACTIVE",
            ~active_assignment_exists,
        )
        .order_by(
            Driver.name.asc(),
            Driver.id.asc(),
        )
        .limit(safe_limit)
    )

    cleaned_search = _clean_text(
        search,
        upper=True,
    )

    if cleaned_search:
        like_value = f"%{cleaned_search}%"

        query = query.where(
            or_(
                func.upper(Driver.name).like(like_value),
                func.upper(Driver.identification).like(
                    like_value
                ),
            )
        )

    return list(
        db.session.scalars(query).all()
    )


def get_available_trucks(
    *,
    search: str | None = None,
    limit: int = 50,
) -> list[Truck]:
    """
    Retorna cabezales activos sin asignación activa.
    """
    safe_limit = max(
        1,
        min(limit, 100),
    )

    active_assignment_exists = (
        select(DriverTruckAssignment.id)
        .where(
            DriverTruckAssignment.truck_id == Truck.id,
            DriverTruckAssignment.status == "ACTIVE",
        )
        .exists()
    )

    query = (
        select(Truck)
        .where(
            Truck.status == "ACTIVE",
            ~active_assignment_exists,
        )
        .order_by(
            Truck.plate.asc(),
            Truck.id.asc(),
        )
        .limit(safe_limit)
    )

    cleaned_search = _clean_text(
        search,
        upper=True,
    )

    if cleaned_search:
        query = query.where(
            func.upper(Truck.plate).like(
                f"%{cleaned_search}%"
            )
        )

    return list(
        db.session.scalars(query).all()
    )


# =========================================================
# ACTUALIZACIONES PROGRAMADAS
# =========================================================
def refresh_expired_transport_documents(
    *,
    commit: bool = True,
) -> dict[str, int]:
    """
    Actualiza vencimientos mediante UPDATE directo.

    No carga todos los documentos en memoria.
    Es apropiado para una tarea programada diaria.
    """
    today = date.today()

    expired_driver_documents = (
        db.session.query(DriverDocument)
        .filter(
            DriverDocument.no_expiry.is_(False),
            DriverDocument.expiry_date.is_not(None),
            DriverDocument.expiry_date < today,
            DriverDocument.status.notin_(
                {
                    "EXPIRED",
                    "NOT_APPLICABLE",
                }
            ),
        )
        .update(
            {
                DriverDocument.status: "EXPIRED",
                DriverDocument.updated_at: datetime.utcnow(),
            },
            synchronize_session=False,
        )
    )

    valid_driver_documents = (
        db.session.query(DriverDocument)
        .filter(
            DriverDocument.no_expiry.is_(False),
            DriverDocument.expiry_date.is_not(None),
            DriverDocument.expiry_date >= today,
            DriverDocument.status == "EXPIRED",
        )
        .update(
            {
                DriverDocument.status: "VALID",
                DriverDocument.updated_at: datetime.utcnow(),
            },
            synchronize_session=False,
        )
    )

    expired_truck_documents = (
        db.session.query(TruckDocument)
        .filter(
            TruckDocument.no_expiry.is_(False),
            TruckDocument.expiry_date.is_not(None),
            TruckDocument.expiry_date < today,
            TruckDocument.status.notin_(
                {
                    "EXPIRED",
                    "NOT_APPLICABLE",
                }
            ),
        )
        .update(
            {
                TruckDocument.status: "EXPIRED",
                TruckDocument.updated_at: datetime.utcnow(),
            },
            synchronize_session=False,
        )
    )

    valid_truck_documents = (
        db.session.query(TruckDocument)
        .filter(
            TruckDocument.no_expiry.is_(False),
            TruckDocument.expiry_date.is_not(None),
            TruckDocument.expiry_date >= today,
            TruckDocument.status == "EXPIRED",
        )
        .update(
            {
                TruckDocument.status: "VALID",
                TruckDocument.updated_at: datetime.utcnow(),
            },
            synchronize_session=False,
        )
    )

    expired_apm = (
        db.session.query(DriverApmRecord)
        .filter(
            DriverApmRecord.expiry_date.is_not(None),
            DriverApmRecord.expiry_date < today,
            DriverApmRecord.expiry_mode == "DATE",
        )
        .update(
            {
                DriverApmRecord.expiry_mode: "EXPIRED",
                DriverApmRecord.card_status: "EXPIRED",
                DriverApmRecord.updated_at: datetime.utcnow(),
            },
            synchronize_session=False,
        )
    )

    result = {
        "driver_documents_expired": (
            expired_driver_documents
        ),
        "driver_documents_validated": (
            valid_driver_documents
        ),
        "truck_documents_expired": (
            expired_truck_documents
        ),
        "truck_documents_validated": (
            valid_truck_documents
        ),
        "apm_expired": expired_apm,
    }

    if commit and any(result.values()):
        _commit_or_raise(
            "No fue posible actualizar los vencimientos."
        )

    return result


# =========================================================
# HELPERS PARA CARGA MASIVA FUTURA
# =========================================================
def find_driver_by_identification(
    identification: Any,
) -> Driver | None:
    normalized = _normalize_identification(
        identification
    )

    return db.session.scalar(
        select(Driver)
        .where(
            Driver.identification == normalized
        )
        .limit(1)
    )


def find_truck_by_plate(
    plate: Any,
) -> Truck | None:
    normalized = _normalize_plate(
        plate
    )

    return db.session.scalar(
        select(Truck)
        .where(
            Truck.plate == normalized
        )
        .limit(1)
    )


def process_driver_batch(
    rows: Iterable[dict[str, Any]],
    *,
    user_id: int,
) -> dict[str, Any]:
    """
    Base transaccional para carga masiva.

    No ejecuta una consulta para verificar relaciones completas.
    Busca únicamente por identificación indexada.

    Procesa en una sola transacción. Si una fila falla, se revierte
    todo el lote para evitar una carga parcial silenciosa.
    """
    created = 0
    updated = 0
    errors: list[str] = []

    for row_number, row in enumerate(
        rows,
        start=2,
    ):
        try:
            identification = _normalize_identification(
                row.get("identification")
            )

            driver = db.session.scalar(
                select(Driver)
                .where(
                    Driver.identification == identification
                )
                .limit(1)
            )

            if driver is None:
                create_driver(
                    row,
                    user_id=user_id,
                    commit=False,
                )
                created += 1
            else:
                update_driver(
                    driver,
                    row,
                    user_id=user_id,
                    commit=False,
                )
                updated += 1

        except TransportServiceError as exc:
            errors.append(
                f"Fila {row_number}: {exc}"
            )

    if errors:
        db.session.rollback()

        return {
            "ok": False,
            "created": 0,
            "updated": 0,
            "errors": errors,
        }

    _commit_or_raise(
        "No fue posible completar la carga masiva de choferes."
    )

    return {
        "ok": True,
        "created": created,
        "updated": updated,
        "errors": [],
    }