# app/blueprints/tica/services.py

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from zoneinfo import ZoneInfo
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO, Iterable
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_

from app.extensions import db
from app.models.tica import (
    TicaDestination,
    TicaDriver,
    TicaGeneratedFile,
    TicaImportBatch,
    TicaTransporter,
)


# =========================================================
# Constantes del módulo
# =========================================================

TRM_ENCODING = "ISO-8859-1"
TRM_EXTENSION = ".TRM"
CR_TIMEZONE = ZoneInfo("America/Costa_Rica")

MOVEMENT_ENTRY = "E"
MOVEMENT_EXIT = "S"
ALLOWED_MOVEMENT_TYPES = {
    MOVEMENT_ENTRY,
    MOVEMENT_EXIT,
}

MOVEMENT_FILE_LABELS = {
    MOVEMENT_ENTRY: "ENTRADA",
    MOVEMENT_EXIT: "SALIDA",
}

IMPORT_TRANSPORTERS = "TRANSPORTERS"
IMPORT_DRIVERS = "DRIVERS"
IMPORT_DESTINATIONS = "DESTINATIONS"

ALLOWED_IMPORT_TYPES = {
    IMPORT_TRANSPORTERS,
    IMPORT_DRIVERS,
    IMPORT_DESTINATIONS,
}

IMPORT_STATUS_PROCESSING = "PROCESSING"
IMPORT_STATUS_COMPLETED = "COMPLETED"
IMPORT_STATUS_PARTIAL = "PARTIAL"
IMPORT_STATUS_ERROR = "ERROR"

GENERATED_STATUS_CREATED = "CREATED"
GENERATED_STATUS_WRITTEN = "WRITTEN"
GENERATED_STATUS_ERROR = "ERROR"

SEARCH_LIMIT_DEFAULT = 20
SEARCH_LIMIT_MAX = 50
IMPORT_ERROR_LIMIT = 500

_IDENTIFICATION_CLEAN_RE = re.compile(r"[^0-9A-Z]+")
_CODE_CLEAN_RE = re.compile(r"[^0-9A-Z_-]+")
_FILE_COMPONENT_RE = re.compile(r"[^0-9A-Z_-]+")
_CONTAINER_CLEAN_RE = re.compile(r"[^0-9A-Z]+")
_PLATE_CLEAN_RE = re.compile(r"[^0-9A-Z-]+")


# =========================================================
# Excepciones controladas
# =========================================================

class TicaServiceError(Exception):
    """Error controlado del módulo TICA."""

    def __init__(
        self,
        message: str,
        *,
        code: str = "TICA_SERVICE_ERROR",
    ):
        super().__init__(message)
        self.message = message
        self.code = code


class TicaValidationError(TicaServiceError):
    """Error de validación de datos de entrada."""



class TicaImportError(TicaServiceError):
    """Error procesando una carga masiva."""


# =========================================================
# DTO de generación
# =========================================================

@dataclass(frozen=True)
class TicaGenerationData:
    movement_type: str
    trip_number: str
    dua_ordinal: int

    transporter_identification: str

    driver_identification: str
    driver_name: str
    plate: str

    weight: Decimal
    packages: int

    destination_code: str
    movement_date: date

    container_number: str
    seal_number: str

    transporter_id: int | None = None
    driver_id: int | None = None
    destination_id: int | None = None

    transporter_name_snapshot: str | None = None
    destination_name_snapshot: str | None = None


@dataclass(frozen=True)
class TicaGeneratedResult:
    """
    Resultado del archivo TRM generado en memoria.

    El servidor no conoce la ruta final elegida por el usuario.
    El navegador entrega el archivo para que se guarde manualmente
    en C:\\tica\\tr\\out.
    """

    file_name: str
    content_text: str
    content_bytes: bytes
    generated_at: datetime


@dataclass
class ImportSummary:
    total_rows: int = 0
    created_rows: int = 0
    updated_rows: int = 0
    skipped_rows: int = 0
    error_rows: int = 0
    errors: list[dict[str, Any]] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []

    @property
    def status(self) -> str:
        if self.error_rows and not (
            self.created_rows
            or self.updated_rows
            or self.skipped_rows
        ):
            return IMPORT_STATUS_ERROR

        if self.error_rows:
            return IMPORT_STATUS_PARTIAL

        return IMPORT_STATUS_COMPLETED

    def as_dict(self) -> dict[str, Any]:
        return {
            "total_rows": self.total_rows,
            "created_rows": self.created_rows,
            "updated_rows": self.updated_rows,
            "skipped_rows": self.skipped_rows,
            "error_rows": self.error_rows,
            "status": self.status,
            "errors": self.errors or [],
        }


# =========================================================
# Normalización general
# =========================================================

def excel_value_to_text(value: Any) -> str:
    """
    Convierte valores numéricos de Excel a texto sin agregar ".0".
    """
    if value is None:
        return ""

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")

    return str(value)


def normalize_text(
    value: Any,
    *,
    uppercase: bool = True,
    max_length: int | None = None,
) -> str:
    """
    Convierte cualquier valor a texto limpio.

    - Elimina espacios repetidos.
    - Opcionalmente transforma a mayúsculas.
    - Permite limitar longitud.
    """
    if value is None:
        return ""

    text_value = " ".join(str(value).strip().split())

    if uppercase:
        text_value = text_value.upper()

    if max_length is not None:
        text_value = text_value[:max_length]

    return text_value


def normalize_identification(value: Any) -> str:
    """
    Normaliza cédulas e identificadores.

    Conserva números y letras para no imponer una regla que la
    base de datos actual no exige.
    """
    normalized = normalize_text(
        excel_value_to_text(value),
        uppercase=True,
        max_length=50,
    )

    return _IDENTIFICATION_CLEAN_RE.sub("", normalized)


def normalize_plate(value: Any) -> str:
    normalized = normalize_text(
        excel_value_to_text(value),
        uppercase=True,
        max_length=30,
    )

    return _PLATE_CLEAN_RE.sub("", normalized)


def normalize_destination_code(value: Any) -> str:
    normalized = normalize_text(
        excel_value_to_text(value),
        uppercase=True,
        max_length=50,
    )

    return _CODE_CLEAN_RE.sub("", normalized)


def normalize_container_number(value: Any) -> str:
    """
    El archivo real utiliza el contenedor sin guiones:
    FCIU5752715.
    """
    normalized = normalize_text(
        value,
        uppercase=True,
        max_length=30,
    )

    return _CONTAINER_CLEAN_RE.sub("", normalized)


def normalize_seal_number(value: Any) -> str:
    return normalize_text(
        value,
        uppercase=True,
        max_length=120,
    )


def normalize_trip_number(value: Any) -> str:
    """
    El viaje se ingresa manualmente.

    Se conserva como texto para no perder ceros iniciales,
    pero únicamente permite números.
    """
    normalized = normalize_text(
        excel_value_to_text(value),
        uppercase=False,
        max_length=60,
    )

    if not normalized:
        return ""

    if not normalized.isdigit():
        raise TicaValidationError(
            "El viaje debe contener únicamente números.",
            code="INVALID_TRIP_NUMBER",
        )

    return normalized


def normalize_movement_type(value: Any) -> str:
    movement_type = normalize_text(
        value,
        uppercase=True,
        max_length=1,
    )

    if movement_type not in ALLOWED_MOVEMENT_TYPES:
        raise TicaValidationError(
            "El tipo de movimiento debe ser Entrada o Salida.",
            code="INVALID_MOVEMENT_TYPE",
        )

    return movement_type


def parse_positive_integer(
    value: Any,
    *,
    field_name: str,
    minimum: int = 1,
) -> int:
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError):
        raise TicaValidationError(
            f"{field_name} debe ser un número entero.",
            code="INVALID_INTEGER",
        )

    if parsed < minimum:
        raise TicaValidationError(
            f"{field_name} debe ser mayor o igual a {minimum}.",
            code="INTEGER_OUT_OF_RANGE",
        )

    return parsed


def parse_nonnegative_integer(
    value: Any,
    *,
    field_name: str,
) -> int:
    return parse_positive_integer(
        value,
        field_name=field_name,
        minimum=0,
    )


def parse_nonnegative_decimal(
    value: Any,
    *,
    field_name: str,
) -> Decimal:
    normalized = str(value or "").strip().replace(",", ".")

    try:
        parsed = Decimal(normalized)
    except (InvalidOperation, ValueError):
        raise TicaValidationError(
            f"{field_name} debe ser un número válido.",
            code="INVALID_DECIMAL",
        )

    if parsed < 0:
        raise TicaValidationError(
            f"{field_name} no puede ser negativo.",
            code="NEGATIVE_DECIMAL",
        )

    return parsed


def decimal_to_xml(value: Decimal) -> str:
    """
    Evita notación científica y ceros decimales innecesarios.

    Ejemplos:
    13712.000 -> 13712
    13712.500 -> 13712.5
    """
    normalized = format(value, "f")

    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")

    return normalized or "0"


def parse_movement_date(value: Any | None) -> date:
    """
    Obtiene la fecha del movimiento usando Costa Rica como zona horaria.
    """
    if value in (None, ""):
        return datetime.now(CR_TIMEZONE).date()

    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(CR_TIMEZONE).date()
        return value.date()

    if isinstance(value, date):
        return value

    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise TicaValidationError(
            "La fecha del movimiento no es válida.",
            code="INVALID_MOVEMENT_DATE",
        ) from exc


def normalize_boolean(value: Any, *, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default

    normalized = normalize_text(
        value,
        uppercase=True,
    )

    if normalized in {"SI", "SÍ", "TRUE", "1", "YES", "ACTIVO"}:
        return True

    if normalized in {"NO", "FALSE", "0", "INACTIVO"}:
        return False

    return default


def xml_value(value: Any) -> str:
    """
    Escapa caracteres especiales XML y valida que el valor pueda
    representarse en ISO-8859-1.
    """
    text_value = str(value if value is not None else "")

    try:
        text_value.encode(TRM_ENCODING)
    except UnicodeEncodeError as exc:
        raise TicaValidationError(
            (
                "Uno de los valores contiene caracteres que no pueden "
                "guardarse en ISO-8859-1."
            ),
            code="UNSUPPORTED_CHARACTER",
        ) from exc

    return xml_escape(
        text_value,
        entities={
            '"': "&quot;",
            "'": "&apos;",
        },
    )


# =========================================================
# Validación de datos para TRM
# =========================================================

def build_generation_data(
    *,
    movement_type: Any,
    trip_number: Any,
    dua_ordinal: Any,
    transporter_identification: Any,
    driver_identification: Any,
    driver_name: Any,
    plate: Any,
    weight: Any,
    packages: Any,
    destination_code: Any,
    container_number: Any,
    seal_number: Any,
    movement_date: Any | None = None,
    transporter_id: int | None = None,
    driver_id: int | None = None,
    destination_id: int | None = None,
    transporter_name_snapshot: Any | None = None,
    destination_name_snapshot: Any | None = None,
) -> TicaGenerationData:
    """
    Valida y normaliza todos los datos antes de generar el TRM.
    """
    normalized_movement_type = normalize_movement_type(
        movement_type
    )

    normalized_trip = normalize_trip_number(
        trip_number
    )

    normalized_ordinal = parse_positive_integer(
        dua_ordinal if dua_ordinal not in (None, "") else 1,
        field_name="Ordinal de DUA",
        minimum=1,
    )

    normalized_transporter_id = normalize_identification(
        transporter_identification
    )

    normalized_driver_id = normalize_identification(
        driver_identification
    )

    normalized_driver_name = normalize_text(
        driver_name,
        uppercase=True,
        max_length=180,
    )

    normalized_plate = normalize_plate(
        plate
    )

    normalized_weight = parse_nonnegative_decimal(
        weight,
        field_name="Peso",
    )

    normalized_packages = parse_nonnegative_integer(
        packages,
        field_name="Bultos",
    )

    normalized_destination = normalize_destination_code(
        destination_code
    )

    normalized_container = normalize_container_number(
        container_number
    )

    normalized_seal = normalize_seal_number(
        seal_number
    )

    normalized_date = parse_movement_date(
        movement_date
    )

    required_values = {
        "Viaje": normalized_trip,
        "Cédula del transportista": normalized_transporter_id,
        "Cédula del chofer": normalized_driver_id,
        "Nombre del chofer": normalized_driver_name,
        "Placa": normalized_plate,
        "Código del destino": normalized_destination,
        "Contenedor": normalized_container,
        "Marchamo": normalized_seal,
    }

    missing_fields = [
        field_name
        for field_name, field_value in required_values.items()
        if not field_value
    ]

    if missing_fields:
        raise TicaValidationError(
            "Faltan campos requeridos: "
            + ", ".join(missing_fields)
            + ".",
            code="REQUIRED_FIELDS_MISSING",
        )

    return TicaGenerationData(
        movement_type=normalized_movement_type,
        trip_number=normalized_trip,
        dua_ordinal=normalized_ordinal,
        transporter_identification=normalized_transporter_id,
        driver_identification=normalized_driver_id,
        driver_name=normalized_driver_name,
        plate=normalized_plate,
        weight=normalized_weight,
        packages=normalized_packages,
        destination_code=normalized_destination,
        movement_date=normalized_date,
        container_number=normalized_container,
        seal_number=normalized_seal,
        transporter_id=transporter_id,
        driver_id=driver_id,
        destination_id=destination_id,
        transporter_name_snapshot=normalize_text(
            transporter_name_snapshot,
            uppercase=True,
            max_length=180,
        ) or None,
        destination_name_snapshot=normalize_text(
            destination_name_snapshot,
            uppercase=True,
            max_length=180,
        ) or None,
    )


# =========================================================
# Generación exacta del TRM
# =========================================================

def build_trm_text(data: TicaGenerationData) -> str:
    """
    Genera el contenido exacto del archivo TRM.

    Reglas confirmadas:
    - ISO-8859-1.
    - CRLF.
    - El ordinal se replica en cinco lugares.
    - E/S se replica en TRTENTSAL y TRTMOVPRCTPO.
    - Se conserva el orden de nodos del archivo suministrado.
    """
    movement_type = xml_value(data.movement_type)
    ordinal = xml_value(data.dua_ordinal)

    lines = [
        '<?xml version="1.0" encoding="ISO-8859-1"?>',
        (
            '<ROOT xmlns="http://www.hacienda.go.cr/TICA" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        ),
        "  <ROWSET_TRTMOVPORTON>",
        f'    <ROW_TRTMOVPORTON NUM="{ordinal}">',
        f"      <TRTVJID>{xml_value(data.trip_number)}</TRTVJID>",
        f"      <TRTMOVORD>{ordinal}</TRTMOVORD>",
        "      <TRTMOVRCT>E077</TRTMOVRCT>",
        f"      <TRTENTSAL>{movement_type}</TRTENTSAL>",
        "      <TRTMOVPAISTRANS>188</TRTMOVPAISTRANS>",
        "      <TRTMOVTRANSIDTIPO>J</TRTMOVTRANSIDTIPO>",
        (
            "      <TRTMOVTRANSID>"
            f"{xml_value(data.transporter_identification)}"
            "</TRTMOVTRANSID>"
        ),
        "      <TRTMOVCHOFERIDTIPO>F</TRTMOVCHOFERIDTIPO>",
        (
            "      <TRTMOVCHOFERID>"
            f"{xml_value(data.driver_identification)}"
            "</TRTMOVCHOFERID>"
        ),
        (
            "      <TRTMOVCHOFERNOMBRE>"
            f"{xml_value(data.driver_name)}"
            "</TRTMOVCHOFERNOMBRE>"
        ),
        "      <TRTNACMAT>188</TRTNACMAT>",
        f"      <TRTMOVMAT>{xml_value(data.plate)}</TRTMOVMAT>",
        "      <TRTMOVMATR>0</TRTMOVMATR>",
        "      <TRTMOVMATZ>1</TRTMOVMATZ>",
        (
            "      <TRTMOVPESO>"
            f"{xml_value(decimal_to_xml(data.weight))}"
            "</TRTMOVPESO>"
        ),
        f"      <TRTMOVBULT>{xml_value(data.packages)}</TRTMOVBULT>",
        (
            "      <TRTRTADST>"
            f"{xml_value(data.destination_code)}"
            "</TRTRTADST>"
        ),
        (
            "      <TRTMOVFCHDESCARGA>"
            f"{xml_value(data.movement_date.isoformat())}"
            "</TRTMOVFCHDESCARGA>"
        ),
        "      <TRTMOVPORPRC>N</TRTMOVPORPRC>",
        "      <ROWSET_TRTMOVPORTACTAS>",
        f'        <ROW_TRTMOVPORTACTAS NUM="{ordinal}">',
        "          <TRTACTOBS>0</TRTACTOBS>",
        "          <TRTACTID>1</TRTACTID>",
        "        </ROW_TRTMOVPORTACTAS>",
        "      </ROWSET_TRTMOVPORTACTAS>",
        "      <ROWSET_TRTMOVPORTCONT>",
        f'        <ROW_TRTMOVPORTCONT NUM="{ordinal}">',
        (
            "          <TRTMOVCONT>"
            f"{xml_value(data.container_number)}"
            "</TRTMOVCONT>"
        ),
        "          <TRTMOVCSOLPERM>",
        "          </TRTMOVCSOLPERM>",
        "        </ROW_TRTMOVPORTCONT>",
        "      </ROWSET_TRTMOVPORTCONT>",
        "      <ROWSET_TRTMOVPORTPREC>",
        f'        <ROW_TRTMOVPORTPREC NUM="{ordinal}">',
        (
            "          <TRTMOVPRCTPO>"
            f"{movement_type}"
            "</TRTMOVPRCTPO>"
        ),
        (
            "          <TRTMOVPRC>"
            f"{xml_value(data.seal_number)}"
            "</TRTMOVPRC>"
        ),
        "        </ROW_TRTMOVPORTPREC>",
        "      </ROWSET_TRTMOVPORTPREC>",
        "    </ROW_TRTMOVPORTON>",
        "  </ROWSET_TRTMOVPORTON>",
        "</ROOT>",
    ]

    return "\r\n".join(lines)


def encode_trm_text(content_text: str) -> bytes:
    try:
        return content_text.encode(
            TRM_ENCODING,
            errors="strict",
        )
    except UnicodeEncodeError as exc:
        raise TicaValidationError(
            (
                "El archivo contiene caracteres que no pueden "
                "codificarse en ISO-8859-1."
            ),
            code="TRM_ENCODING_ERROR",
        ) from exc


def build_trm_file_name(data: TicaGenerationData) -> str:
    movement_label = MOVEMENT_FILE_LABELS[
        data.movement_type
    ]

    trip_component = _FILE_COMPONENT_RE.sub(
        "",
        data.trip_number.upper(),
    )

    container_component = _FILE_COMPONENT_RE.sub(
        "",
        data.container_number.upper(),
    )

    if not trip_component or not container_component:
        raise TicaValidationError(
            "No se pudo construir un nombre válido para el archivo.",
            code="INVALID_FILE_NAME",
        )

    return (
        f"{movement_label}_"
        f"{trip_component}_"
        f"{container_component}"
        f"{TRM_EXTENSION}"
    )


def generate_trm_in_memory(
    data: TicaGenerationData,
) -> TicaGeneratedResult:
    """
    Genera el archivo TRM completamente en memoria.

    No escribe archivos en Render.
    No accede al disco local del usuario.
    No realiza consultas ni commits de base de datos.

    La ruta Flask será responsable de:
    1. guardar el contenido en tica_generated_files;
    2. hacer commit;
    3. devolver el archivo al navegador.
    """
    content_text = build_trm_text(data)
    content_bytes = encode_trm_text(content_text)
    file_name = build_trm_file_name(data)

    return TicaGeneratedResult(
        file_name=file_name,
        content_text=content_text,
        content_bytes=content_bytes,
        generated_at=datetime.utcnow(),
    )


def build_generated_file_record(
    *,
    data: TicaGenerationData,
    result: TicaGeneratedResult,
    created_by_user_id: int | None,
    status: str = GENERATED_STATUS_CREATED,
    error_message: str | None = None,
) -> TicaGeneratedFile:
    """
    Construye el registro histórico del archivo generado.

    Guarda:
    - los datos utilizados;
    - el nombre del archivo;
    - el contenido TRM completo;
    - el usuario;
    - la fecha de generación.

    No hace add(), flush() ni commit().
    """
    return TicaGeneratedFile(
        transporter_id=data.transporter_id,
        driver_id=data.driver_id,
        destination_id=data.destination_id,

        movement_type=data.movement_type,
        trip_number=data.trip_number,
        dua_ordinal=data.dua_ordinal,

        transporter_name_snapshot=(
            data.transporter_name_snapshot
        ),
        transporter_identification=(
            data.transporter_identification
        ),

        driver_name=data.driver_name,
        driver_identification=(
            data.driver_identification
        ),
        plate=data.plate,

        destination_name_snapshot=(
            data.destination_name_snapshot
        ),
        destination_code=data.destination_code,

        weight=data.weight,
        packages=data.packages,
        movement_date=data.movement_date,

        container_number=data.container_number,
        seal_number=data.seal_number,

        file_name=result.file_name,

        # El servidor no conoce la ruta final escogida
        # por el usuario en su computadora.
        output_path=None,

        # Copia histórica exacta del archivo generado.
        content_text=result.content_text,

        status=status,

        error_message=(
            normalize_text(
                error_message,
                uppercase=False,
            )[:4000]
            if error_message
            else None
        ),

        # El modelo actual tiene written_at.
        # Se usa como fecha en que el TRM quedó generado
        # y almacenado correctamente en la base.
        written_at=result.generated_at,

        created_by_user_id=created_by_user_id,
        created_at=result.generated_at,
    )

def rebuild_generated_file_bytes(
    generated_file: TicaGeneratedFile,
) -> bytes:
    """
    Reconstruye los bytes de un archivo histórico.

    Permite descargar nuevamente el mismo TRM aunque TICA ya lo haya
    consumido y eliminado de C:\\tica\\tr\\out.
    """
    content_text = generated_file.content_text or ""

    if not content_text:
        raise TicaServiceError(
            "El registro no contiene una copia del archivo TRM.",
            code="TRM_CONTENT_NOT_FOUND",
        )

    return encode_trm_text(content_text)


# =========================================================
# Búsquedas ágiles de catálogos
# =========================================================

def _resolve_search_limit(value: Any) -> int:
    try:
        limit = int(value)
    except (TypeError, ValueError):
        limit = SEARCH_LIMIT_DEFAULT

    return max(
        1,
        min(limit, SEARCH_LIMIT_MAX),
    )


def search_transporters(
    query_text: str,
    *,
    limit: int = SEARCH_LIMIT_DEFAULT,
    active_only: bool = True,
) -> list[dict[str, Any]]:
    query_text = normalize_text(
        query_text,
        uppercase=True,
        max_length=100,
    )

    if len(query_text) < 2:
        return []

    normalized_identification = normalize_identification(
        query_text
    )

    search_pattern = f"%{query_text}%"
    identification_pattern = (
        f"%{normalized_identification}%"
        if normalized_identification
        else None
    )

    query = db.session.query(
        TicaTransporter.id,
        TicaTransporter.name,
        TicaTransporter.identification_number,
        TicaTransporter.is_active,
    )

    if active_only:
        query = query.filter(
            TicaTransporter.is_active.is_(True)
        )

    filters = [
        func.upper(
            TicaTransporter.name
        ).like(search_pattern),
    ]

    if identification_pattern:
        filters.append(
            TicaTransporter.identification_number.like(
                identification_pattern
            )
        )

    rows = (
        query
        .filter(or_(*filters))
        .order_by(
            TicaTransporter.name.asc(),
            TicaTransporter.id.asc(),
        )
        .limit(_resolve_search_limit(limit))
        .all()
    )

    return [
        {
            "id": row.id,
            "name": row.name,
            "identification_number": (
                row.identification_number
            ),
            "is_active": bool(row.is_active),
        }
        for row in rows
    ]


def search_drivers(
    *,
    transporter_id: int,
    query_text: str = "",
    limit: int = SEARCH_LIMIT_DEFAULT,
    active_only: bool = True,
) -> list[dict[str, Any]]:
    try:
        resolved_transporter_id = int(transporter_id)
    except (TypeError, ValueError):
        return []

    query_text = normalize_text(
        query_text,
        uppercase=True,
        max_length=100,
    )

    normalized_identification = normalize_identification(
        query_text
    )

    normalized_plate = normalize_plate(
        query_text
    )

    query = db.session.query(
        TicaDriver.id,
        TicaDriver.transporter_id,
        TicaDriver.name,
        TicaDriver.identification_number,
        TicaDriver.plate,
        TicaDriver.is_active,
    ).filter(
        TicaDriver.transporter_id
        == resolved_transporter_id
    )

    if active_only:
        query = query.filter(
            TicaDriver.is_active.is_(True)
        )

    if query_text:
        filters = [
            func.upper(
                TicaDriver.name
            ).like(f"%{query_text}%"),
        ]

        if normalized_identification:
            filters.append(
                TicaDriver.identification_number.like(
                    f"%{normalized_identification}%"
                )
            )

        if normalized_plate:
            filters.append(
                func.upper(TicaDriver.plate).like(
                    f"%{normalized_plate}%"
                )
            )

        query = query.filter(
            or_(*filters)
        )

    rows = (
        query
        .order_by(
            TicaDriver.name.asc(),
            TicaDriver.id.asc(),
        )
        .limit(_resolve_search_limit(limit))
        .all()
    )

    return [
        {
            "id": row.id,
            "transporter_id": row.transporter_id,
            "name": row.name,
            "identification_number": (
                row.identification_number
            ),
            "plate": row.plate,
            "is_active": bool(row.is_active),
        }
        for row in rows
    ]


def search_destinations(
    query_text: str,
    *,
    limit: int = SEARCH_LIMIT_DEFAULT,
    active_only: bool = True,
) -> list[dict[str, Any]]:
    query_text = normalize_text(
        query_text,
        uppercase=True,
        max_length=100,
    )

    if len(query_text) < 2:
        return []

    normalized_code = normalize_destination_code(
        query_text
    )

    filters = [
        func.upper(
            TicaDestination.name
        ).like(f"%{query_text}%"),
    ]

    if normalized_code:
        filters.append(
            func.upper(
                TicaDestination.code
            ).like(f"%{normalized_code}%")
        )

    query = db.session.query(
        TicaDestination.id,
        TicaDestination.name,
        TicaDestination.code,
        TicaDestination.is_active,
    )

    if active_only:
        query = query.filter(
            TicaDestination.is_active.is_(True)
        )

    rows = (
        query
        .filter(or_(*filters))
        .order_by(
            TicaDestination.name.asc(),
            TicaDestination.id.asc(),
        )
        .limit(_resolve_search_limit(limit))
        .all()
    )

    return [
        {
            "id": row.id,
            "name": row.name,
            "code": row.code,
            "is_active": bool(row.is_active),
        }
        for row in rows
    ]


# =========================================================
# Crear o actualizar catálogos
# =========================================================

def create_or_update_transporter(
    *,
    name: Any,
    identification_number: Any,
    user_id: int | None,
    is_active: bool = True,
) -> tuple[TicaTransporter, bool]:
    normalized_name = normalize_text(
        name,
        uppercase=True,
        max_length=180,
    )

    normalized_identification = normalize_identification(
        identification_number
    )

    if not normalized_name:
        raise TicaValidationError(
            "El nombre del transportista es obligatorio.",
            code="TRANSPORTER_NAME_REQUIRED",
        )

    if not normalized_identification:
        raise TicaValidationError(
            "La cédula del transportista es obligatoria.",
            code="TRANSPORTER_IDENTIFICATION_REQUIRED",
        )

    transporter = (
        TicaTransporter.query
        .filter(
            TicaTransporter.identification_number
            == normalized_identification
        )
        .first()
    )

    created = transporter is None

    if transporter is None:
        transporter = TicaTransporter(
            name=normalized_name,
            identification_number=(
                normalized_identification
            ),
            is_active=bool(is_active),
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.session.add(transporter)
    else:
        transporter.name = normalized_name
        transporter.is_active = bool(is_active)
        transporter.updated_by_user_id = user_id
        transporter.updated_at = datetime.utcnow()

    return transporter, created


def create_or_update_driver(
    *,
    transporter_id: Any,
    name: Any,
    identification_number: Any,
    plate: Any,
    user_id: int | None,
    is_active: bool = True,
) -> tuple[TicaDriver, bool]:
    try:
        resolved_transporter_id = int(transporter_id)
    except (TypeError, ValueError):
        raise TicaValidationError(
            "El transportista seleccionado no es válido.",
            code="INVALID_TRANSPORTER",
        )

    transporter_exists = db.session.query(
        TicaTransporter.id
    ).filter(
        TicaTransporter.id == resolved_transporter_id
    ).scalar()

    if not transporter_exists:
        raise TicaValidationError(
            "El transportista seleccionado no existe.",
            code="TRANSPORTER_NOT_FOUND",
        )

    normalized_name = normalize_text(
        name,
        uppercase=True,
        max_length=180,
    )

    normalized_identification = normalize_identification(
        identification_number
    )

    normalized_plate = normalize_plate(
        plate
    )

    if not normalized_name:
        raise TicaValidationError(
            "El nombre del chofer es obligatorio.",
            code="DRIVER_NAME_REQUIRED",
        )

    if not normalized_identification:
        raise TicaValidationError(
            "La cédula del chofer es obligatoria.",
            code="DRIVER_IDENTIFICATION_REQUIRED",
        )

    if not normalized_plate:
        raise TicaValidationError(
            "La placa es obligatoria.",
            code="DRIVER_PLATE_REQUIRED",
        )

    driver = (
        TicaDriver.query
        .filter(
            TicaDriver.identification_number
            == normalized_identification
        )
        .first()
    )

    created = driver is None

    if driver is None:
        driver = TicaDriver(
            transporter_id=resolved_transporter_id,
            name=normalized_name,
            identification_number=(
                normalized_identification
            ),
            plate=normalized_plate,
            is_active=bool(is_active),
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.session.add(driver)
    else:
        driver.transporter_id = resolved_transporter_id
        driver.name = normalized_name
        driver.plate = normalized_plate
        driver.is_active = bool(is_active)
        driver.updated_by_user_id = user_id
        driver.updated_at = datetime.utcnow()

    return driver, created


def create_or_update_destination(
    *,
    name: Any,
    code: Any,
    user_id: int | None,
    is_active: bool = True,
) -> tuple[TicaDestination, bool]:
    normalized_name = normalize_text(
        name,
        uppercase=True,
        max_length=180,
    )

    normalized_code = normalize_destination_code(
        code
    )

    if not normalized_name:
        raise TicaValidationError(
            "El nombre de la ubicación es obligatorio.",
            code="DESTINATION_NAME_REQUIRED",
        )

    if not normalized_code:
        raise TicaValidationError(
            "El código de la ubicación es obligatorio.",
            code="DESTINATION_CODE_REQUIRED",
        )

    destination = (
        TicaDestination.query
        .filter(
            TicaDestination.code == normalized_code
        )
        .first()
    )

    created = destination is None

    if destination is None:
        destination = TicaDestination(
            name=normalized_name,
            code=normalized_code,
            is_active=bool(is_active),
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.session.add(destination)
    else:
        destination.name = normalized_name
        destination.is_active = bool(is_active)
        destination.updated_by_user_id = user_id
        destination.updated_at = datetime.utcnow()

    return destination, created


# =========================================================
# Excel: utilidades
# =========================================================

def _open_excel_workbook(
    file_source: BinaryIO | bytes,
):
    try:
        if isinstance(file_source, bytes):
            stream = io.BytesIO(file_source)
        else:
            stream = file_source

            try:
                stream.seek(0)
            except (AttributeError, OSError):
                pass

        return load_workbook(
            filename=stream,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise TicaImportError(
            "No se pudo leer el archivo Excel.",
            code="INVALID_EXCEL_FILE",
        ) from exc


def _normalize_header(value: Any) -> str:
    return (
        normalize_text(
            value,
            uppercase=False,
        )
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
    )


def _read_excel_rows(
    file_source: BinaryIO | bytes,
    *,
    required_headers: Iterable[str],
) -> list[dict[str, Any]]:
    """
    Lee un Excel en modo read_only.

    Devuelve solo filas con al menos un valor.
    """
    workbook = _open_excel_workbook(
        file_source
    )

    try:
        worksheet = workbook.active
        row_iterator = worksheet.iter_rows(
            values_only=True
        )

        try:
            raw_headers = next(row_iterator)
        except StopIteration:
            raise TicaImportError(
                "El archivo Excel está vacío.",
                code="EMPTY_EXCEL_FILE",
            )

        headers = [
            _normalize_header(header)
            for header in raw_headers
        ]

        required_header_set = {
            _normalize_header(header)
            for header in required_headers
        }

        missing_headers = sorted(
            required_header_set - set(headers)
        )

        if missing_headers:
            raise TicaImportError(
                (
                    "Faltan columnas requeridas: "
                    + ", ".join(missing_headers)
                    + "."
                ),
                code="EXCEL_HEADERS_MISSING",
            )

        rows: list[dict[str, Any]] = []

        for excel_row_number, raw_row in enumerate(
            row_iterator,
            start=2,
        ):
            row_data = {
                headers[index]: raw_row[index]
                if index < len(raw_row)
                else None
                for index in range(len(headers))
                if headers[index]
            }

            if not any(
                value not in (None, "")
                for value in row_data.values()
            ):
                continue

            row_data["_excel_row"] = excel_row_number
            rows.append(row_data)

        return rows

    finally:
        workbook.close()


def _append_import_error(
    summary: ImportSummary,
    *,
    row_number: int,
    message: str,
) -> None:
    summary.error_rows += 1

    if len(summary.errors or []) < IMPORT_ERROR_LIMIT:
        summary.errors.append({
            "row": row_number,
            "message": message,
        })


# =========================================================
# Carga masiva: transportistas
# =========================================================

def import_transporters_from_excel(
    file_source: BinaryIO | bytes,
    *,
    user_id: int | None,
) -> ImportSummary:
    rows = _read_excel_rows(
        file_source,
        required_headers={
            "transportista_nombre",
            "transportista_cedula",
        },
    )

    summary = ImportSummary(
        total_rows=len(rows)
    )

    normalized_rows: list[dict[str, Any]] = []
    identifications: set[str] = set()

    for row in rows:
        row_number = int(row["_excel_row"])

        name = normalize_text(
            row.get("transportista_nombre"),
            uppercase=True,
            max_length=180,
        )

        identification = normalize_identification(
            row.get("transportista_cedula")
        )

        if not name or not identification:
            _append_import_error(
                summary,
                row_number=row_number,
                message=(
                    "Nombre y cédula del transportista "
                    "son obligatorios."
                ),
            )
            continue

        normalized_rows.append({
            "row_number": row_number,
            "name": name,
            "identification": identification,
            "is_active": normalize_boolean(
                row.get("activo"),
                default=True,
            ),
        })

        identifications.add(
            identification
        )

    existing_by_identification = {}

    if identifications:
        existing_rows = (
            TicaTransporter.query
            .filter(
                TicaTransporter.identification_number.in_(
                    identifications
                )
            )
            .all()
        )

        existing_by_identification = {
            item.identification_number: item
            for item in existing_rows
        }

    seen_in_file: set[str] = set()

    for row in normalized_rows:
        identification = row["identification"]

        if identification in seen_in_file:
            summary.skipped_rows += 1
            continue

        seen_in_file.add(
            identification
        )

        transporter = existing_by_identification.get(
            identification
        )

        if transporter is None:
            transporter = TicaTransporter(
                name=row["name"],
                identification_number=identification,
                is_active=row["is_active"],
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )

            db.session.add(transporter)
            existing_by_identification[
                identification
            ] = transporter

            summary.created_rows += 1
        else:
            transporter.name = row["name"]
            transporter.is_active = row["is_active"]
            transporter.updated_by_user_id = user_id
            transporter.updated_at = datetime.utcnow()

            summary.updated_rows += 1

    return summary


# =========================================================
# Carga masiva: choferes
# =========================================================

def import_drivers_from_excel(
    file_source: BinaryIO | bytes,
    *,
    user_id: int | None,
) -> ImportSummary:
    rows = _read_excel_rows(
        file_source,
        required_headers={
            "chofer_nombre",
            "chofer_cedula",
            "placa",
            "transportista_cedula",
        },
    )

    summary = ImportSummary(
        total_rows=len(rows)
    )

    normalized_rows: list[dict[str, Any]] = []
    driver_identifications: set[str] = set()
    transporter_identifications: set[str] = set()

    for row in rows:
        row_number = int(row["_excel_row"])

        driver_name = normalize_text(
            row.get("chofer_nombre"),
            uppercase=True,
            max_length=180,
        )

        driver_identification = normalize_identification(
            row.get("chofer_cedula")
        )

        plate = normalize_plate(
            row.get("placa")
        )

        transporter_identification = (
            normalize_identification(
                row.get("transportista_cedula")
            )
        )

        if not all([
            driver_name,
            driver_identification,
            plate,
            transporter_identification,
        ]):
            _append_import_error(
                summary,
                row_number=row_number,
                message=(
                    "Nombre, cédula del chofer, placa y "
                    "cédula del transportista son obligatorios."
                ),
            )
            continue

        normalized_rows.append({
            "row_number": row_number,
            "name": driver_name,
            "identification": driver_identification,
            "plate": plate,
            "transporter_identification": (
                transporter_identification
            ),
            "is_active": normalize_boolean(
                row.get("activo"),
                default=True,
            ),
        })

        driver_identifications.add(
            driver_identification
        )

        transporter_identifications.add(
            transporter_identification
        )

    transporters_by_identification = {}

    if transporter_identifications:
        transporter_rows = (
            TicaTransporter.query
            .filter(
                TicaTransporter.identification_number.in_(
                    transporter_identifications
                )
            )
            .all()
        )

        transporters_by_identification = {
            item.identification_number: item
            for item in transporter_rows
        }

    existing_drivers_by_identification = {}

    if driver_identifications:
        existing_driver_rows = (
            TicaDriver.query
            .filter(
                TicaDriver.identification_number.in_(
                    driver_identifications
                )
            )
            .all()
        )

        existing_drivers_by_identification = {
            item.identification_number: item
            for item in existing_driver_rows
        }

    seen_in_file: set[str] = set()

    for row in normalized_rows:
        identification = row["identification"]

        if identification in seen_in_file:
            summary.skipped_rows += 1
            continue

        seen_in_file.add(
            identification
        )

        transporter = transporters_by_identification.get(
            row["transporter_identification"]
        )

        if transporter is None:
            _append_import_error(
                summary,
                row_number=row["row_number"],
                message=(
                    "No existe el transportista con cédula "
                    f"{row['transporter_identification']}."
                ),
            )
            continue

        driver = existing_drivers_by_identification.get(
            identification
        )

        if driver is None:
            driver = TicaDriver(
                transporter_id=transporter.id,
                name=row["name"],
                identification_number=identification,
                plate=row["plate"],
                is_active=row["is_active"],
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )

            db.session.add(driver)

            existing_drivers_by_identification[
                identification
            ] = driver

            summary.created_rows += 1
        else:
            driver.transporter_id = transporter.id
            driver.name = row["name"]
            driver.plate = row["plate"]
            driver.is_active = row["is_active"]
            driver.updated_by_user_id = user_id
            driver.updated_at = datetime.utcnow()

            summary.updated_rows += 1

    return summary


# =========================================================
# Carga masiva: destinos
# =========================================================

def import_destinations_from_excel(
    file_source: BinaryIO | bytes,
    *,
    user_id: int | None,
) -> ImportSummary:
    rows = _read_excel_rows(
        file_source,
        required_headers={
            "ubicacion_nombre",
            "ubicacion_codigo",
        },
    )

    summary = ImportSummary(
        total_rows=len(rows)
    )

    normalized_rows: list[dict[str, Any]] = []
    codes: set[str] = set()

    for row in rows:
        row_number = int(row["_excel_row"])

        name = normalize_text(
            row.get("ubicacion_nombre"),
            uppercase=True,
            max_length=180,
        )

        code = normalize_destination_code(
            row.get("ubicacion_codigo")
        )

        if not name or not code:
            _append_import_error(
                summary,
                row_number=row_number,
                message=(
                    "Nombre y código de la ubicación "
                    "son obligatorios."
                ),
            )
            continue

        normalized_rows.append({
            "row_number": row_number,
            "name": name,
            "code": code,
            "is_active": normalize_boolean(
                row.get("activo"),
                default=True,
            ),
        })

        codes.add(code)

    existing_by_code = {}

    if codes:
        existing_rows = (
            TicaDestination.query
            .filter(
                TicaDestination.code.in_(codes)
            )
            .all()
        )

        existing_by_code = {
            item.code: item
            for item in existing_rows
        }

    seen_in_file: set[str] = set()

    for row in normalized_rows:
        code = row["code"]

        if code in seen_in_file:
            summary.skipped_rows += 1
            continue

        seen_in_file.add(code)

        destination = existing_by_code.get(
            code
        )

        if destination is None:
            destination = TicaDestination(
                name=row["name"],
                code=code,
                is_active=row["is_active"],
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )

            db.session.add(destination)
            existing_by_code[code] = destination

            summary.created_rows += 1
        else:
            destination.name = row["name"]
            destination.is_active = row["is_active"]
            destination.updated_by_user_id = user_id
            destination.updated_at = datetime.utcnow()

            summary.updated_rows += 1

    return summary


def run_catalog_import(
    *,
    import_type: str,
    file_source: BinaryIO | bytes,
    original_file_name: str,
    user_id: int | None,
) -> tuple[TicaImportBatch, ImportSummary]:
    """
    Ejecuta una carga masiva y actualiza su registro histórico.

    No hace commit. La ruta controla la transacción.
    """
    normalized_type = normalize_text(
        import_type,
        uppercase=True,
        max_length=30,
    )

    if normalized_type not in ALLOWED_IMPORT_TYPES:
        raise TicaImportError(
            "Tipo de importación inválido.",
            code="INVALID_IMPORT_TYPE",
        )

    batch = TicaImportBatch(
        import_type=normalized_type,
        original_file_name=normalize_text(
            original_file_name,
            uppercase=False,
            max_length=255,
        ) or "archivo.xlsx",
        status=IMPORT_STATUS_PROCESSING,
        created_by_user_id=user_id,
        created_at=datetime.utcnow(),
    )

    db.session.add(batch)
    db.session.flush()

    try:
        if normalized_type == IMPORT_TRANSPORTERS:
            summary = import_transporters_from_excel(
                file_source,
                user_id=user_id,
            )

        elif normalized_type == IMPORT_DRIVERS:
            summary = import_drivers_from_excel(
                file_source,
                user_id=user_id,
            )

        else:
            summary = import_destinations_from_excel(
                file_source,
                user_id=user_id,
            )

        batch.total_rows = summary.total_rows
        batch.created_rows = summary.created_rows
        batch.updated_rows = summary.updated_rows
        batch.skipped_rows = summary.skipped_rows
        batch.error_rows = summary.error_rows
        batch.status = summary.status
        batch.result_json = summary.as_dict()
        batch.completed_at = datetime.utcnow()

        return batch, summary

    except Exception as exc:
        batch.status = IMPORT_STATUS_ERROR
        batch.error_message = str(exc)[:4000]
        batch.completed_at = datetime.utcnow()

        raise


# =========================================================
# Plantillas Excel
# =========================================================

def _workbook_to_bytes(workbook: Workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return stream.getvalue()


def build_transporters_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transportistas"

    worksheet.append([
        "transportista_nombre",
        "transportista_cedula",
        "activo",
    ])

    worksheet.append([
        "TRANSPORTES EJEMPLO",
        "3101000000",
        "SI",
    ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:C1"

    content = _workbook_to_bytes(workbook)
    workbook.close()

    return content


def build_drivers_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Choferes"

    worksheet.append([
        "chofer_nombre",
        "chofer_cedula",
        "placa",
        "transportista_cedula",
        "activo",
    ])

    worksheet.append([
        "NOMBRE DEL CHOFER",
        "101110111",
        "ABC123",
        "3101000000",
        "SI",
    ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:E1"

    content = _workbook_to_bytes(workbook)
    workbook.close()

    return content


def build_destinations_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ubicaciones"

    worksheet.append([
        "ubicacion_nombre",
        "ubicacion_codigo",
        "activo",
    ])

    worksheet.append([
        "DESTINO EJEMPLO",
        "H100",
        "SI",
    ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:C1"

    content = _workbook_to_bytes(workbook)
    workbook.close()

    return content