# app/blueprints/dispatch/routes.py

from datetime import datetime, date, time

from flask import render_template, request, redirect, url_for, flash, session, abort, jsonify
from flask_login import login_required, current_user
from app.models.container import Container, ContainerPosition
from app.models.container_classification import ContainerClassification
from app.models.dispatch import DispatchAssignment, UserNotification, GpsDevice, GpsAssignment
from app.blueprints.dispatch import dispatch_bp
from app.models.yard import YardBay
from app.extensions import db
from sqlalchemy.orm import selectinload, joinedload
from app.models.site import Site, UserSite
from app.models.dispatch import (
    DispatchContainerSize,
    ShippingLine,
    DispatchRequest,
    DispatchRequestLine,
)
from io import BytesIO
from datetime import datetime, date, time, timedelta
from flask import send_file
from app.models.eir import EIR
from app.models.chassis import Chassis

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, legal, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from app.services.notifications import create_notifications_for_roles, notification_url
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import traceback


def _allowed_sites_for_user(user):
    if getattr(user, "role", None) == "admin":
        return Site.query.filter_by(is_active=True).order_by(Site.name.asc()).all()

    return (
        db.session.query(Site)
        .join(UserSite, UserSite.site_id == Site.id)
        .filter(UserSite.user_id == user.id, Site.is_active == True)  # noqa: E712
        .order_by(Site.name.asc())
        .all()
    )


def _get_active_site_id():
    return session.get("active_site_id")


def _set_active_site_id(site_id: int):
    session["active_site_id"] = int(site_id)


def _ensure_active_site():
    allowed = _allowed_sites_for_user(current_user)

    if not allowed:
        abort(403)

    active_id = _get_active_site_id()
    allowed_ids = {s.id for s in allowed}

    if not active_id or active_id not in allowed_ids:
        _set_active_site_id(allowed[0].id)
        active_id = allowed[0].id

    return active_id


def _parse_date(value: str):
    value = (value or "").strip()
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def _parse_time(value: str):
    value = (value or "").strip()
    if not value:
        return None
    return datetime.strptime(value, "%H:%M").time()

def _recompute_dispatch_status(req: DispatchRequest):
    lines = DispatchRequestLine.query.filter_by(request_id=req.id).all()

    any_assigned = False
    all_assigned = True

    for line in lines:
        assigned_count = len(line.assignments)
        qty = int(line.quantity or 0)

        if assigned_count <= 0:
            line.status = "PENDIENTE"
            all_assigned = False
        elif assigned_count >= qty:
            line.status = "ASIGNADA"
            any_assigned = True
        else:
            line.status = "PARCIAL"
            any_assigned = True
            all_assigned = False

    if not lines:
        req.status = "PENDIENTE"
    elif all_assigned:
        req.status = "ASIGNADA"
    elif any_assigned:
        req.status = "PARCIAL"
    else:
        req.status = "PENDIENTE"

    req.updated_at = datetime.utcnow()

@dispatch_bp.get("/notifications/<int:notification_id>/read")
@login_required
def read_notification(notification_id: int):
    site_id = _ensure_active_site()

    notification = UserNotification.query.get_or_404(notification_id)

    if notification.user_id != current_user.id:
        abort(403)

    if notification.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    if not notification.is_read:
        notification.is_read = True
        notification.read_at = datetime.utcnow()
        db.session.commit()

    endpoint, params = notification_url(notification)

    if endpoint:
        return redirect(url_for(endpoint, **params))

    return redirect(url_for("dispatch.pending_requests"))


@dispatch_bp.post("/notifications/mark-read")
@login_required
def mark_notifications_read():
    site_id = _ensure_active_site()

    query = UserNotification.query.filter(
        UserNotification.user_id == current_user.id,
        UserNotification.is_read == False,  # noqa: E712
    )

    if site_id:
        query = query.filter(UserNotification.site_id == site_id)

    notifications = query.all()

    now = datetime.utcnow()

    for notification in notifications:
        notification.is_read = True
        notification.read_at = now

    db.session.commit()

    return jsonify({
        "ok": True,
        "marked": len(notifications),
    })


@dispatch_bp.get("/")
@login_required
def index():
    return redirect(url_for("dispatch.pending_requests"))


@dispatch_bp.route("/new", methods=["GET", "POST"])
@login_required
def new_request():
    site_id = _ensure_active_site()

    sizes = (
        DispatchContainerSize.query
        .filter_by(is_active=True)
        .order_by(DispatchContainerSize.sort_order.asc())
        .all()
    )

    shipping_lines = (
        ShippingLine.query
        .filter_by(is_active=True)
        .order_by(ShippingLine.sort_order.asc())
        .all()
    )

    if request.method == "GET":
        return render_template(
            "dispatch/new_request.html",
            sizes=sizes,
            shipping_lines=shipping_lines,
        )

    request_type = (request.form.get("request_type") or "DESPACHO").strip().upper()
    booking = (request.form.get("booking") or "").strip().upper() or None
    shipping_line = (request.form.get("shipping_line") or "").strip().upper()

    client_name = (request.form.get("client_name") or "").strip().upper() or None
    product_name = (request.form.get("product_name") or "").strip().upper() or None

    chassis_type = (request.form.get("chassis_type") or "").strip().upper() or None
    port_out = (request.form.get("port_out") or "").strip().upper() or None
    requires_gps = (request.form.get("requires_gps") or "NO").strip().upper() == "SI"
    special_instructions = (request.form.get("special_instructions") or "").strip().upper() or None

    line_sizes = request.form.getlist("line_size[]")
    line_quantities = request.form.getlist("line_quantity[]")
    line_dates = request.form.getlist("line_date[]")
    line_times = request.form.getlist("line_time[]")
    line_conditions = request.form.getlist("line_condition[]")

    if request_type not in {"DESPACHO", "VACIO"}:
        flash("Tipo de solicitud inválido.", "danger")
        return redirect(url_for("dispatch.new_request"))

    if not shipping_line:
        flash("Debe seleccionar una naviera.", "danger")
        return redirect(url_for("dispatch.new_request"))

    valid_sizes = {s.code for s in sizes}
    valid_shipping_lines = {s.code for s in shipping_lines}

    if shipping_line not in valid_shipping_lines:
        flash("Naviera inválida.", "danger")
        return redirect(url_for("dispatch.new_request"))

    clean_lines = []

    for idx, size_code in enumerate(line_sizes):
        size_code = (size_code or "").strip().upper()

        if not size_code:
            continue

        try:
            quantity = int(line_quantities[idx])
        except Exception:
            quantity = 0

        load_date = _parse_date(line_dates[idx] if idx < len(line_dates) else "")
        load_time = _parse_time(line_times[idx] if idx < len(line_times) else "")

        condition = (
            line_conditions[idx]
            if idx < len(line_conditions)
            else ("VACIO" if request_type == "VACIO" else "CARGADO")
        )
        condition = (condition or "").strip().upper()

        if size_code not in valid_sizes:
            flash(f"Tamaño inválido: {size_code}", "danger")
            return redirect(url_for("dispatch.new_request"))

        if quantity < 1 or quantity > 20:
            flash("La cantidad debe estar entre 1 y 20.", "danger")
            return redirect(url_for("dispatch.new_request"))

        if not load_date:
            flash("Cada línea debe tener fecha de carga.", "danger")
            return redirect(url_for("dispatch.new_request"))

        if condition not in {"CARGADO", "VACIO"}:
            condition = "VACIO" if request_type == "VACIO" else "CARGADO"

        clean_lines.append({
            "container_size": size_code,
            "quantity": quantity,
            "load_date": load_date,
            "load_time": load_time,
            "condition_type": condition,
        })

    if not clean_lines:
        flash("Debe agregar al menos una línea de solicitud.", "danger")
        return redirect(url_for("dispatch.new_request"))

    req = DispatchRequest(
        site_id=site_id,
        request_type=request_type,
        booking=booking,
        shipping_line=shipping_line,
        client_name=client_name,
        product_name=product_name,
        chassis_type=chassis_type,
        port_out=port_out,
        requires_gps=requires_gps,
        special_instructions=special_instructions,
        status="PENDIENTE",
        requested_by_user_id=current_user.id,
    )

    db.session.add(req)
    db.session.flush()

    for line in clean_lines:
        db.session.add(
            DispatchRequestLine(
                request_id=req.id,
                container_size=line["container_size"],
                quantity=line["quantity"],
                load_date=line["load_date"],
                load_time=line["load_time"],
                condition_type=line["condition_type"],
                status="PENDIENTE",
            )
        )

    create_notifications_for_roles(
        site_id=site_id,
        roles={"patio", "inspeccion"},
        title=f"Nueva solicitud #{req.id}",
        message=f"Solicitud {request_type} creada para {shipping_line}.",
        related_type="DISPATCH_REQUEST",
        related_id=req.id,
        exclude_user_ids={current_user.id},
    )

    if requires_gps:
        create_notifications_for_roles(
            site_id=site_id,
            roles={"tracking"},
            title=f"Solicitud GPS #{req.id}",
            message=f"La solicitud #{req.id} requiere GPS.",
            related_type="GPS_REQUEST",
            related_id=req.id,
            exclude_user_ids={current_user.id},
        )

    db.session.commit()

    flash(f"Solicitud #{req.id} creada correctamente.", "success")
    return redirect(url_for("dispatch.pending_requests"))


@dispatch_bp.get("/pending")
@login_required
def pending_requests():
    site_id = _ensure_active_site()

    q = (request.args.get("q") or "").strip().upper()

    next_move_subq = (
        db.session.query(
            DispatchRequestLine.request_id.label("request_id"),
            db.func.min(DispatchRequestLine.load_date).label("next_load_date"),
            db.func.min(DispatchRequestLine.load_time).label("next_load_time"),
        )
        .group_by(DispatchRequestLine.request_id)
        .subquery()
    )

    query = (
        DispatchRequest.query
        .options(
            selectinload(DispatchRequest.lines)
            .selectinload(DispatchRequestLine.assignments)
        )
        .outerjoin(
            next_move_subq,
            next_move_subq.c.request_id == DispatchRequest.id,
        )
        .filter(
            DispatchRequest.site_id == site_id,
            DispatchRequest.status.in_(["PENDIENTE", "PARCIAL"]),
        )
    )

    if q:
        query = query.filter(
            db.func.upper(
                db.func.coalesce(DispatchRequest.booking, "")
            ).like(f"%{q}%")
        )

    requests = (
        query
        .order_by(
            next_move_subq.c.next_load_date.asc().nulls_last(),
            next_move_subq.c.next_load_time.asc().nulls_last(),
            DispatchRequest.requested_at.asc(),
        )
        .all()
    )

    return render_template(
        "dispatch/pending_requests.html",
        requests=requests,
        q=q,
    )
@dispatch_bp.get("/request/<int:request_id>")
@login_required
def request_detail(request_id: int):
    site_id = _ensure_active_site()

    req = (
        DispatchRequest.query
        .options(
            selectinload(DispatchRequest.lines)
            .selectinload(DispatchRequestLine.assignments)
        )
        .filter(DispatchRequest.id == request_id)
        .first_or_404()
    )

    if req.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    latest_classification_subquery = (
        db.session.query(
            ContainerClassification.container_id,
            db.func.max(ContainerClassification.classified_at).label("max_classified_at")
        )
        .filter(ContainerClassification.site_id == site_id)
        .group_by(ContainerClassification.container_id)
        .subquery()
    )

    line_data = []

    for line in req.lines:
        assigned_count = len(line.assignments)
        pending_count = max(int(line.quantity or 0) - assigned_count, 0)

        if req.request_type == "VACIO":
            status_filter = "PARA_EVACUAR"
        else:
            status_filter = "NORMAL"

        available_rows = (
            db.session.query(Container, ContainerPosition, YardBay, ContainerClassification)
            .outerjoin(ContainerPosition, ContainerPosition.container_id == Container.id)
            .outerjoin(YardBay, YardBay.id == ContainerPosition.bay_id)
            .outerjoin(
                latest_classification_subquery,
                latest_classification_subquery.c.container_id == Container.id
            )
            .outerjoin(
                ContainerClassification,
                db.and_(
                    ContainerClassification.container_id == Container.id,
                    ContainerClassification.classified_at == latest_classification_subquery.c.max_classified_at,
                )
            )
            .filter(
                Container.site_id == site_id,
                Container.is_in_yard == True,  # noqa: E712
                Container.size == line.container_size,
                db.func.coalesce(Container.dispatch_status, "NORMAL") == status_filter,
            )
            .order_by(Container.code.asc())
            .all()
        )

        available = []

        for c, pos, bay, cls in available_rows:
            shipping_line = ((cls.shipping_line if cls else "") or "").strip().upper()

            if req.shipping_line and shipping_line and shipping_line != req.shipping_line:
                continue

            available.append({
                "id": c.id,
                "code": c.code,
                "size": c.size,
                "shipping_line": shipping_line or "SIN NAVIERA",
                "gate_in_origin_port": c.gate_in_origin_port or "",
                "classification": ((cls.final_classification if cls else "") or "").strip().upper(),
                "notes": ((cls.summary_text if cls else "") or c.status_notes or "").strip(),
                "dispatch_status": c.dispatch_status or "NORMAL",
                "position": None if not pos else {
                    "bay_code": bay.code if bay else None,
                    "depth_row": pos.depth_row,
                    "tier": pos.tier,
                },
            })

        line_data.append({
            "line": line,
            "assigned_count": assigned_count,
            "pending_count": pending_count,
            "available": available,
            "status_filter": status_filter,
        })

    return render_template(
        "dispatch/request_detail.html",
        req=req,
        line_data=line_data,
    )

@dispatch_bp.post("/request/<int:request_id>/delete")
@login_required
def delete_pending_request(request_id: int):
    site_id = _ensure_active_site()

    allowed_roles = {"admin", "supervision", "despachador"}
    user_role = (getattr(current_user, "role", "") or "").strip().lower()

    if user_role not in allowed_roles:
        abort(403)

    req = (
        DispatchRequest.query
        .options(
            selectinload(DispatchRequest.lines)
            .selectinload(DispatchRequestLine.assignments)
        )
        .filter(DispatchRequest.id == request_id)
        .first_or_404()
    )

    if req.site_id != site_id and user_role != "admin":
        abort(403)

    has_assignments = any(
        len(line.assignments or []) > 0
        for line in req.lines
    )

    if has_assignments:
        flash("No se puede eliminar una solicitud que ya tiene contenedores asignados.", "danger")
        return redirect(url_for("dispatch.pending_requests"))

    if req.status not in {"PENDIENTE"}:
        flash("Solo se pueden eliminar solicitudes pendientes sin asignaciones.", "danger")
        return redirect(url_for("dispatch.pending_requests"))

    request_id_deleted = req.id

    db.session.delete(req)
    db.session.commit()

    flash(f"Solicitud #{request_id_deleted} eliminada correctamente.", "success")
    return redirect(url_for("dispatch.pending_requests"))

@dispatch_bp.post("/request/<int:request_id>/assign/<int:line_id>")
@login_required
def assign_containers(request_id: int, line_id: int):
    site_id = _ensure_active_site()

    req = DispatchRequest.query.get_or_404(request_id)

    if req.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    line = DispatchRequestLine.query.get_or_404(line_id)

    if line.request_id != req.id:
        abort(404)

    selected_ids = request.form.getlist("container_ids[]")

    if not selected_ids:
        flash("Debe seleccionar al menos un contenedor.", "warning")
        return redirect(url_for("dispatch.request_detail", request_id=req.id))

    already_assigned = len(line.assignments)
    pending_count = max(int(line.quantity or 0) - already_assigned, 0)

    if pending_count <= 0:
        flash("Esta línea ya está completamente asignada.", "info")
        return redirect(url_for("dispatch.request_detail", request_id=req.id))

    selected_ids = [int(x) for x in selected_ids if str(x).isdigit()]
    selected_ids = selected_ids[:pending_count]

    request_type = (req.request_type or "DESPACHO").strip().upper()

    if request_type == "VACIO":
        allowed_status = "PARA_EVACUAR"
        next_status = "EVACUAR_SOLICITADO"
    else:
        allowed_status = "NORMAL"
        next_status = "PARA_DESPACHO"

    containers = (
        Container.query
        .filter(
            Container.id.in_(selected_ids),
            Container.site_id == site_id,
            Container.is_in_yard == True,  # noqa: E712
            Container.size == line.container_size,
            db.func.coalesce(Container.dispatch_status, "NORMAL") == allowed_status,
        )
        .all()
    )

    if not containers:
        if request_type == "VACIO":
            flash("No se encontraron contenedores en estado Evacuar para asignar a esta solicitud de vacío.", "danger")
        else:
            flash("No se encontraron contenedores disponibles para asignar a esta solicitud de despacho.", "danger")

        return redirect(url_for("dispatch.request_detail", request_id=req.id))

    assigned_codes = []

    for c in containers:
        assignment_notes = (
            request.form.get(f"assignment_notes_{c.id}") or ""
        ).strip()

        assignment = DispatchAssignment(
            request_line_id=line.id,
            container_id=c.id,
            assigned_by_user_id=current_user.id,
            status="ASIGNADO",
            assignment_notes=assignment_notes or None,
        )

        db.session.add(assignment)

        c.dispatch_status = next_status
        c.dispatch_marked_at = datetime.utcnow()
        c.dispatch_marked_by_user_id = current_user.id

        assigned_codes.append(c.code)

    db.session.flush()

    pending_gps = (
        GpsAssignment.query
        .filter(
            GpsAssignment.site_id == site_id,
            GpsAssignment.dispatch_request_id == req.id,
            GpsAssignment.dispatch_request_line_id == line.id,
            GpsAssignment.status == "ASIGNADO",
            db.or_(
                GpsAssignment.dispatch_assignment_id.is_(None),
                GpsAssignment.container_id.is_(None),
            ),
        )
        .order_by(GpsAssignment.assigned_at.asc(), GpsAssignment.id.asc())
        .first()
    )

    if pending_gps:
        first_assignment = (
            DispatchAssignment.query
            .filter_by(request_line_id=line.id)
            .order_by(DispatchAssignment.assigned_at.asc(), DispatchAssignment.id.asc())
            .first()
        )

        if first_assignment:
            pending_gps.dispatch_assignment_id = first_assignment.id
            pending_gps.container_id = first_assignment.container_id
            pending_gps.chassis_id = first_assignment.chassis_id
            pending_gps.updated_at = datetime.utcnow()

    total_assigned_after = already_assigned + len(containers)

    if total_assigned_after >= int(line.quantity or 0):
        line.status = "ASIGNADA"
    else:
        line.status = "PARCIAL"

    all_lines = DispatchRequestLine.query.filter_by(request_id=req.id).all()
    assigned_lines = [l for l in all_lines if l.status == "ASIGNADA"]
    partial_lines = [l for l in all_lines if l.status == "PARCIAL"]

    if len(assigned_lines) == len(all_lines):
        req.status = "ASIGNADA"
    elif assigned_lines or partial_lines:
        req.status = "PARCIAL"
    else:
        req.status = "PENDIENTE"

    req.updated_at = datetime.utcnow()

    create_notifications_for_roles(
        site_id=site_id,
        roles={"despachador"},
        title=f"Contenedor asignado solicitud #{req.id}",
        message="Se asignaron contenedores: " + ", ".join(assigned_codes),
        related_type="CONTAINER_ASSIGNED",
        related_id=req.id,
        exclude_user_ids={current_user.id},
    )

    db.session.commit()

    

    flash(f"Se asignaron {len(containers)} contenedores correctamente.", "success")
    return redirect(url_for("dispatch.request_detail", request_id=req.id))

@dispatch_bp.get("/assigned")
@login_required
def assigned_requests():
    site_id = _ensure_active_site()

    q = (request.args.get("q") or "").strip().upper()
    container_q = (request.args.get("container_q") or "").strip().upper()

    try:
        page = int(request.args.get("page") or 1)
    except Exception:
        page = 1

    try:
        per_page = int(request.args.get("per_page") or 10)
    except Exception:
        per_page = 10

    if page < 1:
        page = 1

    allowed_per_page = {5, 10, 20, 50, 100}
    if per_page not in allowed_per_page:
        per_page = 10

    query = (
        DispatchRequest.query
        .options(
            selectinload(DispatchRequest.lines)
            .selectinload(DispatchRequestLine.assignments)
            .joinedload(DispatchAssignment.container)
        )
        .filter(
            DispatchRequest.site_id == site_id,
            DispatchRequest.status.in_(["PARCIAL", "ASIGNADA"])
        )
    )

    if q:
        query = query.filter(
            db.func.upper(
                db.func.coalesce(DispatchRequest.booking, "")
            ).like(f"%{q}%")
        )

    if container_q:
        query = (
            query
            .join(
                DispatchRequestLine,
                DispatchRequestLine.request_id == DispatchRequest.id
            )
            .join(
                DispatchAssignment,
                DispatchAssignment.request_line_id == DispatchRequestLine.id
            )
            .join(
                Container,
                Container.id == DispatchAssignment.container_id
            )
            .filter(
                db.func.upper(Container.code).like(f"%{container_q}%")
            )
            .distinct()
        )

    pagination = (
        query
        .order_by(
            DispatchRequest.updated_at.desc(),
            DispatchRequest.requested_at.desc()
        )
        .paginate(
            page=page,
            per_page=per_page,
            error_out=False,
        )
    )

    requests = pagination.items
    request_ids = [r.id for r in requests]

    gps_rows = []

    if request_ids:
        gps_rows = (
            GpsAssignment.query
            .options(joinedload(GpsAssignment.gps_device))
            .filter(
                GpsAssignment.site_id == site_id,
                GpsAssignment.dispatch_request_id.in_(request_ids),
                GpsAssignment.status == "ASIGNADO",
            )
            .all()
        )

    gps_by_assignment_id = {}
    gps_by_line_id = {}

    for gps in gps_rows:
        if gps.dispatch_assignment_id:
            gps_by_assignment_id[gps.dispatch_assignment_id] = gps

        if gps.dispatch_request_line_id:
            gps_by_line_id[gps.dispatch_request_line_id] = gps

    return render_template(
        "dispatch/assigned_requests.html",
        requests=requests,
        q=q,
        container_q=container_q,
        per_page=per_page,
        pagination=pagination,
        gps_by_assignment_id=gps_by_assignment_id,
        gps_by_line_id=gps_by_line_id,
    )

@dispatch_bp.get("/agenda")
@login_required
def agenda():
    site_id = _ensure_active_site()

    import pytz

    cr_tz = pytz.timezone("America/Costa_Rica")
    today = datetime.now(cr_tz).date()

    lines = (
        DispatchRequestLine.query
        .options(
            joinedload(DispatchRequestLine.request),
            selectinload(DispatchRequestLine.assignments)
            .joinedload(DispatchAssignment.container)
        )
        .join(
            DispatchRequest,
            DispatchRequest.id == DispatchRequestLine.request_id
        )
        .filter(
            DispatchRequest.site_id == site_id,
            DispatchRequest.status != "CANCELADA",
            DispatchRequestLine.load_date >= today
        )
        .order_by(
            DispatchRequestLine.load_date.asc(),
            DispatchRequestLine.load_time.asc()
        )
        .all()
    )

    return render_template(
        "dispatch/agenda.html",
        lines=lines,
    )

@dispatch_bp.get("/prelist")
@login_required
def prelist():
    site_id = _ensure_active_site()

    from datetime import timedelta
    import pytz

    cr_tz = pytz.timezone("America/Costa_Rica")
    now_cr = datetime.now(cr_tz)

    today = now_cr.date()
    tomorrow = today + timedelta(days=1)
    current_time = now_cr.time()

    lines = (
        DispatchRequestLine.query
        .options(
            joinedload(DispatchRequestLine.request),
            selectinload(DispatchRequestLine.assignments)
            .joinedload(DispatchAssignment.container)
        )
        .join(
            DispatchRequest,
            DispatchRequest.id == DispatchRequestLine.request_id
        )
        .filter(
            DispatchRequest.site_id == site_id,
            DispatchRequest.status != "CANCELADA",
            DispatchRequestLine.load_date.in_([today, tomorrow])
        )
        .order_by(
            DispatchRequestLine.load_date.asc(),
            DispatchRequestLine.load_time.asc()
        )
        .all()
    )

    prelist_lines = []

    for line in lines:
        if line.load_date == tomorrow:
            prelist_lines.append(line)
            continue

        if line.load_date == today:
            if line.load_time is None or line.load_time > current_time:
                prelist_lines.append(line)

    prelist_lines.sort(
        key=lambda x: (
            x.load_date,
            x.load_time or datetime.min.time()
        )
    )

    dispatch_lines = [
        line for line in prelist_lines
        if ((line.request.request_type or "").strip().upper() == "DESPACHO")
    ]

    empty_lines = [
        line for line in prelist_lines
        if ((line.request.request_type or "").strip().upper() == "VACIO")
    ]

    return render_template(
        "dispatch/prelist.html",
        lines=prelist_lines,
        dispatch_lines=dispatch_lines,
        empty_lines=empty_lines,
    )

@dispatch_bp.post("/assignment/<int:assignment_id>/release")
@login_required
def release_assignment(assignment_id: int):
    site_id = _ensure_active_site()

    assignment = DispatchAssignment.query.get_or_404(assignment_id)
    line = DispatchRequestLine.query.get_or_404(assignment.request_line_id)
    req = DispatchRequest.query.get_or_404(line.request_id)

    if req.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    container = Container.query.get_or_404(assignment.container_id)

    request_type = (req.request_type or "DESPACHO").strip().upper()

    if request_type == "VACIO":
        container.dispatch_status = "PARA_EVACUAR"
    else:
        container.dispatch_status = "NORMAL"

    container.dispatch_marked_at = None
    container.dispatch_marked_by_user_id = None
    container.mounted_at = None
    container.mounted_by_user_id = None

    if hasattr(container, "is_mounted"):
        container.is_mounted = False

    db.session.delete(assignment)
    db.session.flush()

    _recompute_dispatch_status(req)

    db.session.commit()

    flash(f"Contenedor {container.code} liberado correctamente.", "success")
    return redirect(url_for("dispatch.assigned_requests"))

@dispatch_bp.post("/line/<int:line_id>/release-pending")
@login_required
def release_pending_line(line_id: int):
    site_id = _ensure_active_site()

    line = DispatchRequestLine.query.get_or_404(line_id)
    req = DispatchRequest.query.get_or_404(line.request_id)

    if req.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    assigned_count = len(line.assignments)
    qty = int(line.quantity or 0)
    pending_count = max(qty - assigned_count, 0)

    if pending_count <= 0:
        flash("Esta línea no tiene pendientes por liberar.", "info")
        return redirect(url_for("dispatch.assigned_requests"))

    line.quantity = qty - 1

    if line.quantity <= 0:
        db.session.delete(line)
        db.session.flush()

    _recompute_dispatch_status(req)

    db.session.commit()

    flash("Pendiente liberado correctamente.", "success")
    return redirect(url_for("dispatch.assigned_requests"))

@dispatch_bp.post("/assignment/<int:assignment_id>/reschedule")
@login_required
def reschedule_assignment(assignment_id: int):
    site_id = _ensure_active_site()

    assignment = DispatchAssignment.query.get_or_404(assignment_id)
    line = DispatchRequestLine.query.get_or_404(assignment.request_line_id)
    req = DispatchRequest.query.get_or_404(line.request_id)

    if req.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    new_date = _parse_date(request.form.get("load_date") or "")
    new_time = _parse_time(request.form.get("load_time") or "")

    if not new_date:
        flash("Debe indicar la nueva fecha.", "danger")
        return redirect(url_for("dispatch.assigned_requests"))

    assigned_count = len(line.assignments)

    if assigned_count <= 1 and int(line.quantity or 0) <= 1:
        line.load_date = new_date
        line.load_time = new_time
    else:
        line.quantity = max(int(line.quantity or 1) - 1, 1)

        new_line = DispatchRequestLine(
            request_id=req.id,
            container_size=line.container_size,
            quantity=1,
            load_date=new_date,
            load_time=new_time,
            condition_type=line.condition_type,
            status="ASIGNADA",
        )

        db.session.add(new_line)
        db.session.flush()

        assignment.request_line_id = new_line.id

    req.updated_at = datetime.utcnow()

    _recompute_dispatch_status(req)

    db.session.commit()

    flash("Asignación reagendada correctamente.", "success")
    return redirect(url_for("dispatch.assigned_requests"))

@dispatch_bp.get("/prelist/pdf")
@login_required
def prelist_pdf():
    site_id = _ensure_active_site()

    import pytz

    cr_tz = pytz.timezone("America/Costa_Rica")
    now_cr = datetime.now(cr_tz)

    today = now_cr.date()
    tomorrow = today + timedelta(days=1)
    current_time = now_cr.time()

    active_site = Site.query.get(site_id)
    site_name = active_site.name if active_site else f"Predio {site_id}"

    lines = (
        DispatchRequestLine.query
        .options(
            joinedload(DispatchRequestLine.request),
            selectinload(DispatchRequestLine.assignments)
            .joinedload(DispatchAssignment.container)
        )
        .join(
            DispatchRequest,
            DispatchRequest.id == DispatchRequestLine.request_id
        )
        .filter(
            DispatchRequest.site_id == site_id,
            DispatchRequest.status != "CANCELADA",
            DispatchRequest.request_type == "DESPACHO",
            DispatchRequestLine.load_date.in_([today, tomorrow])
        )
        .order_by(
            DispatchRequestLine.load_date.asc(),
            DispatchRequestLine.load_time.asc().nulls_last(),
            DispatchRequest.shipping_line.asc(),
            DispatchRequest.booking.asc().nulls_last(),
        )
        .all()
    )

    prelist_lines = []

    for line in lines:
        if line.load_date == tomorrow:
            prelist_lines.append(line)
            continue

        if line.load_date == today:
            if line.load_time is None or line.load_time > current_time:
                prelist_lines.append(line)

    request_ids = []
    line_ids = []
    assignment_ids = []
    container_ids = []
    chassis_ids = set()

    for line in prelist_lines:
        if line.request_id:
            request_ids.append(line.request_id)

        line_ids.append(line.id)

        for a in line.assignments:
            assignment_ids.append(a.id)

            if a.container_id:
                container_ids.append(a.container_id)

            if a.chassis_id:
                chassis_ids.add(a.chassis_id)

    gps_rows = []

    if request_ids:
        gps_rows = (
            GpsAssignment.query
            .options(joinedload(GpsAssignment.gps_device))
            .filter(
                GpsAssignment.site_id == site_id,
                GpsAssignment.dispatch_request_id.in_(request_ids),
                GpsAssignment.status == "ASIGNADO",
            )
            .all()
        )

    gps_by_assignment_id = {}
    gps_by_line_id = {}

    for gps in gps_rows:
        if gps.dispatch_assignment_id:
            gps_by_assignment_id[gps.dispatch_assignment_id] = gps

        if gps.dispatch_request_line_id:
            gps_by_line_id[gps.dispatch_request_line_id] = gps

    eirs = []

    if container_ids:
        eirs = (
            EIR.query
            .filter(
                EIR.site_id == site_id,
                EIR.container_id.in_(container_ids),
                EIR.chassis_id.isnot(None),
                EIR.status.in_(["PENDING", "CONFIRMED"]),
            )
            .order_by(EIR.container_id.asc(), EIR.id.desc())
            .all()
        )

    eir_by_container = {}

    for eir in eirs:
        if eir.container_id not in eir_by_container:
            eir_by_container[eir.container_id] = eir

        if eir.chassis_id:
            chassis_ids.add(eir.chassis_id)

    chassis_by_id = {}

    if chassis_ids:
        chassis_rows = (
            Chassis.query
            .filter(Chassis.id.in_(list(chassis_ids)))
            .all()
        )

        chassis_by_id = {
            ch.id: ch
            for ch in chassis_rows
        }

    def _container_type(size_value):
        size_value = (size_value or "").strip().upper()

        if size_value.startswith("20"):
            return "20"

        if size_value.startswith("45"):
            return "45"

        if size_value.startswith("40"):
            return "40"

        return size_value or ""

    def _format_from_request_or_chassis(tipo, req, chassis=None):
        axles = None

        if chassis and getattr(chassis, "axles", None):
            axles = chassis.axles
        else:
            chassis_type = (getattr(req, "chassis_type", "") or "").strip().upper()

            if "2" in chassis_type:
                axles = 2
            elif "3" in chassis_type:
                axles = 3

        if tipo and axles:
            return f"{tipo}x{axles}"

        return ""

    def _format_date_no_year(value):
        if not value:
            return ""

        return value.strftime("%d/%m")

    def _format_time(value):
        if not value:
            return ""

        return value.strftime("%I:%M %p")

    def _short_site(value):
        value = (value or "").strip().upper()

        if "COYOL" in value:
            return "COYOL"

        if "CALDERA" in value:
            return "CALDERA"

        if "LIMON" in value or "LIMÓN" in value:
            return "LIMON"

        return value[:10]

    

    def _gps_text(req, line, assignment=None):
        gps = None

        if assignment:
            gps = gps_by_assignment_id.get(assignment.id)

        if not gps:
            gps = gps_by_line_id.get(line.id)

        if gps and gps.gps_device:
            return gps.gps_device.gps_number or "SI"

        if getattr(req, "requires_gps", False):
            return "SI"

        return "NO"

    def _detail_text(req):
        request_type = (req.request_type or "").strip().upper()

        if request_type == "DESPACHO":
            return "CARGA"

        if request_type == "VACIO":
            return "VACIO"

        return request_type or "—"

    data = [[
        "PREDIO",
        "NAVIERA",
        "CONTENEDOR",
        "CHASIS",
        "TIPO",
        "FORMATO",
        "FECHA",
        "HORA",
        "GPS",
        "CLIENTE / PLANTA",
        "PRODUCTO",
        "DESTINO",
        "COMENTARIO",
        "DETALLES",
    ]]

    tomorrow_after_11_rows = []

    for line in prelist_lines:
        req = line.request
        assignments = list(line.assignments or [])

        for a in assignments:
            container = a.container
            eir = eir_by_container.get(a.container_id)

            chassis_id = a.chassis_id or (eir.chassis_id if eir else None)
            chassis = chassis_by_id.get(chassis_id) if chassis_id else None

            container_size = container.size if container else line.container_size
            tipo = _container_type(container_size)

            formato = _format_from_request_or_chassis(tipo, req, chassis)

            row = [
                _short_site(site_name),
                req.shipping_line or "",
                container.code if container else "",
                chassis.chassis_number if chassis else "",
                container_size or "",
                formato,
                _format_date_no_year(line.load_date),
                _format_time(line.load_time),
                _gps_text(req, line, a),
                req.client_name or "",
                req.product_name or "",
                req.port_out or "",
                a.assignment_notes or "",
                _detail_text(req),
            ]

            data.append(row)

            is_tomorrow_after_11 = (
                line.load_date == tomorrow
                and line.load_time is not None
                and line.load_time >= time(11, 0)
            )

            if is_tomorrow_after_11:
                tomorrow_after_11_rows.append(len(data) - 1)

        assigned_count = len(assignments)
        pending_count = max(int(line.quantity or 0) - assigned_count, 0)

        for _ in range(pending_count):
            pending_tipo = _container_type(line.container_size)
            pending_formato = _format_from_request_or_chassis(pending_tipo, req, None)

            row = [
                _short_site(site_name),
                req.shipping_line or "",
                "",
                "",
                line.container_size or "",
                pending_formato,
                _format_date_no_year(line.load_date),
                _format_time(line.load_time),
                _gps_text(req, line, None),
                req.client_name or "",
                req.product_name or "",
                req.port_out or "",
                "PENDIENTE DE ASIGNAR",
                _detail_text(req),
            ]

            data.append(row)

            is_tomorrow_after_11 = (
                line.load_date == tomorrow
                and line.load_time is not None
                and line.load_time >= time(11, 0)
            )

            if is_tomorrow_after_11:
                tomorrow_after_11_rows.append(len(data) - 1)

    if len(data) == 1:
        data.append([
            _short_site(site_name),
            "—",
            "—",
            "—",
            "—",
            "—",
            "—",
            "—",
            "NO",
            "Sin registros",
            "—",
            "—",
            "—",
            "—",
        ])

    bio = BytesIO()

    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(legal),
        rightMargin=4,
        leftMargin=4,
        topMargin=6,
        bottomMargin=6,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Normal"]
    title_style.fontName = "Helvetica-Bold"
    title_style.fontSize = 12
    title_style.leading = 13

    small_style = styles["Normal"]
    small_style.fontSize = 8
    small_style.leading = 9

    elements = []

    printed_at = now_cr.strftime("%d/%m/%Y %I:%M %p")

    elements.append(Paragraph("<b>PRELISTA OPERATIVA</b>", title_style))
    elements.append(Paragraph(f"Impreso: {printed_at}", small_style))
    elements.append(Spacer(1, 4))

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            44,   # Predio
            45,   # Naviera
            72,   # Contenedor
            64,   # Chasis
            40,   # Tipo
            44,   # Formato
            38,   # Fecha
            52,   # Hora
            55,   # GPS
            116,  # Cliente / Planta
            112,  # Producto
            112,  # Destino
            130,  # Comentario
            58,   # Detalles
        ],
    )

    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.black),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.3),

        ("FONTSIZE", (0, 1), (-1, -1), 6.8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),

        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (9, 1), (12, -1), "LEFT"),

        ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),

        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
            colors.white,
            colors.HexColor("#F3F4F6"),
        ]),
    ]

    for row_idx in tomorrow_after_11_rows:
        table_style.extend([
            ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.black),
            ("TEXTCOLOR", (0, row_idx), (-1, row_idx), colors.white),
            ("FONTNAME", (0, row_idx), (-1, row_idx), "Helvetica-Bold"),
        ])

    table.setStyle(TableStyle(table_style))

    elements.append(table)
    doc.build(elements)

    bio.seek(0)

    response = send_file(
        bio,
        as_attachment=False,
        download_name="prelista_operativa.pdf",
        mimetype="application/pdf",
    )

    response.headers["Content-Disposition"] = "inline; filename=prelista_operativa.pdf"

    return response

@dispatch_bp.get("/gps")
@login_required
def gps_dashboard():
    return redirect(url_for("dispatch.gps_inventory"))


@dispatch_bp.get("/gps/inventory")
@login_required
def gps_inventory():
    site_id = _ensure_active_site()

    gps_rows = (
        GpsDevice.query
        .filter(GpsDevice.site_id == site_id)
        .order_by(GpsDevice.gps_number.asc())
        .all()
    )

    return render_template(
        "dispatch/gps_inventory.html",
        gps_rows=gps_rows,
    )


@dispatch_bp.post("/gps/inventory/create")
@login_required
def gps_create_device():
    site_id = _ensure_active_site()

    gps_number = (request.form.get("gps_number") or "").strip().upper()
    battery_range = (request.form.get("battery_range") or "").strip()
    notes = (request.form.get("notes") or "").strip().upper() or None

    if not gps_number:
        flash("Debe indicar el número de GPS.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    existing = GpsDevice.query.filter_by(gps_number=gps_number).first()
    if existing:
        flash("Ese GPS ya existe en el inventario.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    gps = GpsDevice(
        site_id=site_id,
        gps_number=gps_number,
        status="DISPONIBLE",
        battery_range=battery_range or None,
        notes=notes,
        is_active=True,
    )

    db.session.add(gps)
    db.session.commit()

    flash(f"GPS {gps_number} agregado correctamente.", "success")
    return redirect(url_for("dispatch.gps_inventory"))

@dispatch_bp.get("/gps/inventory/template")
@login_required
def gps_inventory_template():
    _ensure_active_site()

    wb = Workbook()
    ws = wb.active
    ws.title = "GPS"

    headers = [
        "gps_number",
        "current_location",
        "battery_range",
        "status",
        "is_active",
        "notes",
    ]

    ws.append(headers)

    examples = [
        ["GPS001", "BODEGA", "80-100", "DISPONIBLE", "SI", "GPS NUEVO"],
        ["GPS002", "TALLER", "40-60", "MANTENIMIENTO", "SI", "REVISAR BATERÍA"],
    ]

    for row in examples:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25   # location
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 45

    status_validation = DataValidation(
        type="list",
        formula1='"DISPONIBLE,MANTENIMIENTO,FUERA_SERVICIO"',
        allow_blank=True,
    )
    active_validation = DataValidation(
        type="list",
        formula1='"SI,NO"',
        allow_blank=True,
    )

    ws.add_data_validation(status_validation)
    ws.add_data_validation(active_validation)

    status_validation.add("D2:D500")
    active_validation.add("E2:E500")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="plantilla_carga_gps.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@dispatch_bp.post("/gps/inventory/bulk-upload")
@login_required
def gps_inventory_bulk_upload():
    site_id = _ensure_active_site()

    file = request.files.get("gps_file")

    if not file or not file.filename:
        flash("Debe seleccionar un archivo Excel.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    if not file.filename.lower().endswith(".xlsx"):
        flash("El archivo debe ser formato .xlsx.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        flash("No se pudo leer el archivo Excel. Verifique que sea una plantilla válida.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    expected_headers = {
        "gps_number",
        "current_location",
        "battery_range",
        "status",
        "is_active",
        "notes",
    }

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)

    if not first_row:
        flash("La plantilla está vacía.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    header_map = {}

    for idx, value in enumerate(first_row):
        header = (str(value or "").strip().lower())
        if header:
            header_map[header] = idx

    if "gps_number" not in header_map:
        flash("La plantilla debe contener la columna obligatoria gps_number.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    unknown_headers = set(header_map.keys()) - expected_headers

    if unknown_headers:
        flash(
            "La plantilla contiene columnas no permitidas: "
            + ", ".join(sorted(unknown_headers)),
            "danger",
        )
        return redirect(url_for("dispatch.gps_inventory"))

    valid_statuses = {"DISPONIBLE", "MANTENIMIENTO", "FUERA_DE_SERVICIO"}

    rows_to_process = []
    seen_gps_numbers = set()
    errors = []
    skipped_count = 0

    def get_value(row_values, header):
        idx = header_map.get(header)

        if idx is None:
            return None

        if idx >= len(row_values):
            return None

        value = row_values[idx]

        if value is None:
            return None

        return str(value).strip()

    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        gps_number = (get_value(row_values, "gps_number") or "").strip().upper()

        if not gps_number:
            skipped_count += 1
            continue

        if gps_number in seen_gps_numbers:
            errors.append(f"Fila {row_idx}: GPS duplicado dentro del archivo ({gps_number}).")
            skipped_count += 1
            continue

        seen_gps_numbers.add(gps_number)

        current_location = get_value(row_values, "current_location")
        battery_range = get_value(row_values, "battery_range")
        status = (get_value(row_values, "status") or "").strip().upper()
        is_active_raw = (get_value(row_values, "is_active") or "").strip().upper()
        notes = get_value(row_values, "notes")

        if status and status not in valid_statuses:
            errors.append(
                f"Fila {row_idx}: estado inválido para {gps_number}. "
                "Use DISPONIBLE, MANTENIMIENTO o FUERA_DE_SERVICIO."
            )
            skipped_count += 1
            continue

        is_active_value = None

        if is_active_raw:
            if is_active_raw in {"SI", "SÍ", "TRUE", "1", "ACTIVO"}:
                is_active_value = True
            elif is_active_raw in {"NO", "FALSE", "0", "INACTIVO"}:
                is_active_value = False
            else:
                errors.append(f"Fila {row_idx}: is_active inválido para {gps_number}. Use SI o NO.")
                skipped_count += 1
                continue

        rows_to_process.append({
            "row_idx": row_idx,
            "gps_number": gps_number,
            "current_location": current_location,
            "battery_range": battery_range,
            "status": status,
            "is_active": is_active_value,
            "notes": notes,
        })

    if errors:
        flash(
            "No se realizó la carga porque hay errores en la plantilla. "
            "Corrija el archivo e intente nuevamente.",
            "danger",
        )

        for error in errors[:10]:
            flash(error, "warning")

        if len(errors) > 10:
            flash(f"Hay {len(errors) - 10} errores adicionales.", "warning")

        return redirect(url_for("dispatch.gps_inventory"))

    gps_numbers = [row["gps_number"] for row in rows_to_process]

    existing_gps_rows = []

    if gps_numbers:
        existing_gps_rows = (
            GpsDevice.query
            .filter(
                GpsDevice.site_id == site_id,
                GpsDevice.gps_number.in_(gps_numbers),
            )
            .all()
        )

    gps_by_number = {
        gps.gps_number: gps
        for gps in existing_gps_rows
    }

    existing_gps_ids = [
        gps.id
        for gps in existing_gps_rows
    ]

    assigned_gps_ids = set()

    if existing_gps_ids:
        assigned_rows = (
            GpsAssignment.query
            .filter(
                GpsAssignment.gps_device_id.in_(existing_gps_ids),
                GpsAssignment.status == "ASIGNADO",
            )
            .all()
        )

        assigned_gps_ids = {
            row.gps_device_id
            for row in assigned_rows
        }

    created_count = 0
    updated_count = 0

    try:
        for row in rows_to_process:
            gps_number = row["gps_number"]
            gps = gps_by_number.get(gps_number)

            if gps:
                if row["status"]:
                    if gps.id in assigned_gps_ids and row["status"] != "DISPONIBLE":
                        errors.append(
                            f"Fila {row['row_idx']}: GPS {gps_number} está asignado. "
                            "No se puede cambiar a mantenimiento o fuera de servicio."
                        )
                        continue

                    gps.status = row["status"]

                if row["current_location"] is not None:
                    gps.current_location = row["current_location"].upper() or None

                if row["battery_range"] is not None:
                    gps.battery_range = row["battery_range"] or None

                if row["notes"] is not None:
                    gps.notes = row["notes"].upper() or None

                if row["is_active"] is not None:
                    gps.is_active = row["is_active"]

                gps.updated_at = datetime.utcnow()
                updated_count += 1

            else:
                gps = GpsDevice(
                    site_id=site_id,
                    gps_number=gps_number,
                    current_location=(row["current_location"].upper() if row["current_location"] else None),
                    status=row["status"] or "DISPONIBLE",
                    battery_range=row["battery_range"] or None,
                    notes=(row["notes"].upper() if row["notes"] else None),
                    is_active=True if row["is_active"] is None else row["is_active"],
                )

                db.session.add(gps)
                gps_by_number[gps_number] = gps
                created_count += 1

        if errors:
            db.session.rollback()

            flash(
                "No se realizó la carga porque hay errores en la plantilla. "
                "Corrija el archivo e intente nuevamente.",
                "danger",
            )

            for error in errors[:10]:
                flash(error, "warning")

            if len(errors) > 10:
                flash(f"Hay {len(errors) - 10} errores adicionales.", "warning")

            return redirect(url_for("dispatch.gps_inventory"))

        db.session.commit()

    except Exception as e:
        db.session.rollback()

        traceback.print_exc()

        print("=" * 80)
        print(type(e).__name__)
        print(e)
        print("=" * 80)

        flash(str(e), "danger")

        return redirect(url_for("dispatch.gps_inventory"))
    flash(
        f"Carga masiva completada. Creados: {created_count}. "
        f"Actualizados: {updated_count}. Filas omitidas: {skipped_count}.",
        "success",
    )

    return redirect(url_for("dispatch.gps_inventory"))

@dispatch_bp.post("/gps/device/<int:gps_id>/update")
@login_required
def gps_update_device(gps_id: int):
    site_id = _ensure_active_site()

    gps = GpsDevice.query.get_or_404(gps_id)

    if gps.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    battery_range = (request.form.get("battery_range") or "").strip()
    notes = (request.form.get("notes") or "").strip().upper() or None

    gps.battery_range = battery_range or None
    gps.notes = notes
    gps.updated_at = datetime.utcnow()

    db.session.commit()

    flash(f"GPS {gps.gps_number} actualizado.", "success")
    return redirect(url_for("dispatch.gps_inventory"))


@dispatch_bp.post("/gps/device/<int:gps_id>/status")
@login_required
def gps_change_status(gps_id: int):
    site_id = _ensure_active_site()

    gps = GpsDevice.query.get_or_404(gps_id)

    if gps.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    new_status = (request.form.get("status") or "").strip().upper()

    if new_status not in {"DISPONIBLE", "MANTENIMIENTO", "FUERA_SERVICIO"}:
        flash("Estado inválido para GPS.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    active_assignment = (
        GpsAssignment.query
        .filter(
            GpsAssignment.gps_device_id == gps.id,
            GpsAssignment.status == "ASIGNADO",
        )
        .first()
    )

    if active_assignment and new_status != "DISPONIBLE":
        flash("No se puede cambiar a mantenimiento o fuera de servicio un GPS asignado.", "danger")
        return redirect(url_for("dispatch.gps_inventory"))

    gps.status = new_status
    gps.updated_at = datetime.utcnow()

    db.session.commit()

    flash(f"GPS {gps.gps_number} cambiado a {new_status}.", "success")
    return redirect(url_for("dispatch.gps_inventory"))


@dispatch_bp.get("/gps/requests")
@login_required
def gps_requests():
    site_id = _ensure_active_site()

    lines = (
        DispatchRequestLine.query
        .options(
            joinedload(DispatchRequestLine.request),
            selectinload(DispatchRequestLine.assignments)
            .joinedload(DispatchAssignment.container),
        )
        .join(DispatchRequest, DispatchRequest.id == DispatchRequestLine.request_id)
        .filter(
            DispatchRequest.site_id == site_id,
            DispatchRequest.requires_gps == True,  # noqa: E712
            DispatchRequest.status != "CANCELADA",
        )
        .order_by(
            DispatchRequestLine.load_date.asc(),
            DispatchRequestLine.load_time.asc().nulls_last(),
            DispatchRequest.booking.asc().nulls_last(),
        )
        .all()
    )

    active_gps_rows = (
        GpsAssignment.query
        .filter(
            GpsAssignment.site_id == site_id,
            GpsAssignment.status == "ASIGNADO",
        )
        .all()
    )

    assigned_line_ids = {
        row.dispatch_request_line_id
        for row in active_gps_rows
        if row.dispatch_request_line_id
    }

    available_gps = (
        GpsDevice.query
        .filter(
            GpsDevice.site_id == site_id,
            GpsDevice.status == "DISPONIBLE",
            GpsDevice.is_active == True,  # noqa: E712
        )
        .order_by(GpsDevice.gps_number.asc())
        .all()
    )

    pending_lines = [
        line for line in lines
        if line.id not in assigned_line_ids
    ]

    return render_template(
        "dispatch/gps_requests.html",
        lines=pending_lines,
        available_gps=available_gps,
    )


@dispatch_bp.post("/gps/assign/<int:line_id>")
@login_required
def gps_assign(line_id: int):
    site_id = _ensure_active_site()

    line = DispatchRequestLine.query.get_or_404(line_id)
    req = DispatchRequest.query.get_or_404(line.request_id)

    if req.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    if not getattr(req, "requires_gps", False):
        flash("Esta solicitud no requiere GPS.", "danger")
        return redirect(url_for("dispatch.gps_requests"))

    gps_id = request.form.get("gps_device_id", type=int)

    if not gps_id:
        flash("Debe seleccionar un GPS.", "danger")
        return redirect(url_for("dispatch.gps_requests"))

    gps = GpsDevice.query.get_or_404(gps_id)

    if gps.site_id != site_id:
        flash("El GPS seleccionado no pertenece al predio activo.", "danger")
        return redirect(url_for("dispatch.gps_requests"))

    if gps.status != "DISPONIBLE" or not gps.is_active:
        flash("Solo se puede asignar un GPS disponible.", "danger")
        return redirect(url_for("dispatch.gps_requests"))

    existing = (
        GpsAssignment.query
        .filter(
            GpsAssignment.dispatch_request_line_id == line.id,
            GpsAssignment.status == "ASIGNADO",
        )
        .first()
    )

    if existing:
        flash("Esta línea ya tiene un GPS asignado.", "warning")
        return redirect(url_for("dispatch.gps_requests"))

    first_assignment = None
    assignments = list(line.assignments or [])

    if assignments:
        first_assignment = assignments[0]

    gps_assignment = GpsAssignment(
        site_id=site_id,
        dispatch_request_id=req.id,
        dispatch_request_line_id=line.id,
        dispatch_assignment_id=first_assignment.id if first_assignment else None,
        gps_device_id=gps.id,
        container_id=first_assignment.container_id if first_assignment else None,
        chassis_id=first_assignment.chassis_id if first_assignment else None,
        status="ASIGNADO",
        assigned_by_user_id=current_user.id,
    )

    gps.status = "ASIGNADO"
    gps.updated_at = datetime.utcnow()

    db.session.add(gps_assignment)
    db.session.flush()

    container_code = ""
    if first_assignment and first_assignment.container:
        container_code = first_assignment.container.code or ""

    message = f"GPS {gps.gps_number} asignado a solicitud #{req.id}."
    if container_code:
        message += f" Contenedor: {container_code}."

    create_notifications_for_roles(
        site_id=site_id,
        roles={"supervision", "despachador"},
        title=f"GPS asignado solicitud #{req.id}",
        message=message,
        related_type="GPS_ASSIGNED",
        related_id=req.id,
        exclude_user_ids={current_user.id},
    )

    db.session.commit()

    flash(f"GPS {gps.gps_number} asignado correctamente.", "success")
    return redirect(url_for("dispatch.gps_requests"))


@dispatch_bp.get("/gps/assigned")
@login_required
def gps_assigned():
    site_id = _ensure_active_site()

    rows = (
        GpsAssignment.query
        .options(
            joinedload(GpsAssignment.gps_device),
            joinedload(GpsAssignment.request),
            joinedload(GpsAssignment.line),
            joinedload(GpsAssignment.assignment)
            .joinedload(DispatchAssignment.container),
        )
        .filter(
            GpsAssignment.site_id == site_id,
            GpsAssignment.status == "ASIGNADO",
        )
        .order_by(GpsAssignment.assigned_at.desc())
        .all()
    )

    chassis_ids = {
        row.chassis_id
        for row in rows
        if row.chassis_id
    }

    chassis_by_id = {}

    if chassis_ids:
        chassis_rows = (
            Chassis.query
            .filter(Chassis.id.in_(list(chassis_ids)))
            .all()
        )

        chassis_by_id = {
            ch.id: ch
            for ch in chassis_rows
        }

    return render_template(
        "dispatch/gps_assigned.html",
        rows=rows,
        chassis_by_id=chassis_by_id,
    )


@dispatch_bp.post("/gps/assignment/<int:gps_assignment_id>/release")
@login_required
def gps_release_assignment(gps_assignment_id: int):
    site_id = _ensure_active_site()

    gps_assignment = GpsAssignment.query.get_or_404(gps_assignment_id)

    if gps_assignment.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    if gps_assignment.status != "ASIGNADO":
        flash("Esta asignación GPS ya no está activa.", "warning")
        return redirect(url_for("dispatch.gps_assigned"))

    gps = GpsDevice.query.get_or_404(gps_assignment.gps_device_id)

    gps_assignment.status = "LIBERADO"
    gps_assignment.released_at = datetime.utcnow()
    gps_assignment.updated_at = datetime.utcnow()

    gps.status = "DISPONIBLE"
    gps.updated_at = datetime.utcnow()

    db.session.commit()

    flash(f"GPS {gps.gps_number} liberado correctamente.", "success")
    return redirect(url_for("dispatch.gps_assigned"))

@dispatch_bp.post("/request/<int:request_id>/reschedule")
@login_required
def reschedule_pending_request(request_id: int):
    site_id = _ensure_active_site()

    role = (current_user.role or "").strip().lower()
    if role not in {"admin", "supervision", "despachador"}:
        abort(403)

    req = DispatchRequest.query.get_or_404(request_id)

    if req.site_id != site_id and role != "admin":
        abort(403)

    if req.status != "PENDIENTE":
        flash("Solo se pueden reagendar solicitudes pendientes.", "warning")
        return redirect(url_for("dispatch.pending_requests"))

    new_date_raw = (request.form.get("load_date") or "").strip()
    new_time_raw = (request.form.get("load_time") or "").strip()

    new_date = _parse_date(new_date_raw)
    new_time = _parse_time(new_time_raw)

    if not new_date:
        flash("Debe indicar una fecha válida para reagendar.", "danger")
        return redirect(url_for("dispatch.pending_requests"))

    for line in req.lines:
        line.load_date = new_date
        line.load_time = new_time
        db.session.add(line)

    req.updated_at = datetime.utcnow()
    db.session.add(req)

    create_notifications_for_roles(
        site_id=site_id,
        roles={"patio", "inspeccion"},
        title=f"Solicitud #{req.id} reagendada",
        message=f"La solicitud #{req.id} fue reagendada para {new_date.strftime('%d/%m/%Y')}"
                + (f" {new_time.strftime('%I:%M %p')}" if new_time else "."),
        related_type="DISPATCH_REQUEST",
        related_id=req.id,
        exclude_user_ids={current_user.id},
    )

    db.session.commit()

    flash(f"Solicitud #{req.id} reagendada correctamente.", "success")
    return redirect(url_for("dispatch.pending_requests"))

@dispatch_bp.post("/request/<int:request_id>/line/<int:line_id>/reschedule")
@login_required
def reschedule_pending_request_line(request_id: int, line_id: int):
    site_id = _ensure_active_site()

    role = (current_user.role or "").strip().lower()
    if role not in {"admin", "supervision", "despachador"}:
        abort(403)

    req = DispatchRequest.query.get_or_404(request_id)

    if req.site_id != site_id and role != "admin":
        abort(403)

    if req.status != "PENDIENTE":
        flash("Solo se pueden reagendar líneas de solicitudes pendientes.", "warning")
        return redirect(url_for("dispatch.pending_requests"))

    line = DispatchRequestLine.query.get_or_404(line_id)

    if line.request_id != req.id:
        abort(404)

    if line.assignments:
        flash("No se puede reagendar una línea que ya tiene contenedores asignados.", "warning")
        return redirect(url_for("dispatch.pending_requests"))

    new_date_raw = (request.form.get("load_date") or "").strip()
    new_time_raw = (request.form.get("load_time") or "").strip()

    new_date = _parse_date(new_date_raw)
    new_time = _parse_time(new_time_raw)

    if not new_date:
        flash("Debe indicar una fecha válida para reagendar.", "danger")
        return redirect(url_for("dispatch.pending_requests"))

    old_date = line.load_date
    old_time = line.load_time

    line.load_date = new_date
    line.load_time = new_time

    req.updated_at = datetime.utcnow()

    db.session.add(line)
    db.session.add(req)

    old_txt = old_date.strftime("%d/%m/%Y") if old_date else "sin fecha"
    if old_time:
        old_txt += f" {old_time.strftime('%I:%M %p')}"

    new_txt = new_date.strftime("%d/%m/%Y")
    if new_time:
        new_txt += f" {new_time.strftime('%I:%M %p')}"

    create_notifications_for_roles(
        site_id=site_id,
        roles={"patio", "inspeccion"},
        title=f"Línea reagendada solicitud #{req.id}",
        message=(
            f"{line.quantity} x {line.container_size} fue reagendada "
            f"de {old_txt} a {new_txt}."
        ),
        related_type="DISPATCH_REQUEST",
        related_id=req.id,
        exclude_user_ids={current_user.id},
    )

    db.session.commit()

    flash(f"Línea de solicitud #{req.id} reagendada correctamente.", "success")
    return redirect(url_for("dispatch.pending_requests"))

@dispatch_bp.post("/assignment/<int:assignment_id>/carrier-reported")
@login_required
def toggle_assignment_carrier_reported(assignment_id: int):
    site_id = _ensure_active_site()

    role = (current_user.role or "").strip().lower()

    if role not in {"admin", "supervision", "despachador"}:
        abort(403)

    assignment = DispatchAssignment.query.get_or_404(assignment_id)
    line = DispatchRequestLine.query.get_or_404(assignment.request_line_id)
    req = DispatchRequest.query.get_or_404(line.request_id)

    if req.site_id != site_id and role != "admin":
        abort(403)

    carrier_reported = request.form.get("carrier_reported") == "1"

    assignment.carrier_reported = carrier_reported

    db.session.commit()

    flash("Referencia de reporte a naviera actualizada.", "success")
    return redirect(url_for("dispatch.assigned_requests"))