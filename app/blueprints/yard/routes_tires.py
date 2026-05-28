from datetime import datetime
from io import BytesIO

from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_login import login_required, current_user
from sqlalchemy import text

from app.blueprints.yard import yard_bp
from app.extensions import db
from app.models.chassis import Chassis
from app.models.chassis_tire import ChassisTire
from app.models.tire import Tire
from app.models.tire_retread_event import TireRetreadEvent

from .routes import (
    _ensure_active_site,
    allowed_positions_for,
    _translate_tire_position,
    _calc_tire_state_from_mm,
    _normalize_tire_status,
    _sync_tire_master_state,
    _open_tire_retread_event,
    _close_tire_retread_event,
)


@yard_bp.get("/llantas")
@login_required
def tires_list():
    _ensure_active_site()

    q = (request.args.get("q") or "").strip()
    color = (request.args.get("color") or "").strip().upper()
    mounted = (request.args.get("mounted") or "").strip().upper()

    filters = []
    params = {}

    if q:
        filters.append("""
            (
                t.tire_number ILIKE :q
                OR COALESCE(t.brand, '') ILIKE :q
                OR COALESCE(t.last_marchamo, '') ILIKE :q
                OR COALESCE(ch.chassis_number, '') ILIKE :q
            )
        """)
        params["q"] = f"%{q}%"

    if color == "VERDE":
        filters.append("""
            (
                (ct.id IS NOT NULL AND ct.is_flat = FALSE AND ct.estrias_mm BETWEEN 9 AND 12)
                OR
                (ct.id IS NULL AND COALESCE(t.last_is_flat, FALSE) = FALSE AND t.last_estrias_mm BETWEEN 9 AND 12)
            )
        """)
    elif color == "AMARILLO":
        filters.append("""
            (
                (ct.id IS NOT NULL AND ct.is_flat = FALSE AND ct.estrias_mm BETWEEN 4 AND 8)
                OR
                (ct.id IS NULL AND COALESCE(t.last_is_flat, FALSE) = FALSE AND t.last_estrias_mm BETWEEN 4 AND 8)
            )
        """)
    elif color == "ROJO":
        filters.append("""
            (
                (ct.id IS NOT NULL AND (ct.is_flat = TRUE OR ct.estrias_mm <= 3))
                OR
                (ct.id IS NULL AND (COALESCE(t.last_is_flat, FALSE) = TRUE OR t.last_estrias_mm <= 3))
            )
        """)

    if mounted == "SI":
        filters.append("ct.id IS NOT NULL")
    elif mounted == "NO":
        filters.append("ct.id IS NULL")

    where_sql = ""
    if filters:
        where_sql = "WHERE " + " AND ".join(filters)

    sql = text(f"""
        SELECT
            t.id AS tire_id,
            t.tire_number,
            t.brand,
            t.model,
            t.size,
            t.notes,
            t.status,

            ct.id AS chassis_tire_id,
            ct.marchamo AS mounted_marchamo,
            ct.estrias_mm AS mounted_estrias_mm,
            ct.is_flat AS mounted_is_flat,
            ct.tire_state AS mounted_tire_state,

            t.last_marchamo,
            t.last_estrias_mm,
            t.last_is_flat,
            t.last_tire_state,

            ch.id AS chassis_id,
            ch.chassis_number,
            ct.position_code,

            CASE
                WHEN ct.id IS NOT NULL THEN
                    CASE
                        WHEN ct.is_flat = TRUE THEN 'ROJO'
                        WHEN ct.estrias_mm IS NULL THEN ''
                        WHEN ct.estrias_mm <= 3 THEN 'ROJO'
                        WHEN ct.estrias_mm BETWEEN 4 AND 8 THEN 'AMARILLO'
                        WHEN ct.estrias_mm BETWEEN 9 AND 12 THEN 'VERDE'
                        ELSE ''
                    END
                ELSE
                    CASE
                        WHEN t.last_is_flat = TRUE THEN 'ROJO'
                        WHEN t.last_estrias_mm IS NULL THEN ''
                        WHEN t.last_estrias_mm <= 3 THEN 'ROJO'
                        WHEN t.last_estrias_mm BETWEEN 4 AND 8 THEN 'AMARILLO'
                        WHEN t.last_estrias_mm BETWEEN 9 AND 12 THEN 'VERDE'
                        ELSE ''
                    END
            END AS estado_color
        FROM yard_gate_alamo.tires t
        LEFT JOIN yard_gate_alamo.chassis_tires ct
          ON ct.tire_id = t.id
        LEFT JOIN yard_gate_alamo.chassis ch
          ON ch.id = ct.chassis_id
        {where_sql}
        ORDER BY t.tire_number ASC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    for r in rows:
        estrias_mm = r["mounted_estrias_mm"] if r["chassis_tire_id"] else r["last_estrias_mm"]
        is_flat = bool(r["mounted_is_flat"]) if r["chassis_tire_id"] else bool(r["last_is_flat"])
        tire_state = r["mounted_tire_state"] if r["chassis_tire_id"] else r["last_tire_state"]
        marchamo = r["mounted_marchamo"] if r["chassis_tire_id"] else r["last_marchamo"]

        if r["chassis_tire_id"]:
            ubicacion = r["position_code"] or ""
            position_label = _translate_tire_position(r["position_code"])
        else:
            if (r["status"] or "") == "RECAUCHE":
                ubicacion = "RECAUCHE"
                position_label = "Recauche"
            elif (r["status"] or "") == "DESECHADA":
                ubicacion = "DESECHADA"
                position_label = "Desechada"
            else:
                ubicacion = "TALLER_BODEGA"
                position_label = "Taller/Bodega"

        items.append({
            "tire_id": r["tire_id"],
            "tire_number": r["tire_number"] or "",
            "brand": r["brand"] or "",
            "model": r["model"] or "",
            "size": r["size"] or "",
            "notes": r["notes"] or "",
            "marchamo": marchamo or "",
            "estrias_mm": estrias_mm,
            "is_flat": is_flat,
            "tire_state": tire_state or "",
            "chassis_id": r["chassis_id"],
            "chassis_number": r["chassis_number"] or "",
            "position_code": ubicacion or "",
            "position_label": position_label or "",
            "estado_color": r["estado_color"] or "",
            "is_mounted": bool(r["chassis_tire_id"]),
            "status": r["status"] or "EN_TALLER_BODEGA",
        })

    return render_template(
        "yard/tires_list.html",
        items=items,
        total=len(items),
        q=q,
        color=color,
        mounted=mounted,
    )


@yard_bp.get("/llantas/nueva")
@login_required
def tire_create_view():
    return render_template(
        "yard/tire_form.html",
        mode="create",
        tire=None,
    )


@yard_bp.post("/llantas/nueva")
@login_required
def tire_create_post():
    tire_number = (request.form.get("tire_number") or "").strip().upper()
    brand = (request.form.get("brand") or "").strip().upper()
    model = (request.form.get("model") or "").strip().upper()
    size = (request.form.get("size") or "").strip().upper()
    notes = (request.form.get("notes") or "").strip()
    status = _normalize_tire_status(request.form.get("status"))

    if not tire_number:
        flash("Debes ingresar el número de llanta.", "danger")
        return redirect(url_for("yard.tire_create_view"))

    existing = Tire.query.filter_by(tire_number=tire_number).first()
    if existing:
        flash("Ya existe una llanta con ese número.", "danger")
        return redirect(url_for("yard.tire_create_view"))

    tire = Tire(
        tire_number=tire_number,
        brand=brand or None,
        model=model or None,
        size=size or None,
        notes=notes or None,
        status=status,
    )
    db.session.add(tire)
    db.session.commit()

    flash(f"Llanta {tire_number} creada correctamente.", "success")
    return redirect(url_for("yard.tires_list"))


@yard_bp.get("/llantas/<int:tire_id>")
@login_required
def tire_detail_view(tire_id: int):
    tire = Tire.query.get_or_404(tire_id)

    row = db.session.execute(text("""
        SELECT
            t.id AS tire_id,
            t.tire_number,
            t.brand,
            t.model,
            t.size,
            t.notes,
            t.status,

            ct.id AS chassis_tire_id,
            ct.marchamo AS mounted_marchamo,
            ct.estrias_mm AS mounted_estrias_mm,
            ct.is_flat AS mounted_is_flat,
            ct.tire_state AS mounted_tire_state,

            t.last_marchamo,
            t.last_estrias_mm,
            t.last_is_flat,
            t.last_tire_state,

            ch.id AS chassis_id,
            ch.chassis_number,
            ct.position_code
        FROM yard_gate_alamo.tires t
        LEFT JOIN yard_gate_alamo.chassis_tires ct
          ON ct.tire_id = t.id
        LEFT JOIN yard_gate_alamo.chassis ch
          ON ch.id = ct.chassis_id
        WHERE t.id = :tire_id
        LIMIT 1
    """), {"tire_id": tire_id}).mappings().first()

    item = None
    if row:
        marchamo = row["mounted_marchamo"] if row["chassis_tire_id"] else row["last_marchamo"]
        estrias_mm = row["mounted_estrias_mm"] if row["chassis_tire_id"] else row["last_estrias_mm"]
        is_flat = bool(row["mounted_is_flat"]) if row["chassis_tire_id"] else bool(row["last_is_flat"])
        tire_state = row["mounted_tire_state"] if row["chassis_tire_id"] else row["last_tire_state"]

        item = {
            "tire_id": row["tire_id"],
            "tire_number": row["tire_number"] or "",
            "brand": row["brand"] or "",
            "model": row["model"] or "",
            "size": row["size"] or "",
            "notes": row["notes"] or "",
            "marchamo": marchamo or "",
            "estrias_mm": estrias_mm,
            "is_flat": is_flat,
            "tire_state": tire_state or "",
            "chassis_id": row["chassis_id"],
            "chassis_number": row["chassis_number"] or "",
            "position_code": row["position_code"] or "",
            "position_label": _translate_tire_position(row["position_code"]),
            "status": row["status"] or "EN_TALLER_BODEGA",
        }

    return render_template(
        "yard/tire_form.html",
        mode="edit",
        tire=item,
    )


@yard_bp.post("/llantas/<int:tire_id>/editar")
@login_required
def tire_edit_post(tire_id: int):
    tire = Tire.query.get_or_404(tire_id)
    previous_status = tire.status

    tire_number = (request.form.get("tire_number") or "").strip().upper()
    brand = (request.form.get("brand") or "").strip().upper()
    model = (request.form.get("model") or "").strip().upper()
    size = (request.form.get("size") or "").strip().upper()
    notes = (request.form.get("notes") or "").strip()

    status = _normalize_tire_status(request.form.get("status"))

    marchamo = (request.form.get("marchamo") or "").strip()
    chassis_number = (request.form.get("chassis_number") or "").strip().upper()
    position_code = (request.form.get("position_code") or "").strip().upper()
    confirm_replace = (request.form.get("confirm_replace") or "").strip() == "1"

    estrias_mm_raw = (request.form.get("estrias_mm") or "").strip()
    is_flat = (request.form.get("is_flat") or "").strip() == "1"

    if not tire_number:
        flash("Debes ingresar el número de llanta.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    duplicate = Tire.query.filter(Tire.tire_number == tire_number, Tire.id != tire.id).first()
    if duplicate:
        flash("Ya existe otra llanta con ese número.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    estrias_mm = None
    if estrias_mm_raw:
        try:
            estrias_mm = int(estrias_mm_raw)
        except Exception:
            flash("Las estrías deben ser numéricas.", "danger")
            return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

        if estrias_mm < 1 or estrias_mm > 12:
            flash("Las estrías deben estar entre 1 y 12.", "danger")
            return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    tire_state = _calc_tire_state_from_mm(estrias_mm, is_flat)

    tire.tire_number = tire_number
    tire.brand = brand or None
    tire.model = model or None
    tire.size = size or None
    tire.notes = notes or None

    _sync_tire_master_state(
        tire,
        marchamo=marchamo,
        estrias_mm=estrias_mm,
        is_flat=is_flat,
        tire_state=tire_state,
    )

    current_mount = ChassisTire.query.filter_by(tire_id=tire.id).first()

    if not chassis_number and not position_code:
        if current_mount:
            _sync_tire_master_state(
                tire,
                marchamo=current_mount.marchamo,
                estrias_mm=current_mount.estrias_mm,
                is_flat=bool(current_mount.is_flat),
                tire_state=current_mount.tire_state,
            )
            db.session.delete(current_mount)

        tire.status = status if status in {"EN_TALLER_BODEGA", "RECAUCHE", "DESECHADA"} else "EN_TALLER_BODEGA"

        if tire.status == "RECAUCHE" and previous_status != "RECAUCHE":
            _open_tire_retread_event(
                tire_id=tire.id,
                previous_estrias_mm=tire.last_estrias_mm,
                previous_marchamo=tire.last_marchamo,
                user_id=current_user.id,
                notes="Salida a recauche desde ver llanta",
            )

        elif previous_status == "RECAUCHE" and tire.status in {"EN_TALLER_BODEGA", "DESECHADA"}:
            _close_tire_retread_event(
                tire_id=tire.id,
                new_estrias_mm=tire.last_estrias_mm,
                new_marchamo=tire.last_marchamo,
                user_id=current_user.id,
                final_status=tire.status,
            )

        db.session.add(tire)
        db.session.commit()

        flash(f"Llanta {tire_number} actualizada correctamente.", "success")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    if not chassis_number:
        flash("Debes indicar el número de chasis.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    if not position_code:
        flash("Debes indicar la posición de la llanta.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    target_chassis = Chassis.query.filter_by(chassis_number=chassis_number).first()
    if not target_chassis:
        flash("El número de chasis no existe.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    allowed = set(allowed_positions_for(int(target_chassis.axles or 2)))
    if position_code not in allowed:
        flash("La posición no es válida para ese chasis.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    if tire.status != "EN_TALLER_BODEGA":
        flash(
            f"La llanta {tire.tire_number} no puede montarse porque su estado actual es {tire.status}. "
            f"Solo se pueden montar llantas en EN_TALLER_BODEGA.",
            "danger"
        )
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    if current_mount and (
        current_mount.chassis_id != target_chassis.id
        or current_mount.position_code != position_code
    ):
        flash(
            f"La llanta {tire.tire_number} ya está asignada al chasis "
            f"{current_mount.chassis.chassis_number if current_mount.chassis else current_mount.chassis_id} "
            f"en la posición {_translate_tire_position(current_mount.position_code)}.",
            "danger"
        )
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    occupied = ChassisTire.query.filter_by(
        chassis_id=target_chassis.id,
        position_code=position_code
    ).first()

    if occupied and occupied.tire_id != tire.id:
        existing_tire = Tire.query.get(occupied.tire_id) if occupied.tire_id else None
        existing_tire_number = existing_tire.tire_number if existing_tire else "SIN NÚMERO"
        translated_pos = _translate_tire_position(position_code)

        if not confirm_replace:
            flash(
                f"La posición {translated_pos} del chasis {target_chassis.chassis_number} "
                f"ya está ocupada por la llanta {existing_tire_number}. "
                f"Si deseas reemplazarla, confirma la operación.",
                "warning"
            )
            return redirect(
                url_for(
                    "yard.tire_detail_view",
                    tire_id=tire.id,
                )
            )

        if existing_tire:
            existing_tire.status = "EN_TALLER_BODEGA"
            _sync_tire_master_state(
                existing_tire,
                marchamo=occupied.marchamo,
                estrias_mm=occupied.estrias_mm,
                is_flat=bool(occupied.is_flat),
                tire_state=occupied.tire_state,
            )

        db.session.delete(occupied)
        db.session.flush()

    mount_row = current_mount
    if not mount_row:
        mount_row = ChassisTire(
            chassis_id=target_chassis.id,
            position_code=position_code,
            tire_id=tire.id,
            installed_at=datetime.utcnow(),
        )

    mount_row.chassis_id = target_chassis.id
    mount_row.position_code = position_code
    mount_row.tire_id = tire.id
    mount_row.marchamo = marchamo or None
    mount_row.estrias_mm = estrias_mm
    mount_row.is_flat = is_flat
    mount_row.tire_state = tire_state
    mount_row.updated_at = datetime.utcnow()

    _sync_tire_master_state(
        tire,
        marchamo=marchamo,
        estrias_mm=estrias_mm,
        is_flat=is_flat,
        tire_state=tire_state,
    )

    db.session.add(mount_row)

    tire.status = "ASIGNADA"
    db.session.add(tire)

    db.session.commit()

    flash(
        f"Llanta {tire.tire_number} asignada al chasis {target_chassis.chassis_number} "
        f"en la posición {_translate_tire_position(position_code)}.",
        "success"
    )
    return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))


@yard_bp.get("/llantas/import")
@login_required
def tires_import_view():
    return render_template("yard/tires_import.html")


@yard_bp.get("/llantas/export")
@login_required
def tires_export():
    _ensure_active_site()

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.tires_list"))

    rows = (
        Tire.query
        .order_by(Tire.tire_number.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Llantas"

    headers = [
        "id (no borrar si vas a actualizar)",
        "tire_number (número de llanta)",
        "brand (marca) [opcional]",
        "model (modelo) [opcional]",
        "size (tamaño) [opcional]",
        "status (ASIGNADA/EN_TALLER_BODEGA/RECAUCHE/DESECHADA)",
        "notes (notas) [opcional]",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for t in rows:
        ws.append([
            t.id,
            t.tire_number or "",
            t.brand or "",
            t.model or "",
            t.size or "",
            t.status or "EN_TALLER_BODEGA",
            t.notes or "",
        ])

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 40

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="llantas_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@yard_bp.post("/llantas/import")
@login_required
def tires_import_post():
    _ensure_active_site()

    f = request.files.get("file")
    if not f or not f.filename:
        flash("Sube un archivo Excel.", "danger")
        return redirect(url_for("yard.tires_import_view"))

    try:
        from openpyxl import load_workbook
    except Exception:
        flash("Falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.tires_import_view"))

    wb = load_workbook(f, data_only=True)
    ws = wb.active

    imported = 0
    updated = 0
    errors = []

    staged = []
    ids_to_find = []
    tire_numbers_to_find = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tire_id_raw = row[0] if len(row) > 0 else None
        tire_number = (str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else "")
        brand = (str(row[2]).strip().upper() if len(row) > 2 and row[2] is not None else "")
        model = (str(row[3]).strip().upper() if len(row) > 3 and row[3] is not None else "")
        size = (str(row[4]).strip().upper() if len(row) > 4 and row[4] is not None else "")
        status = _normalize_tire_status(row[5] if len(row) > 5 else None)
        notes = (str(row[6]).strip() if len(row) > 6 and row[6] is not None else "")

        if not any([tire_id_raw, tire_number, brand, model, size, status, notes]):
            continue

        if not tire_number:
            errors.append(f"Fila {idx}: falta tire_number.")
            continue

        tire_id = None
        if tire_id_raw not in (None, ""):
            try:
                tire_id = int(tire_id_raw)
            except Exception:
                errors.append(f"Fila {idx}: id inválido ({tire_id_raw}).")
                continue

        staged.append({
            "idx": idx,
            "id": tire_id,
            "tire_number": tire_number,
            "brand": brand or None,
            "model": model or None,
            "size": size or None,
            "status": status,
            "notes": notes or None,
        })

        if tire_id:
            ids_to_find.append(tire_id)
        tire_numbers_to_find.append(tire_number)

    if not staged:
        flash(f"No se importó nada. Errores: {len(errors)}", "danger")
        session["tires_import_errors"] = errors[:200]
        return redirect(url_for("yard.tires_import_view"))

    existing_by_id = {}
    if ids_to_find:
        rows_by_id = Tire.query.filter(Tire.id.in_(ids_to_find)).all()
        existing_by_id = {t.id: t for t in rows_by_id}

    existing_by_number = {}
    if tire_numbers_to_find:
        rows_by_number = Tire.query.filter(Tire.tire_number.in_(tire_numbers_to_find)).all()
        existing_by_number = {t.tire_number: t for t in rows_by_number}

    used_numbers_in_file = set()

    for item in staged:
        tire_id = item["id"]
        tire_number = item["tire_number"]

        if tire_number in used_numbers_in_file:
            errors.append(f"Fila {item['idx']}: tire_number repetido en el mismo archivo ({tire_number}).")
            continue

        used_numbers_in_file.add(tire_number)

        tire = None

        if tire_id and tire_id in existing_by_id:
            tire = existing_by_id[tire_id]
        elif tire_number in existing_by_number:
            tire = existing_by_number[tire_number]

        if tire:
            tire.tire_number = tire_number
            tire.brand = item["brand"]
            tire.model = item["model"]
            tire.size = item["size"]
            tire.status = item["status"]
            tire.notes = item["notes"]
            db.session.add(tire)
            updated += 1
        else:
            tire = Tire(
                tire_number=tire_number,
                brand=item["brand"],
                model=item["model"],
                size=item["size"],
                status=item["status"],
                notes=item["notes"],
            )
            db.session.add(tire)
            imported += 1

    db.session.commit()

    if errors:
        flash(f"Importado: {imported} | Actualizado: {updated} | Errores: {len(errors)}", "warning")
        session["tires_import_errors"] = errors[:200]
    else:
        flash(f"Importado: {imported} | Actualizado: {updated}", "success")
        session.pop("tires_import_errors", None)

    return redirect(url_for("yard.tires_list"))


@yard_bp.get("/llantas/recauche")
@login_required
def tire_retread_report_view():
    _ensure_active_site()
    return render_template("yard/tire_retread_report.html")


@yard_bp.get("/api/llantas/recauche-report")
@login_required
def tire_retread_report():
    _ensure_active_site()

    tire_number = (request.args.get("tire_number") or "").strip().upper()

    filters = []
    params = {}

    if tire_number:
        filters.append("t.tire_number ILIKE :tire_number")
        params["tire_number"] = f"%{tire_number}%"

    where_sql = ""
    if filters:
        where_sql = " AND " + " AND ".join(filters)

    sql = text(f"""
        SELECT
            e.id,
            t.tire_number,

            e.sent_at,
            e.returned_at,

            e.previous_estrias_mm,
            e.new_estrias_mm,

            e.previous_marchamo,
            e.new_marchamo,

            e.event_status,

            u1.username AS sent_by_user,
            u2.username AS returned_by_user

        FROM yard_gate_alamo.tire_retread_events e
        JOIN yard_gate_alamo.tires t
          ON t.id = e.tire_id

        LEFT JOIN yard_gate_alamo.users u1
          ON u1.id = e.sent_by

        LEFT JOIN yard_gate_alamo.users u2
          ON u2.id = e.returned_by

        WHERE 1=1
        {where_sql}

        ORDER BY e.sent_at DESC NULLS LAST, e.id DESC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []

    for r in rows:
        if r["event_status"] == "RETURNED":
            status_final = "RECAUCHADA"
        elif r["event_status"] == "SCRAPPED":
            status_final = "DESECHADA"
        else:
            status_final = "EN PROCESO"

        items.append({
            "tire_number": r["tire_number"] or "",
            "sent_at": r["sent_at"].strftime("%d/%m/%Y %I:%M %p") if r["sent_at"] else "",
            "returned_at": r["returned_at"].strftime("%d/%m/%Y %I:%M %p") if r["returned_at"] else "",
            "before_mm": r["previous_estrias_mm"],
            "after_mm": r["new_estrias_mm"],
            "old_marchamo": r["previous_marchamo"] or "",
            "new_marchamo": r["new_marchamo"] or "",
            "status_final": status_final,
            "sent_by": r["sent_by_user"] or "",
            "returned_by": r["returned_by_user"] or "",
        })

    return jsonify({"ok": True, "items": items})