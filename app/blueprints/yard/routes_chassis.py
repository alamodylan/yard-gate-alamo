from datetime import datetime
from io import BytesIO

from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, abort, session
from flask_login import login_required, current_user
from sqlalchemy import text

from app.blueprints.yard import yard_bp
from app.extensions import db
from app.models.site import Site
from app.models.chassis import Chassis, ChassisInventory
from app.models.chassis_tire import ChassisTire
from app.models.tire import Tire, TireReading, TirePosition
from app.services.audit import audit_log

from .routes import (
    _ensure_active_site,
    CHASSIS_NUM_RE,
    CHASSIS_STATUSES,
    CHASSIS_KINDS,
    SIDE_TO_POSITION,
    POSITION_TO_SIDE,
    TIRE_STATES,
    _norm_enum,
    classify_chassis_number,
    allowed_positions_for,
    _translate_tire_position,
    _calc_tire_state_from_mm,
    _calc_tire_state_from_data,
    _sync_tire_master_state,
    _maybe_register_tire_retread,
    _normalize_structure_status_for_db,
    _normalize_twistlocks_status_for_db,
    _normalize_landing_gear_status_for_db,
    _normalize_lights_status_for_db,
    _normalize_mudflap_status_for_db,
    _normalize_position_for_tire_master,
    _save_grouped_tire_readings,
    _insert_dynamic,
    _fetch_last_final_eir_for_chassis,
    _build_workshop_ticket_text,
)


# =========================
# Chassis pages
# =========================

@yard_bp.get("/chassis")
@login_required
def chassis_list():
    site_id = _ensure_active_site()

    rows = (
        Chassis.query
        .filter(
            Chassis.site_id == site_id,
            Chassis.is_in_yard.is_(True),
        )
        .order_by(Chassis.chassis_number.asc())
        .all()
    )

    items = []
    for ch in rows:
        ubicacion = ch.site.name if getattr(ch, "site", None) else "Predio"

        items.append({
            "id": ch.id,
            "chassis_number": ch.chassis_number,
            "plate": ch.plate,
            "chassis_kind": ch.chassis_kind,
            "length_ft": ch.length_ft,
            "axles": ch.axles,
            "status": ch.status,
            "is_in_yard": bool(ch.is_in_yard),
            "ubicacion": ubicacion,
            "site_name": ch.site.name if getattr(ch, "site", None) else "",
        })

    sites = Site.query.order_by(Site.name.asc()).all()

    return render_template(
        "yard/chassis_list.html",
        rows=rows,
        items=items,
        sites=sites
    )


@yard_bp.get("/chassis/dashboard")
@login_required
def chassis_dashboard():
    site_id = _ensure_active_site()

    base = Chassis.query.filter(
        Chassis.site_id == site_id,
        Chassis.is_in_yard.is_(True),
    )

    counts = {
        "40FT_2AX": base.filter(Chassis.type_code == "40FT_2AX").count(),
        "40FT_3AX": base.filter(Chassis.type_code == "40FT_3AX").count(),
        "20FT_2AX": base.filter(Chassis.type_code == "20FT_2AX").count(),
        "20FT_3AX": base.filter(Chassis.type_code == "20FT_3AX").count(),
    }

    total = base.count()
    unknown = base.filter((Chassis.type_code.is_(None)) | (Chassis.type_code == "UNKNOWN")).count()

    return render_template(
        "yard/chassis_dashboard.html",
        total=total,
        unknown=unknown,
        counts=counts
    )


@yard_bp.get("/chassis/import")
@login_required
def chassis_import_view():
    return render_template("yard/chassis_import.html")


@yard_bp.post("/chassis/import")
@login_required
def chassis_import_post():
    site_id = _ensure_active_site()

    f = request.files.get("file")
    if not f or not f.filename:
        flash("Sube un archivo Excel.", "danger")
        return redirect(url_for("yard.chassis_import_view"))

    try:
        from openpyxl import load_workbook
    except Exception:
        flash("Falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.chassis_import_view"))

    wb = load_workbook(f, data_only=True)
    ws = wb.active

    imported = 0
    updated = 0
    errors = []

    sites = Site.query.all()
    sites_by_name = {(s.name or "").strip().upper(): s for s in sites}

    staged = []
    numbers = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        chassis_number = (str(row[0]).strip() if row and row[0] is not None else "")
        plate = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else None)
        length_ft = row[2] if len(row) > 2 else None
        axles = row[3] if len(row) > 3 else None
        type_code = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else None)

        status = _norm_enum(row[5]) if len(row) > 5 and row[5] is not None else ""
        chassis_kind = _norm_enum(row[6]) if len(row) > 6 and row[6] is not None else ""
        predio_name = (str(row[7]).strip() if len(row) > 7 and row[7] is not None else "")

        if not CHASSIS_NUM_RE.match(chassis_number):
            errors.append(f"Fila {idx}: chassis_number inválido ({chassis_number})")
            continue

        d_len, d_ax, d_type = classify_chassis_number(chassis_number)

        if (not length_ft) or (not axles):
            if d_len is None or d_ax is None:
                errors.append(f"Fila {idx}: prefijo no reconocido ({chassis_number})")
                continue
            length_ft = int(length_ft) if length_ft else d_len
            axles = int(axles) if axles else d_ax

        if not type_code:
            type_code = d_type

        try:
            length_ft = int(length_ft)
            axles = int(axles)
        except Exception:
            errors.append(f"Fila {idx}: length_ft/axles inválidos")
            continue

        if length_ft not in (20, 40, 45) or axles not in (2, 3):
            errors.append(f"Fila {idx}: fuera de rango length_ft={length_ft} axles={axles}")
            continue

        if not status:
            status = "BUENO"
        if status not in CHASSIS_STATUSES:
            errors.append(f"Fila {idx}: status inválido ({status})")
            continue

        if not chassis_kind:
            chassis_kind = "CHASIS"
        if chassis_kind not in CHASSIS_KINDS:
            errors.append(f"Fila {idx}: tipo inválido ({chassis_kind})")
            continue

        target_site_id = site_id
        if predio_name:
            s = sites_by_name.get(predio_name.strip().upper())
            if not s:
                errors.append(f"Fila {idx}: predio no existe ({predio_name})")
                continue
            target_site_id = s.id

        staged.append({
            "idx": idx,
            "chassis_number": chassis_number,
            "plate": plate,
            "length_ft": length_ft,
            "axles": axles,
            "type_code": type_code,
            "status": status,
            "chassis_kind": chassis_kind,
            "site_id": target_site_id,
        })
        numbers.append(chassis_number)

    if not staged:
        flash(f"No se importó nada. Errores: {len(errors)}", "danger")
        session["chassis_import_errors"] = errors[:200]
        return redirect(url_for("yard.chassis_import_view"))

    existing_rows = (
        Chassis.query
        .filter(Chassis.chassis_number.in_(numbers))
        .all()
    )
    existing_map = {c.chassis_number: c for c in existing_rows}

    for item in staged:
        chassis_number = item["chassis_number"]
        plate = item["plate"]
        length_ft = item["length_ft"]
        axles = item["axles"]
        type_code = item["type_code"]
        status = item["status"]
        chassis_kind = item["chassis_kind"]
        target_site_id = item["site_id"]

        existing = existing_map.get(chassis_number)

        if existing:
            existing.site_id = target_site_id
            existing.plate = plate
            existing.length_ft = length_ft
            existing.axles = axles
            existing.type_code = type_code
            existing.status = status
            existing.chassis_kind = chassis_kind
            existing.has_plate = True if plate else False
            existing.is_in_yard = True
            db.session.add(existing)
            updated += 1
        else:
            ch = Chassis(
                site_id=target_site_id,
                chassis_number=chassis_number,
                plate=plate,
                length_ft=length_ft,
                axles=axles,
                type_code=type_code,
                status=status,
                chassis_kind=chassis_kind,
                has_plate=True if plate else False,
                is_in_yard=True,
            )
            db.session.add(ch)
            imported += 1

    db.session.commit()

    if errors:
        flash(f"Importado: {imported} | Actualizado: {updated} | Errores: {len(errors)}", "warning")
        session["chassis_import_errors"] = errors[:200]
    else:
        flash(f"Importado: {imported} | Actualizado: {updated}", "success")
        session.pop("chassis_import_errors", None)

    return redirect(url_for("yard.chassis_list"))


@yard_bp.get("/chassis/export")
@login_required
def chassis_export():
    site_id = _ensure_active_site()

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.chassis_list"))

    rows = (
        Chassis.query
        .filter(
            Chassis.site_id == site_id,
            Chassis.is_in_yard.is_(True),
        )
        .order_by(Chassis.chassis_number.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Chassis"

    headers = [
        "chassis_number (número de chasis 5 dígitos)",
        "plate (placa) [opcional]",
        "length_ft (largo en pies: 20/40/45) [opcional]",
        "axles (ejes: 2/3) [opcional]",
        "type_code (tipo: 20FT_2AX/20FT_3AX/40FT_2AX/40FT_3AX) [opcional]",
        "status (BUENO/DAÑADO/FUERA_DE_SERVICIO/ATADO) [opcional]",
        "chassis_kind (CHASIS/LOW_BOY/TANQUETA/PLANA/CARRETA) [opcional]",
        "predio (nombre del predio / Site.name) [opcional]",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for ch in rows:
        ws.append([
            ch.chassis_number,
            ch.plate or "",
            getattr(ch, "length_ft", "") or "",
            getattr(ch, "axles", "") or "",
            ch.type_code or "",
            getattr(ch, "status", "") or "BUENO",
            getattr(ch, "chassis_kind", "") or "CHASIS",
            (ch.site.name if getattr(ch, "site", None) else ""),
        ])

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 26

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "chassis_import_template.xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@yard_bp.get("/chassis/<int:chassis_id>")
@login_required
def chassis_detail(chassis_id: int):
    site_id = _ensure_active_site()
    ch = Chassis.query.get_or_404(chassis_id)

    if ch.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    axles = int(getattr(ch, "axles", 2) or 2)
    length_ft = int(getattr(ch, "length_ft", 40) or 40)

    sites = Site.query.order_by(Site.name.asc()).all()

    return render_template(
        "yard/chassis_detail.html",
        ch=ch,
        axles=axles,
        length_ft=length_ft,
        sites=sites,
        statuses=sorted(CHASSIS_STATUSES),
        kinds=sorted(CHASSIS_KINDS),
    )


# =========================
# Chassis tires API
# =========================

@yard_bp.get("/api/chassis/<int:chassis_id>/tires")
@login_required
def api_chassis_tires_get(chassis_id: int):
    _ensure_active_site()
    ch = Chassis.query.get_or_404(chassis_id)

    axles = int(getattr(ch, "axles", 2) or 2)
    allowed = set(allowed_positions_for(axles))

    rows = ChassisTire.query.filter_by(chassis_id=ch.id).all()

    positions = {}
    for p in allowed:
        positions[p] = {
            "marchamo": None,
            "tire_state": "OK",
            "tire_id": None,
            "tire_number": None,
            "brand": None,
            "model": None,
            "size": None,
            "notes": None,
            "status": None,
            "estrias_mm": None,
            "is_flat": False,
        }

    for r in rows:
        pos = (r.position_code or "").strip().upper()
        if pos not in allowed:
            continue

        positions[pos] = {
            "marchamo": r.marchamo,
            "tire_state": (r.tire_state or "OK").upper(),
            "tire_id": r.tire.id if r.tire else None,
            "tire_number": r.tire.tire_number if r.tire else None,
            "brand": r.tire.brand if r.tire else None,
            "model": r.tire.model if r.tire else None,
            "size": r.tire.size if r.tire else None,
            "notes": r.tire.notes if r.tire else None,
            "status": r.tire.status if r.tire else None,
            "estrias_mm": getattr(r, "estrias_mm", None),
            "is_flat": bool(getattr(r, "is_flat", False)),
        }

    return jsonify({
        "ok": True,
        "chassis": {
            "id": ch.id,
            "chassis_number": ch.chassis_number,
            "plate": ch.plate,
            "axles": ch.axles,
            "status": ch.status,
            "site_id": ch.site_id,
            "type_code": getattr(ch, "type_code", None),
        },
        "positions": positions
    })


@yard_bp.get("/api/llantas/disponibles")
@login_required
def api_tires_available():
    _ensure_active_site()

    q = (request.args.get("q") or "").strip().upper()

    filters = ["t.status = 'EN_TALLER_BODEGA'"]
    params = {}

    if q:
        filters.append("""
            (
                t.tire_number ILIKE :q
                OR COALESCE(t.brand, '') ILIKE :q
                OR COALESCE(t.model, '') ILIKE :q
                OR COALESCE(t.size, '') ILIKE :q
            )
        """)
        params["q"] = f"%{q}%"

    sql = text(f"""
        SELECT
            t.id,
            t.tire_number,
            t.brand,
            t.model,
            t.size,
            t.notes,
            t.status,
            t.last_marchamo,
            t.last_estrias_mm,
            t.last_is_flat
        FROM yard_gate_alamo.tires t
        LEFT JOIN yard_gate_alamo.chassis_tires ct
          ON ct.tire_id = t.id
        WHERE {" AND ".join(filters)}
        ORDER BY t.tire_number ASC
        LIMIT 100
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    for r in rows:
        items.append({
            "id": r["id"],
            "tire_number": r["tire_number"] or "",
            "brand": r["brand"] or "",
            "model": r["model"] or "",
            "size": r["size"] or "",
            "notes": r["notes"] or "",
            "status": r["status"] or "EN_TALLER_BODEGA",
            "last_marchamo": r["last_marchamo"] or "",
            "last_estrias_mm": r["last_estrias_mm"],
            "last_is_flat": bool(r["last_is_flat"]) if r["last_is_flat"] is not None else False,
        })

    return jsonify({
        "ok": True,
        "items": items
    })


@yard_bp.post("/api/chassis/<int:chassis_id>/tires")
@login_required
def api_chassis_tires_set(chassis_id: int):
    site_id = _ensure_active_site()
    ch = Chassis.query.get_or_404(chassis_id)

    if ch.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    axles = int(getattr(ch, "axles", 2) or 2)
    allowed = set(allowed_positions_for(axles))

    data = request.get_json(silent=True) or {}
    action = (data.get("action") or "legacy_update").strip().lower()

    pos = (data.get("position_code") or "").strip().upper()
    if pos not in allowed:
        return jsonify({"ok": False, "error": "INVALID_POSITION"}), 400

    confirm_replace = bool(data.get("confirm_replace"))

    row = ChassisTire.query.filter_by(chassis_id=ch.id, position_code=pos).first()

    if action == "unassign":
        if not row:
            return jsonify({"ok": False, "error": "POSITION_EMPTY"}), 400

        tire = Tire.query.get(row.tire_id) if row.tire_id else None
        if tire:
            tire.status = "EN_TALLER_BODEGA"
            _sync_tire_master_state(
                tire,
                marchamo=row.marchamo,
                estrias_mm=row.estrias_mm,
                is_flat=bool(row.is_flat),
                tire_state=row.tire_state,
            )

        db.session.delete(row)
        db.session.commit()

        return jsonify({"ok": True, "action": "unassign"})

    if action == "assign_existing":
        tire_id = data.get("tire_id")

        if not tire_id:
            return jsonify({"ok": False, "error": "TIRE_ID_REQUIRED"}), 400

        tire = Tire.query.get(tire_id)
        if not tire:
            return jsonify({"ok": False, "error": "TIRE_NOT_FOUND"}), 404

        marchamo = (data.get("marchamo") or "").strip()
        if not marchamo:
            marchamo = (tire.last_marchamo or "").strip()

        estrias_mm_raw = data.get("estrias_mm")
        is_flat_raw = data.get("is_flat")
        if is_flat_raw in (None, "",):
            is_flat = bool(tire.last_is_flat)
        else:
            is_flat = bool(is_flat_raw)

        if tire.status != "EN_TALLER_BODEGA":
            return jsonify({
                "ok": False,
                "error": "TIRE_STATUS_NOT_ALLOWED",
                "detail": f"La llanta está en estado {tire.status}"
            }), 400

        existing_mount = ChassisTire.query.filter_by(tire_id=tire.id).first()
        if existing_mount and (
            existing_mount.chassis_id != ch.id or existing_mount.position_code != pos
        ):
            return jsonify({
                "ok": False,
                "error": "TIRE_ALREADY_ASSIGNED",
                "detail": (
                    f"La llanta ya está asignada al chasis "
                    f"{existing_mount.chassis.chassis_number if existing_mount.chassis else existing_mount.chassis_id} "
                    f"en la posición {existing_mount.position_code}"
                )
            }), 400

        estrias_mm = None
        if estrias_mm_raw not in (None, "",):
            try:
                estrias_mm = int(estrias_mm_raw)
            except Exception:
                return jsonify({"ok": False, "error": "INVALID_ESTRIAS_MM"}), 400

            if estrias_mm < 1 or estrias_mm > 12:
                return jsonify({"ok": False, "error": "ESTRIAS_OUT_OF_RANGE"}), 400
        else:
            estrias_mm = tire.last_estrias_mm

        if row and row.tire_id != tire.id:
            old_tire = Tire.query.get(row.tire_id) if row.tire_id else None

            if not confirm_replace:
                return jsonify({
                    "ok": False,
                    "error": "POSITION_OCCUPIED",
                    "detail": (
                        f"La posición {_translate_tire_position(pos)} ya está ocupada"
                    )
                }), 409

            if old_tire:
                old_tire.status = "EN_TALLER_BODEGA"
                _sync_tire_master_state(
                    old_tire,
                    marchamo=row.marchamo,
                    estrias_mm=row.estrias_mm,
                    is_flat=bool(row.is_flat),
                    tire_state=row.tire_state,
                )

            db.session.delete(row)
            db.session.flush()
            row = None

        tire_state = _calc_tire_state_from_mm(estrias_mm, is_flat)

        previous_estrias_mm = row.estrias_mm if row else None
        previous_marchamo = row.marchamo if row else None
        previous_tire_id = row.tire_id if row else None

        if not row:
            row = ChassisTire(
                chassis_id=ch.id,
                position_code=pos,
                tire_id=tire.id,
                installed_at=datetime.utcnow(),
            )

        row.chassis_id = ch.id
        row.position_code = pos
        row.tire_id = tire.id
        row.marchamo = marchamo or None
        row.estrias_mm = estrias_mm
        row.is_flat = is_flat
        row.tire_state = tire_state
        row.updated_at = datetime.utcnow()

        _sync_tire_master_state(
            tire,
            marchamo=marchamo,
            estrias_mm=estrias_mm,
            is_flat=is_flat,
            tire_state=tire_state,
        )

        if previous_tire_id == tire.id:
            _maybe_register_tire_retread(
                tire_id=tire.id,
                previous_estrias_mm=previous_estrias_mm,
                new_estrias_mm=estrias_mm,
                previous_marchamo=previous_marchamo,
                new_marchamo=marchamo,
                created_by=current_user.id,
            )

        tire.status = "ASIGNADA"

        db.session.add(row)
        db.session.add(tire)
        db.session.commit()

        return jsonify({
            "ok": True,
            "action": "assign_existing",
            "tire_state": tire_state
        })

    if action == "create_and_assign":
        tire_number = (data.get("tire_number") or "").strip().upper()
        brand = (data.get("brand") or "").strip().upper()
        model = (data.get("model") or "").strip().upper()
        size = (data.get("size") or "").strip().upper()
        notes = (data.get("notes") or "").strip()
        marchamo = (data.get("marchamo") or "").strip()

        estrias_mm_raw = data.get("estrias_mm")
        is_flat = bool(data.get("is_flat"))

        if not tire_number:
            return jsonify({"ok": False, "error": "TIRE_NUMBER_REQUIRED"}), 400

        existing_tire = Tire.query.filter_by(tire_number=tire_number).first()
        if existing_tire:
            return jsonify({
                "ok": False,
                "error": "TIRE_NUMBER_ALREADY_EXISTS"
            }), 409

        estrias_mm = None
        if estrias_mm_raw not in (None, "",):
            try:
                estrias_mm = int(estrias_mm_raw)
            except Exception:
                return jsonify({"ok": False, "error": "INVALID_ESTRIAS_MM"}), 400

            if estrias_mm < 1 or estrias_mm > 12:
                return jsonify({"ok": False, "error": "ESTRIAS_OUT_OF_RANGE"}), 400

        if row and row.tire_id:
            old_tire = Tire.query.get(row.tire_id)

            if not confirm_replace:
                return jsonify({
                    "ok": False,
                    "error": "POSITION_OCCUPIED",
                    "detail": (
                        f"La posición {_translate_tire_position(pos)} ya está ocupada"
                    )
                }), 409

            if old_tire:
                old_tire.status = "EN_TALLER_BODEGA"
                _sync_tire_master_state(
                    old_tire,
                    marchamo=row.marchamo,
                    estrias_mm=row.estrias_mm,
                    is_flat=bool(row.is_flat),
                    tire_state=row.tire_state,
                )

            db.session.delete(row)
            db.session.flush()
            row = None

        tire = Tire(
            tire_number=tire_number,
            brand=brand or None,
            model=model or None,
            size=size or None,
            notes=notes or None,
            status="ASIGNADA",
        )
        db.session.add(tire)
        db.session.flush()

        tire_state = _calc_tire_state_from_mm(estrias_mm, is_flat)

        if not row:
            row = ChassisTire(
                chassis_id=ch.id,
                position_code=pos,
                tire_id=tire.id,
                installed_at=datetime.utcnow(),
            )

        row.chassis_id = ch.id
        row.position_code = pos
        row.tire_id = tire.id
        row.marchamo = marchamo or None
        row.estrias_mm = estrias_mm
        row.is_flat = is_flat
        row.tire_state = tire_state
        row.updated_at = datetime.utcnow()

        _sync_tire_master_state(
            tire,
            marchamo=marchamo,
            estrias_mm=estrias_mm,
            is_flat=is_flat,
            tire_state=tire_state,
        )

        db.session.add(row)
        db.session.commit()

        return jsonify({
            "ok": True,
            "action": "create_and_assign",
            "tire_state": tire_state,
            "tire_id": tire.id
        })

    marchamo = (data.get("marchamo") or "").strip()
    tire_number = (data.get("tire_number") or "").strip().upper()
    brand = (data.get("brand") or "").strip().upper()

    estrias_mm_raw = data.get("estrias_mm")
    is_flat = bool(data.get("is_flat"))

    estrias_mm = None
    if estrias_mm_raw not in (None, "",):
        try:
            estrias_mm = int(estrias_mm_raw)
        except Exception:
            return jsonify({"ok": False, "error": "INVALID_ESTRIAS_MM"}), 400

        if estrias_mm < 1 or estrias_mm > 12:
            return jsonify({"ok": False, "error": "ESTRIAS_OUT_OF_RANGE"}), 400

    tire_state = _calc_tire_state_from_data(estrias_mm, is_flat)

    tire = None
    if tire_number:
        tire = Tire.query.filter_by(tire_number=tire_number).first()
        if not tire:
            tire = Tire(
                tire_number=tire_number,
                brand=brand or None,
                status="ASIGNADA"
            )
            db.session.add(tire)
            db.session.flush()
        else:
            if brand and (tire.brand != brand):
                tire.brand = brand
            if tire.status == "EN_TALLER_BODEGA":
                tire.status = "ASIGNADA"
            db.session.add(tire)

    previous_estrias_mm = row.estrias_mm if row else None
    previous_marchamo = row.marchamo if row else None
    previous_tire_id = row.tire_id if row else None

    if not row:
        row = ChassisTire(chassis_id=ch.id, position_code=pos)

    row.marchamo = marchamo or None
    row.estrias_mm = estrias_mm
    row.is_flat = is_flat
    row.tire_state = tire_state
    row.tire_id = tire.id if tire else None
    row.updated_at = datetime.utcnow()

    if tire:
        _sync_tire_master_state(
            tire,
            marchamo=marchamo,
            estrias_mm=estrias_mm,
            is_flat=is_flat,
            tire_state=tire_state,
        )

    current_tire_id = tire.id if tire else None
    if previous_tire_id and current_tire_id and previous_tire_id == current_tire_id:
        _maybe_register_tire_retread(
            tire_id=current_tire_id,
            previous_estrias_mm=previous_estrias_mm,
            new_estrias_mm=estrias_mm,
            previous_marchamo=previous_marchamo,
            new_marchamo=marchamo,
            created_by=current_user.id,
        )

    if previous_tire_id and current_tire_id and previous_tire_id != current_tire_id:
        old_tire = Tire.query.get(previous_tire_id)
        if old_tire:
            old_tire.status = "EN_TALLER_BODEGA"
            _sync_tire_master_state(
                old_tire,
                marchamo=previous_marchamo,
                estrias_mm=previous_estrias_mm,
                is_flat=bool(row.is_flat) if row else False,
                tire_state=row.tire_state if row else "OK",
            )

    db.session.add(row)
    db.session.commit()

    return jsonify({"ok": True, "action": "legacy_update", "tire_state": tire_state})


@yard_bp.post("/api/chassis/<int:chassis_id>/classify")
@login_required
def api_chassis_classify(chassis_id: int):
    site_id = _ensure_active_site()
    ch = Chassis.query.get_or_404(chassis_id)

    if ch.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    data = request.get_json(silent=True) or {}

    structure_status = _normalize_structure_status_for_db(
        _norm_enum(data.get("structure_status"))
    )
    twistlocks_status = _normalize_twistlocks_status_for_db(
        _norm_enum(data.get("twistlocks_status"))
    )
    landing_gear_status = _normalize_landing_gear_status_for_db(
        _norm_enum(data.get("landing_gear_status"))
    )
    lights_status = _normalize_lights_status_for_db(
        _norm_enum(data.get("lights_status"))
    )
    mudflap_status = _normalize_mudflap_status_for_db(
        _norm_enum(data.get("mudflap_status"))
    )

    plate_text = (data.get("plate_text") or "").strip()
    comments = (data.get("comments") or "").strip()
    damage_summary = (data.get("damage_summary") or "").strip()

    tires = data.get("tires") or []
    axles = int(getattr(ch, "axles", 2) or 2)
    allowed = set(allowed_positions_for(axles))

    tire_lines = []
    any_tire_issue = False
    grouped_tire_readings = {}

    for t in tires:
        pos = (t.get("position_code") or "").strip().upper()
        if pos not in allowed:
            continue

        ingreso_marchamo = (t.get("ingreso_marchamo") or "").strip()
        marchamo_check = (t.get("marchamo_check") or "OK").strip().upper()
        estrias_mm_raw = t.get("estrias_mm")
        is_flat = bool(t.get("is_flat"))

        estrias_mm = None
        if estrias_mm_raw not in (None, "",):
            try:
                estrias_mm = int(estrias_mm_raw)
            except Exception:
                estrias_mm = None

        tire_state = _calc_tire_state_from_data(estrias_mm, is_flat)

        if marchamo_check not in {"OK", "DISTINTO", "NO_TIENE", "ILEGIBLE"}:
            marchamo_check = "OK"

        if tire_state not in TIRE_STATES:
            tire_state = "OK"

        master_pos = _normalize_position_for_tire_master(pos)
        if master_pos:
            grp = grouped_tire_readings.setdefault(master_pos, {
                "states": [],
                "pressures": [],
                "comments": [],
                "seal_issue": False,
                "seal_2": None,
            })

            grp["states"].append(tire_state)

            if marchamo_check != "OK":
                grp["seal_issue"] = True

            detail_parts = [
                f"{pos}",
                f"MARCHAMO={marchamo_check}",
                f"STATE={tire_state}",
            ]
            if ingreso_marchamo:
                detail_parts.append(f"INGRESO={ingreso_marchamo}")
            if estrias_mm not in (None, ""):
                detail_parts.append(f"MM={estrias_mm}")
            if is_flat:
                detail_parts.append("FLAT=SI")

            grp["comments"].append(" | ".join(detail_parts))

        row = ChassisTire.query.filter_by(chassis_id=ch.id, position_code=pos).first()
        if row:
            row.estrias_mm = estrias_mm
            row.is_flat = is_flat
            row.tire_state = tire_state
            row.updated_at = datetime.utcnow()
            db.session.add(row)

        if marchamo_check != "OK":
            any_tire_issue = True
            if marchamo_check == "DISTINTO":
                tire_lines.append(f"{pos}: MARCHAMO DE INGRESO DISTINTO - REVISAR")
            elif marchamo_check == "NO_TIENE":
                tire_lines.append(f"{pos}: MARCHAMO DE INGRESO NO TIENE - REVISAR")
            elif marchamo_check == "ILEGIBLE":
                tire_lines.append(f"{pos}: MARCHAMO DE INGRESO ILEGIBLE - REVISAR")

        if is_flat:
            any_tire_issue = True
            tire_lines.append(f"{pos}: PINCHADA (DESINFLADA)")
        elif tire_state != "OK":
            any_tire_issue = True
            tire_lines.append(f"{pos}: ESTADO {tire_state} (MM={estrias_mm if estrias_mm is not None else '—'})")

    _save_grouped_tire_readings(
        site_id=site_id,
        chassis_id=ch.id,
        axles=axles,
        grouped_readings=grouped_tire_readings,
        user_id=current_user.id,
        event_type="EIR_IN",
        event_id=None,
    )

    structure_lines = []

    if structure_status in {"GOLPE", "DOBLADO", "SOLDADURA"}:
        structure_lines.append(f"Estructura: {structure_status}")

    if twistlocks_status in {"DANADOS"}:
        structure_lines.append(f"Twistlocks: {twistlocks_status}")

    if landing_gear_status in {"DANADAS"}:
        structure_lines.append(f"Pata de apoyo: {landing_gear_status}")

    if lights_status in {"IZQ_DANADA", "DER_DANADA"}:
        structure_lines.append(f"Luces: {lights_status}")

    if mudflap_status in {"NO_TRAE"}:
        structure_lines.append(f"Faldones: {mudflap_status}")

    if damage_summary:
        structure_lines.append(f"Resumen: {damage_summary}")

    needs_workshop = bool(structure_lines) or bool(any_tire_issue)

    last_eir = _fetch_last_final_eir_for_chassis(ch.id)
    eir_prev_id = int(last_eir["id"]) if last_eir and last_eir.get("id") else None

    inspection_id = _insert_dynamic("yard_gate_alamo", "chassis_inspections", {
        "site_id": site_id,
        "chassis_id": ch.id,
        "inspected_at": datetime.utcnow(),
        "inspected_by_user_id": current_user.id,
        "structure_status": structure_status or None,
        "twistlocks_status": twistlocks_status or None,
        "landing_gear_status": landing_gear_status or None,
        "lights_status": lights_status or None,
        "mudflap_status": mudflap_status or None,
        "plate_text": plate_text or None,
        "comments": comments or None,
        "needs_workshop": needs_workshop,
        "damage_summary": (damage_summary or None),
    })

    ch.site_id = site_id
    ch.is_in_yard = True
    db.session.add(ch)

    inv = ChassisInventory.query.filter_by(site_id=site_id, chassis_id=ch.id).first()
    if not inv:
        inv = ChassisInventory(
            site_id=site_id,
            chassis_id=ch.id,
            chassis_code=ch.chassis_number,
            is_in_yard=True,
        )
    else:
        inv.site_id = site_id
        inv.chassis_code = ch.chassis_number
        inv.is_in_yard = True

    db.session.add(inv)

    ticket_id = None
    if needs_workshop:
        body = _build_workshop_ticket_text(
            chassis_number=ch.chassis_number,
            axles=axles,
            structure_lines=structure_lines,
            tire_lines=tire_lines,
            eir_prev_id=eir_prev_id
        )

        ticket_id = _insert_dynamic("yard_gate_alamo", "workshop_tickets", {
            "site_id": site_id,
            "chassis_id": ch.id,
            "inspection_id": inspection_id,
            "created_at": datetime.utcnow(),
            "created_by_user_id": current_user.id,
            "status": "OPEN",
            "ticket_type": "CHASSIS_DAMAGE",
            "payload_text": body,
            "notes": body,
        })

        audit_log(
            current_user.id,
            "WORKSHOP_TICKET_CREATED_FROM_CHASSIS_CLASSIFICATION",
            "workshop_ticket",
            ticket_id,
            {"site_id": site_id, "chassis_id": ch.id, "eir_prev_id": eir_prev_id},
        )

    audit_log(
        current_user.id,
        "CHASSIS_CLASSIFIED",
        "chassis",
        ch.id,
        {"site_id": site_id, "needs_workshop": needs_workshop, "eir_prev_id": eir_prev_id},
    )

    db.session.commit()

    return jsonify({
        "ok": True,
        "needs_workshop": needs_workshop,
        "ticket_id": ticket_id,
        "eir_prev_id": eir_prev_id
    })


@yard_bp.get("/api/chassis/<int:chassis_id>/axle-seals")
@login_required
def get_chassis_axle_seals(chassis_id):
    chassis = Chassis.query.get_or_404(chassis_id)

    readings = (
        db.session.query(TireReading, TirePosition)
        .join(TirePosition, TireReading.tire_position_id == TirePosition.id)
        .filter(
            TireReading.chassis_id == chassis_id,
            TireReading.event_type == "CHASSIS_DETAIL"
        )
        .all()
    )

    result = {}

    for reading, pos in readings:
        side_code = POSITION_TO_SIDE.get(pos.position_code)
        if not side_code:
            continue

        result[side_code] = {
            "seal_1": reading.seal_1 or "",
            "seal_2": reading.seal_2 or ""
        }

    return {
        "ok": True,
        "items": result
    }


@yard_bp.post("/api/chassis/<int:chassis_id>/axle-seals")
@login_required
def save_chassis_axle_seals(chassis_id):
    data = request.get_json() or {}

    side_code = (data.get("side_code") or "").strip().upper()
    seal_1 = (data.get("seal_1") or "").strip()
    seal_2 = (data.get("seal_2") or "").strip()

    if side_code not in SIDE_TO_POSITION:
        return {"ok": False, "error": "side_code inválido"}, 400

    chassis = Chassis.query.get_or_404(chassis_id)

    position_code = SIDE_TO_POSITION[side_code]

    position = TirePosition.query.filter_by(position_code=position_code).first()
    if not position:
        return {"ok": False, "error": f"Posición {position_code} no encontrada"}, 400

    reading = (
        TireReading.query
        .filter_by(
            chassis_id=chassis_id,
            tire_position_id=position.id,
            event_type="CHASSIS_DETAIL"
        )
        .first()
    )

    if reading:
        reading.seal_1 = seal_1
        reading.seal_2 = seal_2
        reading.recorded_at = datetime.utcnow()
        reading.recorded_by_user_id = current_user.id
    else:
        reading = TireReading(
            site_id=chassis.site_id,
            chassis_id=chassis_id,
            tire_position_id=position.id,
            event_type="CHASSIS_DETAIL",
            event_id=None,
            seal_1=seal_1,
            seal_2=seal_2,
            recorded_at=datetime.utcnow(),
            recorded_by_user_id=current_user.id
        )
        db.session.add(reading)

    db.session.commit()

    return {"ok": True}