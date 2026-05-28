import io
from datetime import datetime

import pytz
from flask import render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from sqlalchemy import text

from app.blueprints.yard import yard_bp
from app.extensions import db
from app.models.container import Container
from app.models.movement import Movement
from app.models.site import Site
from app.services.audit import audit_log

from .routes import _ensure_active_site, REPORT_TYPES


CR_TZ = pytz.timezone("America/Costa_Rica")
UTC_TZ = pytz.utc


def _cr_range_to_utc_naive(date_from: str, date_to: str):
    """
    date_from/date_to vienen como YYYY-MM-DD (día CR).
    Convertimos [00:00:00 .. 23:59:59] CR -> UTC naive.
    """
    d1_local_naive = datetime.fromisoformat(date_from + "T00:00:00")
    d2_local_naive = datetime.fromisoformat(date_to + "T23:59:59")

    d1_utc = CR_TZ.localize(d1_local_naive).astimezone(UTC_TZ)
    d2_utc = CR_TZ.localize(d2_local_naive).astimezone(UTC_TZ)

    return d1_utc.replace(tzinfo=None), d2_utc.replace(tzinfo=None)


def _parse_report_filters(args):
    movement_type = (args.get("movement_type") or "").strip().upper()
    if movement_type and movement_type not in REPORT_TYPES:
        movement_type = ""

    date_from = args.get("date_from")
    date_to = args.get("date_to")

    if not date_from or not date_to:
        return None, None, None, "Indica rango de fechas."

    try:
        d1, d2 = _cr_range_to_utc_naive(date_from, date_to)
    except Exception:
        return None, None, None, "Formato de fecha inválido (usa YYYY-MM-DD)."

    if d2 < d1:
        return None, None, None, "El rango de fechas es inválido (Hasta < Desde)."

    return movement_type, d1, d2, None


def _query_report_rows(site_id, movement_type, d1, d2):
    q = (
        db.session.query(Movement, Container)
        .join(Container, Container.id == Movement.container_id)
        .filter(Movement.site_id == site_id)
        .filter(Movement.occurred_at >= d1, Movement.occurred_at <= d2)
    )

    if movement_type:
        q = q.filter(Movement.movement_type == movement_type)

    return q.order_by(Movement.occurred_at.asc()).all()


@yard_bp.get("/reports")
@login_required
def reports_view():
    return render_template("yard/reports.html", rows=None, movement_type="", date_from="", date_to="")


@yard_bp.get("/reports/run")
@login_required
def reports_run():
    site_id = _ensure_active_site()

    movement_type, d1, d2, err = _parse_report_filters(request.args)
    if err:
        flash(err, "danger")
        return redirect(url_for("yard.reports_view"))

    rows = _query_report_rows(site_id, movement_type, d1, d2)

    audit_log(
        current_user.id,
        "REPORT_RUN",
        "report",
        None,
        {
            "from": request.args.get("date_from"),
            "to": request.args.get("date_to"),
            "movement_type": movement_type or "ALL",
            "site_id": site_id,
        },
    )
    db.session.commit()

    return render_template(
        "yard/reports.html",
        rows=rows,
        date_from=request.args.get("date_from"),
        date_to=request.args.get("date_to"),
        movement_type=movement_type,
    )


@yard_bp.get("/reports/export")
@login_required
def reports_export():
    site_id = _ensure_active_site()

    movement_type, d1, d2, err = _parse_report_filters(request.args)
    if err:
        flash(err, "danger")
        return redirect(url_for("yard.reports_view"))

    rows = _query_report_rows(site_id, movement_type, d1, d2)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        flash("No se puede exportar: falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.reports_run", **request.args))

    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = ["Fecha/Hora", "Movimiento", "Contenedor", "Ubicación", "Chofer", "Placa"]
    ws.append(headers)

    for mv, c in rows:
        loc = "—"
        if mv.bay_code:
            parts = [mv.bay_code]
            if mv.depth_row:
                parts.append(f"F{int(mv.depth_row):02d}")
            if mv.tier:
                parts.append(f"N{int(mv.tier)}")
            loc = " ".join(parts)

        ws.append([
            mv.occurred_at.strftime("%Y-%m-%d %H:%M:%S") if mv.occurred_at else "",
            mv.movement_type or "",
            c.code if c else "",
            loc,
            mv.driver_name or "",
            mv.truck_plate or "",
        ])

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for cell in ws[get_column_letter(col_idx)]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)

    audit_log(
        current_user.id,
        "REPORT_EXPORTED",
        "report",
        None,
        {
            "from": request.args.get("date_from"),
            "to": request.args.get("date_to"),
            "movement_type": movement_type or "ALL",
            "rows": len(rows),
            "site_id": site_id,
        },
    )
    db.session.commit()

    mt = movement_type or "ALL"
    fname = f"reportes_{mt}_{request.args.get('date_from')}_a_{request.args.get('date_to')}.xlsx"

    return send_file(
        buff,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================
# Reportes nuevos / dashboard
# =========================

@yard_bp.get("/reportes")
@login_required
def reports_dashboard():
    site_id = _ensure_active_site()
    active_site = Site.query.get(site_id)

    return render_template(
        "yard/reports_dashboard.html",
        active_site=active_site,
    )


@yard_bp.get("/reportes/chasis-fuera")
@login_required
def report_chassis_outside():
    _ensure_active_site()

    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    origin_site_id = (request.args.get("origin_site_id") or "").strip()

    filters = []
    params = {}

    if date_from:
        filters.append("""
            COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date) >= :date_from
        """)
        params["date_from"] = date_from

    if date_to:
        filters.append("""
            COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date) <= :date_to
        """)
        params["date_to"] = date_to

    if origin_site_id:
        try:
            origin_site_id_int = int(origin_site_id)
            filters.append("e.site_id = :origin_site_id")
            params["origin_site_id"] = origin_site_id_int
        except Exception:
            origin_site_id = ""

    extra_where = ""
    if filters:
        extra_where = " AND " + " AND ".join(filters)

    sql = text(f"""
        SELECT
            c.id,
            c.chassis_number,
            c.plate,
            c.length_ft,
            c.axles,
            s.name AS origin_name,
            e.destination,
            COALESCE(e.inventory_out_at, e.finalized_at, e.trip_date) AS departure_at,
            CURRENT_DATE - COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date) AS days_out,
            c.status
        FROM yard_gate_alamo.chassis c
        LEFT JOIN LATERAL (
            SELECT *
            FROM yard_gate_alamo.eirs e
            WHERE e.chassis_id = c.id
              AND e.status = 'CONFIRMED'
            ORDER BY
                e.inventory_out_at DESC NULLS LAST,
                e.finalized_at DESC NULLS LAST,
                e.trip_date DESC,
                e.id DESC
            LIMIT 1
        ) e ON TRUE
        LEFT JOIN yard_gate_alamo.sites s
            ON s.id = e.site_id
        WHERE c.is_in_yard = false
          {extra_where}
        ORDER BY
            days_out DESC NULLS LAST,
            departure_at DESC NULLS LAST,
            c.chassis_number ASC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    total = 0
    for r in rows:
        total += 1

        departure_at = r["departure_at"]
        days_out = r["days_out"]

        if departure_at and hasattr(departure_at, "strftime"):
            departure_at_str = (
                departure_at.strftime("%d/%m/%Y %I:%M %p")
                if hasattr(departure_at, "hour")
                else departure_at.strftime("%d/%m/%Y")
            )
        else:
            departure_at_str = ""

        items.append({
            "id": r["id"],
            "chassis_number": r["chassis_number"],
            "plate": r["plate"] or "",
            "length_ft": r["length_ft"] or "",
            "axles": r["axles"] or "",
            "origin_name": r["origin_name"] or "",
            "destination": r["destination"] or "Fuera de patio",
            "departure_at": departure_at,
            "departure_at_str": departure_at_str,
            "days_out": int(days_out) if days_out is not None else 0,
            "status": r["status"] or "",
        })

    sites = (
        Site.query
        .filter(Site.id.in_([2, 3, 4]))
        .order_by(Site.name.asc())
        .all()
    )

    return render_template(
        "yard/report_chassis_outside.html",
        items=items,
        total=total,
        sites=sites,
        date_from=date_from,
        date_to=date_to,
        origin_site_id=origin_site_id,
    )


@yard_bp.get("/reportes/movimientos-contenedor")
@login_required
def report_container_movements():
    site_id = _ensure_active_site()

    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    movement_type = (request.args.get("movement_type") or "").strip().upper()

    valid_types = {"", "GATE_IN", "GATE_OUT"}
    if movement_type not in valid_types:
        movement_type = ""

    filters = ["mv.site_id = :site_id", "mv.movement_type IN ('GATE_IN', 'GATE_OUT')"]
    params = {"site_id": site_id}

    if date_from:
        filters.append("mv.occurred_at::date >= :date_from")
        params["date_from"] = date_from

    if date_to:
        filters.append("mv.occurred_at::date <= :date_to")
        params["date_to"] = date_to

    if movement_type:
        filters.append("mv.movement_type = :movement_type")
        params["movement_type"] = movement_type

    where_sql = " AND ".join(filters)

    sql = text(f"""
        SELECT
            mv.id AS movement_id,
            mv.occurred_at,
            mv.movement_type,
            c.id AS container_id,
            c.code AS container_code,

            -- Chasis para GATE_IN
            ch_in.chassis_number AS chassis_gate_in,

            -- Chasis para GATE_OUT
            ch_out.chassis_number AS chassis_gate_out,

            -- Origen para GATE_IN = destino del último EIR confirmado anterior
            prev_eir.destination AS origin_name,

            -- Destino para GATE_OUT
            eir_out.destination AS destination_name

        FROM yard_gate_alamo.movements mv
        JOIN yard_gate_alamo.containers c
          ON c.id = mv.container_id

        -- Chasis del GATE_IN (vía tire_readings)
        LEFT JOIN LATERAL (
            SELECT tr.chassis_id
            FROM yard_gate_alamo.tire_readings tr
            WHERE tr.event_type = 'GATE_IN'
              AND tr.event_id = mv.id
              AND tr.chassis_id IS NOT NULL
            ORDER BY tr.recorded_at DESC NULLS LAST, tr.id DESC
            LIMIT 1
        ) tr_in ON TRUE
        LEFT JOIN yard_gate_alamo.chassis ch_in
          ON ch_in.id = tr_in.chassis_id

        -- EIR del GATE_OUT actual
        LEFT JOIN yard_gate_alamo.eirs eir_out
          ON eir_out.gate_out_movement_id = mv.id
         AND eir_out.status = 'CONFIRMED'

        LEFT JOIN yard_gate_alamo.chassis ch_out
          ON ch_out.id = eir_out.chassis_id

        -- EIR anterior del mismo contenedor para resolver ORIGEN del GATE_IN
        LEFT JOIN LATERAL (
            SELECT e_prev.destination
            FROM yard_gate_alamo.eirs e_prev
            WHERE e_prev.container_id = mv.container_id
              AND e_prev.status = 'CONFIRMED'
              AND COALESCE(e_prev.inventory_out_at, e_prev.finalized_at, e_prev.created_at) < mv.occurred_at
            ORDER BY
              COALESCE(e_prev.inventory_out_at, e_prev.finalized_at, e_prev.created_at) DESC,
              e_prev.id DESC
            LIMIT 1
        ) prev_eir ON TRUE

        WHERE {where_sql}
        ORDER BY mv.occurred_at DESC, mv.id DESC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    for r in rows:
        movement_type_row = (r["movement_type"] or "").upper()

        chassis_number = ""
        origin_name = ""
        destination_name = ""

        if movement_type_row == "GATE_IN":
            chassis_number = r["chassis_gate_in"] or ""
            origin_name = r["origin_name"] or ""
            destination_name = ""
        elif movement_type_row == "GATE_OUT":
            chassis_number = r["chassis_gate_out"] or ""
            origin_name = ""
            destination_name = r["destination_name"] or ""

        occurred_at = r["occurred_at"]
        occurred_at_str = occurred_at.strftime("%d/%m/%Y %I:%M %p") if occurred_at else ""

        items.append({
            "movement_id": r["movement_id"],
            "container_id": r["container_id"],
            "container_code": r["container_code"] or "",
            "occurred_at": occurred_at,
            "occurred_at_str": occurred_at_str,
            "movement_type": movement_type_row,
            "chassis_number": chassis_number,
            "origin_name": origin_name,
            "destination_name": destination_name,
        })

    return render_template(
        "yard/report_container_movements.html",
        items=items,
        total=len(items),
        date_from=date_from,
        date_to=date_to,
        movement_type=movement_type,
    )


@yard_bp.get("/reportes/movimientos-chasis")
@login_required
def report_chassis_movements():
    site_id = _ensure_active_site()

    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    movement_type = (request.args.get("movement_type") or "").strip().upper()

    valid_types = {"", "GATE_IN", "GATE_OUT"}
    if movement_type not in valid_types:
        movement_type = ""

    filters_in = ["ci.site_id = :site_id"]
    filters_out = ["e.site_id = :site_id", "e.status = 'CONFIRMED'", "e.chassis_id IS NOT NULL"]
    params = {"site_id": site_id}

    if date_from:
        filters_in.append("ci.inspected_at::date >= :date_from")
        filters_out.append("COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date, e.created_at::date) >= :date_from")
        params["date_from"] = date_from

    if date_to:
        filters_in.append("ci.inspected_at::date <= :date_to")
        filters_out.append("COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date, e.created_at::date) <= :date_to")
        params["date_to"] = date_to

    if movement_type == "GATE_IN":
        enable_in = True
        enable_out = False
    elif movement_type == "GATE_OUT":
        enable_in = False
        enable_out = True
    else:
        enable_in = True
        enable_out = True

    sql_parts = []

    if enable_in:
        sql_parts.append(f"""
            SELECT
                ci.inspected_at AS event_at,
                'GATE_IN' AS movement_type,
                ch.id AS chassis_id,
                ch.chassis_number,
                ch.plate,
                ch.length_ft,
                ch.axles,
                COALESCE(prev_eir.destination, '') AS origin_name,
                '' AS destination_name,
                COALESCE(u.username, '') AS username
            FROM yard_gate_alamo.chassis_inspections ci
            JOIN yard_gate_alamo.chassis ch
              ON ch.id = ci.chassis_id
            LEFT JOIN yard_gate_alamo.users u
              ON u.id = ci.inspected_by_user_id
            LEFT JOIN LATERAL (
                SELECT e_prev.destination
                FROM yard_gate_alamo.eirs e_prev
                WHERE e_prev.chassis_id = ci.chassis_id
                  AND e_prev.status = 'CONFIRMED'
                  AND COALESCE(e_prev.inventory_out_at, e_prev.finalized_at, e_prev.created_at) < ci.inspected_at
                ORDER BY
                  COALESCE(e_prev.inventory_out_at, e_prev.finalized_at, e_prev.created_at) DESC,
                  e_prev.id DESC
                LIMIT 1
            ) prev_eir ON TRUE
            WHERE {" AND ".join(filters_in)}
        """)

    if enable_out:
        sql_parts.append(f"""
            SELECT
                COALESCE(e.inventory_out_at, e.finalized_at, e.created_at) AS event_at,
                'GATE_OUT' AS movement_type,
                ch.id AS chassis_id,
                ch.chassis_number,
                ch.plate,
                ch.length_ft,
                ch.axles,
                COALESCE(s.name, '') AS origin_name,
                COALESCE(e.destination, '') AS destination_name,
                COALESCE(u.username, '') AS username
            FROM yard_gate_alamo.eirs e
            JOIN yard_gate_alamo.chassis ch
              ON ch.id = e.chassis_id
            LEFT JOIN yard_gate_alamo.sites s
              ON s.id = e.site_id
            LEFT JOIN yard_gate_alamo.users u
              ON u.id = COALESCE(e.last_edited_by_user_id, e.created_by_user_id)
            WHERE {" AND ".join(filters_out)}
        """)

    sql = text(f"""
        SELECT *
        FROM (
            {" UNION ALL ".join(sql_parts)}
        ) x
        ORDER BY x.event_at DESC, x.chassis_number ASC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    for r in rows:
        event_at = r["event_at"]
        event_at_str = event_at.strftime("%d/%m/%Y %I:%M %p") if event_at else ""

        items.append({
            "event_at": event_at,
            "event_at_str": event_at_str,
            "movement_type": r["movement_type"] or "",
            "chassis_id": r["chassis_id"],
            "chassis_number": r["chassis_number"] or "",
            "plate": r["plate"] or "",
            "length_ft": r["length_ft"] or "",
            "axles": r["axles"] or "",
            "origin_name": r["origin_name"] or "",
            "destination_name": r["destination_name"] or "",
            "username": r["username"] or "",
        })

    return render_template(
        "yard/report_chassis_movements.html",
        items=items,
        total=len(items),
        date_from=date_from,
        date_to=date_to,
        movement_type=movement_type,
    )