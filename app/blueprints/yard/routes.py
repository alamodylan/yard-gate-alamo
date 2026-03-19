# app/blueprints/yard/routes.py
import re
import os
import io
from datetime import datetime, timedelta, date
from io import BytesIO
import json
from sqlalchemy import or_

import pytz
import requests
from sqlalchemy import text, bindparam  # ✅ para SQL directo (predios EIR/chasis)

from flask import (
    render_template,
    request,
    redirect,
    url_for,
    flash,
    jsonify,
    send_file,
    session,
    abort,
)
from flask_login import login_required, current_user

from app.blueprints.yard import yard_bp
from app.extensions import db
from app.models.yard import YardBlock, YardBay
from app.models.container import Container, ContainerPosition
from app.models.movement import Movement, MovementPhoto
from app.models.ticket import TicketPrint
from app.models.site import Site, UserSite
from app.services.audit import audit_log
from app.services.yard_logic import find_first_free_slot
from app.services.storage import get_storage, build_photo_key
from app.services.ticketing import build_ticket_payload
from app.models.eir import EIR, EIRContainerDamage
from app.models.chassis import ChassisInventory
from app.models.chassis import Chassis
from app.models.tire import Tire
from app.models.chassis_tire import ChassisTire
from app.models.tire_retread_event import TireRetreadEvent

CR_TZ = pytz.timezone("America/Costa_Rica")
UTC_TZ = pytz.utc

CONTAINER_RE = re.compile(r"^[A-Z]{4}-\d{6}-\d$")
SIZES = [
    "40HC",
    "40ST",
    "40RF",
    "40OT",
    "20ST",
    "20OT",
    "20RF",
    "20TQ",
    "45HC",
]
APP_NAME = "Yard Gate Álamo"

REPORT_TYPES = {"GATE_IN", "GATE_OUT", "MOVE"}

CHASSIS_NUM_RE = re.compile(r"^\d{5}$")
TIRE_STATES = {"OK", "GASTADA", "PINCHADA", "CAMBIAR", "NO_APTA"}

# ✅ Predios reales (ATM no es predio; MAERSK queda igual)
PREDIO_CODES = {"COYOL", "CALDERA", "LIMON"}


# =========================
# Multi-predio helpers (site)
# =========================
def _allowed_sites_for_user(user):
    if getattr(user, "role", None) == "admin":
        return Site.query.filter_by(is_active=True).order_by(Site.name.asc()).all()

    return (
        db.session.query(Site)
        .join(UserSite, UserSite.site_id == Site.id)
        .filter(UserSite.user_id == user.id, Site.is_active == True)  # noqa: E712
        .order_by(Site.name.asc())
        .all()
    )


def _get_active_site_id():
    return session.get("active_site_id")


def _set_active_site_id(site_id: int):
    session["active_site_id"] = int(site_id)


def _ensure_active_site():
    allowed = _allowed_sites_for_user(current_user)
    if not allowed:
        abort(403)

    active_id = _get_active_site_id()
    allowed_ids = {s.id for s in allowed}

    if not active_id or active_id not in allowed_ids:
        _set_active_site_id(allowed[0].id)
        active_id = allowed[0].id

    return active_id


# ==========================================================
# Predio activo (helper para templates dinámicos)
# ==========================================================
def _active_site():
    site_id = session.get("active_site_id")
    if not site_id:
        return None
    return Site.query.get(site_id)


def _active_site_key():
    site = _active_site()
    if not site:
        return ""
    value = getattr(site, "code", None) or getattr(site, "name", None) or ""
    return value.strip().upper()


@yard_bp.app_context_processor
def inject_active_site():
    return {"active_site_key": _active_site_key()}


def _is_predio_site(site_id: int) -> bool:
    """
    True si el predio activo es COYOL/CALDERA/LIMON.
    MAERSK y cualquier otro se queda con el flujo actual.
    """
    s = Site.query.get(site_id)
    return bool(s and (s.code or "").upper() in PREDIO_CODES)


def _register_ticket_print(site_id: int, movement_id: int, printed_by_user_id: int, payload: str) -> TicketPrint:
    row = TicketPrint(
        site_id=site_id,
        movement_id=movement_id,
        printed_by_user_id=printed_by_user_id,
        ticket_payload=payload,
        printed_at=datetime.utcnow(),
    )
    db.session.add(row)
    return row


# =========================
# Dashboard Sites
# =========================
@yard_bp.get("/sites")
@login_required
def sites_dashboard():
    allowed = _allowed_sites_for_user(current_user)
    active_id = _get_active_site_id()
    return render_template("yard/sites.html", sites=allowed, active_site_id=active_id)


@yard_bp.post("/sites/select")
@login_required
def sites_select():
    site_id = request.form.get("site_id")
    if not site_id or not str(site_id).isdigit():
        flash("Predio inválido.", "danger")
        return redirect(url_for("yard.sites_dashboard"))

    site_id = int(site_id)
    allowed_ids = {s.id for s in _allowed_sites_for_user(current_user)}
    if site_id not in allowed_ids:
        flash("No tienes acceso a ese predio.", "danger")
        return redirect(url_for("yard.sites_dashboard"))

    _set_active_site_id(site_id)
    return redirect(url_for("yard.map_view"))


@yard_bp.get("/")
@login_required
def home():
    allowed = _allowed_sites_for_user(current_user)
    if len(allowed) == 1:
        _set_active_site_id(allowed[0].id)
        return redirect(url_for("yard.map_view"))
    return redirect(url_for("yard.sites_dashboard"))


@yard_bp.get("/map")
@login_required
def map_view():
    site_id = _ensure_active_site()

    blocks = (
        YardBlock.query
        .filter_by(site_id=site_id)
        .order_by(YardBlock.code.asc())
        .all()
    )

    selected_block = (request.args.get("block") or "A").upper()
    if selected_block not in {"A", "B", "C", "D"}:
        selected_block = "A"

    return render_template("yard/map.html", blocks=blocks, selected_block=selected_block)


@yard_bp.get("/bay/<string:bay_code>")
@login_required
def bay_detail_view(bay_code: str):
    site_id = _ensure_active_site()

    bay_code = bay_code.upper()
    bay = YardBay.query.filter_by(code=bay_code, is_active=True, site_id=site_id).first_or_404()

    rows = (
        db.session.query(Container, ContainerPosition)
        .join(ContainerPosition, ContainerPosition.container_id == Container.id)
        .filter(ContainerPosition.bay_id == bay.id, Container.site_id == site_id)
        .order_by(ContainerPosition.depth_row.asc(), ContainerPosition.tier.asc())
        .all()
    )

    items = []
    for c, p in rows:
        items.append(
            {
                "id": c.id,
                "code": c.code,
                "size": c.size,
                "depth_row": p.depth_row,
                "tier": p.tier,
            }
        )

    return render_template("yard/bay_detail.html", bay=bay, items=items)


# =========================
# APIs mapa / bandeja
# =========================
@yard_bp.get("/api/yard/containers-in-yard")
@login_required
def api_containers_in_yard():
    """
    Para bandeja (mapa):
    Retorna contenedores en patio con su posición actual.
    """
    site_id = _ensure_active_site()

    rows = (
        db.session.query(Container, ContainerPosition, YardBay)
        .join(ContainerPosition, ContainerPosition.container_id == Container.id)
        .join(YardBay, YardBay.id == ContainerPosition.bay_id)
        .filter(Container.is_in_yard == True, Container.site_id == site_id)  # noqa: E712
        .order_by(YardBay.code.asc(), ContainerPosition.depth_row.asc(), ContainerPosition.tier.asc())
        .all()
    )

    payload = []
    for c, p, bay in rows:
        payload.append(
            {
                "id": c.id,
                "code": c.code,
                "size": c.size,
                "year": c.year,
                "status_notes": c.status_notes,
                "position": {
                    "bay_code": bay.code,
                    "depth_row": p.depth_row,
                    "tier": p.tier,
                },
            }
        )

    return jsonify({"rows": payload})


@yard_bp.get("/api/yard/bays")
@login_required
def api_bays_by_block():
    site_id = _ensure_active_site()

    block_code = (request.args.get("block") or "").upper()
    block = YardBlock.query.filter_by(code=block_code, site_id=site_id).first()
    if not block:
        return jsonify({"bays": []})

    bays = (
        YardBay.query.filter_by(block_id=block.id, is_active=True, site_id=site_id)
        .order_by(YardBay.bay_number.asc())
        .all()
    )
    return jsonify({"bays": [{"id": b.id, "bay_number": b.bay_number, "code": b.code} for b in bays]})


@yard_bp.get("/api/yard/map")
@login_required
def api_yard_map():
    """
    Devuelve las estibas del bloque con conteo (used/capacity).
    Ideal: incluye x,y,w,h y límites para permitir layout visual real en frontend.
    """
    site_id = _ensure_active_site()

    block_code = (request.args.get("block") or "A").upper()
    block = YardBlock.query.filter_by(code=block_code, site_id=site_id).first()
    if not block:
        return jsonify({"error": "Bloque inválido"}), 400

    bays = (
        YardBay.query.filter_by(block_id=block.id, is_active=True, site_id=site_id)
        .order_by(YardBay.bay_number.asc())
        .all()
    )

    counts = dict(
        db.session.query(ContainerPosition.bay_id, db.func.count(ContainerPosition.container_id))
        .join(YardBay, YardBay.id == ContainerPosition.bay_id)
        .filter(YardBay.site_id == site_id)
        .group_by(ContainerPosition.bay_id)
        .all()
    )

    payload = []
    for b in bays:
        capacity = b.max_depth_rows * b.max_tiers
        used = int(counts.get(b.id, 0))
        payload.append(
            {
                "id": b.id,
                "code": b.code,
                "bay_number": b.bay_number,
                "used": used,
                "capacity": capacity,
                "max_depth_rows": b.max_depth_rows,
                "max_tiers": b.max_tiers,
                "x": b.x,
                "y": b.y,
                "w": b.w,
                "h": b.h,
            }
        )

    return jsonify({"block": block_code, "bays": payload})


@yard_bp.get("/api/yard/block/<string:block_code>/availability")
@login_required
def api_block_availability(block_code: str):
    """
    Disponibilidad por estiba (verde/rojo) basada en capacidad total.
    """
    site_id = _ensure_active_site()

    block_code = (block_code or "").upper()
    block = YardBlock.query.filter_by(code=block_code, site_id=site_id).first()
    if not block:
        return jsonify({"error": "Bloque inválido"}), 400

    bays = (
        YardBay.query.filter_by(block_id=block.id, is_active=True, site_id=site_id)
        .order_by(YardBay.bay_number.asc())
        .all()
    )

    counts = dict(
        db.session.query(ContainerPosition.bay_id, db.func.count(ContainerPosition.container_id))
        .join(YardBay, YardBay.id == ContainerPosition.bay_id)
        .filter(YardBay.site_id == site_id)
        .group_by(ContainerPosition.bay_id)
        .all()
    )

    payload = []
    for b in bays:
        capacity = b.max_depth_rows * b.max_tiers
        used = int(counts.get(b.id, 0))
        free = capacity - used
        payload.append(
            {
                "id": b.id,
                "code": b.code,
                "bay_number": b.bay_number,
                "used": used,
                "capacity": capacity,
                "free": free,
                "available": free > 0,
            }
        )

    return jsonify({"block": block_code, "bays": payload})


@yard_bp.get("/api/yard/bays/<string:bay_code>/last-available")
@login_required
def api_bay_last_available(bay_code: str):
    """
    Devuelve la sugerencia de slot según la regla REAL del sistema:
    - más adentro primero (depth_row más alto)
    - tier automático (1..max)
    """
    site_id = _ensure_active_site()

    bay_code = (bay_code or "").upper()
    bay = YardBay.query.filter_by(code=bay_code, is_active=True, site_id=site_id).first()
    if not bay:
        return jsonify({"error": "Estiba inválida"}), 400

    slot = find_first_free_slot(bay.id)
    if not slot:
        return jsonify({"ok": False, "error": "BAY_FULL"}), 409

    depth_row, tier = slot
    return jsonify({"ok": True, "bay_code": bay.code, "depth_row": depth_row, "tier": tier})


@yard_bp.get("/api/yard/bays/<string:bay_code>/rows-availability")
@login_required
def api_bay_rows_availability(bay_code: str):
    """
    Disponibilidad por FILAS para una estiba.
    Ideal: incluye suggested_tier por fila para evitar request extra.
    """
    site_id = _ensure_active_site()

    bay_code = (bay_code or "").upper()
    bay = YardBay.query.filter_by(code=bay_code, is_active=True, site_id=site_id).first()
    if not bay:
        return jsonify({"error": "Estiba inválida"}), 400

    max_levels = int(bay.max_tiers or 4)

    counts_by_row = dict(
        db.session.query(ContainerPosition.depth_row, db.func.count(ContainerPosition.container_id))
        .join(Container, Container.id == ContainerPosition.container_id)
        .filter(ContainerPosition.bay_id == bay.id, Container.site_id == site_id)
        .group_by(ContainerPosition.depth_row)
        .all()
    )

    rows = []
    for row_num in range(1, int(bay.max_depth_rows or 1) + 1):
        used = int(counts_by_row.get(row_num, 0))
        is_full = used >= max_levels

        suggested_tier = None
        if not is_full:
            occ_tiers = {
                int(t[0]) for t in db.session.query(ContainerPosition.tier)
                .join(Container, Container.id == ContainerPosition.container_id)
                .filter(
                    ContainerPosition.bay_id == bay.id,
                    ContainerPosition.depth_row == row_num,
                    Container.site_id == site_id
                )
                .all()
            }
            for t in range(1, max_levels + 1):
                if t not in occ_tiers:
                    suggested_tier = t
                    break

        rows.append(
            {
                "row": row_num,
                "levels_used": used,
                "max_levels": max_levels,
                "is_full": is_full,
                "suggested_tier": suggested_tier,
            }
        )

    return jsonify({"ok": True, "bay_code": bay.code, "rows": rows})


@yard_bp.get("/api/yard/bays/<string:bay_code>/row/<int:row_number>/suggest-tier")
@login_required
def api_bay_row_suggest_tier(bay_code: str, row_number: int):
    """
    Sugerir tier exacto dentro de una fila específica (1..max_tiers).
    Mantenerlo aunque rows-availability tenga suggested_tier, para “fuente de verdad”.
    """
    site_id = _ensure_active_site()

    bay_code = (bay_code or "").upper()
    bay = YardBay.query.filter_by(code=bay_code, is_active=True, site_id=site_id).first()
    if not bay:
        return jsonify({"error": "Estiba inválida"}), 400

    if row_number < 1 or row_number > int(bay.max_depth_rows or 0):
        return jsonify({"ok": False, "error": "ROW_OUT_OF_RANGE"}), 400

    occupied = (
        db.session.query(ContainerPosition.tier)
        .join(Container, Container.id == ContainerPosition.container_id)
        .filter(
            ContainerPosition.bay_id == bay.id,
            ContainerPosition.depth_row == row_number,
            Container.site_id == site_id
        )
        .all()
    )
    occupied_tiers = {int(t[0]) for t in occupied}

    for tier in range(1, int(bay.max_tiers or 1) + 1):
        if tier not in occupied_tiers:
            return jsonify({"ok": True, "bay_code": bay.code, "depth_row": row_number, "tier": tier})

    return jsonify({"ok": False, "error": "ROW_FULL"}), 409


@yard_bp.post("/api/yard/place")
@login_required
def api_place_container():
    """
    Coloca un contenedor en una estiba.
    Compatibilidad:
      - Viejo: { "container_id": 123, "to_bay_code": "A07" } -> AUTO (find_first_free_slot)
      - Nuevo: { "container_id": 123, "to_bay_code": "A07", "to_depth_row": 10, "to_tier": 2 } -> EXACTO
    """
    site_id = _ensure_active_site()

    data = request.get_json(silent=True) or {}
    container_id = data.get("container_id")
    to_bay_code = (data.get("to_bay_code") or "").upper()

    if not container_id or not to_bay_code:
        return jsonify({"error": "Datos incompletos"}), 400

    c = Container.query.get(container_id)
    if not c or not c.is_in_yard or c.site_id != site_id:
        return jsonify({"error": "Contenedor no existe o no está en patio (predio actual)"}), 400

    to_bay = YardBay.query.filter_by(code=to_bay_code, is_active=True, site_id=site_id).first()
    if not to_bay:
        return jsonify({"error": "Estiba destino inválida"}), 400

    # Lock de estiba para evitar colisiones en concurrencia (sin unique constraint DB)
    db.session.query(YardBay).filter(YardBay.id == to_bay.id).with_for_update().one()

    to_depth_row = data.get("to_depth_row")
    to_tier = data.get("to_tier")

    if to_depth_row is not None and to_tier is not None:
        try:
            depth_row = int(to_depth_row)
            tier = int(to_tier)
        except Exception:
            return jsonify({"error": "Fila/Nivel inválidos"}), 400

        if not (1 <= depth_row <= to_bay.max_depth_rows) or not (1 <= tier <= to_bay.max_tiers):
            return jsonify({"error": "Fila/Nivel fuera de rango"}), 400

        occupied = ContainerPosition.query.filter_by(
            bay_id=to_bay.id,
            depth_row=depth_row,
            tier=tier
        ).first()
        if occupied:
            return jsonify({"error": "Slot ocupado"}), 409
    else:
        slot = find_first_free_slot(to_bay.id)
        if not slot:
            return jsonify({"error": "Estiba llena"}), 409
        depth_row, tier = slot

    old_pos = ContainerPosition.query.filter_by(container_id=c.id).first()
    old = None
    if old_pos:
        old_bay = YardBay.query.get(old_pos.bay_id)
        old = {
            "bay_code": old_bay.code if old_bay else None,
            "depth_row": old_pos.depth_row,
            "tier": old_pos.tier,
        }

    ContainerPosition.query.filter_by(container_id=c.id).delete()
    db.session.add(
        ContainerPosition(
            container_id=c.id,
            bay_id=to_bay.id,
            depth_row=depth_row,
            tier=tier,
            placed_by_user_id=current_user.id,
        )
    )

    mv = Movement(
        site_id=site_id,
        container_id=c.id,
        movement_type="MOVE",
        occurred_at=datetime.utcnow(),
        bay_code=to_bay.code,
        depth_row=depth_row,
        tier=tier,
        created_by_user_id=current_user.id,
        notes="PLACED_BY_BLOCK_UI",
    )
    db.session.add(mv)

    audit_log(
        current_user.id,
        "CONTAINER_PLACED",
        "container",
        c.id,
        {
            "from": old,
            "to": {"bay_code": to_bay.code, "depth_row": depth_row, "tier": tier},
            "rule": "AUTO_LAST_AVAILABLE" if (to_depth_row is None or to_tier is None) else "MANUAL_EXACT",
            "site_id": site_id,
        },
    )

    db.session.commit()
    return jsonify({"ok": True, "bay_code": to_bay.code, "depth_row": depth_row, "tier": tier})


@yard_bp.post("/api/yard/move")
@login_required
def api_move_container():
    """
    Drag & drop (detalle de estiba): mueve contenedor a otra estiba.
    Payload:
      { "container_id": 123, "to_bay_code": "B07", "mode": "auto" }
      o manual:
      { "container_id": 123, "to_bay_code": "B07", "mode": "manual", "depth_row": 1, "tier": 2 }
    """
    site_id = _ensure_active_site()

    data = request.get_json(silent=True) or {}
    container_id = data.get("container_id")
    to_bay_code = (data.get("to_bay_code") or "").upper()
    mode = (data.get("mode") or "auto").lower()

    if not container_id or not to_bay_code:
        return jsonify({"error": "Datos incompletos"}), 400

    c = Container.query.get(container_id)
    if not c or not c.is_in_yard or c.site_id != site_id:
        return jsonify({"error": "Contenedor no existe o no está en patio (predio actual)"}), 400

    to_bay = YardBay.query.filter_by(code=to_bay_code, is_active=True, site_id=site_id).first()
    if not to_bay:
        return jsonify({"error": "Estiba destino inválida"}), 400

    # Lock de estiba destino
    db.session.query(YardBay).filter(YardBay.id == to_bay.id).with_for_update().one()

    if mode == "manual":
        try:
            depth_row = int(data.get("depth_row"))
            tier = int(data.get("tier"))
        except Exception:
            return jsonify({"error": "Fila/Nivel inválidos"}), 400

        if not (1 <= depth_row <= to_bay.max_depth_rows) or not (1 <= tier <= to_bay.max_tiers):
            return jsonify({"error": "Fila/Nivel fuera de rango"}), 400

        occupied = ContainerPosition.query.filter_by(bay_id=to_bay.id, depth_row=depth_row, tier=tier).first()
        if occupied:
            return jsonify({"error": "Slot ocupado"}), 409
    else:
        slot = find_first_free_slot(to_bay.id)
        if not slot:
            return jsonify({"error": "Estiba llena"}), 409
        depth_row, tier = slot

    old_pos = ContainerPosition.query.filter_by(container_id=c.id).first()
    old = None
    if old_pos:
        old_bay = YardBay.query.get(old_pos.bay_id)
        old = {
            "bay_code": old_bay.code if old_bay else None,
            "depth_row": old_pos.depth_row,
            "tier": old_pos.tier,
        }

    ContainerPosition.query.filter_by(container_id=c.id).delete()
    db.session.add(
        ContainerPosition(
            container_id=c.id,
            bay_id=to_bay.id,
            depth_row=depth_row,
            tier=tier,
            placed_by_user_id=current_user.id,
        )
    )

    mv = Movement(
        site_id=site_id,
        container_id=c.id,
        movement_type="MOVE",
        occurred_at=datetime.utcnow(),
        bay_code=to_bay.code,
        depth_row=depth_row,
        tier=tier,
        created_by_user_id=current_user.id,
        notes=None,
    )
    db.session.add(mv)

    audit_log(
        current_user.id,
        "CONTAINER_MOVED",
        "container",
        c.id,
        {"from": old, "to": {"bay_code": to_bay.code, "depth_row": depth_row, "tier": tier}, "site_id": site_id},
    )

    db.session.commit()
    return jsonify({"ok": True, "bay_code": to_bay.code, "depth_row": depth_row, "tier": tier})


@yard_bp.get("/api/yard/bays/<string:bay_code>/row/<int:row_number>/containers")
@login_required
def api_bay_row_containers(bay_code: str, row_number: int):
    site_id = _ensure_active_site()

    bay_code = (bay_code or "").upper()
    bay = YardBay.query.filter_by(code=bay_code, is_active=True, site_id=site_id).first()
    if not bay:
        return jsonify({"ok": False, "error": "BAY_NOT_FOUND"}), 404

    if row_number < 1 or row_number > int(bay.max_depth_rows or 0):
        return jsonify({"ok": False, "error": "ROW_OUT_OF_RANGE"}), 400

    rows = (
        db.session.query(Container, ContainerPosition)
        .join(ContainerPosition, ContainerPosition.container_id == Container.id)
        .filter(
            ContainerPosition.bay_id == bay.id,
            ContainerPosition.depth_row == row_number,
            Container.site_id == site_id
        )
        .order_by(ContainerPosition.tier.asc())
        .all()
    )

    items = []
    for c, p in rows:
        items.append({
            "id": c.id,
            "code": c.code,
            "size": c.size,
            "tier": p.tier
        })

    return jsonify({"ok": True, "bay_code": bay.code, "depth_row": row_number, "containers": items})

def _get_container_prefill_data(code: str) -> dict:
    """
    Busca datos base del contenedor y su última clasificación conocida.
    Retorna estructura lista para autocompletar Gate In.
    """
    normalized_code = (code or "").strip().upper()
    if not normalized_code:
        return {"found": False}

    container = (
        Container.query
        .filter(Container.code == normalized_code)
        .order_by(Container.id.desc())
        .first()
    )

    last_class = db.session.execute(text("""
        SELECT
            cc.container_id,
            cc.shipping_line,
            cc.max_gross_kg,
            cc.tare_kg,
            cc.manufacture_year,
            cc.summary_text,
            cc.notes,
            cc.classified_at
        FROM yard_gate_alamo.container_classifications cc
        JOIN (
            SELECT container_id, MAX(classified_at) AS max_classified_at
            FROM yard_gate_alamo.container_classifications
            WHERE container_id IN (
                SELECT id FROM yard_gate_alamo.containers WHERE code = :code
            )
            GROUP BY container_id
        ) x
          ON x.container_id = cc.container_id
         AND x.max_classified_at = cc.classified_at
        ORDER BY cc.classified_at DESC
        LIMIT 1
    """), {"code": normalized_code}).mappings().first()

    if not container and not last_class:
        return {"found": False}

    data = {
        "found": True,
        "container_id": container.id if container else None,
        "code": normalized_code,
        "size": container.size if container else None,
        "year": container.year if container else None,
        "status_notes": container.status_notes if container else None,
        "site_id": container.site_id if container else None,
        "is_in_yard": bool(container.is_in_yard) if container else False,

        "shipping_line": last_class["shipping_line"] if last_class else None,
        "max_gross_kg": last_class["max_gross_kg"] if last_class else None,
        "tare_kg": last_class["tare_kg"] if last_class else None,
        "manufacture_year": last_class["manufacture_year"] if last_class else None,
        "summary_text": last_class["summary_text"] if last_class else None,
        "classification_notes": last_class["notes"] if last_class else None,
    }

    if not data["year"] and data["manufacture_year"]:
        data["year"] = data["manufacture_year"]

    return data
# =========================
# Gate In / Gate Out
# =========================
@yard_bp.get("/gate-in")
@login_required
def gate_in_view():
    site_id = _ensure_active_site()

    blocks = (
        YardBlock.query
        .filter_by(site_id=site_id)
        .order_by(YardBlock.code.asc())
        .all()
    )

    sql_rows = db.session.execute(text("""
        SELECT
            c.id,
            c.chassis_number,
            c.plate,
            c.axles,
            c.status,
            c.site_id,
            c.type_code
        FROM yard_gate_alamo.chassis c
        ORDER BY c.chassis_number ASC
    """)).mappings().all()

    chassis_rows = [dict(r) for r in sql_rows]

    return render_template(
        "yard/gate_in.html",
        blocks=blocks,
        sizes=SIZES,
        chassis_rows=chassis_rows,
    )

@yard_bp.get("/api/container-prefill/<code>")
@login_required
def api_container_prefill(code: str):
    _ensure_active_site()

    normalized_code = (code or "").strip().upper()

    if not CONTAINER_RE.match(normalized_code):
        return jsonify({
            "ok": False,
            "found": False,
            "error": "INVALID_CONTAINER_CODE"
        }), 400

    payload = _get_container_prefill_data(normalized_code)

    return jsonify({
        "ok": True,
        **payload
    })

@yard_bp.post("/gate-in")
@login_required
def gate_in_post():
    site_id = _ensure_active_site()
    active_site = Site.query.get(site_id)

    code = (request.form.get("container_code") or "").strip().upper()
    size = (request.form.get("size") or "").strip()
    year_raw = (request.form.get("year") or "").strip()

    # Notas "compat" (textarea)
    status_notes_extra = (request.form.get("status_notes") or "").strip()

    driver_name = (request.form.get("driver_name") or "").strip()
    driver_id_doc = (request.form.get("driver_id_doc") or "").strip()
    truck_plate = (request.form.get("truck_plate") or "").strip()

    block_code = (request.form.get("block") or "").strip().upper()
    bay_number_raw = (request.form.get("bay_number") or "").strip()

    placement_mode = (request.form.get("placement_mode") or "auto").strip().lower()
    depth_row_raw = (request.form.get("depth_row") or "").strip()
    tier_raw = (request.form.get("tier") or "").strip()

    # =========================
    # Clasificación contenedor
    # =========================
    summary_text = (request.form.get("summary_text") or "").strip()
    classification_notes = (request.form.get("classification_notes") or "").strip()

    shipping_line = (request.form.get("shipping_line") or "").strip().upper()
    shipping_line_other = (request.form.get("shipping_line_other") or "").strip().upper()
    if shipping_line == "VASI":
        shipping_line = shipping_line_other or ""

    max_gross_hidden = (request.form.get("max_gross_kg_hidden") or "").strip()
    max_gross_other = (request.form.get("max_gross_kg") or "").strip()
    max_gross_kg = None
    if max_gross_hidden:
        try:
            max_gross_kg = int(max_gross_hidden)
        except ValueError:
            max_gross_kg = None
    elif max_gross_other:
        try:
            max_gross_kg = int(max_gross_other)
        except ValueError:
            max_gross_kg = None

    tare_raw = (request.form.get("tare_kg") or "").strip()
    tare_kg = None
    if tare_raw:
        try:
            tare_kg = int(tare_raw)
        except ValueError:
            tare_kg = None

    # Este needs_workshop sigue siendo el del contenedor
    needs_workshop = (request.form.get("needs_workshop") or "0").strip() == "1"

    final_status_notes = summary_text or ""
    if status_notes_extra:
        final_status_notes = (final_status_notes + (", " if final_status_notes else "") + status_notes_extra).strip()

    # =========================
    # Clasificación chasis
    # =========================
    chassis_id_raw = (request.form.get("chassis_id") or "").strip()
    chassis_tire_checks_json_raw = (request.form.get("chassis_tire_checks_json") or "{}").strip()
    chassis_inspection_json_raw = (request.form.get("chassis_inspection_json") or "{}").strip()

    selected_chassis = None
    chassis_tire_checks = {}
    chassis_inspection = {}

    if chassis_id_raw:
        if not str(chassis_id_raw).isdigit():
            flash("Chasis inválido.", "danger")
            return redirect(url_for("yard.gate_in_view"))

        selected_chassis = Chassis.query.get(int(chassis_id_raw))
        if not selected_chassis:
            flash("El chasis seleccionado no existe.", "danger")
            return redirect(url_for("yard.gate_in_view"))

        # Regla:
        # - Se puede seleccionar desde el maestro general
        # - Pero NO puede ingresar aquí si está activo en inventario de otro predio
        # - Si está activo en este mismo predio, también se bloquea por doble ingreso
        active_inv = (
            ChassisInventory.query
            .filter_by(chassis_id=selected_chassis.id, is_in_yard=True)
            .first()
        )

        if active_inv:
            inv_site = Site.query.get(active_inv.site_id)
            inv_site_name = inv_site.name if inv_site else f"ID {active_inv.site_id}"

            if active_inv.site_id == site_id:
                flash(
                    f"El chasis {selected_chassis.chassis_number} ya se encuentra en inventario de este predio.",
                    "danger"
                )
                return redirect(url_for("yard.gate_in_view"))

            flash(
                f"El chasis {selected_chassis.chassis_number} está activo en inventario del predio {inv_site_name}. "
                f"Primero debe realizarse el Gate Out / EIR de salida en ese predio.",
                "danger"
            )
            return redirect(url_for("yard.gate_in_view"))

        try:
            parsed_tires = json.loads(chassis_tire_checks_json_raw or "{}")
            chassis_tire_checks = parsed_tires if isinstance(parsed_tires, dict) else {}
        except Exception:
            flash("La clasificación de llantas del chasis viene dañada.", "danger")
            return redirect(url_for("yard.gate_in_view"))

        try:
            parsed_inspection = json.loads(chassis_inspection_json_raw or "{}")
            chassis_inspection = parsed_inspection if isinstance(parsed_inspection, dict) else {}
        except Exception:
            flash("La clasificación estructural del chasis viene dañada.", "danger")
            return redirect(url_for("yard.gate_in_view"))

    # =========================
    # Validaciones contenedor
    # =========================
    if not CONTAINER_RE.match(code):
        flash("Formato de contenedor inválido. Debe ser AAAA-000000-0.", "danger")
        return redirect(url_for("yard.gate_in_view"))

    if size not in SIZES:
        flash("Tamaño inválido.", "danger")
        return redirect(url_for("yard.gate_in_view"))

    year = None
    if year_raw:
        try:
            year = int(year_raw)
            if year < 1950 or year > (datetime.utcnow().year + 1):
                raise ValueError()
        except ValueError:
            flash("Año inválido.", "danger")
            return redirect(url_for("yard.gate_in_view"))

    if block_code not in {"A", "B", "C", "D"}:
        flash("Bloque inválido.", "danger")
        return redirect(url_for("yard.gate_in_view"))

    try:
        bay_number = int(bay_number_raw)
        if not (1 <= bay_number <= 15):
            raise ValueError()
    except ValueError:
        flash("Estiba inválida (1..15).", "danger")
        return redirect(url_for("yard.gate_in_view"))

    block = YardBlock.query.filter_by(code=block_code, site_id=site_id).first()
    bay = (
        YardBay.query.filter_by(block_id=block.id, bay_number=bay_number, is_active=True, site_id=site_id).first()
        if block else None
    )
    if not bay:
        flash("Estiba no encontrada.", "danger")
        return redirect(url_for("yard.gate_in_view"))

    # Lock de estiba durante asignación
    db.session.query(YardBay).filter(YardBay.id == bay.id).with_for_update().one()

    existing_here = Container.query.filter_by(site_id=site_id, code=code).first()

    other_in_yard = (
        Container.query
        .filter(Container.code == code, Container.is_in_yard == True, Container.site_id != site_id)  # noqa: E712
        .first()
    )
    if other_in_yard:
        flash("Este contenedor está en patio, pero en otro predio.", "danger")
        return redirect(url_for("yard.gate_in_view"))

    if existing_here and existing_here.is_in_yard:
        flash("Este contenedor ya está en patio.", "danger")
        return redirect(url_for("yard.gate_in_view"))

    # =========================
    # Upsert Container
    # =========================
    # Buscar primero en este predio
    if existing_here:
        c = existing_here
    else:
        # Si no existe en este predio, buscamos el contenedor globalmente por código
        c = (
            Container.query
            .filter(Container.code == code)
            .order_by(Container.id.desc())
            .first()
        )

    if not c:
        c = Container(
            code=code,
            size=size,
            year=year,
            status_notes=final_status_notes or None,
            is_in_yard=True,
            site_id=site_id
        )
        db.session.add(c)
        db.session.flush()
    else:
        # Se mueve al predio actual y entra a patio
        c.site_id = site_id
        c.is_in_yard = True

        # Solo actualizar si realmente viene valor
        if size:
            c.size = size
        if year is not None:
            c.year = year
        if final_status_notes:
            c.status_notes = final_status_notes

        db.session.add(c)
        db.session.flush()
    # =========================
    # Guardar clasificación contenedor
    # =========================
    should_insert_class = any([
        bool(shipping_line),
        bool(summary_text),
        max_gross_kg is not None,
        tare_kg is not None,
        year is not None,
        bool(classification_notes),
        bool(needs_workshop),
    ])

    if should_insert_class:
        shipping_line_db = (shipping_line or "").strip().upper()

        if not shipping_line_db:
            shipping_line_db = "ATM"

        sql_ins = text("""
            INSERT INTO yard_gate_alamo.container_classifications
            (site_id, container_id, classified_at, classified_by_user_id,
             shipping_line, max_gross_kg, tare_kg, manufacture_year,
             needs_workshop, summary_text, notes)
            VALUES
            (:site_id, :container_id, NOW(), :uid,
             :shipping_line, :max_gross_kg, :tare_kg, :manufacture_year,
             :needs_workshop, :summary_text, :notes)
        """)
        db.session.execute(sql_ins, {
            "site_id": site_id,
            "container_id": c.id,
            "uid": current_user.id,
            "shipping_line": shipping_line_db,
            "max_gross_kg": max_gross_kg,
            "tare_kg": tare_kg,
            "manufacture_year": year,
            "needs_workshop": bool(needs_workshop),
            "summary_text": (summary_text or None),
            "notes": (classification_notes or None),
        })

    # =========================
    # Slot / posición
    # =========================
    if placement_mode == "manual":
        try:
            depth_row = int(depth_row_raw)
            tier = int(tier_raw)
        except ValueError:
            db.session.rollback()
            flash("Fila/Nivel inválidos.", "danger")
            return redirect(url_for("yard.gate_in_view"))

        if not (1 <= depth_row <= bay.max_depth_rows) or not (1 <= tier <= bay.max_tiers):
            db.session.rollback()
            flash("Fila/Nivel fuera de rango.", "danger")
            return redirect(url_for("yard.gate_in_view"))

        occupied = ContainerPosition.query.filter_by(bay_id=bay.id, depth_row=depth_row, tier=tier).first()
        if occupied:
            db.session.rollback()
            flash("Ese slot ya está ocupado.", "danger")
            return redirect(url_for("yard.gate_in_view"))
    else:
        slot = find_first_free_slot(bay.id)
        if not slot:
            db.session.rollback()
            flash(f"La estiba {bay.code} está llena.", "danger")
            return redirect(url_for("yard.gate_in_view"))
        depth_row, tier = slot

    ContainerPosition.query.filter_by(container_id=c.id).delete()
    db.session.add(
        ContainerPosition(
            container_id=c.id,
            bay_id=bay.id,
            depth_row=depth_row,
            tier=tier,
            placed_by_user_id=current_user.id,
        )
    )

    mv = Movement(
        site_id=site_id,
        container_id=c.id,
        movement_type="GATE_IN",
        occurred_at=datetime.utcnow(),
        bay_code=bay.code,
        depth_row=depth_row,
        tier=tier,
        driver_name=driver_name or None,
        driver_id_doc=driver_id_doc or None,
        truck_plate=truck_plate or None,
        notes=final_status_notes or None,
        created_by_user_id=current_user.id,
        created_at=datetime.utcnow(),
    )
    db.session.add(mv)
    db.session.flush()

    # =========================
    # Procesar clasificación de chasis
    # =========================
    workshop_ticket_id = None
    chassis_classification_ticket_payload = None

    if selected_chassis:
        axles = int(getattr(selected_chassis, "axles", 2) or 2)
        allowed = set(allowed_positions_for(axles))

        structure_status = _normalize_structure_status_for_db(
            _norm_enum(chassis_inspection.get("structure_status"))
        )
        twistlocks_status = _normalize_twistlocks_status_for_db(
            _norm_enum(chassis_inspection.get("twistlocks_status"))
        )
        landing_gear_status = _normalize_landing_gear_status_for_db(
            _norm_enum(chassis_inspection.get("landing_gear_status"))
        )
        lights_status = _normalize_lights_status_for_db(
            _norm_enum(chassis_inspection.get("lights_status"))
        )
        mudflap_status = _normalize_mudflap_status_for_db(
            _norm_enum(chassis_inspection.get("mudflap_status"))
        )

        plate_text = (chassis_inspection.get("plate_text") or "").strip()
        plate_validation_status = _norm_enum(chassis_inspection.get("plate_validation_status"))

        damage_summary = (chassis_inspection.get("damage_summary") or "").strip()
        comments = (chassis_inspection.get("comments") or "").strip()
        driver_comments = (chassis_inspection.get("driver_comments") or "").strip()

        tire_lines = []
        ticket_alert_lines = []
        ticket_tire_rows = []
        any_tire_issue = False
        grouped_tire_readings = {}

        tire_state_labels = {
            "GASTADA": "REGULAR",
            "PINCHADA": "DESINFLADA",
            "CAMBIAR": "MAL ESTADO",
            "NO_APTA": "ROJA",
            "OK": "VERDE",
        }

        for pos, item in (chassis_tire_checks or {}).items():
            pos = (pos or "").strip().upper()
            if pos not in allowed:
                continue

            item = item or {}
            seal_status = _norm_enum(item.get("seal_status")) or "OK"
            tire_number_status = _norm_enum(item.get("tire_number_status")) or "OK"
            pressure_psi = item.get("pressure_psi")

            estrias_mm_raw = item.get("estrias_mm")
            is_flat = bool(item.get("is_flat"))

            estrias_mm = None
            if estrias_mm_raw not in (None, ""):
                try:
                    estrias_mm = int(estrias_mm_raw)
                except Exception:
                    estrias_mm = None

            tire_state = _calc_tire_state_from_data(estrias_mm, is_flat)

            if seal_status not in {"OK", "DISTINTO"}:
                seal_status = "OK"

            if tire_number_status not in {"OK", "DISTINTO"}:
                tire_number_status = "OK"

            if tire_state not in TIRE_STATES:
                tire_state = "OK"

            row = ChassisTire.query.filter_by(
                chassis_id=selected_chassis.id,
                position_code=pos
            ).first()

            if not row:
                row = ChassisTire(
                    chassis_id=selected_chassis.id,
                    position_code=pos,
                )
                db.session.add(row)
                db.session.flush()

            marchamo_config = row.marchamo
            tire_number_config = row.tire.tire_number if row.tire else None

            master_pos = _normalize_position_for_tire_master(pos)
            if master_pos:
                grp = grouped_tire_readings.setdefault(master_pos, {
                    "states": [],
                    "pressures": [],
                    "comments": [],
                    "seal_issue": False,
                    "seal_2": None,
                })

                grp["states"].append(tire_state)

                if pressure_psi not in (None, ""):
                    grp["pressures"].append(pressure_psi)

                if seal_status != "OK":
                    grp["seal_issue"] = True

                detail_parts = [
                    f"{pos}",
                    f"SEAL={seal_status}",
                    f"STATE={tire_state}",
                ]
                if estrias_mm not in (None, ""):
                    detail_parts.append(f"MM={estrias_mm}")
                if is_flat:
                    detail_parts.append("FLAT=SI")
                if pressure_psi not in (None, ""):
                    detail_parts.append(f"PSI={pressure_psi}")

                grp["comments"].append(" | ".join(detail_parts))

            row.estrias_mm = estrias_mm
            row.is_flat = is_flat
            row.tire_state = tire_state
            row.updated_at = datetime.utcnow()
            db.session.add(row)

            if seal_status == "DISTINTO":
                any_tire_issue = True
                tire_lines.append(f"{pos}: MARCHAMO DISTINTO")
                ticket_alert_lines.append(f"MARCHAMO DISTINTO EN {pos}")

            if tire_number_status == "DISTINTO":
                any_tire_issue = True
                tire_lines.append(f"{pos}: NUMERO DE LLANTA DISTINTO")
                ticket_alert_lines.append(f"NUMERO DE LLANTA DISTINTO EN {pos}")

            if is_flat:
                any_tire_issue = True
                tire_lines.append(f"{pos}: PINCHADA (DESINFLADA)")
                ticket_alert_lines.append(f"LLANTA PINCHADA EN {pos}")
            elif tire_state != "OK":
                any_tire_issue = True
                tire_lines.append(
                    f"{pos}: ESTADO {tire_state} ({tire_state_labels.get(tire_state, tire_state)})"
                )

            ticket_tire_rows.append({
                "pos": pos,
                "marchamo_config": marchamo_config,
                "seal_status": seal_status,
                "tire_number_config": tire_number_config,
                "tire_number_status": tire_number_status,
                "estrias_mm": estrias_mm,
                "is_flat": is_flat,
                "tire_state": tire_state,
                "pressure_psi": pressure_psi,
            })

        _save_grouped_tire_readings(
            site_id=site_id,
            chassis_id=selected_chassis.id,
            axles=axles,
            grouped_readings=grouped_tire_readings,
            user_id=current_user.id,
            event_type="GATE_IN",
            event_id=mv.id,
        )

        structure_lines = []

        if structure_status in {"GOLPE", "DOBLADO", "SOLDADURA"}:
            structure_lines.append(f"Estructura: {structure_status}")

        if twistlocks_status in {"DANADOS"}:
            structure_lines.append(f"Twistlocks: {twistlocks_status}")

        if landing_gear_status in {"DANADAS"}:
            structure_lines.append(f"Patas: {landing_gear_status}")

        if lights_status in {"IZQ_DANADA", "DER_DANADA"}:
            structure_lines.append(f"Luces: {lights_status}")

        if mudflap_status in {"NO_TRAE"}:
            structure_lines.append(f"Faldones: {mudflap_status}")

        if plate_validation_status in {"DISTINTA", "NO_TRAE"}:
            line = f"Placa: {plate_validation_status}"
            if plate_text:
                line += f" (CONFIGURADA: {plate_text})"
            structure_lines.append(line)

        if damage_summary:
            structure_lines.append(f"Resumen: {damage_summary}")

        if comments:
            structure_lines.append(f"Chequeador: {comments}")

        if driver_comments:
            structure_lines.append(f"Chofer: {driver_comments}")

        chassis_needs_workshop_manual = bool(chassis_inspection.get("needs_workshop"))
        needs_workshop_chassis = bool(structure_lines) or bool(any_tire_issue) or chassis_needs_workshop_manual

        inspection_id = _insert_dynamic("yard_gate_alamo", "chassis_inspections", {
            "site_id": site_id,
            "chassis_id": selected_chassis.id,
            "inspected_at": datetime.utcnow(),
            "inspected_by_user_id": current_user.id,
            "structure_status": structure_status or None,
            "twistlocks_status": twistlocks_status or None,
            "landing_gear_status": landing_gear_status or None,
            "lights_status": lights_status or None,
            "mudflap_status": mudflap_status or None,
            "plate_text": plate_text or None,
            "plate_validation_status": plate_validation_status or None,
            "comments": comments or None,
            "driver_comments": driver_comments or None,
            "needs_workshop": needs_workshop_chassis,
            "damage_summary": damage_summary or None,
            "movement_id": mv.id,
        })

        selected_chassis.site_id = site_id
        selected_chassis.is_in_yard = True
        selected_chassis.updated_at = datetime.utcnow()
        db.session.add(selected_chassis)

        inv = ChassisInventory.query.filter_by(chassis_id=selected_chassis.id).first()
        if not inv:
            inv = ChassisInventory(
                site_id=site_id,
                chassis_id=selected_chassis.id,
                chassis_code=selected_chassis.chassis_number,
                is_in_yard=True,
            )
        else:
            inv.site_id = site_id
            inv.chassis_code = selected_chassis.chassis_number
            inv.is_in_yard = True
            inv.updated_at = datetime.utcnow()
        db.session.add(inv)

        username = (
            getattr(current_user, "name", None)
            or getattr(current_user, "username", None)
            or getattr(current_user, "email", None)
            or f"USER {current_user.id}"
        )

        chassis_classification_ticket_payload = _build_chassis_gate_in_ticket_text(
            site_name=(active_site.name if active_site else ""),
            username=username,
            occurred_at=mv.occurred_at or datetime.utcnow(),
            chassis_number=selected_chassis.chassis_number,
            plate=selected_chassis.plate,
            structure_status=structure_status,
            twistlocks_status=twistlocks_status,
            landing_gear_status=landing_gear_status,
            lights_status=lights_status,
            mudflap_status=mudflap_status,
            plate_validation_status=plate_validation_status,
            damage_summary=damage_summary or None,
            comments=comments or None,
            driver_comments=driver_comments or None,
            tire_rows=ticket_tire_rows,
            alert_lines=ticket_alert_lines,
        )

        if needs_workshop_chassis:
            last_eir = _fetch_last_final_eir_for_chassis(selected_chassis.id)
            eir_prev_id = int(last_eir["id"]) if last_eir and last_eir.get("id") else None

            body = _build_workshop_ticket_text(
                chassis_number=selected_chassis.chassis_number,
                axles=axles,
                structure_lines=structure_lines,
                tire_lines=tire_lines,
                eir_prev_id=eir_prev_id
            )

            workshop_ticket_id = _insert_dynamic("yard_gate_alamo", "workshop_tickets", {
                "site_id": site_id,
                "chassis_id": selected_chassis.id,
                "inspection_id": inspection_id,
                "created_at": datetime.utcnow(),
                "created_by_user_id": current_user.id,
                "status": "OPEN",
                "ticket_type": "CHASSIS_DAMAGE",
                "movement_id": mv.id,
                "payload_text": body,
                "notes": body,
            })

            audit_log(
                current_user.id,
                "WORKSHOP_TICKET_CREATED_FROM_GATE_IN",
                "workshop_ticket",
                workshop_ticket_id,
                {
                    "site_id": site_id,
                    "movement_id": mv.id,
                    "chassis_id": selected_chassis.id,
                    "container_id": c.id,
                },
            )

        audit_log(
            current_user.id,
            "CHASSIS_CLASSIFIED_FROM_GATE_IN",
            "chassis",
            selected_chassis.id,
            {
                "site_id": site_id,
                "movement_id": mv.id,
                "container_id": c.id,
                "needs_workshop": needs_workshop_chassis,
                "classification_ticket": bool(chassis_classification_ticket_payload),
            },
        )

    # =========================
    # Fotos
    # =========================
    storage = get_storage()
    photos = request.files.getlist("photos") or []

    for f in photos:
        if not f or not f.filename:
            continue
        try:
            key = build_photo_key(c.code, mv.id, f.filename)
            url = storage.upload_fileobj(
                f,
                key,
                f.mimetype or "application/octet-stream"
            )
            db.session.add(
                MovementPhoto(
                    movement_id=mv.id,
                    photo_type="CONTAINER",
                    url=url
                )
            )
        except Exception as e:
            db.session.add(
                MovementPhoto(
                    movement_id=mv.id,
                    photo_type="UPLOAD_ERROR",
                    url=str(e)
                )
            )

    audit_log(
        current_user.id,
        "GATE_IN_CREATED",
        "container",
        c.id,
        {
            "container_code": c.code,
            "bay": bay.code,
            "depth_row": depth_row,
            "tier": tier,
            "site_id": site_id,
            "chassis_id": selected_chassis.id if selected_chassis else None,
            "workshop_ticket_id": workshop_ticket_id,
        },
    )

    db.session.commit()

    if workshop_ticket_id:
        flash(
            f"Gate In registrado: {c.code} en {bay.code} F{depth_row:02d} N{tier}. "
            f"Se generó ticket de taller para el chasis {selected_chassis.chassis_number}.",
            "success",
        )
        return redirect(url_for("yard.gate_in_view"))

    flash(f"Gate In registrado: {c.code} en {bay.code} F{depth_row:02d} N{tier}.", "success")
    return redirect(url_for("yard.gate_in_view"))


@yard_bp.get("/gate-out")
@login_required
def gate_out_view():
    site_id = _ensure_active_site()

    active_site = Site.query.get(site_id)
    site_code = (active_site.code or "").upper() if active_site else ""

    # ==========================================================
    # ✅ PREDIOS: COYOL / CALDERA / LIMON -> flujo nuevo EIR
    # ==========================================================
    if site_code in {"COYOL", "CALDERA", "LIMON"}:
        sql_last_class = text("""
            SELECT DISTINCT ON (cc.container_id)
                cc.container_id,
                cc.shipping_line
            FROM yard_gate_alamo.container_classifications cc
            WHERE cc.site_id = :site_id
            ORDER BY cc.container_id, cc.classified_at DESC NULLS LAST, cc.id DESC
        """)
        class_rows = db.session.execute(sql_last_class, {"site_id": site_id}).mappings().all()
        shipping_line_map = {int(r["container_id"]): (r["shipping_line"] or "").strip().upper() for r in class_rows}

        containers_raw = (
            db.session.query(Container, ContainerPosition, YardBay)
            .join(ContainerPosition, ContainerPosition.container_id == Container.id)
            .join(YardBay, YardBay.id == ContainerPosition.bay_id)
            .filter(Container.is_in_yard == True, Container.site_id == site_id)  # noqa: E712
            .order_by(YardBay.code.asc(), ContainerPosition.depth_row.asc(), ContainerPosition.tier.asc())
            .all()
        )

        containers = []
        for c, p, b in containers_raw:
            containers.append({
                "container": c,
                "position": p,
                "bay": b,
                "shipping_line": shipping_line_map.get(c.id, ""),
            })

        chassis_rows = (
            Chassis.query
            .filter(
                Chassis.site_id == site_id,
                Chassis.is_in_yard == True  # noqa: E712
            )
            .order_by(Chassis.chassis_number.asc())
            .all()
        )

        eirs_draft = (
            EIR.query
            .filter_by(site_id=site_id, status="DRAFT")
            .order_by(EIR.id.desc())
            .limit(200)
            .all()
        )

        return render_template(
            "yard/gate_out_predios.html",
            containers=containers,
            chassis_rows=chassis_rows,
            eirs_draft=eirs_draft,
        )

    # ==========================================================
    # 🔹 MAERSK: flujo viejo intacto
    # ==========================================================
    containers = (
        db.session.query(Container, ContainerPosition, YardBay)
        .join(ContainerPosition, ContainerPosition.container_id == Container.id)
        .join(YardBay, YardBay.id == ContainerPosition.bay_id)
        .filter(Container.is_in_yard == True, Container.site_id == site_id)  # noqa: E712
        .order_by(YardBay.code.asc(), ContainerPosition.depth_row.asc(), ContainerPosition.tier.asc())
        .all()
    )
    return render_template("yard/gate_out.html", rows=containers)


@yard_bp.post("/gate-out")
@login_required
def gate_out_post():
    site_id = _ensure_active_site()
    active_site = Site.query.get(site_id)
    site_code = (active_site.code or "").upper() if active_site else ""
    is_predio = site_code in {"COYOL", "CALDERA", "LIMON"}

    # ==========================================================
    # ✅ PREDIOS: Gate Out / EIR
    # ==========================================================
    if is_predio:
        mode = (request.form.get("mode") or "create").strip().lower()  # create | link
        eir_id_raw = (request.form.get("eir_id") or "").strip()

        has_chassis = (request.form.get("has_chassis") or "0").strip() == "1"
        has_container = (request.form.get("has_container") or "0").strip() == "1"
        is_reefer = (request.form.get("is_reefer") or "0").strip() == "1"
        has_genset = (request.form.get("has_genset") or "0").strip() == "1"

        chassis_id_raw = (request.form.get("chassis_id") or "").strip()
        container_id_raw = (request.form.get("container_id") or "").strip()

        terminal_name = (request.form.get("terminal_name") or (active_site.name if active_site else "")).strip()
        trip_date_raw = (request.form.get("trip_date") or "").strip()
        trip_time_raw = (request.form.get("trip_time") or "").strip()
        carrier = (request.form.get("carrier") or "ATM").strip() or "ATM"
        origin = (request.form.get("origin") or site_code).strip()
        destination = (request.form.get("destination") or "").strip()
        operation_type = (request.form.get("operation_type") or "").strip().upper()

        driver_name = (request.form.get("driver_name") or "").strip()
        driver_id_doc = (request.form.get("driver_id_doc") or "").strip()
        truck_plate = (request.form.get("truck_plate") or "").strip()

        shipping_line = (request.form.get("shipping_line") or "").strip().upper()
        container_seal = (request.form.get("container_seal") or "").strip()
        general_notes = (request.form.get("notes") or "").strip()

        # Sección chasis
        chassis_lights_status = (request.form.get("chassis_lights_status") or "").strip().upper()
        chassis_lights_detail = (request.form.get("chassis_lights_detail") or "").strip()
        chassis_twistlocks_status = (request.form.get("chassis_twistlocks_status") or "").strip().upper()
        chassis_twistlocks_detail = (request.form.get("chassis_twistlocks_detail") or "").strip()
        chassis_mudflaps_status = (request.form.get("chassis_mudflaps_status") or "").strip().upper()
        chassis_mudflaps_detail = (request.form.get("chassis_mudflaps_detail") or "").strip()
        chassis_landing_gear_status = (request.form.get("chassis_landing_gear_status") or "").strip().upper()
        chassis_landing_gear_detail = (request.form.get("chassis_landing_gear_detail") or "").strip()
        chassis_structure_status = (request.form.get("chassis_structure_status") or "").strip().upper()
        chassis_structure_detail = (request.form.get("chassis_structure_detail") or "").strip()

        # Sección reefer
        rf_running_status = (request.form.get("rf_running_status") or "").strip().upper()
        rf_temperature = (request.form.get("rf_temperature") or "").strip()
        rf_genset = (request.form.get("rf_genset") or "").strip().upper()
        rf_plug = (request.form.get("rf_plug") or "").strip()
        rf_cord = (request.form.get("rf_cord") or "").strip()
        rf_computer = (request.form.get("rf_computer") or "").strip()
        rf_fuel = (request.form.get("rf_fuel") or "").strip()
        rf_hourmeter = (request.form.get("rf_hourmeter") or "").strip()
        rf_alternator = (request.form.get("rf_alternator") or "").strip()
        rf_battery = (request.form.get("rf_battery") or "").strip()
        rf_notes = (request.form.get("rf_notes") or "").strip()

        damage_points_raw = (request.form.get("container_damage_points_json") or "[]").strip()

        # -------------------------
        # Validaciones mínimas
        # -------------------------
        terminal_name = terminal_name or (active_site.name if active_site else site_code or "")
        origin = origin or (active_site.name if active_site else site_code or "")
        carrier = carrier or "ATM"

        if trip_date_raw:
            try:
                trip_date = datetime.strptime(trip_date_raw, "%Y-%m-%d").date()
            except Exception:
                flash("Fecha inválida. Usa el selector de fecha.", "danger")
                return redirect(url_for("yard.gate_out_view"))
        else:
            trip_date = datetime.utcnow().date()

        trip_time = None
        if trip_time_raw:
            try:
                trip_time = datetime.strptime(trip_time_raw, "%H:%M").time()
            except Exception:
                flash("Hora inválida. Usa el selector de hora.", "danger")
                return redirect(url_for("yard.gate_out_view"))

        if operation_type and operation_type not in {"EXPORTACION", "IMPORTACION"}:
            flash("Tipo de operación inválido.", "danger")
            return redirect(url_for("yard.gate_out_view"))

        # -------------------------
        # Parse daños visuales
        # -------------------------
        try:
            damage_points = json.loads(damage_points_raw or "[]")
            if not isinstance(damage_points, list):
                damage_points = []
        except Exception:
            flash("Los daños del contenedor vienen dañados en el formulario.", "danger")
            return redirect(url_for("yard.gate_out_view"))

        # -------------------------
        # Validar / cargar contenedor
        # -------------------------
        c = None
        bay_code = None
        depth_row = None
        tier = None
        container_size = None
        container_snapshot = None

        if has_container:
            if not container_id_raw or not str(container_id_raw).isdigit():
                flash("Debes seleccionar un contenedor.", "danger")
                return redirect(url_for("yard.gate_out_view"))

            c = Container.query.get(int(container_id_raw))
            if not c or not c.is_in_yard or c.site_id != site_id:
                flash("Contenedor no válido o no está en patio en este predio.", "danger")
                return redirect(url_for("yard.gate_out_view"))

            pos = ContainerPosition.query.filter_by(container_id=c.id).first()
            if pos:
                bay = YardBay.query.get(pos.bay_id)
                bay_code = bay.code if bay else None
                depth_row = pos.depth_row
                tier = pos.tier

            container_size = getattr(c, "size", None)

            # Intentar jalar naviera de la última clasificación del Gate In
            if not shipping_line:
                sql_last_class = text("""
                    SELECT shipping_line
                    FROM yard_gate_alamo.container_classifications
                    WHERE site_id = :site_id
                      AND container_id = :container_id
                    ORDER BY classified_at DESC NULLS LAST, id DESC
                    LIMIT 1
                """)
                row_class = db.session.execute(sql_last_class, {
                    "site_id": site_id,
                    "container_id": c.id,
                }).mappings().first()
                if row_class and row_class.get("shipping_line"):
                    shipping_line = (row_class.get("shipping_line") or "").strip().upper()

            container_snapshot = {
                "container_id": c.id,
                "container_code": c.code,
                "size": container_size,
                "shipping_line": shipping_line or None,
                "seal": container_seal or None,
                "position": {
                    "bay_code": bay_code,
                    "depth_row": depth_row,
                    "tier": tier,
                },
                "damage_count": len(damage_points),
            }

        # -------------------------
        # Validar / cargar chasis
        # -------------------------
        ch = None
        chassis_snapshot = None

        if has_chassis:
            if not chassis_id_raw or not str(chassis_id_raw).isdigit():
                flash("Debes seleccionar un chasis.", "danger")
                return redirect(url_for("yard.gate_out_view"))

            ch = Chassis.query.get(int(chassis_id_raw))
            if not ch:
                flash("Chasis inválido.", "danger")
                return redirect(url_for("yard.gate_out_view"))

            if ch.site_id != site_id or not ch.is_in_yard:
                flash("Ese chasis no está disponible en este predio.", "danger")
                return redirect(url_for("yard.gate_out_view"))

            # Snapshot de llantas / marchamos
            tire_rows = (
                ChassisTire.query
                .filter_by(chassis_id=ch.id)
                .order_by(ChassisTire.position_code.asc())
                .all()
            )

            tires_snapshot = []
            for tr in tire_rows:
                tires_snapshot.append({
                    "position_code": tr.position_code,
                    "marchamo": tr.marchamo,
                    "tire_state": tr.tire_state,
                    "tire_number": tr.tire.tire_number if tr.tire else None,
                    "brand": tr.tire.brand if tr.tire else None,
                    "estrias_mm": getattr(tr, "estrias_mm", None),
                    "is_flat": bool(getattr(tr, "is_flat", False)),
                })

            chassis_snapshot = {
                "chassis_id": ch.id,
                "chassis_number": ch.chassis_number,
                "plate": ch.plate,
                "axles": ch.axles,
                "type_code": getattr(ch, "type_code", None),
                "inspection": {
                    "lights": {
                        "status": chassis_lights_status or "OK",
                        "detail": chassis_lights_detail or None,
                    },
                    "twist_locks": {
                        "status": chassis_twistlocks_status or "OK",
                        "detail": chassis_twistlocks_detail or None,
                    },
                    "mudflaps": {
                        "status": chassis_mudflaps_status or "OK",
                        "detail": chassis_mudflaps_detail or None,
                    },
                    "landing_gear": {
                        "status": chassis_landing_gear_status or "OK",
                        "detail": chassis_landing_gear_detail or None,
                    },
                    "structure": {
                        "status": chassis_structure_status or "OK",
                        "detail": chassis_structure_detail or None,
                    },
                },
                "tires": tires_snapshot,
            }

        # -------------------------
        # Reefer snapshot
        # -------------------------
        reefer_snapshot = None
        if is_reefer:
            reefer_snapshot = {
                "running_status": rf_running_status or None,
                "temperature": rf_temperature or None,
                "genset": rf_genset or None,
                "plug": rf_plug or None,
                "cord": rf_cord or None,
                "computer": rf_computer or None,
                "fuel": rf_fuel or None,
                "hourmeter": rf_hourmeter or None,
                "alternator": rf_alternator or None,
                "battery": rf_battery or None,
                "notes": rf_notes or None,
            }

        # -------------------------
        # Si no trae ningún equipo, bloquear
        # -------------------------
        if not has_container and not has_chassis:
            flash("Debes indicar al menos un equipo: chasis o contenedor.", "danger")
            return redirect(url_for("yard.gate_out_view"))

        # -------------------------
        # Crear Movement si hay contenedor
        # -------------------------
        # -------------------------
        # En predios NO se crea Movement todavía
        # El Movement real se crea hasta CONFIRMAR el EIR
        # -------------------------
        mv = None

        # -------------------------
        # Crear o ligar EIR
        # -------------------------
        if mode == "link":
            if not eir_id_raw or not str(eir_id_raw).isdigit():
                db.session.rollback()
                flash("Selecciona un EIR PENDING válido para ligar.", "danger")
                return redirect(url_for("yard.gate_out_view"))

            eir = EIR.query.get(int(eir_id_raw))
            if not eir or eir.site_id != site_id:
                db.session.rollback()
                flash("Ese EIR no corresponde a este predio.", "danger")
                return redirect(url_for("yard.gate_out_view"))

            if eir.status not in {"PENDING", "EDITING"}:
                db.session.rollback()
                flash("Solo puedes ligar EIRs en estado PENDING o EDITING.", "danger")
                return redirect(url_for("yard.gate_out_view"))

            # Limpiar daños viejos para reescribir
            EIRContainerDamage.query.filter_by(eir_id=eir.id).delete()
        else:
            eir = EIR(
                site_id=site_id,
                created_by_user_id=current_user.id,
                terminal_name=terminal_name or "",
                trip_date=trip_date,
                carrier=carrier or "ATM",
                origin=origin or "",
                destination=destination or "",
                has_chassis=bool(has_chassis and ch),
                chassis_id=ch.id if ch else None,
                has_container=bool(has_container and c),
                container_id=c.id if c else None,
                is_reefer=bool(is_reefer),
                has_genset=bool(has_genset),
                status="PENDING",
            )
            
            db.session.add(eir)
            db.session.flush()

        # -------------------------
        # Completar / actualizar EIR
        # -------------------------
        eir.terminal_name = terminal_name or ""
        eir.trip_date = trip_date
        eir.trip_time = trip_time
        eir.carrier = carrier or "ATM"
        eir.origin = origin or ""
        eir.destination = destination or ""
        eir.operation_type = operation_type or None

        eir.driver_name = driver_name or None
        eir.driver_id_doc = driver_id_doc or None
        eir.truck_plate = truck_plate or None

        eir.has_chassis = bool(has_chassis and ch)
        eir.chassis_id = ch.id if ch else None
        eir.chassis_plate = ch.plate if ch and ch.plate else None

        eir.has_container = bool(has_container and c)
        eir.container_id = c.id if c else None
        eir.container_size = container_size if c else None
        eir.shipping_line = shipping_line or None
        eir.container_seal = container_seal or None

        eir.is_reefer = bool(is_reefer)
        eir.has_genset = bool(has_genset)

        eir.general_notes = general_notes or None
        eir.chassis_snapshot_json = chassis_snapshot
        eir.container_snapshot_json = container_snapshot
        eir.reefer_snapshot_json = reefer_snapshot
        eir.gate_out_movement_id = None

        # Guardado inicial del EIR en estado PENDING
        now_utc = datetime.utcnow()
        eir.status = "PENDING"
        eir.updated_at = now_utc
        eir.pdf_generated_at = now_utc
        eir.finalized_at = None
        eir.inventory_out_at = None
        eir.editable_until = None

        # -------------------------
        # Guardar daños visuales
        # -------------------------
        for item in damage_points:
            side = (item.get("side") or "").strip().upper()
            damage_type = (item.get("damage_type") or "").strip().upper()

            try:
                x = float(item.get("x"))
                y = float(item.get("y"))
            except Exception:
                continue

            if side not in {"LEFT", "RIGHT", "FRONT", "REAR", "ROOF", "INTERIOR"}:
                continue

            if damage_type not in {"A", "R", "G", "M", "C", "F", "H", "Q"}:
                continue

            dmg = EIRContainerDamage(
                eir_id=eir.id,
                side=side,
                damage_type=damage_type,
                x=x,
                y=y,
                notes=(item.get("notes") or "").strip() or None,
                created_by_user_id=current_user.id,
            )
            db.session.add(dmg)


        audit_log(
            current_user.id,
            "EIR_PENDING_SAVED",
            "eir",
            eir.id,
            {
                "site_id": site_id,
                "eir_id": eir.id,
                "movement_id": mv.id if mv else None,
                "container_code": c.code if c else None,
                "chassis_id": ch.id if ch else None,
                "damage_count": len(damage_points),
                "is_reefer": bool(is_reefer),
            },
        )

        db.session.commit()

        flash(
            f"EIR #{eir.id} guardado correctamente en estado PENDIENTE. "
            f"Debes confirmarlo para aplicar la salida de inventario.",
            "success",
        )

        return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

    # ==========================================================
    # 🔹 MAERSK: flujo viejo intacto
    # ==========================================================
    container_id = request.form.get("container_id")
    driver_name = (request.form.get("driver_name") or "").strip()
    driver_id_doc = (request.form.get("driver_id_doc") or "").strip()
    truck_plate = (request.form.get("truck_plate") or "").strip()
    notes = (request.form.get("notes") or "").strip()

    if not container_id or not str(container_id).isdigit():
        flash("Selecciona un contenedor.", "danger")
        return redirect(url_for("yard.gate_out_view"))

    c = Container.query.get(int(container_id))
    if not c or not c.is_in_yard or c.site_id != site_id:
        flash("Contenedor no válido o ya salió (predio actual).", "danger")
        return redirect(url_for("yard.gate_out_view"))

    pos = ContainerPosition.query.filter_by(container_id=c.id).first()
    bay_code = None
    depth_row = None
    tier = None
    if pos:
        bay = YardBay.query.get(pos.bay_id)
        bay_code = bay.code if bay else None
        depth_row = pos.depth_row
        tier = pos.tier

    mv = Movement(
        site_id=site_id,
        container_id=c.id,
        movement_type="GATE_OUT",
        occurred_at=datetime.utcnow(),
        bay_code=bay_code,
        depth_row=depth_row,
        tier=tier,
        driver_name=driver_name or None,
        driver_id_doc=driver_id_doc or None,
        truck_plate=truck_plate or None,
        notes=notes or None,
        created_by_user_id=current_user.id,
        created_at=datetime.utcnow(),
    )
    db.session.add(mv)
    db.session.flush()

    storage = get_storage()
    photos = request.files.getlist("photos") or []
    for f in photos:
        if not f or not f.filename:
            continue
        key = build_photo_key(c.code, mv.id, f.filename)
        url = storage.upload_fileobj(f, key, f.mimetype or "application/octet-stream")
        db.session.add(MovementPhoto(movement_id=mv.id, photo_type="DRIVER_ID", url=url))

    ContainerPosition.query.filter_by(container_id=c.id).delete()
    c.is_in_yard = False

    audit_log(
        current_user.id,
        "GATE_OUT_CREATED",
        "container",
        c.id,
        {
            "container_code": c.code,
            "from_bay": bay_code,
            "depth_row": depth_row,
            "tier": tier,
            "site_id": site_id
        },
    )

    db.session.commit()
    flash(f"Gate Out registrado: {c.code}", "success")
    return redirect(url_for("yard.ticket_view", movement_id=mv.id))

def _cr_range_to_utc_naive(date_from: str, date_to: str):
    """
    date_from/date_to vienen como YYYY-MM-DD (día CR).
    Convertimos [00:00:00 .. 23:59:59] CR -> UTC naive.
    """
    d1_local_naive = datetime.fromisoformat(date_from + "T00:00:00")
    d2_local_naive = datetime.fromisoformat(date_to + "T23:59:59")

    d1_utc = CR_TZ.localize(d1_local_naive).astimezone(UTC_TZ)
    d2_utc = CR_TZ.localize(d2_local_naive).astimezone(UTC_TZ)

    return d1_utc.replace(tzinfo=None), d2_utc.replace(tzinfo=None)


def _parse_report_filters(args):
    movement_type = (args.get("movement_type") or "").strip().upper()
    if movement_type and movement_type not in REPORT_TYPES:
        movement_type = ""

    date_from = args.get("date_from")
    date_to = args.get("date_to")

    if not date_from or not date_to:
        return None, None, None, "Indica rango de fechas."

    try:
        d1, d2 = _cr_range_to_utc_naive(date_from, date_to)
    except Exception:
        return None, None, None, "Formato de fecha inválido (usa YYYY-MM-DD)."

    if d2 < d1:
        return None, None, None, "El rango de fechas es inválido (Hasta < Desde)."

    return movement_type, d1, d2, None


def _query_report_rows(site_id, movement_type, d1, d2):
    q = (
        db.session.query(Movement, Container)
        .join(Container, Container.id == Movement.container_id)
        .filter(Movement.site_id == site_id)
        .filter(Movement.occurred_at >= d1, Movement.occurred_at <= d2)
    )

    if movement_type:
        q = q.filter(Movement.movement_type == movement_type)

    return q.order_by(Movement.occurred_at.asc()).all()




# =========================
# EIR - Listado / Detalle / PDF
# =========================
@yard_bp.get("/eir")
@login_required
def eir_list_view():
    site_id = _ensure_active_site()

    q = (request.args.get("q") or "").strip().upper()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    status = (request.args.get("status") or "").strip().upper()

    query = (
        EIR.query
        .filter(EIR.site_id == site_id)
        .outerjoin(Container, Container.id == EIR.container_id)
        .outerjoin(Chassis, Chassis.id == EIR.chassis_id)
    )

    if q:
        query = query.filter(
            or_(
                Container.code.ilike(f"%{q}%"),
                Chassis.chassis_number.ilike(f"%{q}%")
            )
        )

    if date_from:
        try:
            d_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            query = query.filter(EIR.trip_date >= d_from)
        except Exception:
            flash("Fecha desde inválida.", "danger")
            return redirect(url_for("yard.eir_list_view"))

    if date_to:
        try:
            d_to = datetime.strptime(date_to, "%Y-%m-%d").date()
            query = query.filter(EIR.trip_date <= d_to)
        except Exception:
            flash("Fecha hasta inválida.", "danger")
            return redirect(url_for("yard.eir_list_view"))

    if status:
        query = query.filter(EIR.status == status)

    rows = query.order_by(EIR.id.desc()).all()

    return render_template(
        "yard/eir_list.html",
        rows=rows,
        q=q,
        date_from=date_from,
        date_to=date_to,
        status=status,
    )


@yard_bp.get("/eir/<int:eir_id>")
@login_required
def eir_detail_view(eir_id: int):
    site_id = _ensure_active_site()

    eir = EIR.query.get_or_404(eir_id)
    if eir.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    return render_template("yard/eir_detail.html", eir=eir)


@yard_bp.get("/eir/<int:eir_id>/pdf")
@login_required
def eir_pdf_view(eir_id: int):
    site_id = _ensure_active_site()

    eir = EIR.query.get_or_404(eir_id)
    if eir.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    # Por ahora esto renderiza una vista imprimible HTML.
    # Luego, si quieres PDF real, aquí metemos WeasyPrint o xhtml2pdf.
    return render_template("yard/eir_pdf.html", eir=eir)

@yard_bp.post("/eir/<int:eir_id>/revert")
@login_required
def eir_revert_view(eir_id: int):
    site_id = _ensure_active_site()

    eir = EIR.query.get_or_404(eir_id)
    if eir.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    # Solo EIR FINAL puede revertirse
    if eir.status != "CONFIRMED":
        flash("Solo se puede revertir un EIR en estado CONFIRMADO.", "danger")
        return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

    now_utc = datetime.now(UTC_TZ)

    editable_until = eir.editable_until
    if editable_until and editable_until.tzinfo is None:
        editable_until = UTC_TZ.localize(editable_until)

    # Validar ventana de 24 horas
    if not eir.editable_until or now_utc > eir.editable_until:
        flash("La ventana de 24 horas para revertir este EIR ya venció.", "danger")
        return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

    reverted_anything = False

    # ==========================================================
    # Restaurar contenedor a inventario / patio
    # ==========================================================
    if eir.has_container and eir.container_id:
        c = Container.query.get(eir.container_id)
        if c:
            c.is_in_yard = True
            db.session.add(c)
            reverted_anything = True

            # Restaurar posición si está en snapshot
            snap = eir.container_snapshot_json or {}
            pos = snap.get("position") or {}
            bay_code = (pos.get("bay_code") or "").strip().upper()
            depth_row = pos.get("depth_row")
            tier = pos.get("tier")

            if bay_code and depth_row and tier:
                bay = YardBay.query.filter_by(
                    code=bay_code,
                    site_id=eir.site_id,
                    is_active=True
                ).first()

                if bay:
                    # Si no existe una posición actual, la recreamos
                    existing_pos = ContainerPosition.query.filter_by(container_id=c.id).first()
                    if not existing_pos:
                        slot_taken = ContainerPosition.query.filter_by(
                            bay_id=bay.id,
                            depth_row=depth_row,
                            tier=tier
                        ).first()

                        # Solo restaurar al slot exacto si no está ocupado
                        if not slot_taken:
                            db.session.add(
                                ContainerPosition(
                                    container_id=c.id,
                                    bay_id=bay.id,
                                    depth_row=depth_row,
                                    tier=tier,
                                    placed_by_user_id=current_user.id,
                                )
                            )

    # ==========================================================
    # Restaurar chasis a inventario
    # ==========================================================
    if eir.has_chassis and eir.chassis_id:
        ch = Chassis.query.get(eir.chassis_id)
        if ch:
            ch.site_id = eir.site_id
            ch.is_in_yard = True
            db.session.add(ch)
            reverted_anything = True

    # ==========================================================
    # Marcar EIR como revertido
    # ==========================================================
    eir.status = "REVERTED"
    eir.reverted_at = now_utc
    eir.reverted_by_user_id = current_user.id
    eir.inventory_restored_at = now_utc
    eir.updated_at = now_utc

    # No borramos movement, no borramos daños, no borramos snapshot.
    # El EIR queda como evidencia histórica de una salida anulada.

    audit_log(
        current_user.id,
        "EIR_REVERTED",
        "eir",
        eir.id,
        {
            "site_id": eir.site_id,
            "eir_id": eir.id,
            "container_id": eir.container_id,
            "chassis_id": eir.chassis_id,
            "movement_id": eir.gate_out_movement_id,
            "reverted_anything": reverted_anything,
        },
    )

    db.session.commit()
    flash(f"EIR #{eir.id} revertido correctamente. El equipo volvió a inventario.", "success")
    return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

@yard_bp.post("/eir/<int:eir_id>/confirm")
@login_required
def eir_confirm_view(eir_id: int):
    site_id = _ensure_active_site()

    eir = EIR.query.get_or_404(eir_id)
    if eir.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    if eir.status != "PENDING":
        flash("Solo se puede confirmar un EIR en estado PENDIENTE.", "danger")
        return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

    c = None
    bay_code = depth_row = tier = None

    if eir.has_container and eir.container_id:
        c = Container.query.get(eir.container_id)
        if not c or c.site_id != site_id or not c.is_in_yard:
            flash("El contenedor ya no está disponible en inventario para confirmar este EIR.", "danger")
            return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

        pos = ContainerPosition.query.filter_by(container_id=c.id).first()
        if pos:
            bay = YardBay.query.get(pos.bay_id)
            bay_code = bay.code if bay else None
            depth_row = pos.depth_row
            tier = pos.tier

    ch = None
    if eir.has_chassis and eir.chassis_id:
        ch = Chassis.query.get(eir.chassis_id)
        if not ch or ch.site_id != site_id or not ch.is_in_yard:
            flash("El chasis ya no está disponible en inventario para confirmar este EIR.", "danger")
            return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

    if not c and not ch:
        flash("Este EIR no tiene equipo válido para confirmar.", "danger")
        return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

    mv = Movement(
        site_id=site_id,
        container_id=c.id if c else None,
        movement_type="GATE_OUT",
        occurred_at=datetime.utcnow(),
        bay_code=bay_code,
        depth_row=depth_row,
        tier=tier,
        driver_name=eir.driver_name or None,
        driver_id_doc=eir.driver_id_doc or None,
        truck_plate=eir.truck_plate or None,
        notes=eir.general_notes or None,
        created_by_user_id=current_user.id,
        created_at=datetime.utcnow(),
    )
    db.session.add(mv)
    db.session.flush()

    if c:
        ContainerPosition.query.filter_by(container_id=c.id).delete()
        c.is_in_yard = False
        db.session.add(c)

    if ch:
        ch.is_in_yard = False
        db.session.add(ch)

        inv_rows = ChassisInventory.query.filter_by(chassis_id=ch.id).all()
        for inv in inv_rows:
            inv.is_in_yard = False
            inv.updated_at = datetime.utcnow()
            db.session.add(inv)

    now_utc = datetime.utcnow()
    eir.gate_out_movement_id = mv.id
    eir.status = "CONFIRMED"
    eir.finalized_at = now_utc
    eir.inventory_out_at = now_utc
    eir.editable_until = now_utc + timedelta(hours=24)
    eir.updated_at = now_utc
    eir.last_edited_at = now_utc
    eir.last_edited_by_user_id = current_user.id

    audit_log(
        current_user.id,
        "EIR_CONFIRMED",
        "eir",
        eir.id,
        {
            "site_id": site_id,
            "eir_id": eir.id,
            "movement_id": mv.id,
            "container_id": eir.container_id,
            "chassis_id": eir.chassis_id,
            "chassis_only": bool(ch and not c),
        },
    )

    db.session.commit()
    flash(f"EIR #{eir.id} confirmado correctamente. Se aplicó el Gate Out.", "success")
    return redirect(url_for("yard.eir_detail_view", eir_id=eir.id))

@yard_bp.get("/reports")
@login_required
def reports_view():
    return render_template("yard/reports.html", rows=None, movement_type="", date_from="", date_to="")


@yard_bp.get("/reports/run")
@login_required
def reports_run():
    site_id = _ensure_active_site()

    movement_type, d1, d2, err = _parse_report_filters(request.args)
    if err:
        flash(err, "danger")
        return redirect(url_for("yard.reports_view"))

    rows = _query_report_rows(site_id, movement_type, d1, d2)

    audit_log(
        current_user.id,
        "REPORT_RUN",
        "report",
        None,
        {
            "from": request.args.get("date_from"),
            "to": request.args.get("date_to"),
            "movement_type": movement_type or "ALL",
            "site_id": site_id,
        },
    )
    db.session.commit()

    return render_template(
        "yard/reports.html",
        rows=rows,
        date_from=request.args.get("date_from"),
        date_to=request.args.get("date_to"),
        movement_type=movement_type,
    )


@yard_bp.get("/reports/export")
@login_required
def reports_export():
    site_id = _ensure_active_site()

    movement_type, d1, d2, err = _parse_report_filters(request.args)
    if err:
        flash(err, "danger")
        return redirect(url_for("yard.reports_view"))

    rows = _query_report_rows(site_id, movement_type, d1, d2)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        flash("No se puede exportar: falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.reports_run", **request.args))

    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = ["Fecha/Hora", "Movimiento", "Contenedor", "Ubicación", "Chofer", "Placa"]
    ws.append(headers)

    for mv, c in rows:
        loc = "—"
        if mv.bay_code:
            parts = [mv.bay_code]
            if mv.depth_row:
                parts.append(f"F{int(mv.depth_row):02d}")
            if mv.tier:
                parts.append(f"N{int(mv.tier)}")
            loc = " ".join(parts)

        ws.append([
            mv.occurred_at.strftime("%Y-%m-%d %H:%M:%S") if mv.occurred_at else "",
            mv.movement_type or "",
            c.code if c else "",
            loc,
            mv.driver_name or "",
            mv.truck_plate or "",
        ])

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for cell in ws[get_column_letter(col_idx)]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)

    audit_log(
        current_user.id,
        "REPORT_EXPORTED",
        "report",
        None,
        {
            "from": request.args.get("date_from"),
            "to": request.args.get("date_to"),
            "movement_type": movement_type or "ALL",
            "rows": len(rows),
            "site_id": site_id,
        },
    )
    db.session.commit()

    mt = movement_type or "ALL"
    fname = f"reportes_{mt}_{request.args.get('date_from')}_a_{request.args.get('date_to')}.xlsx"

    return send_file(
        buff,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================
# Tickets / impresión
# =========================
@yard_bp.post("/print/<int:movement_id>")
@login_required
def print_ticket(movement_id: int):
    site_id = _ensure_active_site()

    mv = Movement.query.get_or_404(movement_id)
    if mv.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    c = Container.query.get_or_404(mv.container_id)

    payload = build_ticket_payload("Yard Gate Álamo", mv, c)

    agent_url = os.environ.get("PRINT_AGENT_URL")  # ej: http://10.0.0.50:9109/print
    if not agent_url:
        return jsonify({"error": "PRINT_AGENT_URL no configurado"}), 500

    r = requests.post(
        agent_url,
        json={"printer": "EPSON_M188D", "payload": payload},
        timeout=10,
    )

    if r.status_code != 200:
        return jsonify({"error": "Print agent error", "detail": r.text}), 502

    _register_ticket_print(site_id, mv.id, current_user.id, payload)
    audit_log(current_user.id, "TICKET_PRINTED_AGENT", "movement", mv.id, {"container": c.code, "site_id": site_id})
    db.session.commit()

    return jsonify({"ok": True})


@yard_bp.get("/ticket/<int:movement_id>")
@login_required
def ticket_view(movement_id: int):
    site_id = _ensure_active_site()

    mv = Movement.query.get_or_404(movement_id)
    if mv.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    c = Container.query.get_or_404(mv.container_id)

    payload = build_ticket_payload(APP_NAME, mv, c)
    _register_ticket_print(site_id, mv.id, current_user.id, payload)
    audit_log(current_user.id, "TICKET_PRINTED", "movement", mv.id, {"container": c.code, "site_id": site_id})
    db.session.commit()

    return render_template("yard/ticket.html", mv=mv, c=c, payload=payload)


@yard_bp.get("/ticket/reprint/<int:print_id>")
@login_required
def ticket_reprint(print_id: int):
    site_id = _ensure_active_site()

    tp = TicketPrint.query.get_or_404(print_id)
    if tp.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    mv = Movement.query.get_or_404(tp.movement_id)
    c = Container.query.get_or_404(mv.container_id)

    audit_log(
        current_user.id,
        "TICKET_REPRINTED",
        "ticket_print",
        tp.id,
        {"movement_id": mv.id, "container": c.code, "site_id": site_id},
    )
    db.session.commit()

    return render_template("yard/ticket.html", mv=mv, c=c, payload=tp.ticket_payload, is_reprint=True)


# =========================
# Chassis helpers
# =========================

CHASSIS_STATUSES = {"BUENO", "DAÑADO", "FUERA_DE_SERVICIO", "ATADO"}
CHASSIS_KINDS = {"CHASIS", "LOW_BOY", "TANQUETA", "PLANA", "CARRETA"}

def _norm_enum(val):
    return (val or "").strip().upper().replace(" ", "_")

def _calc_tire_state_from_data(estrias_mm, is_flat=False):
    if is_flat:
        return "PINCHADA"

    if estrias_mm in (None, "",):
        return "OK"

    try:
        mm = int(estrias_mm)
    except Exception:
        return "OK"

    if 9 <= mm <= 12:
        return "OK"         # verde
    if 4 <= mm <= 8:
        return "GASTADA"    # amarillo
    if 1 <= mm <= 3:
        return "NO_APTA"    # rojo

    return "OK"

def classify_chassis_number(num: str):
    prefix = (num or "")[:2]
    if prefix == "40":
        return 40, 2, "40FT_2AX"
    if prefix == "43":
        return 40, 3, "40FT_3AX"
    if prefix == "20":
        return 20, 2, "20FT_2AX"
    if prefix == "23":
        return 20, 3, "20FT_3AX"
    return None, None, "UNKNOWN"

def _normalize_lights_status_for_db(value: str | None) -> str | None:
    v = (value or "").strip().upper()

    mapping = {
        "OK": "OK",

        # Valores del frontend / lógica actual
        "UNA_DANADA": "IZQ_DANADA",
        "UNA_DAÑADA": "IZQ_DANADA",
        "AMBAS_DANADAS": "IZQ_DANADA",
        "AMBAS_DAÑADAS": "IZQ_DANADA",

        # Si alguna vez mandan lado explícito
        "IZQ_DANADA": "IZQ_DANADA",
        "IZQ_DAÑADA": "IZQ_DANADA",
        "DER_DANADA": "DER_DANADA",
        "DER_DAÑADA": "DER_DANADA",
    }

    return mapping.get(v, "OK" if v else None)


def _build_chassis_gate_in_ticket_text(
    *,
    site_name: str,
    username: str,
    occurred_at: datetime,
    chassis_number: str,
    plate: str | None,
    structure_status: str | None,
    twistlocks_status: str | None,
    landing_gear_status: str | None,
    lights_status: str | None,
    mudflap_status: str | None,
    plate_validation_status: str | None,
    damage_summary: str | None,
    comments: str | None,
    driver_comments: str | None,
    tire_rows: list[dict],
    alert_lines: list[str],
) -> str:
    dt_local = occurred_at.replace(tzinfo=UTC_TZ).astimezone(CR_TZ)

    lines = []
    lines.append(APP_NAME)
    lines.append("CLASIFICACION CHASIS - GATE IN")
    lines.append(dt_local.strftime("%d/%m/%Y %I:%M %p"))
    if site_name:
        lines.append(f"PREDIO: {site_name}")
    if username:
        lines.append(f"USUARIO: {username}")

    lines.append("--------------------------------")
    lines.append(f"CHASIS: {chassis_number}")
    lines.append(f"PLACA: {plate or 'SIN PLACA'}")

    lines.append("--------------------------------")
    lines.append("ESTRUCTURA")
    lines.append(f"Estructura: {structure_status or '—'}")
    lines.append(f"Twistlocks: {twistlocks_status or '—'}")
    lines.append(f"Patas: {landing_gear_status or '—'}")
    lines.append(f"Luces: {lights_status or '—'}")
    lines.append(f"Faldones: {mudflap_status or '—'}")
    lines.append(f"Placa: {plate_validation_status or '—'}")

    if damage_summary:
        lines.append(f"Resumen: {damage_summary}")
    if comments:
        lines.append(f"Chequeador: {comments}")
    if driver_comments:
        lines.append(f"Chofer: {driver_comments}")

    lines.append("--------------------------------")
    lines.append("LLANTAS")

    for row in tire_rows:
        estrias_txt = row.get("estrias_mm")
        if estrias_txt in (None, ""):
            estrias_txt = "—"

        flat_txt = "SI" if row.get("is_flat") else "NO"

        lines.append(
            f"{row['pos']} | M:{row['marchamo_config'] or '—'} | "
            f"M ING:{row['seal_status']} | "
            f"L:{row['tire_number_config'] or '—'} | "
            f"L ING:{row['tire_number_status']} | "
            f"MM:{estrias_txt} | PINCH:{flat_txt} | "
            f"EST:{row['tire_state']}"
        )

    if alert_lines:
        lines.append("--------------------------------")
        lines.append("ALERTAS")
        for x in alert_lines:
            lines.append(f"! {x}")

    lines.append("--------------------------------")
    return "\n".join(lines).strip()

def _normalize_structure_status_for_db(value: str | None) -> str | None:
    v = (value or "").strip().upper()

    mapping = {
        "OK": "OK",

        "DANO_LEVE": "GOLPE",
        "DAÑO_LEVE": "GOLPE",
        "DANO_GRAVE": "DOBLADO",
        "DAÑO_GRAVE": "DOBLADO",

        "GOLPE": "GOLPE",
        "DOBLADO": "DOBLADO",
        "SOLDADURA": "SOLDADURA",

        "DAÑADO": "DOBLADO",
        "DANADO": "DOBLADO",
        "FUERA_DE_SERVICIO": "DOBLADO",
        "ATADO": "SOLDADURA",
    }

    return mapping.get(v, "OK" if v else None)


def _normalize_twistlocks_status_for_db(value: str | None) -> str | None:
    v = (value or "").strip().upper()

    mapping = {
        "OK": "BIEN",
        "BIEN": "BIEN",

        "DANO_LEVE": "DANADOS",
        "DAÑO_LEVE": "DANADOS",
        "DANO_GRAVE": "DANADOS",
        "DAÑO_GRAVE": "DANADOS",
        "DAÑADO": "DANADOS",
        "DANADO": "DANADOS",
        "FUERA_DE_SERVICIO": "DANADOS",
        "ATADO": "DANADOS",

        "DANADOS": "DANADOS",
        "DAÑADOS": "DANADOS",
    }

    return mapping.get(v, "BIEN" if v else None)


def _normalize_landing_gear_status_for_db(value: str | None) -> str | None:
    v = (value or "").strip().upper()

    mapping = {
        "OK": "OK",
        "DANADAS": "DANADAS",
        "DAÑADAS": "DANADAS",

        "DANO_LEVE": "DANADAS",
        "DAÑO_LEVE": "DANADAS",
        "DANO_GRAVE": "DANADAS",
        "DAÑO_GRAVE": "DANADAS",
        "DAÑADO": "DANADAS",
        "DANADO": "DANADAS",
        "FUERA_DE_SERVICIO": "DANADAS",
        "ATADO": "DANADAS",
    }

    return mapping.get(v, "OK" if v else None)


def _normalize_lights_status_for_db(value: str | None) -> str | None:
    v = (value or "").strip().upper()

    mapping = {
        "OK": "OK",

        "UNA_DANADA": "IZQ_DANADA",
        "UNA_DAÑADA": "IZQ_DANADA",
        "AMBAS_DANADAS": "IZQ_DANADA",
        "AMBAS_DAÑADAS": "IZQ_DANADA",

        "IZQ_DANADA": "IZQ_DANADA",
        "IZQ_DAÑADA": "IZQ_DANADA",
        "DER_DANADA": "DER_DANADA",
        "DER_DAÑADA": "DER_DANADA",
    }

    return mapping.get(v, "OK" if v else None)


def _normalize_mudflap_status_for_db(value: str | None) -> str | None:
    v = (value or "").strip().upper()

    mapping = {
        "OK": "OK",
        "NO_TRAE": "NO_TRAE",

        "DANADO": "NO_TRAE",
        "DAÑADO": "NO_TRAE",
        "FUERA_DE_SERVICIO": "NO_TRAE",
        "ATADO": "NO_TRAE",
    }

    return mapping.get(v, "OK" if v else None)

def allowed_positions_for(axles: int):
    if axles == 2:
        return [
            "AX1_L_IN", "AX1_L_OUT", "AX1_R_IN", "AX1_R_OUT",
            "AX2_L_IN", "AX2_L_OUT", "AX2_R_IN", "AX2_R_OUT",
        ]
    if axles == 3:
        return [
            "AX1_L_IN", "AX1_L_OUT", "AX1_R_IN", "AX1_R_OUT",
            "AX2_L_IN", "AX2_L_OUT", "AX2_R_IN", "AX2_R_OUT",
            "AX3_L_IN", "AX3_L_OUT", "AX3_R_IN", "AX3_R_OUT",
        ]
    return []

def _normalize_position_for_tire_master(pos: str) -> str | None:
    """
    Convierte posiciones detalladas del frontend/chassis_tires a la posición
    maestra que existe en yard_gate_alamo.tire_positions.

    Ejemplos:
    AX1_L_IN  -> A1_IN
    AX1_R_IN  -> A1_IN
    AX2_L_OUT -> A2_OUT
    AX3_R_OUT -> A3_OUT

    También acepta ya-normalizadas:
    A1_IN, A1_OUT, A2_IN, A2_OUT, etc.
    """
    value = (pos or "").strip().upper()

    m_simple = re.match(r"^A([1-3])_(IN|OUT)$", value)
    if m_simple:
        return value

    m = re.match(r"^AX([1-3])_(L|R)_(IN|OUT)$", value)
    if not m:
        return None

    axle = m.group(1)
    inout = m.group(3)
    return f"A{axle}_{inout}"


def _resolve_tire_position_id(axles: int, pos: str) -> int:
    """
    Busca el id real en yard_gate_alamo.tire_positions usando:
    - axle_count (2 o 3)
    - position_code normalizado (A1_IN, A1_OUT, etc.)
    """
    master_pos = _normalize_position_for_tire_master(pos)
    if not master_pos:
        raise ValueError(f"Posición inválida para tire_positions: {pos}")

    sql = text("""
        SELECT id
        FROM yard_gate_alamo.tire_positions
        WHERE axle_count = :axles
          AND position_code = :position_code
        LIMIT 1
    """)
    row = db.session.execute(sql, {
        "axles": int(axles),
        "position_code": master_pos,
    }).mappings().first()

    if not row:
        raise ValueError(
            f"No existe tire_positions para axle_count={axles}, position_code={master_pos}"
        )

    return int(row["id"])


def _condition_from_tire_states(states: list[str]) -> str:
    """
    Mapea varios estados internos de llanta al valor permitido por
    tire_readings.condition:
    - OK
    - DESGASTADA
    - REPARABLE
    """
    normalized = {(x or "").strip().upper() for x in states if x}

    if normalized & {"PINCHADA", "CAMBIAR", "NO_APTA"}:
        return "REPARABLE"
    if "GASTADA" in normalized:
        return "DESGASTADA"
    return "OK"


def _pick_valid_pressure(values: list) -> float | None:
    """
    Devuelve la primera presión válida (> 0). Si ninguna sirve, None.
    """
    for v in values:
        try:
            num = float(v)
            if num > 0:
                return num
        except Exception:
            continue
    return None


def _insert_tire_reading_row(
    *,
    site_id: int,
    chassis_id: int,
    axles: int,
    pos: str,
    event_type: str,
    event_id,
    seal_1,
    seal_2,
    pressure_psi,
    condition: str | None,
    comments: str | None,
    user_id: int,
):
    """
    Inserta una fila REAL en yard_gate_alamo.tire_readings usando tire_positions.id.
    """
    allowed_event_types = {"GATE_IN", "EIR_OUT", "EIR_IN"}
    resolved_event_type = event_type if event_type in allowed_event_types else "GATE_IN"

    tire_position_id = _resolve_tire_position_id(axles, pos)

    resolved_pressure = pressure_psi
    try:
        if resolved_pressure not in (None, ""):
            resolved_pressure = float(resolved_pressure)
        else:
            resolved_pressure = None
    except Exception:
        resolved_pressure = None

    # Constraint real:
    # ck_pressure_gate_in -> si event_type='GATE_IN', pressure_psi > 0
    if resolved_event_type == "GATE_IN" and (resolved_pressure is None or resolved_pressure <= 0):
        resolved_pressure = 1.0

    payload = {
        "site_id": site_id,
        "chassis_id": chassis_id,
        "event_type": resolved_event_type,
        "event_id": event_id,
        "tire_position_id": tire_position_id,
        "seal_1": seal_1,
        "seal_2": seal_2,
        "pressure_psi": resolved_pressure,
        "condition": condition,
        "comments": comments,
        "recorded_at": datetime.utcnow(),
        "recorded_by_user_id": user_id,
    }

    _insert_dynamic("yard_gate_alamo", "tire_readings", payload)


def _save_grouped_tire_readings(
    *,
    site_id: int,
    chassis_id: int,
    axles: int,
    grouped_readings: dict,
    user_id: int,
    event_type: str,
    event_id=None,
):
    """
    Guarda lecturas agrupadas por posición maestra de tire_positions.
    Ejemplo:
      AX1_L_IN + AX1_R_IN -> A1_IN
      AX1_L_OUT + AX1_R_OUT -> A1_OUT
    """
    for master_pos, group in grouped_readings.items():
        states = group.get("states", [])
        pressures = group.get("pressures", [])
        comments_list = group.get("comments", [])
        seal_issue = bool(group.get("seal_issue"))
        second_seal = group.get("seal_2")

        condition = _condition_from_tire_states(states)
        pressure = _pick_valid_pressure(pressures)

        _insert_tire_reading_row(
            site_id=site_id,
            chassis_id=chassis_id,
            axles=axles,
            pos=master_pos,
            event_type=event_type,
            event_id=event_id,
            seal_1="DISTINTO" if seal_issue else None,
            seal_2=second_seal,
            pressure_psi=pressure,
            condition=condition,
            comments=" || ".join(comments_list) if comments_list else None,
            user_id=user_id,
        )

# =========================
# Chassis classification helpers (sin modelos nuevos)
# =========================

MARCHAMO_CHECK = {"OK", "DISTINTO", "NO_TIENE", "ILEGIBLE"}

# Traducciones pedidas
LABELS_ES = {
    "landing_gear": "Pata de apoyo",
    "mudflap": "Faldones",
}

def _map_tire_condition_for_db(tire_state: str | None) -> str | None:
    """
    Mapea el estado interno de llanta a los únicos valores permitidos por
    yard_gate_alamo.tire_readings.condition:
    - OK
    - DESGASTADA
    - REPARABLE
    """
    value = (tire_state or "").strip().upper()

    mapping = {
        "OK": "OK",
        "GASTADA": "DESGASTADA",
        "PINCHADA": "REPARABLE",
        "CAMBIAR": "REPARABLE",
        "NO_APTA": "REPARABLE",
    }
    return mapping.get(value, "OK")

def _get_table_columns(schema: str, table: str) -> set[str]:
    sql = text("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = :schema AND table_name = :table
    """)
    rows = db.session.execute(sql, {"schema": schema, "table": table}).fetchall()
    return {r[0] for r in rows}

def _insert_dynamic(schema: str, table: str, values: dict) -> int | None:
    """
    Inserta solo las columnas que existan en la tabla (evita romper por schema distinto).
    Retorna id si la tabla tiene id y la DB lo devuelve, si no, None.
    """
    cols = _get_table_columns(schema, table)
    payload = {k: v for k, v in values.items() if k in cols}

    if not payload:
        return None

    col_list = ", ".join(payload.keys())
    param_list = ", ".join([f":{k}" for k in payload.keys()])

    # Si existe 'id' intentamos RETURNING id, si no, sin returning.
    if "id" in cols:
        sql = text(f"""
            INSERT INTO {schema}.{table} ({col_list})
            VALUES ({param_list})
            RETURNING id
        """)
        new_id = db.session.execute(sql, payload).scalar()
        return int(new_id) if new_id is not None else None

    sql = text(f"""
        INSERT INTO {schema}.{table} ({col_list})
        VALUES ({param_list})
    """)
    db.session.execute(sql, payload)
    return None

def _fetch_last_final_eir_for_chassis(chassis_id: int):
    """
    Trae el último EIR FINAL (o estado equivalente) donde participó ese chasis.
    No asumimos nombres perfectos: usamos columnas existentes.
    """
    cols = _get_table_columns("yard_gate_alamo", "eirs")
    status_col = "status" if "status" in cols else None
    updated_col = "updated_at" if "updated_at" in cols else ("created_at" if "created_at" in cols else None)

    where_status = ""
    if status_col:
        # soporta tu flujo actual
        where_status = f"AND COALESCE(e.{status_col}, '') IN ('CONFIRMED','FINAL','CERRADO','POR COBRAR','PENDIENTE COBRO','ABIERTO','ASIGNADO')"

    order_by = f"ORDER BY e.{updated_col} DESC NULLS LAST, e.id DESC" if updated_col else "ORDER BY e.id DESC"

    sql = text(f"""
        SELECT e.*
        FROM yard_gate_alamo.eirs e
        WHERE e.chassis_id = :cid
        {where_status}
        {order_by}
        LIMIT 1
    """)
    row = db.session.execute(sql, {"cid": chassis_id}).mappings().first()
    return row  # dict-like o None

def _last_confirmed_eir_destination_by_chassis_ids(chassis_ids: list[int]) -> dict[int, str]:
    """
    Trae el último destination de EIR CONFIRMED por chasis.
    Retorna: {chassis_id: "DESTINO"}
    """
    if not chassis_ids:
        return {}

    sql = (
        text(
            """
            SELECT DISTINCT ON (chassis_id)
                chassis_id,
                destination
            FROM yard_gate_alamo.eirs
            WHERE chassis_id IN :ids
              AND status = 'CONFIRMED'
            ORDER BY chassis_id, id DESC
            """
        )
        .bindparams(bindparam("ids", expanding=True))
    )

    rows = db.session.execute(sql, {"ids": chassis_ids}).mappings().all()

    out = {}
    for r in rows:
        out[int(r["chassis_id"])] = (r["destination"] or "").strip()

    return out

def _build_workshop_ticket_text(
    chassis_number: str,
    axles: int,
    structure_lines: list[str],
    tire_lines: list[str],
    eir_prev_id: int | None
) -> str:
    out = []
    out.append(f"CHASIS: {chassis_number}")
    out.append(f"EJES: {axles}")
    if eir_prev_id:
        out.append(f"CONCILIAR CONTRA EIR ANTERIOR: #{eir_prev_id}")

    if structure_lines:
        out.append("")
        out.append("DAÑOS / OBSERVACIONES (ESTRUCTURA):")
        out.extend([f"- {x}" for x in structure_lines])

    if tire_lines:
        out.append("")
        out.append("LLANTAS / MARCHAMOS:")
        out.extend([f"- {x}" for x in tire_lines])

    return "\n".join(out).strip()

def _save_tire_reading(site_id: int, chassis_id: int, pos: str, ingreso_marchamo: str | None,
                       check: str, tire_state: str, user_id: int, estrias_mm=None,
                       is_flat=False, event_type: str = "GATE_IN", pressure_psi=None,
                       event_id=None, second_seal=None):
    """
    Función legacy. Se mantiene temporalmente para no romper imports o referencias viejas,
    pero el flujo correcto ahora usa:
    - _normalize_position_for_tire_master
    - _insert_tire_reading_row
    - _save_grouped_tire_readings
    """
    master_pos = _normalize_position_for_tire_master(pos)
    if not master_pos:
        return

    _insert_tire_reading_row(
        site_id=site_id,
        chassis_id=chassis_id,
        axles=2 if "AX3" not in (pos or "").upper() else 3,
        pos=master_pos,
        event_type=event_type,
        event_id=event_id,
        seal_1=ingreso_marchamo or (check if check != "OK" else None),
        seal_2=second_seal,
        pressure_psi=pressure_psi,
        condition=_condition_from_tire_states([tire_state]),
        comments=(
            f"POS {pos} | CHECK_MARCHAMO {check}"
            + (f" | MM {estrias_mm}" if estrias_mm not in (None, "") else "")
            + (" | PINCHADA SI" if is_flat else "")
        ),
        user_id=user_id,
    )

def _maybe_register_tire_retread(
    *,
    tire_id: int | None,
    previous_estrias_mm,
    new_estrias_mm,
    previous_marchamo: str | None,
    new_marchamo: str | None,
    created_by: int | None,
):
    """
    Registra un recauche cuando la misma llanta pasa de estrías bajas a 12 mm.
    Regla actual:
    - misma llanta
    - estrías anteriores <= 4
    - nuevas estrías = 12
    - evita registrar si ya estaba en 12
    """
    if not tire_id:
        return

    try:
        old_mm = int(previous_estrias_mm) if previous_estrias_mm not in (None, "",) else None
    except Exception:
        old_mm = None

    try:
        new_mm = int(new_estrias_mm) if new_estrias_mm not in (None, "",) else None
    except Exception:
        new_mm = None

    if old_mm is None or new_mm is None:
        return

    if not (old_mm <= 4 and new_mm == 12 and old_mm != 12):
        return

    event = TireRetreadEvent(
        tire_id=tire_id,
        previous_estrias_mm=old_mm,
        new_estrias_mm=new_mm,
        previous_marchamo=(previous_marchamo or None),
        new_marchamo=(new_marchamo or None),
        created_by=created_by,
    )
    db.session.add(event)

# =========================
# Chassis pages
# =========================

@yard_bp.get("/chassis")
@login_required
def chassis_list():
    site_id = _ensure_active_site()

    rows = (
        Chassis.query
        .filter(
            Chassis.site_id == site_id,
            Chassis.is_in_yard.is_(True),
        )
        .order_by(Chassis.chassis_number.asc())
        .all()
    )
    items = []
    for ch in rows:
        ubicacion = ch.site.name if getattr(ch, "site", None) else "Predio"

        items.append({
            "id": ch.id,
            "chassis_number": ch.chassis_number,
            "plate": ch.plate,
            "chassis_kind": ch.chassis_kind,
            "length_ft": ch.length_ft,
            "axles": ch.axles,
            "status": ch.status,
            "is_in_yard": bool(ch.is_in_yard),
            "ubicacion": ubicacion,
            "site_name": ch.site.name if getattr(ch, "site", None) else "",
        })

    sites = Site.query.order_by(Site.name.asc()).all()
    return render_template(
        "yard/chassis_list.html",
        rows=rows,
        items=items,
        sites=sites
    )


@yard_bp.get("/chassis/dashboard")
@login_required
def chassis_dashboard():
    site_id = _ensure_active_site()
    base = Chassis.query.filter(
        Chassis.site_id == site_id,
        Chassis.is_in_yard.is_(True),
    )

    counts = {
        "40FT_2AX": base.filter(Chassis.type_code == "40FT_2AX").count(),
        "40FT_3AX": base.filter(Chassis.type_code == "40FT_3AX").count(),
        "20FT_2AX": base.filter(Chassis.type_code == "20FT_2AX").count(),
        "20FT_3AX": base.filter(Chassis.type_code == "20FT_3AX").count(),
    }
    total = base.count()
    unknown = base.filter((Chassis.type_code.is_(None)) | (Chassis.type_code == "UNKNOWN")).count()

    return render_template("yard/chassis_dashboard.html", total=total, unknown=unknown, counts=counts)


@yard_bp.get("/chassis/import")
@login_required
def chassis_import_view():
    return render_template("yard/chassis_import.html")


@yard_bp.post("/chassis/import")
@login_required
def chassis_import_post():
    site_id = _ensure_active_site()

    f = request.files.get("file")
    if not f or not f.filename:
        flash("Sube un archivo Excel.", "danger")
        return redirect(url_for("yard.chassis_import_view"))

    try:
        from openpyxl import load_workbook
    except Exception:
        flash("Falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.chassis_import_view"))

    wb = load_workbook(f, data_only=True)
    ws = wb.active

    imported = 0
    updated = 0
    errors = []

    # -------------------------
    # 0) Cache de Sites
    # -------------------------
    sites = Site.query.all()
    sites_by_name = {(s.name or "").strip().upper(): s for s in sites}

    # -------------------------
    # 1) Leer Excel a memoria
    # -------------------------
    staged = []
    numbers = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        chassis_number = (str(row[0]).strip() if row and row[0] is not None else "")
        plate = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else None)
        length_ft = row[2] if len(row) > 2 else None
        axles = row[3] if len(row) > 3 else None
        type_code = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else None)

        # NUEVO: status / tipo / predio
        status = _norm_enum(row[5]) if len(row) > 5 and row[5] is not None else ""
        chassis_kind = _norm_enum(row[6]) if len(row) > 6 and row[6] is not None else ""
        predio_name = (str(row[7]).strip() if len(row) > 7 and row[7] is not None else "")

        if not CHASSIS_NUM_RE.match(chassis_number):
            errors.append(f"Fila {idx}: chassis_number inválido ({chassis_number})")
            continue

        # fallback por prefijo SOLO si falta algo
        d_len, d_ax, d_type = classify_chassis_number(chassis_number)

        if (not length_ft) or (not axles):
            if d_len is None or d_ax is None:
                errors.append(f"Fila {idx}: prefijo no reconocido ({chassis_number})")
                continue
            length_ft = int(length_ft) if length_ft else d_len
            axles = int(axles) if axles else d_ax

        if not type_code:
            type_code = d_type

        try:
            length_ft = int(length_ft)
            axles = int(axles)
        except Exception:
            errors.append(f"Fila {idx}: length_ft/axles inválidos")
            continue

        if length_ft not in (20, 40, 45) or axles not in (2, 3):
            errors.append(f"Fila {idx}: fuera de rango length_ft={length_ft} axles={axles}")
            continue

        # status default + validación
        if not status:
            status = "BUENO"
        if status not in CHASSIS_STATUSES:
            errors.append(f"Fila {idx}: status inválido ({status})")
            continue

        # chassis_kind default + validación
        if not chassis_kind:
            chassis_kind = "CHASIS"
        if chassis_kind not in CHASSIS_KINDS:
            errors.append(f"Fila {idx}: tipo inválido ({chassis_kind})")
            continue

        # predio (site_id) - si viene vacío, se usa el predio activo
        target_site_id = site_id
        if predio_name:
            s = sites_by_name.get(predio_name.strip().upper())
            if not s:
                errors.append(f"Fila {idx}: predio no existe ({predio_name})")
                continue
            target_site_id = s.id

        staged.append({
            "idx": idx,
            "chassis_number": chassis_number,
            "plate": plate,
            "length_ft": length_ft,
            "axles": axles,
            "type_code": type_code,
            "status": status,
            "chassis_kind": chassis_kind,
            "site_id": target_site_id,
        })
        numbers.append(chassis_number)

    if not staged:
        flash(f"No se importó nada. Errores: {len(errors)}", "danger")
        session["chassis_import_errors"] = errors[:200]
        return redirect(url_for("yard.chassis_import_view"))

    # -------------------------------------------
    # 2) Traer existentes en UNA sola consulta IN
    # -------------------------------------------
    existing_rows = (
        Chassis.query
        .filter(Chassis.chassis_number.in_(numbers))
        .all()
    )
    existing_map = {c.chassis_number: c for c in existing_rows}

    # -------------------------
    # 3) Upsert en memoria
    # -------------------------
    for item in staged:
        chassis_number = item["chassis_number"]
        plate = item["plate"]
        length_ft = item["length_ft"]
        axles = item["axles"]
        type_code = item["type_code"]
        status = item["status"]
        chassis_kind = item["chassis_kind"]
        target_site_id = item["site_id"]

        existing = existing_map.get(chassis_number)

        if existing:
            existing.site_id = target_site_id
            existing.plate = plate
            existing.length_ft = length_ft
            existing.axles = axles
            existing.type_code = type_code
            existing.status = status
            existing.chassis_kind = chassis_kind
            existing.has_plate = True if plate else False
            existing.is_in_yard = True
            db.session.add(existing)
            updated += 1
        else:
            ch = Chassis(
                site_id=target_site_id,
                chassis_number=chassis_number,
                plate=plate,
                length_ft=length_ft,
                axles=axles,
                type_code=type_code,
                status=status,
                chassis_kind=chassis_kind,
                has_plate=True if plate else False,
                is_in_yard=True,
            )
            db.session.add(ch)
            imported += 1

    # -------------------------
    # 4) Un solo commit
    # -------------------------
    db.session.commit()

    if errors:
        flash(f"Importado: {imported} | Actualizado: {updated} | Errores: {len(errors)}", "warning")
        session["chassis_import_errors"] = errors[:200]
    else:
        flash(f"Importado: {imported} | Actualizado: {updated}", "success")
        session.pop("chassis_import_errors", None)

    return redirect(url_for("yard.chassis_list"))


@yard_bp.get("/chassis/export")
@login_required
def chassis_export():
    site_id = _ensure_active_site()

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.chassis_list"))

    # Trae lo que ya existe en este predio
    rows = (
        Chassis.query
        .filter(
            Chassis.site_id == site_id,
            Chassis.is_in_yard.is_(True),
        )
        .order_by(Chassis.chassis_number.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Chassis"

    headers = [
        "chassis_number (número de chasis 5 dígitos)",
        "plate (placa) [opcional]",
        "length_ft (largo en pies: 20/40/45) [opcional]",
        "axles (ejes: 2/3) [opcional]",
        "type_code (tipo: 20FT_2AX/20FT_3AX/40FT_2AX/40FT_3AX) [opcional]",
        "status (BUENO/DAÑADO/FUERA_DE_SERVICIO/ATADO) [opcional]",
        "chassis_kind (CHASIS/LOW_BOY/TANQUETA/PLANA/CARRETA) [opcional]",
        "predio (nombre del predio / Site.name) [opcional]",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for ch in rows:
        ws.append([
            ch.chassis_number,
            ch.plate or "",
            getattr(ch, "length_ft", "") or "",
            getattr(ch, "axles", "") or "",
            ch.type_code or "",
            getattr(ch, "status", "") or "BUENO",
            getattr(ch, "chassis_kind", "") or "CHASIS",
            (ch.site.name if getattr(ch, "site", None) else ""),
        ])

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 26

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "chassis_import_template.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@yard_bp.get("/chassis/<int:chassis_id>")
@login_required
def chassis_detail(chassis_id: int):
    site_id = _ensure_active_site()
    ch = Chassis.query.get_or_404(chassis_id)

    if ch.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    axles = int(getattr(ch, "axles", 2) or 2)
    length_ft = int(getattr(ch, "length_ft", 40) or 40)

    # para el dropdown de predios + status + tipo en la vista detalle
    sites = Site.query.order_by(Site.name.asc()).all()

    return render_template(
        "yard/chassis_detail.html",
        ch=ch,
        axles=axles,
        length_ft=length_ft,
        sites=sites,
        statuses=sorted(CHASSIS_STATUSES),
        kinds=sorted(CHASSIS_KINDS),
    )


# =========================
# Chassis tires API
# =========================

@yard_bp.get("/api/chassis/<int:chassis_id>/tires")
@login_required
def api_chassis_tires_get(chassis_id: int):
    _ensure_active_site()
    ch = Chassis.query.get_or_404(chassis_id)

    axles = int(getattr(ch, "axles", 2) or 2)
    allowed = set(allowed_positions_for(axles))

    rows = ChassisTire.query.filter_by(chassis_id=ch.id).all()

    positions = {}
    for p in allowed:
        positions[p] = {
            "marchamo": None,
            "tire_state": "OK",
            "tire_id": None,
            "tire_number": None,
            "brand": None,
            "model": None,
            "size": None,
            "notes": None,
            "status": None,
            "estrias_mm": None,
            "is_flat": False,
        }

    for r in rows:
        pos = (r.position_code or "").strip().upper()
        if pos not in allowed:
            continue

        positions[pos] = {
            "marchamo": r.marchamo,
            "tire_state": (r.tire_state or "OK").upper(),
            "tire_id": r.tire.id if r.tire else None,
            "tire_number": r.tire.tire_number if r.tire else None,
            "brand": r.tire.brand if r.tire else None,
            "model": r.tire.model if r.tire else None,
            "size": r.tire.size if r.tire else None,
            "notes": r.tire.notes if r.tire else None,
            "status": r.tire.status if r.tire else None,
            "estrias_mm": getattr(r, "estrias_mm", None),
            "is_flat": bool(getattr(r, "is_flat", False)),
        }

    return jsonify({
        "ok": True,
        "chassis": {
            "id": ch.id,
            "chassis_number": ch.chassis_number,
            "plate": ch.plate,
            "axles": ch.axles,
            "status": ch.status,
            "site_id": ch.site_id,
            "type_code": getattr(ch, "type_code", None),
        },
        "positions": positions
    })
@yard_bp.get("/api/llantas/disponibles")
@login_required
def api_tires_available():
    _ensure_active_site()

    q = (request.args.get("q") or "").strip().upper()

    filters = ["t.status = 'EN_TALLER_BODEGA'"]
    params = {}

    if q:
        filters.append("""
            (
                t.tire_number ILIKE :q
                OR COALESCE(t.brand, '') ILIKE :q
                OR COALESCE(t.model, '') ILIKE :q
                OR COALESCE(t.size, '') ILIKE :q
            )
        """)
        params["q"] = f"%{q}%"

    sql = text(f"""
        SELECT
            t.id,
            t.tire_number,
            t.brand,
            t.model,
            t.size,
            t.notes,
            t.status
        FROM yard_gate_alamo.tires t
        LEFT JOIN yard_gate_alamo.chassis_tires ct
          ON ct.tire_id = t.id
        WHERE {" AND ".join(filters)}
        ORDER BY t.tire_number ASC
        LIMIT 100
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    for r in rows:
        items.append({
            "id": r["id"],
            "tire_number": r["tire_number"] or "",
            "brand": r["brand"] or "",
            "model": r["model"] or "",
            "size": r["size"] or "",
            "notes": r["notes"] or "",
            "status": r["status"] or "EN_TALLER_BODEGA",
        })

    return jsonify({
        "ok": True,
        "items": items
    })


@yard_bp.post("/api/chassis/<int:chassis_id>/tires")
@login_required
def api_chassis_tires_set(chassis_id: int):
    site_id = _ensure_active_site()
    ch = Chassis.query.get_or_404(chassis_id)

    if ch.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    axles = int(getattr(ch, "axles", 2) or 2)
    allowed = set(allowed_positions_for(axles))

    data = request.get_json(silent=True) or {}
    action = (data.get("action") or "legacy_update").strip().lower()

    pos = (data.get("position_code") or "").strip().upper()
    if pos not in allowed:
        return jsonify({"ok": False, "error": "INVALID_POSITION"}), 400

    confirm_replace = bool(data.get("confirm_replace"))

    row = ChassisTire.query.filter_by(chassis_id=ch.id, position_code=pos).first()

    # =========================
    # ACTION: unassign
    # =========================
    if action == "unassign":
        if not row:
            return jsonify({"ok": False, "error": "POSITION_EMPTY"}), 400

        tire = Tire.query.get(row.tire_id) if row.tire_id else None
        if tire:
            tire.status = "EN_TALLER_BODEGA"
            db.session.add(tire)

        db.session.delete(row)
        db.session.commit()

        return jsonify({"ok": True, "action": "unassign"})

    # =========================
    # ACTION: assign_existing
    # =========================
    if action == "assign_existing":
        tire_id = data.get("tire_id")
        marchamo = (data.get("marchamo") or "").strip()

        estrias_mm_raw = data.get("estrias_mm")
        is_flat = bool(data.get("is_flat"))

        if not tire_id:
            return jsonify({"ok": False, "error": "TIRE_ID_REQUIRED"}), 400

        tire = Tire.query.get(tire_id)
        if not tire:
            return jsonify({"ok": False, "error": "TIRE_NOT_FOUND"}), 404

        if tire.status != "EN_TALLER_BODEGA":
            return jsonify({
                "ok": False,
                "error": "TIRE_STATUS_NOT_ALLOWED",
                "detail": f"La llanta está en estado {tire.status}"
            }), 400

        existing_mount = ChassisTire.query.filter_by(tire_id=tire.id).first()
        if existing_mount and (
            existing_mount.chassis_id != ch.id or existing_mount.position_code != pos
        ):
            return jsonify({
                "ok": False,
                "error": "TIRE_ALREADY_ASSIGNED",
                "detail": (
                    f"La llanta ya está asignada al chasis "
                    f"{existing_mount.chassis.chassis_number if existing_mount.chassis else existing_mount.chassis_id} "
                    f"en la posición {existing_mount.position_code}"
                )
            }), 400

        estrias_mm = None
        if estrias_mm_raw not in (None, "",):
            try:
                estrias_mm = int(estrias_mm_raw)
            except Exception:
                return jsonify({"ok": False, "error": "INVALID_ESTRIAS_MM"}), 400

            if estrias_mm < 1 or estrias_mm > 12:
                return jsonify({"ok": False, "error": "ESTRIAS_OUT_OF_RANGE"}), 400

        if row and row.tire_id != tire.id:
            old_tire = Tire.query.get(row.tire_id) if row.tire_id else None

            if not confirm_replace:
                return jsonify({
                    "ok": False,
                    "error": "POSITION_OCCUPIED",
                    "detail": (
                        f"La posición {_translate_tire_position(pos)} ya está ocupada"
                    )
                }), 409

            if old_tire:
                old_tire.status = "EN_TALLER_BODEGA"
                db.session.add(old_tire)

            db.session.delete(row)
            db.session.flush()
            row = None

        tire_state = _calc_tire_state_from_mm(estrias_mm, is_flat)

        previous_estrias_mm = row.estrias_mm if row else None
        previous_marchamo = row.marchamo if row else None
        previous_tire_id = row.tire_id if row else None

        if not row:
            row = ChassisTire(
                chassis_id=ch.id,
                position_code=pos,
                tire_id=tire.id,
                installed_at=datetime.utcnow(),
            )

        row.chassis_id = ch.id
        row.position_code = pos
        row.tire_id = tire.id
        row.marchamo = marchamo or None
        row.estrias_mm = estrias_mm
        row.is_flat = is_flat
        row.tire_state = tire_state
        row.updated_at = datetime.utcnow()

        if previous_tire_id == tire.id:
            _maybe_register_tire_retread(
                tire_id=tire.id,
                previous_estrias_mm=previous_estrias_mm,
                new_estrias_mm=estrias_mm,
                previous_marchamo=previous_marchamo,
                new_marchamo=marchamo,
                created_by=current_user.id,
            )

        tire.status = "ASIGNADA"

        db.session.add(row)
        db.session.add(tire)
        db.session.commit()

        return jsonify({
            "ok": True,
            "action": "assign_existing",
            "tire_state": tire_state
        })

    # =========================
    # ACTION: create_and_assign
    # =========================
    if action == "create_and_assign":
        tire_number = (data.get("tire_number") or "").strip().upper()
        brand = (data.get("brand") or "").strip().upper()
        model = (data.get("model") or "").strip().upper()
        size = (data.get("size") or "").strip().upper()
        notes = (data.get("notes") or "").strip()
        marchamo = (data.get("marchamo") or "").strip()

        estrias_mm_raw = data.get("estrias_mm")
        is_flat = bool(data.get("is_flat"))

        if not tire_number:
            return jsonify({"ok": False, "error": "TIRE_NUMBER_REQUIRED"}), 400

        existing_tire = Tire.query.filter_by(tire_number=tire_number).first()
        if existing_tire:
            return jsonify({
                "ok": False,
                "error": "TIRE_NUMBER_ALREADY_EXISTS"
            }), 409

        estrias_mm = None
        if estrias_mm_raw not in (None, "",):
            try:
                estrias_mm = int(estrias_mm_raw)
            except Exception:
                return jsonify({"ok": False, "error": "INVALID_ESTRIAS_MM"}), 400

            if estrias_mm < 1 or estrias_mm > 12:
                return jsonify({"ok": False, "error": "ESTRIAS_OUT_OF_RANGE"}), 400

        if row and row.tire_id:
            old_tire = Tire.query.get(row.tire_id)

            if not confirm_replace:
                return jsonify({
                    "ok": False,
                    "error": "POSITION_OCCUPIED",
                    "detail": (
                        f"La posición {_translate_tire_position(pos)} ya está ocupada"
                    )
                }), 409

            if old_tire:
                old_tire.status = "EN_TALLER_BODEGA"
                db.session.add(old_tire)

            db.session.delete(row)
            db.session.flush()
            row = None

        tire = Tire(
            tire_number=tire_number,
            brand=brand or None,
            model=model or None,
            size=size or None,
            notes=notes or None,
            status="ASIGNADA",
        )
        db.session.add(tire)
        db.session.flush()

        tire_state = _calc_tire_state_from_mm(estrias_mm, is_flat)

        if not row:
            row = ChassisTire(
                chassis_id=ch.id,
                position_code=pos,
                tire_id=tire.id,
                installed_at=datetime.utcnow(),
            )

        row.chassis_id = ch.id
        row.position_code = pos
        row.tire_id = tire.id
        row.marchamo = marchamo or None
        row.estrias_mm = estrias_mm
        row.is_flat = is_flat
        row.tire_state = tire_state
        row.updated_at = datetime.utcnow()

        db.session.add(row)
        db.session.commit()

        return jsonify({
            "ok": True,
            "action": "create_and_assign",
            "tire_state": tire_state,
            "tire_id": tire.id
        })

    # =========================
    # LEGACY UPDATE
    # =========================
    marchamo = (data.get("marchamo") or "").strip()
    tire_number = (data.get("tire_number") or "").strip().upper()
    brand = (data.get("brand") or "").strip().upper()

    estrias_mm_raw = data.get("estrias_mm")
    is_flat = bool(data.get("is_flat"))

    estrias_mm = None
    if estrias_mm_raw not in (None, "",):
        try:
            estrias_mm = int(estrias_mm_raw)
        except Exception:
            return jsonify({"ok": False, "error": "INVALID_ESTRIAS_MM"}), 400

        if estrias_mm < 1 or estrias_mm > 12:
            return jsonify({"ok": False, "error": "ESTRIAS_OUT_OF_RANGE"}), 400

    tire_state = _calc_tire_state_from_data(estrias_mm, is_flat)

    tire = None
    if tire_number:
        tire = Tire.query.filter_by(tire_number=tire_number).first()
        if not tire:
            tire = Tire(
                tire_number=tire_number,
                brand=brand or None,
                status="ASIGNADA"
            )
            db.session.add(tire)
            db.session.flush()
        else:
            if brand and (tire.brand != brand):
                tire.brand = brand
            if tire.status == "EN_TALLER_BODEGA":
                tire.status = "ASIGNADA"
            db.session.add(tire)

    previous_estrias_mm = row.estrias_mm if row else None
    previous_marchamo = row.marchamo if row else None
    previous_tire_id = row.tire_id if row else None

    if not row:
        row = ChassisTire(chassis_id=ch.id, position_code=pos)

    row.marchamo = marchamo or None
    row.estrias_mm = estrias_mm
    row.is_flat = is_flat
    row.tire_state = tire_state
    row.tire_id = tire.id if tire else None
    row.updated_at = datetime.utcnow()

    current_tire_id = tire.id if tire else None
    if previous_tire_id and current_tire_id and previous_tire_id == current_tire_id:
        _maybe_register_tire_retread(
            tire_id=current_tire_id,
            previous_estrias_mm=previous_estrias_mm,
            new_estrias_mm=estrias_mm,
            previous_marchamo=previous_marchamo,
            new_marchamo=marchamo,
            created_by=current_user.id,
        )

    db.session.add(row)
    db.session.commit()

    return jsonify({"ok": True, "action": "legacy_update", "tire_state": tire_state})


@yard_bp.post("/api/chassis/<int:chassis_id>/classify")
@login_required
def api_chassis_classify(chassis_id: int):
    site_id = _ensure_active_site()
    ch = Chassis.query.get_or_404(chassis_id)

    if ch.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    data = request.get_json(silent=True) or {}

    # --------
    # 1) Estructura
    # --------
    structure_status = _normalize_structure_status_for_db(
        _norm_enum(data.get("structure_status"))
    )
    twistlocks_status = _normalize_twistlocks_status_for_db(
        _norm_enum(data.get("twistlocks_status"))
    )
    landing_gear_status = _normalize_landing_gear_status_for_db(
        _norm_enum(data.get("landing_gear_status"))
    )
    lights_status = _normalize_lights_status_for_db(
        _norm_enum(data.get("lights_status"))
    )
    mudflap_status = _normalize_mudflap_status_for_db(
        _norm_enum(data.get("mudflap_status"))
    )
    plate_text = (data.get("plate_text") or "").strip()
    comments = (data.get("comments") or "").strip()
    damage_summary = (data.get("damage_summary") or "").strip()

    # --------
    # 2) Llantas
    # --------
    tires = data.get("tires") or []
    axles = int(getattr(ch, "axles", 2) or 2)
    allowed = set(allowed_positions_for(axles))

    tire_lines = []
    any_tire_issue = False
    grouped_tire_readings = {}

    for t in tires:
        pos = (t.get("position_code") or "").strip().upper()
        if pos not in allowed:
            continue

        ingreso_marchamo = (t.get("ingreso_marchamo") or "").strip()
        marchamo_check = (t.get("marchamo_check") or "OK").strip().upper()
        estrias_mm_raw = t.get("estrias_mm")
        is_flat = bool(t.get("is_flat"))

        estrias_mm = None
        if estrias_mm_raw not in (None, "",):
            try:
                estrias_mm = int(estrias_mm_raw)
            except Exception:
                estrias_mm = None

        tire_state = _calc_tire_state_from_data(estrias_mm, is_flat)

        if marchamo_check not in MARCHAMO_CHECK:
            marchamo_check = "OK"
        if tire_state not in TIRE_STATES:
            tire_state = "OK"

        master_pos = _normalize_position_for_tire_master(pos)
        if master_pos:
            grp = grouped_tire_readings.setdefault(master_pos, {
                "states": [],
                "pressures": [],
                "comments": [],
                "seal_issue": False,
                "seal_2": None,
            })

            grp["states"].append(tire_state)

            if marchamo_check != "OK":
                grp["seal_issue"] = True

            detail_parts = [
                f"{pos}",
                f"MARCHAMO={marchamo_check}",
                f"STATE={tire_state}",
            ]
            if ingreso_marchamo:
                detail_parts.append(f"INGRESO={ingreso_marchamo}")
            if estrias_mm not in (None, ""):
                detail_parts.append(f"MM={estrias_mm}")
            if is_flat:
                detail_parts.append("FLAT=SI")

            grp["comments"].append(" | ".join(detail_parts))

        # Actualizar estado configurado de la llanta del chasis
        row = ChassisTire.query.filter_by(chassis_id=ch.id, position_code=pos).first()
        if row:
            row.estrias_mm = estrias_mm
            row.is_flat = is_flat
            row.tire_state = tire_state
            row.updated_at = datetime.utcnow()
            db.session.add(row)

        if marchamo_check != "OK":
            any_tire_issue = True
            if marchamo_check == "DISTINTO":
                tire_lines.append(f"{pos}: MARCHAMO DE INGRESO DISTINTO - REVISAR")
            elif marchamo_check == "NO_TIENE":
                tire_lines.append(f"{pos}: MARCHAMO DE INGRESO NO TIENE - REVISAR")
            elif marchamo_check == "ILEGIBLE":
                tire_lines.append(f"{pos}: MARCHAMO DE INGRESO ILEGIBLE - REVISAR")

        if is_flat:
            any_tire_issue = True
            tire_lines.append(f"{pos}: PINCHADA (DESINFLADA)")
        elif tire_state != "OK":
            any_tire_issue = True
            tire_lines.append(f"{pos}: ESTADO {tire_state} (MM={estrias_mm if estrias_mm is not None else '—'})")

    _save_grouped_tire_readings(
        site_id=site_id,
        chassis_id=ch.id,
        axles=axles,
        grouped_readings=grouped_tire_readings,
        user_id=current_user.id,
        event_type="EIR_IN",
        event_id=None,
    )

    # --------
    # 3) Determinar si requiere taller
    # --------
    structure_lines = []

    if structure_status in {"GOLPE", "DOBLADO", "SOLDADURA"}:
        structure_lines.append(f"Estructura: {structure_status}")

    if twistlocks_status in {"DANADOS"}:
        structure_lines.append(f"Twistlocks: {twistlocks_status}")

    if landing_gear_status in {"DANADAS"}:
        structure_lines.append(f"{LABELS_ES['landing_gear']}: {landing_gear_status}")

    if lights_status in {"IZQ_DANADA", "DER_DANADA"}:
        structure_lines.append(f"Luces: {lights_status}")

    if mudflap_status in {"NO_TRAE"}:
        structure_lines.append(f"{LABELS_ES['mudflap']}: {mudflap_status}")

    if damage_summary:
        structure_lines.append(f"Resumen: {damage_summary}")

    needs_workshop = bool(structure_lines) or bool(any_tire_issue)

    # --------
    # 4) Conciliación contra EIR anterior
    # --------
    last_eir = _fetch_last_final_eir_for_chassis(ch.id)
    eir_prev_id = int(last_eir["id"]) if last_eir and last_eir.get("id") else None

    # --------
    # 5) Guardar inspección
    # --------
    inspection_id = _insert_dynamic("yard_gate_alamo", "chassis_inspections", {
        "site_id": site_id,
        "chassis_id": ch.id,
        "inspected_at": datetime.utcnow(),
        "inspected_by_user_id": current_user.id,
        "structure_status": structure_status or None,
        "twistlocks_status": twistlocks_status or None,
        "landing_gear_status": landing_gear_status or None,
        "lights_status": lights_status or None,
        "mudflap_status": mudflap_status or None,
        "plate_text": plate_text or None,
        "comments": comments or None,
        "needs_workshop": needs_workshop,
        "damage_summary": (damage_summary or None),
    })

    # --------
    # 6) Ingreso automático al predio
    # --------
    ch.site_id = site_id
    ch.is_in_yard = True
    db.session.add(ch)

    inv = ChassisInventory.query.filter_by(site_id=site_id, chassis_id=ch.id).first()
    if not inv:
        inv = ChassisInventory(
            site_id=site_id,
            chassis_id=ch.id,
            chassis_code=ch.chassis_number,
            is_in_yard=True,
        )
    else:
        inv.site_id = site_id
        inv.chassis_code = ch.chassis_number
        inv.is_in_yard = True
    db.session.add(inv)

    # --------
    # 7) Ticket único a taller
    # --------
    ticket_id = None
    if needs_workshop:
        body = _build_workshop_ticket_text(
            chassis_number=ch.chassis_number,
            axles=axles,
            structure_lines=structure_lines,
            tire_lines=tire_lines,
            eir_prev_id=eir_prev_id
        )

        ticket_id = _insert_dynamic("yard_gate_alamo", "workshop_tickets", {
            "site_id": site_id,
            "chassis_id": ch.id,
            "inspection_id": inspection_id,
            "created_at": datetime.utcnow(),
            "created_by_user_id": current_user.id,
            "status": "OPEN",
            "ticket_type": "CHASSIS_DAMAGE",
            "payload_text": body,
            "notes": body,
        })

        audit_log(
            current_user.id,
            "WORKSHOP_TICKET_CREATED_FROM_CHASSIS_CLASSIFICATION",
            "workshop_ticket",
            ticket_id,
            {"site_id": site_id, "chassis_id": ch.id, "eir_prev_id": eir_prev_id},
        )

    audit_log(
        current_user.id,
        "CHASSIS_CLASSIFIED",
        "chassis",
        ch.id,
        {"site_id": site_id, "needs_workshop": needs_workshop, "eir_prev_id": eir_prev_id},
    )

    db.session.commit()
    return jsonify({"ok": True, "needs_workshop": needs_workshop, "ticket_id": ticket_id, "eir_prev_id": eir_prev_id})

# =========================
# Reportes
# =========================

@yard_bp.get("/reportes")
@login_required
def reports_dashboard():
    site_id = _ensure_active_site()
    active_site = Site.query.get(site_id)

    return render_template(
        "yard/reports_dashboard.html",
        active_site=active_site,
    )


@yard_bp.get("/reportes/chasis-fuera")
@login_required
def report_chassis_outside():
    _ensure_active_site()

    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    origin_site_id = (request.args.get("origin_site_id") or "").strip()

    filters = []
    params = {}

    if date_from:
        filters.append("""
            COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date) >= :date_from
        """)
        params["date_from"] = date_from

    if date_to:
        filters.append("""
            COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date) <= :date_to
        """)
        params["date_to"] = date_to

    if origin_site_id:
        try:
            origin_site_id_int = int(origin_site_id)
            filters.append("e.site_id = :origin_site_id")
            params["origin_site_id"] = origin_site_id_int
        except Exception:
            origin_site_id = ""

    extra_where = ""
    if filters:
        extra_where = " AND " + " AND ".join(filters)

    sql = text(f"""
        SELECT
            c.id,
            c.chassis_number,
            c.plate,
            c.length_ft,
            c.axles,
            s.name AS origin_name,
            e.destination,
            COALESCE(e.inventory_out_at, e.finalized_at, e.trip_date) AS departure_at,
            CURRENT_DATE - COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date) AS days_out,
            c.status
        FROM yard_gate_alamo.chassis c
        LEFT JOIN LATERAL (
            SELECT *
            FROM yard_gate_alamo.eirs e
            WHERE e.chassis_id = c.id
              AND e.status = 'CONFIRMED'
            ORDER BY
                e.inventory_out_at DESC NULLS LAST,
                e.finalized_at DESC NULLS LAST,
                e.trip_date DESC,
                e.id DESC
            LIMIT 1
        ) e ON TRUE
        LEFT JOIN yard_gate_alamo.sites s
            ON s.id = e.site_id
        WHERE c.is_in_yard = false
          {extra_where}
        ORDER BY
            days_out DESC NULLS LAST,
            departure_at DESC NULLS LAST,
            c.chassis_number ASC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    total = 0
    for r in rows:
        total += 1

        departure_at = r["departure_at"]
        days_out = r["days_out"]

        if departure_at and hasattr(departure_at, "strftime"):
            departure_at_str = departure_at.strftime("%d/%m/%Y %I:%M %p") if hasattr(departure_at, "hour") else departure_at.strftime("%d/%m/%Y")
        else:
            departure_at_str = ""

        items.append({
            "id": r["id"],
            "chassis_number": r["chassis_number"],
            "plate": r["plate"] or "",
            "length_ft": r["length_ft"] or "",
            "axles": r["axles"] or "",
            "origin_name": r["origin_name"] or "",
            "destination": r["destination"] or "Fuera de patio",
            "departure_at": departure_at,
            "departure_at_str": departure_at_str,
            "days_out": int(days_out) if days_out is not None else 0,
            "status": r["status"] or "",
        })

    sites = (
        Site.query
        .filter(Site.id.in_([2, 3, 4]))
        .order_by(Site.name.asc())
        .all()
    )

    return render_template(
        "yard/report_chassis_outside.html",
        items=items,
        total=total,
        sites=sites,
        date_from=date_from,
        date_to=date_to,
        origin_site_id=origin_site_id,
    )
@yard_bp.get("/reportes/movimientos-contenedor")
@login_required
def report_container_movements():
    site_id = _ensure_active_site()

    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    movement_type = (request.args.get("movement_type") or "").strip().upper()

    valid_types = {"", "GATE_IN", "GATE_OUT"}
    if movement_type not in valid_types:
        movement_type = ""

    filters = ["mv.site_id = :site_id", "mv.movement_type IN ('GATE_IN', 'GATE_OUT')"]
    params = {"site_id": site_id}

    if date_from:
        filters.append("mv.occurred_at::date >= :date_from")
        params["date_from"] = date_from

    if date_to:
        filters.append("mv.occurred_at::date <= :date_to")
        params["date_to"] = date_to

    if movement_type:
        filters.append("mv.movement_type = :movement_type")
        params["movement_type"] = movement_type

    where_sql = " AND ".join(filters)

    sql = text(f"""
        SELECT
            mv.id AS movement_id,
            mv.occurred_at,
            mv.movement_type,
            c.id AS container_id,
            c.code AS container_code,

            -- Chasis para GATE_IN
            ch_in.chassis_number AS chassis_gate_in,

            -- Chasis para GATE_OUT
            ch_out.chassis_number AS chassis_gate_out,

            -- Origen para GATE_IN = destino del último EIR confirmado anterior
            prev_eir.destination AS origin_name,

            -- Destino para GATE_OUT
            eir_out.destination AS destination_name

        FROM yard_gate_alamo.movements mv
        JOIN yard_gate_alamo.containers c
          ON c.id = mv.container_id

        -- Chasis del GATE_IN (vía tire_readings)
        LEFT JOIN LATERAL (
            SELECT tr.chassis_id
            FROM yard_gate_alamo.tire_readings tr
            WHERE tr.event_type = 'GATE_IN'
              AND tr.event_id = mv.id
              AND tr.chassis_id IS NOT NULL
            ORDER BY tr.recorded_at DESC NULLS LAST, tr.id DESC
            LIMIT 1
        ) tr_in ON TRUE
        LEFT JOIN yard_gate_alamo.chassis ch_in
          ON ch_in.id = tr_in.chassis_id

        -- EIR del GATE_OUT actual
        LEFT JOIN yard_gate_alamo.eirs eir_out
          ON eir_out.gate_out_movement_id = mv.id
         AND eir_out.status = 'CONFIRMED'

        LEFT JOIN yard_gate_alamo.chassis ch_out
          ON ch_out.id = eir_out.chassis_id

        -- EIR anterior del mismo contenedor para resolver ORIGEN del GATE_IN
        LEFT JOIN LATERAL (
            SELECT e_prev.destination
            FROM yard_gate_alamo.eirs e_prev
            WHERE e_prev.container_id = mv.container_id
              AND e_prev.status = 'CONFIRMED'
              AND COALESCE(e_prev.inventory_out_at, e_prev.finalized_at, e_prev.created_at) < mv.occurred_at
            ORDER BY
              COALESCE(e_prev.inventory_out_at, e_prev.finalized_at, e_prev.created_at) DESC,
              e_prev.id DESC
            LIMIT 1
        ) prev_eir ON TRUE

        WHERE {where_sql}
        ORDER BY mv.occurred_at DESC, mv.id DESC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    for r in rows:
        movement_type_row = (r["movement_type"] or "").upper()

        chassis_number = ""
        origin_name = ""
        destination_name = ""

        if movement_type_row == "GATE_IN":
            chassis_number = r["chassis_gate_in"] or ""
            origin_name = r["origin_name"] or ""
            destination_name = ""
        elif movement_type_row == "GATE_OUT":
            chassis_number = r["chassis_gate_out"] or ""
            origin_name = ""
            destination_name = r["destination_name"] or ""

        occurred_at = r["occurred_at"]
        occurred_at_str = occurred_at.strftime("%d/%m/%Y %I:%M %p") if occurred_at else ""

        items.append({
            "movement_id": r["movement_id"],
            "container_id": r["container_id"],
            "container_code": r["container_code"] or "",
            "occurred_at": occurred_at,
            "occurred_at_str": occurred_at_str,
            "movement_type": movement_type_row,
            "chassis_number": chassis_number,
            "origin_name": origin_name,
            "destination_name": destination_name,
        })

    return render_template(
        "yard/report_container_movements.html",
        items=items,
        total=len(items),
        date_from=date_from,
        date_to=date_to,
        movement_type=movement_type,
    )

@yard_bp.get("/reportes/movimientos-chasis")
@login_required
def report_chassis_movements():
    site_id = _ensure_active_site()

    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    movement_type = (request.args.get("movement_type") or "").strip().upper()

    valid_types = {"", "GATE_IN", "GATE_OUT"}
    if movement_type not in valid_types:
        movement_type = ""

    filters_in = ["ci.site_id = :site_id"]
    filters_out = ["e.site_id = :site_id", "e.status = 'CONFIRMED'", "e.chassis_id IS NOT NULL"]
    params = {"site_id": site_id}

    if date_from:
        filters_in.append("ci.inspected_at::date >= :date_from")
        filters_out.append("COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date, e.created_at::date) >= :date_from")
        params["date_from"] = date_from

    if date_to:
        filters_in.append("ci.inspected_at::date <= :date_to")
        filters_out.append("COALESCE(e.inventory_out_at::date, e.finalized_at::date, e.trip_date, e.created_at::date) <= :date_to")
        params["date_to"] = date_to

    if movement_type == "GATE_IN":
        enable_in = True
        enable_out = False
    elif movement_type == "GATE_OUT":
        enable_in = False
        enable_out = True
    else:
        enable_in = True
        enable_out = True

    sql_parts = []

    if enable_in:
        sql_parts.append(f"""
            SELECT
                ci.inspected_at AS event_at,
                'GATE_IN' AS movement_type,
                ch.id AS chassis_id,
                ch.chassis_number,
                ch.plate,
                ch.length_ft,
                ch.axles,
                COALESCE(prev_eir.destination, '') AS origin_name,
                '' AS destination_name,
                COALESCE(u.username, '') AS username
            FROM yard_gate_alamo.chassis_inspections ci
            JOIN yard_gate_alamo.chassis ch
              ON ch.id = ci.chassis_id
            LEFT JOIN yard_gate_alamo.users u
              ON u.id = ci.inspected_by_user_id
            LEFT JOIN LATERAL (
                SELECT e_prev.destination
                FROM yard_gate_alamo.eirs e_prev
                WHERE e_prev.chassis_id = ci.chassis_id
                  AND e_prev.status = 'CONFIRMED'
                  AND COALESCE(e_prev.inventory_out_at, e_prev.finalized_at, e_prev.created_at) < ci.inspected_at
                ORDER BY
                  COALESCE(e_prev.inventory_out_at, e_prev.finalized_at, e_prev.created_at) DESC,
                  e_prev.id DESC
                LIMIT 1
            ) prev_eir ON TRUE
            WHERE {" AND ".join(filters_in)}
        """)

    if enable_out:
        sql_parts.append(f"""
            SELECT
                COALESCE(e.inventory_out_at, e.finalized_at, e.created_at) AS event_at,
                'GATE_OUT' AS movement_type,
                ch.id AS chassis_id,
                ch.chassis_number,
                ch.plate,
                ch.length_ft,
                ch.axles,
                COALESCE(s.name, '') AS origin_name,
                COALESCE(e.destination, '') AS destination_name,
                COALESCE(u.username, '') AS username
            FROM yard_gate_alamo.eirs e
            JOIN yard_gate_alamo.chassis ch
              ON ch.id = e.chassis_id
            LEFT JOIN yard_gate_alamo.sites s
              ON s.id = e.site_id
            LEFT JOIN yard_gate_alamo.users u
              ON u.id = COALESCE(e.last_edited_by_user_id, e.created_by_user_id)
            WHERE {" AND ".join(filters_out)}
        """)

    sql = text(f"""
        SELECT *
        FROM (
            {" UNION ALL ".join(sql_parts)}
        ) x
        ORDER BY x.event_at DESC, x.chassis_number ASC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    for r in rows:
        event_at = r["event_at"]
        event_at_str = event_at.strftime("%d/%m/%Y %I:%M %p") if event_at else ""

        items.append({
            "event_at": event_at,
            "event_at_str": event_at_str,
            "movement_type": r["movement_type"] or "",
            "chassis_id": r["chassis_id"],
            "chassis_number": r["chassis_number"] or "",
            "plate": r["plate"] or "",
            "length_ft": r["length_ft"] or "",
            "axles": r["axles"] or "",
            "origin_name": r["origin_name"] or "",
            "destination_name": r["destination_name"] or "",
            "username": r["username"] or "",
        })

    return render_template(
        "yard/report_chassis_movements.html",
        items=items,
        total=len(items),
        date_from=date_from,
        date_to=date_to,
        movement_type=movement_type,
    )

def _translate_tire_position(position_code: str | None) -> str:
    value = (position_code or "").strip().upper()

    mapping = {
        "AX1_L_IN": "Eje 1 izq int",
        "AX1_L_OUT": "Eje 1 izq ext",
        "AX1_R_IN": "Eje 1 der int",
        "AX1_R_OUT": "Eje 1 der ext",
        "AX2_L_IN": "Eje 2 izq int",
        "AX2_L_OUT": "Eje 2 izq ext",
        "AX2_R_IN": "Eje 2 der int",
        "AX2_R_OUT": "Eje 2 der ext",
        "AX3_L_IN": "Eje 3 izq int",
        "AX3_L_OUT": "Eje 3 izq ext",
        "AX3_R_IN": "Eje 3 der int",
        "AX3_R_OUT": "Eje 3 der ext",
    }
    return mapping.get(value, value)

def _calc_tire_state_from_mm(estrias_mm, is_flat=False):
    if is_flat:
        return "PINCHADA"

    if estrias_mm in (None, "",):
        return "OK"

    try:
        mm = int(estrias_mm)
    except Exception:
        return "OK"

    if 9 <= mm <= 12:
        return "OK"
    if 4 <= mm <= 8:
        return "GASTADA"
    if 1 <= mm <= 3:
        return "NO_APTA"

    return "OK"


def _normalize_tire_status(value: str | None) -> str:
    v = (value or "").strip().upper()
    allowed = {"ASIGNADA", "EN_TALLER_BODEGA", "RECAUCHE", "DESECHADA"}
    return v if v in allowed else "EN_TALLER_BODEGA"

@yard_bp.get("/llantas")
@login_required
def tires_list():
    _ensure_active_site()

    q = (request.args.get("q") or "").strip()
    color = (request.args.get("color") or "").strip().upper()
    mounted = (request.args.get("mounted") or "").strip().upper()

    filters = []
    params = {}

    if q:
        filters.append("""
            (
                t.tire_number ILIKE :q
                OR COALESCE(t.brand, '') ILIKE :q
                OR COALESCE(ct.marchamo, '') ILIKE :q
                OR COALESCE(ch.chassis_number, '') ILIKE :q
            )
        """)
        params["q"] = f"%{q}%"

    if color == "VERDE":
        filters.append("ct.is_flat = FALSE AND ct.estrias_mm BETWEEN 9 AND 12")
    elif color == "AMARILLO":
        filters.append("ct.is_flat = FALSE AND ct.estrias_mm BETWEEN 4 AND 8")
    elif color == "ROJO":
        filters.append("(ct.is_flat = TRUE OR ct.estrias_mm <= 3)")

    if mounted == "SI":
        filters.append("ct.id IS NOT NULL")
    elif mounted == "NO":
        filters.append("ct.id IS NULL")

    where_sql = ""
    if filters:
        where_sql = "WHERE " + " AND ".join(filters)

    sql = text(f"""
        SELECT
            t.id AS tire_id,
            t.tire_number,
            t.brand,
            t.model,
            t.size,
            t.notes,
            t.status,
            ct.id AS chassis_tire_id,
            ct.marchamo,
            ct.estrias_mm,
            ct.is_flat,
            ct.tire_state,
            ch.id AS chassis_id,
            ch.chassis_number,
            ct.position_code,
            CASE
                WHEN ct.is_flat = TRUE THEN 'ROJO'
                WHEN ct.estrias_mm IS NULL THEN ''
                WHEN ct.estrias_mm <= 3 THEN 'ROJO'
                WHEN ct.estrias_mm BETWEEN 4 AND 8 THEN 'AMARILLO'
                WHEN ct.estrias_mm BETWEEN 9 AND 12 THEN 'VERDE'
                ELSE ''
            END AS estado_color
        FROM yard_gate_alamo.tires t
        LEFT JOIN yard_gate_alamo.chassis_tires ct
        ON ct.tire_id = t.id
        LEFT JOIN yard_gate_alamo.chassis ch
        ON ch.id = ct.chassis_id
        {where_sql}
        ORDER BY t.tire_number ASC
    """)

    rows = db.session.execute(sql, params).mappings().all()

    items = []
    for r in rows:
        items.append({
            "tire_id": r["tire_id"],
            "tire_number": r["tire_number"] or "",
            "brand": r["brand"] or "",
            "model": r["model"] or "",
            "size": r["size"] or "",
            "notes": r["notes"] or "",
            "marchamo": r["marchamo"] or "",
            "estrias_mm": r["estrias_mm"],
            "is_flat": bool(r["is_flat"]) if r["is_flat"] is not None else False,
            "tire_state": r["tire_state"] or "",
            "chassis_id": r["chassis_id"],
            "chassis_number": r["chassis_number"] or "",
            "position_code": r["position_code"] or "",
            "position_label": _translate_tire_position(r["position_code"]),
            "estado_color": r["estado_color"] or "",
            "is_mounted": bool(r["chassis_tire_id"]),
            "status": r["status"] or "EN_TALLER_BODEGA",
        })

    return render_template(
        "yard/tires_list.html",
        items=items,
        total=len(items),
        q=q,
        color=color,
        mounted=mounted,
    )

@yard_bp.get("/llantas/nueva")
@login_required
def tire_create_view():
    return render_template(
        "yard/tire_form.html",
        mode="create",
        tire=None,
    )


@yard_bp.post("/llantas/nueva")
@login_required
def tire_create_post():
    tire_number = (request.form.get("tire_number") or "").strip().upper()
    brand = (request.form.get("brand") or "").strip().upper()
    model = (request.form.get("model") or "").strip().upper()
    size = (request.form.get("size") or "").strip().upper()
    notes = (request.form.get("notes") or "").strip()
    status = _normalize_tire_status(request.form.get("status"))

    if not tire_number:
        flash("Debes ingresar el número de llanta.", "danger")
        return redirect(url_for("yard.tire_create_view"))

    existing = Tire.query.filter_by(tire_number=tire_number).first()
    if existing:
        flash("Ya existe una llanta con ese número.", "danger")
        return redirect(url_for("yard.tire_create_view"))

    tire = Tire(
        tire_number=tire_number,
        brand=brand or None,
        model=model or None,
        size=size or None,
        notes=notes or None,
        status=status,
    )
    db.session.add(tire)
    db.session.commit()

    flash(f"Llanta {tire_number} creada correctamente.", "success")
    return redirect(url_for("yard.tires_list"))

@yard_bp.get("/llantas/<int:tire_id>")
@login_required
def tire_detail_view(tire_id: int):
    tire = Tire.query.get_or_404(tire_id)

    row = db.session.execute(text("""
        SELECT
            t.id AS tire_id,
            t.tire_number,
            t.brand,
            t.model,
            t.size,
            t.notes,
            t.status,
            ct.id AS chassis_tire_id,
            ct.marchamo,
            ct.estrias_mm,
            ct.is_flat,
            ct.tire_state,
            ch.id AS chassis_id,
            ch.chassis_number,
            ct.position_code
        FROM yard_gate_alamo.tires t
        LEFT JOIN yard_gate_alamo.chassis_tires ct
          ON ct.tire_id = t.id
        LEFT JOIN yard_gate_alamo.chassis ch
          ON ch.id = ct.chassis_id
        WHERE t.id = :tire_id
        LIMIT 1
    """), {"tire_id": tire_id}).mappings().first()

    item = None
    if row:
        item = {
            "tire_id": row["tire_id"],
            "tire_number": row["tire_number"] or "",
            "brand": row["brand"] or "",
            "model": row["model"] or "",
            "size": row["size"] or "",
            "notes": row["notes"] or "",
            "marchamo": row["marchamo"] or "",
            "estrias_mm": row["estrias_mm"],
            "is_flat": bool(row["is_flat"]) if row["is_flat"] is not None else False,
            "tire_state": row["tire_state"] or "",
            "chassis_id": row["chassis_id"],
            "chassis_number": row["chassis_number"] or "",
            "position_code": row["position_code"] or "",
            "position_label": _translate_tire_position(row["position_code"]),
            "status": row["status"] or "EN_TALLER_BODEGA",
        }

    return render_template(
        "yard/tire_form.html",
        mode="edit",
        tire=item,
    )


@yard_bp.post("/llantas/<int:tire_id>/editar")
@login_required
def tire_edit_post(tire_id: int):
    tire = Tire.query.get_or_404(tire_id)

    tire_number = (request.form.get("tire_number") or "").strip().upper()
    brand = (request.form.get("brand") or "").strip().upper()
    model = (request.form.get("model") or "").strip().upper()
    size = (request.form.get("size") or "").strip().upper()
    notes = (request.form.get("notes") or "").strip()

    status = _normalize_tire_status(request.form.get("status"))

    marchamo = (request.form.get("marchamo") or "").strip()
    chassis_number = (request.form.get("chassis_number") or "").strip().upper()
    position_code = (request.form.get("position_code") or "").strip().upper()
    confirm_replace = (request.form.get("confirm_replace") or "").strip() == "1"

    estrias_mm_raw = (request.form.get("estrias_mm") or "").strip()
    is_flat = (request.form.get("is_flat") or "").strip() == "1"

    if not tire_number:
        flash("Debes ingresar el número de llanta.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    duplicate = Tire.query.filter(Tire.tire_number == tire_number, Tire.id != tire.id).first()
    if duplicate:
        flash("Ya existe otra llanta con ese número.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    estrias_mm = None
    if estrias_mm_raw:
        try:
            estrias_mm = int(estrias_mm_raw)
        except Exception:
            flash("Las estrías deben ser numéricas.", "danger")
            return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

        if estrias_mm < 1 or estrias_mm > 12:
            flash("Las estrías deben estar entre 1 y 12.", "danger")
            return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    # =========================
    # Actualizar datos base
    # =========================
    tire.tire_number = tire_number
    tire.brand = brand or None
    tire.model = model or None
    tire.size = size or None
    tire.notes = notes or None

    # Montaje actual de esta llanta, si existe
    current_mount = ChassisTire.query.filter_by(tire_id=tire.id).first()

    # =========================
    # Caso 1: NO se quiere montar en chasis
    # =========================
    if not chassis_number and not position_code:
        if current_mount:
            db.session.delete(current_mount)

        tire.status = status if status in {"EN_TALLER_BODEGA", "RECAUCHE", "DESECHADA"} else "EN_TALLER_BODEGA"

        db.session.add(tire)
        db.session.commit()

        flash(f"Llanta {tire_number} actualizada correctamente.", "success")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    # =========================
    # Caso 2: se quiere montar en chasis
    # =========================
    if not chassis_number:
        flash("Debes indicar el número de chasis.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    if not position_code:
        flash("Debes indicar la posición de la llanta.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    target_chassis = Chassis.query.filter_by(chassis_number=chassis_number).first()
    if not target_chassis:
        flash("El número de chasis no existe.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    allowed = set(allowed_positions_for(int(target_chassis.axles or 2)))
    if position_code not in allowed:
        flash("La posición no es válida para ese chasis.", "danger")
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    # Regla: solo se puede montar si está en taller/bodega
    if tire.status != "EN_TALLER_BODEGA":
        flash(
            f"La llanta {tire.tire_number} no puede montarse porque su estado actual es {tire.status}. "
            f"Solo se pueden montar llantas en EN_TALLER_BODEGA.",
            "danger"
        )
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    # Si esta llanta ya está asignada a otro chasis, bloquear
    if current_mount and (
        current_mount.chassis_id != target_chassis.id
        or current_mount.position_code != position_code
    ):
        flash(
            f"La llanta {tire.tire_number} ya está asignada al chasis "
            f"{current_mount.chassis.chassis_number if current_mount.chassis else current_mount.chassis_id} "
            f"en la posición {_translate_tire_position(current_mount.position_code)}.",
            "danger"
        )
        return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

    # Verificar si la posición destino está ocupada por otra llanta
    occupied = ChassisTire.query.filter_by(
        chassis_id=target_chassis.id,
        position_code=position_code
    ).first()

    if occupied and occupied.tire_id != tire.id:
        existing_tire = Tire.query.get(occupied.tire_id) if occupied.tire_id else None
        existing_tire_number = existing_tire.tire_number if existing_tire else "SIN NÚMERO"
        translated_pos = _translate_tire_position(position_code)

        if not confirm_replace:
            flash(
                f"La posición {translated_pos} del chasis {target_chassis.chassis_number} "
                f"ya está ocupada por la llanta {existing_tire_number}. "
                f"Si deseas reemplazarla, confirma la operación.",
                "warning"
            )
            return redirect(
                url_for(
                    "yard.tire_detail_view",
                    tire_id=tire.id,
                )
            )

        # Reemplazo confirmado: la vieja pasa a EN_TALLER_BODEGA
        if existing_tire:
            existing_tire.status = "EN_TALLER_BODEGA"
            db.session.add(existing_tire)

        db.session.delete(occupied)
        db.session.flush()

    # Recalcular estado técnico de llanta montada
    tire_state = _calc_tire_state_from_mm(estrias_mm, is_flat)

    # Crear o actualizar montaje
    mount_row = current_mount
    if not mount_row:
        mount_row = ChassisTire(
            chassis_id=target_chassis.id,
            position_code=position_code,
            tire_id=tire.id,
            installed_at=datetime.utcnow(),
        )

    mount_row.chassis_id = target_chassis.id
    mount_row.position_code = position_code
    mount_row.tire_id = tire.id
    mount_row.marchamo = marchamo or None
    mount_row.estrias_mm = estrias_mm
    mount_row.is_flat = is_flat
    mount_row.tire_state = tire_state
    mount_row.updated_at = datetime.utcnow()

    db.session.add(mount_row)

    tire.status = "ASIGNADA"
    db.session.add(tire)

    db.session.commit()

    flash(
        f"Llanta {tire.tire_number} asignada al chasis {target_chassis.chassis_number} "
        f"en la posición {_translate_tire_position(position_code)}.",
        "success"
    )
    return redirect(url_for("yard.tire_detail_view", tire_id=tire.id))

@yard_bp.get("/llantas/import")
@login_required
def tires_import_view():
    return render_template("yard/tires_import.html")

@yard_bp.get("/llantas/export")
@login_required
def tires_export():
    _ensure_active_site()

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.tires_list"))

    rows = (
        Tire.query
        .order_by(Tire.tire_number.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Llantas"

    headers = [
        "id (no borrar si vas a actualizar)",
        "tire_number (número de llanta)",
        "brand (marca) [opcional]",
        "model (modelo) [opcional]",
        "size (tamaño) [opcional]",
        "status (ASIGNADA/EN_TALLER_BODEGA/RECAUCHE/DESECHADA)",
        "notes (notas) [opcional]",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for t in rows:
        ws.append([
            t.id,
            t.tire_number or "",
            t.brand or "",
            t.model or "",
            t.size or "",
            t.status or "EN_TALLER_BODEGA",
            t.notes or "",
        ])

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 40

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="llantas_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@yard_bp.post("/llantas/import")
@login_required
def tires_import_post():
    _ensure_active_site()

    f = request.files.get("file")
    if not f or not f.filename:
        flash("Sube un archivo Excel.", "danger")
        return redirect(url_for("yard.tires_import_view"))

    try:
        from openpyxl import load_workbook
    except Exception:
        flash("Falta openpyxl en requirements.txt", "danger")
        return redirect(url_for("yard.tires_import_view"))

    wb = load_workbook(f, data_only=True)
    ws = wb.active

    imported = 0
    updated = 0
    errors = []

    staged = []
    ids_to_find = []
    tire_numbers_to_find = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tire_id_raw = row[0] if len(row) > 0 else None
        tire_number = (str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else "")
        brand = (str(row[2]).strip().upper() if len(row) > 2 and row[2] is not None else "")
        model = (str(row[3]).strip().upper() if len(row) > 3 and row[3] is not None else "")
        size = (str(row[4]).strip().upper() if len(row) > 4 and row[4] is not None else "")
        status = _normalize_tire_status(row[5] if len(row) > 5 else None)
        notes = (str(row[6]).strip() if len(row) > 6 and row[6] is not None else "")

        # Saltar filas totalmente vacías
        if not any([tire_id_raw, tire_number, brand, model, size, status, notes]):
            continue

        if not tire_number:
            errors.append(f"Fila {idx}: falta tire_number.")
            continue

        tire_id = None
        if tire_id_raw not in (None, ""):
            try:
                tire_id = int(tire_id_raw)
            except Exception:
                errors.append(f"Fila {idx}: id inválido ({tire_id_raw}).")
                continue

        staged.append({
            "idx": idx,
            "id": tire_id,
            "tire_number": tire_number,
            "brand": brand or None,
            "model": model or None,
            "size": size or None,
            "status": status,
            "notes": notes or None,
        })

        if tire_id:
            ids_to_find.append(tire_id)
        tire_numbers_to_find.append(tire_number)

    if not staged:
        flash(f"No se importó nada. Errores: {len(errors)}", "danger")
        session["tires_import_errors"] = errors[:200]
        return redirect(url_for("yard.tires_import_view"))

    existing_by_id = {}
    if ids_to_find:
        rows_by_id = Tire.query.filter(Tire.id.in_(ids_to_find)).all()
        existing_by_id = {t.id: t for t in rows_by_id}

    existing_by_number = {}
    if tire_numbers_to_find:
        rows_by_number = Tire.query.filter(Tire.tire_number.in_(tire_numbers_to_find)).all()
        existing_by_number = {t.tire_number: t for t in rows_by_number}

    used_numbers_in_file = set()

    for item in staged:
        tire_id = item["id"]
        tire_number = item["tire_number"]

        if tire_number in used_numbers_in_file:
            errors.append(f"Fila {item['idx']}: tire_number repetido en el mismo archivo ({tire_number}).")
            continue
        used_numbers_in_file.add(tire_number)

        tire = None

        if tire_id and tire_id in existing_by_id:
            tire = existing_by_id[tire_id]
        elif tire_number in existing_by_number:
            tire = existing_by_number[tire_number]

        if tire:
            tire.tire_number = tire_number
            tire.brand = item["brand"]
            tire.model = item["model"]
            tire.size = item["size"]
            tire.status = item["status"]
            tire.notes = item["notes"]
            db.session.add(tire)
            updated += 1
        else:
            tire = Tire(
                tire_number=tire_number,
                brand=item["brand"],
                model=item["model"],
                size=item["size"],
                status=item["status"],
                notes=item["notes"],
            )
            db.session.add(tire)
            imported += 1

    db.session.commit()

    if errors:
        flash(f"Importado: {imported} | Actualizado: {updated} | Errores: {len(errors)}", "warning")
        session["tires_import_errors"] = errors[:200]
    else:
        flash(f"Importado: {imported} | Actualizado: {updated}", "success")
        session.pop("tires_import_errors", None)

    return redirect(url_for("yard.tires_list"))

@yard_bp.get("/llantas/recauche")
@login_required
def tire_retread_report_view():
    _ensure_active_site()
    return render_template("yard/tire_retread_report.html")

@yard_bp.get("/api/llantas/recauche-report")
@login_required
def tire_retread_report():
    _ensure_active_site()

    sql = text("""
        SELECT
            e.id,
            e.tire_id,
            e.previous_estrias_mm,
            e.new_estrias_mm,
            e.previous_marchamo,
            e.new_marchamo,
            e.created_by,
            e.created_at,
            t.tire_number,
            u.username
        FROM yard_gate_alamo.tire_retread_events e
        LEFT JOIN yard_gate_alamo.tires t
          ON t.id = e.tire_id
        LEFT JOIN yard_gate_alamo.users u
          ON u.id = e.created_by
        ORDER BY e.created_at DESC, e.id DESC
    """)

    rows = db.session.execute(sql).mappings().all()

    items = []
    for r in rows:
        created_at = r["created_at"]
        items.append({
            "id": r["id"],
            "fecha": created_at.strftime("%d/%m/%Y %I:%M %p") if created_at else "",
            "tire_id": r["tire_id"],
            "tire_number": r["tire_number"] or "",
            "estrias_antes": r["previous_estrias_mm"],
            "estrias_despues": r["new_estrias_mm"],
            "marchamo_anterior": r["previous_marchamo"] or "",
            "marchamo_nuevo": r["new_marchamo"] or "",
            "usuario": r["username"] or "",
        })

    return jsonify({"ok": True, "items": items})

