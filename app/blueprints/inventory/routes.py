# app/blueprints/inventory/routes.py

import os
from io import BytesIO
from flask_login import login_required, current_user

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from sqlalchemy import text, bindparam

from app.extensions import db
from app.blueprints.inventory import inventory_bp
from app.models.container import Container, ContainerPosition
from app.models.yard import YardBay
from app.models.movement import Movement, MovementPhoto
from app.models.site import Site, UserSite
from flask import render_template, request, send_file, session, abort, redirect, url_for, flash
from datetime import datetime, date
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from app.models.container_classification import ContainerClassification
from app.services.audit import audit_log
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


# =========================================================
# Utilidades
# =========================================================

def _normalize_public_url(raw: str) -> str | None:
    if not raw:
        return None

    raw = str(raw).strip()

    if not (raw.startswith("http://") or raw.startswith("https://")):
        return None

    public_base = (os.environ.get("R2_PUBLIC_BASE_URL") or os.environ.get("PUBLIC_BASE_URL") or "").rstrip("/")
    if not public_base:
        return raw

    bucket = os.environ.get("R2_BUCKET") or os.environ.get("S3_BUCKET") or ""
    if bucket and f"/{bucket}/" in raw:
        key = raw.split(f"/{bucket}/", 1)[1]
        return f"{public_base}/{key}"

    return raw


# =========================================================
# Multi-predio helpers
# =========================================================

def _allowed_sites_for_user(user):
    if getattr(user, "role", None) == "admin":
        return Site.query.filter_by(is_active=True).order_by(Site.name.asc()).all()

    return (
        db.session.query(Site)
        .join(UserSite, UserSite.site_id == Site.id)
        .filter(UserSite.user_id == user.id, Site.is_active == True)  # noqa: E712
        .order_by(Site.name.asc())
        .all()
    )


def _get_active_site_id():
    return session.get("active_site_id")


def _set_active_site_id(site_id: int):
    session["active_site_id"] = int(site_id)


def _ensure_active_site():
    allowed = _allowed_sites_for_user(current_user)

    if not allowed:
        abort(403)

    active_id = _get_active_site_id()
    allowed_ids = {s.id for s in allowed}

    if not active_id or active_id not in allowed_ids:
        _set_active_site_id(allowed[0].id)
        active_id = allowed[0].id

    return active_id


# =========================================================
# Query principal de inventario
# =========================================================

def _inventory_query(
    site_id: int,
    in_yard: str | None,
    qtext: str,
    shipping_line: str = "",
    origin: str = "",
    size: str = "",
):
    in_yard = (in_yard or "1").strip()
    qtext = (qtext or "").strip().upper()
    shipping_line = (shipping_line or "").strip().upper()
    origin = (origin or "").strip().upper()
    size = (size or "").strip().upper()

    q = (
        db.session.query(Container, ContainerPosition, YardBay)
        .outerjoin(ContainerPosition, ContainerPosition.container_id == Container.id)
        .outerjoin(YardBay, YardBay.id == ContainerPosition.bay_id)
        .filter(Container.site_id == site_id)
    )

    if in_yard == "1":
        q = q.filter(Container.is_in_yard == True)  # noqa: E712
    elif in_yard == "0":
        q = q.filter(Container.is_in_yard == False)  # noqa: E712

    if qtext:
        q = q.filter(db.func.upper(Container.code).like(f"%{qtext}%"))

    if origin:
        q = q.filter(
            db.func.upper(db.func.coalesce(Container.gate_in_origin_port, "")) == origin
        )

    if size:
        q = q.filter(db.func.upper(Container.size) == size)

    if shipping_line:
        q = (
            q.join(
                ContainerClassification,
                ContainerClassification.container_id == Container.id,
            )
            .filter(
                ContainerClassification.site_id == site_id,
                db.func.upper(ContainerClassification.shipping_line) == shipping_line,
            )
        )

    return q.order_by(Container.updated_at.desc())


# =========================================================
# Clasificación más reciente por contenedor
# =========================================================

def _last_classification_by_container_ids(container_ids: list[int]) -> dict[int, dict]:

    if not container_ids:
        return {}

    sql = (
        text(
            """
            SELECT DISTINCT ON (container_id)
                container_id,
                shipping_line,
                max_gross_kg,
                manufacture_year,
                summary_text,
                final_classification,
                classified_at
            FROM yard_gate_alamo.container_classifications
            WHERE container_id IN :ids
            ORDER BY container_id, classified_at DESC, id DESC
            """
        )
        .bindparams(bindparam("ids", expanding=True))
    )

    rows = db.session.execute(sql, {"ids": container_ids}).mappings().all()

    return {int(r["container_id"]): dict(r) for r in rows}


# =========================================================
# Último EIR por contenedor
# =========================================================

def _last_gate_in_by_container_ids(container_ids: list[int]) -> dict[int, dict]:

    if not container_ids:
        return {}

    sql = (
        text(
            """
            SELECT DISTINCT ON (container_id)
                container_id,
                occurred_at AS gate_in_at
            FROM yard_gate_alamo.movements
            WHERE container_id IN :ids
              AND movement_type = 'GATE_IN'
              AND occurred_at IS NOT NULL
            ORDER BY container_id, occurred_at DESC, id DESC
            """
        )
        .bindparams(bindparam("ids", expanding=True))
    )

    rows = db.session.execute(sql, {"ids": container_ids}).mappings().all()

    return {int(r["container_id"]): dict(r) for r in rows}


# =========================================================
# Pantalla principal inventario
# =========================================================

@inventory_bp.get("/inventory")
@login_required
def inventory_index():
    site_id = _ensure_active_site()

    in_yard = (request.args.get("in_yard") or "1").strip()
    qtext = (request.args.get("q") or "").strip().upper()
    shipping_line = (request.args.get("shipping_line") or "").strip().upper()
    origin = (request.args.get("origin") or "").strip().upper()
    size = (request.args.get("size") or "").strip().upper()
    classification = (request.args.get("classification") or "").strip().upper()
    dispatch_status = (request.args.get("dispatch_status") or "").strip().upper()

    rows = _inventory_query(
        site_id,
        in_yard,
        qtext,
        shipping_line,
        origin,
        size,
    ).all()

    container_ids = [c.id for c, _, _ in rows]

    cls_by_container = _last_classification_by_container_ids(container_ids)
    gate_in_by_container = _last_gate_in_by_container_ids(container_ids)

    items = []

    for c, pos, bay in rows:
        cls = cls_by_container.get(c.id)
        gate_in = gate_in_by_container.get(c.id)

        current_classification = ((cls.get("final_classification") if cls else "") or "").strip().upper()
        current_status = (c.dispatch_status or "NORMAL").strip().upper()

        if classification and current_classification != classification:
            continue

        if dispatch_status and current_status != dispatch_status:
            continue

        gate_in_at = gate_in.get("gate_in_at") if gate_in else None
        days_in_yard = None

        if gate_in_at:
            days_in_yard = (datetime.utcnow().date() - gate_in_at.date()).days

        items.append({
            "id": c.id,
            "code": c.code,
            "gate_in_origin_port": c.gate_in_origin_port,
            "size": c.size,
            "year": (cls.get("manufacture_year") if cls else c.year),
            "shipping_line": (cls.get("shipping_line") if cls else ""),
            "max_gross_kg": (cls.get("max_gross_kg") if cls else ""),
            "classification": current_classification,
            "gate_in_at": gate_in_at,
            "days_in_yard": days_in_yard,
            "is_in_yard": bool(c.is_in_yard),
            "evacuation_destination": c.evacuation_destination,
            "evacuation_type": c.evacuation_type,
            "evacuation_notes": c.evacuation_notes,
            "dispatch_status": current_status,
            "status_notes": (cls.get("summary_text") if cls else (c.status_notes or "")),
            "position": None if not pos else {
                "bay_code": bay.code if bay else None,
                "depth_row": pos.depth_row,
                "tier": pos.tier,
            },
        })

    shipping_lines = [
        r[0]
        for r in (
            db.session.query(ContainerClassification.shipping_line)
            .filter(
                ContainerClassification.site_id == site_id,
                ContainerClassification.shipping_line.isnot(None),
                ContainerClassification.shipping_line != "",
            )
            .distinct()
            .order_by(ContainerClassification.shipping_line)
            .all()
        )
    ]

    origin_options = ["CALDERA", "LIMON"]

    size_options = [
        r[0]
        for r in (
            db.session.query(Container.size)
            .filter(
                Container.site_id == site_id,
                Container.size.isnot(None),
                Container.size != "",
            )
            .distinct()
            .order_by(Container.size)
            .all()
        )
    ]

    classification_options = [
        "A+",
        "A-",
        "B+",
        "B-",
        "C",
        "A2",
        "B2",
        "CHATARRA",
    ]

    status_options = [
        "NORMAL",
        "NO_USAR",
        "PARA_DESPACHO",
        "PARA_EVACUAR",
        "EVACUAR_SOLICITADO",
        "DESPACHO_MONTADO",
        "EVACUACION_MONTADA",
    ]

    return render_template(
        "inventory/index.html",
        items=items,
        in_yard=in_yard,
        shipping_lines=shipping_lines,
        shipping_line=shipping_line,
        origin_options=origin_options,
        size_options=size_options,
        classification_options=classification_options,
        status_options=status_options,
        origin=origin,
        size=size,
        classification=classification,
        dispatch_status=dispatch_status,
        q=qtext,
    )


# =========================================================
# Exportar inventario
# =========================================================

@inventory_bp.get("/inventory/export")
@login_required
def inventory_export():
    site_id = _ensure_active_site()

    in_yard = (request.args.get("in_yard") or "1").strip()
    qtext = (request.args.get("q") or "").strip().upper()
    shipping_line = (request.args.get("shipping_line") or "").strip().upper()
    origin = (request.args.get("origin") or "").strip().upper()
    size = (request.args.get("size") or "").strip().upper()
    classification_filter = (request.args.get("classification") or "").strip().upper()
    dispatch_status_filter = (request.args.get("dispatch_status") or "").strip().upper()

    rows = _inventory_query(
        site_id,
        in_yard,
        qtext,
        shipping_line,
        origin,
        size,
    ).all()

    container_ids = [c.id for c, _, _ in rows]

    cls_by_container = _last_classification_by_container_ids(container_ids)
    gate_in_by_container = _last_gate_in_by_container_ids(container_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = [
        "ID",
        "CONTENEDOR",
        "ORIGEN",
        "TAMAÑO",
        "NAVIERA",
        "AÑO",
        "MAX_GROSS_KG",
        "CLASIFICACION",
        "FECHA_INGRESO",
        "DIAS_EN_PREDIO",
        "ESTADO",
        "ESTIBA",
        "FILA",
        "NIVEL",
        "NOTAS",
    ]

    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for c, pos, bay in rows:
        cls = cls_by_container.get(c.id)
        gate_in = gate_in_by_container.get(c.id)
        gate_in_at = gate_in.get("gate_in_at") if gate_in else None

        classification = ((cls.get("final_classification") if cls else "") or "").strip().upper()
        dispatch_status = (c.dispatch_status or "NORMAL").strip().upper()

        if classification_filter and classification != classification_filter:
            continue

        if dispatch_status_filter and dispatch_status != dispatch_status_filter:
            continue

        gate_in_date_str = gate_in_at.strftime("%Y-%m-%d") if gate_in_at else ""

        days_in_yard = ""
        if gate_in_at:
            days_in_yard = (datetime.utcnow().date() - gate_in_at.date()).days

        naviera = (cls.get("shipping_line") if cls else "") or ""
        year = (cls.get("manufacture_year") if cls else c.year) or ""
        max_gross = (cls.get("max_gross_kg") if cls else "") or ""
        notes = (cls.get("summary_text") if cls else (c.status_notes or "")) or ""

        ws.append([
            c.id,
            c.code or "",
            c.gate_in_origin_port or "",
            c.size or "",
            naviera,
            year,
            max_gross,
            classification,
            gate_in_date_str,
            days_in_yard,
            dispatch_status,
            (bay.code if (pos and bay) else "") or "",
            (pos.depth_row if pos else "") or "",
            (pos.tier if pos else "") or "",
            notes,
        ])

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for cell in ws[col_letter]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)

        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    tag = "ALL"

    if in_yard == "1":
        tag = "EN_PATIO"
    elif in_yard == "0":
        tag = "FUERA_PATIO"

    fname = f"inventario_{tag}.xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# Detalle contenedor
# =========================================================

@inventory_bp.get("/inventory/<int:container_id>")
@login_required
def inventory_detail(container_id: int):

    site_id = _ensure_active_site()

    c = Container.query.get_or_404(container_id)

    if c.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    pos = (
        db.session.query(ContainerPosition, YardBay)
        .outerjoin(YardBay, YardBay.id == ContainerPosition.bay_id)
        .filter(ContainerPosition.container_id == c.id)
        .first()
    )

    current_pos = None

    if pos:
        p, bay = pos

        current_pos = {
            "bay_code": bay.code if bay else None,
            "depth_row": p.depth_row,
            "tier": p.tier,
        }

    movements = (
        Movement.query
        .filter(Movement.container_id == c.id, Movement.site_id == site_id)
        .order_by(Movement.occurred_at.desc())
        .all()
    )

    mv_ids = [m.id for m in movements]

    photos_by_mv: dict[int, list[dict]] = {mid: [] for mid in mv_ids}

    if mv_ids:

        photos = (
            MovementPhoto.query
            .filter(MovementPhoto.movement_id.in_(mv_ids))
            .order_by(MovementPhoto.uploaded_at.asc())
            .all()
        )

        for ph in photos:

            if (ph.photo_type or "").upper() == "UPLOAD_ERROR":
                continue

            url_ok = _normalize_public_url(ph.url)

            if not url_ok:
                continue

            photos_by_mv.setdefault(ph.movement_id, []).append(
                {
                    "photo_type": ph.photo_type,
                    "url": url_ok,
                    "uploaded_at": ph.uploaded_at,
                }
            )

    return render_template(
        "inventory/detail.html",
        c=c,
        current_pos=current_pos,
        movements=movements,
        photos_by_mv=photos_by_mv,
    )

@inventory_bp.post("/inventory/<int:container_id>/mark-evacuation")
@login_required
def mark_container_evacuation(container_id: int):
    site_id = _ensure_active_site()

    c = Container.query.get_or_404(container_id)

    if c.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    if not c.is_in_yard:
        flash("Solo se pueden marcar contenedores que están en patio.", "warning")
        return redirect(url_for("inventory.inventory_index"))

    if (c.dispatch_status or "NORMAL") != "NORMAL":
        flash("Solo se pueden marcar como evacuar contenedores disponibles.", "warning")
        return redirect(url_for("inventory.inventory_index"))

    destination = (request.form.get("evacuation_destination") or "").strip().upper()
    custom_destination = (request.form.get("evacuation_destination_other") or "").strip().upper()
    evacuation_type = (request.form.get("evacuation_type") or "").strip().upper()
    evacuation_notes = (request.form.get("evacuation_notes") or "").strip().upper()

    if destination == "OTRO":
        destination = custom_destination

    if not destination:
        flash("Debe indicar el destino de evacuación.", "danger")
        return redirect(url_for("inventory.inventory_index"))

    if evacuation_type not in {"RT", "BARCO", "EVACUACION"}:
        flash("Debe indicar el tipo de evacuación.", "danger")
        return redirect(url_for("inventory.inventory_index"))

    c.dispatch_status = "PARA_EVACUAR"
    c.dispatch_marked_at = datetime.utcnow()
    c.dispatch_marked_by_user_id = current_user.id
    c.evacuation_destination = destination
    c.evacuation_type = evacuation_type
    c.evacuation_notes = evacuation_notes or None

    db.session.commit()

    flash(f"Contenedor {c.code} marcado para evacuar hacia {destination}.", "success")
    return redirect(url_for("inventory.inventory_index"))


@inventory_bp.post("/inventory/<int:container_id>/unmark-evacuation")
@login_required
def unmark_container_evacuation(container_id: int):
    site_id = _ensure_active_site()

    c = Container.query.get_or_404(container_id)

    if c.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    if (c.dispatch_status or "NORMAL") != "PARA_EVACUAR":
        flash("Solo se puede quitar evacuación a contenedores en estado Evacuar.", "warning")
        return redirect(url_for("inventory.inventory_index"))

    c.dispatch_status = "NORMAL"
    c.dispatch_marked_at = None
    c.dispatch_marked_by_user_id = None
    c.evacuation_destination = None
    c.evacuation_type = None
    c.evacuation_notes = None

    db.session.commit()

    flash(f"Contenedor {c.code} volvió a Disponible.", "success")
    return redirect(url_for("inventory.inventory_index"))

@inventory_bp.get("/inventory/evacuation-list")
@login_required
def evacuation_list():
    site_id = _ensure_active_site()

    size_filter = (request.args.get("size") or "").strip().upper()
    shipping_line_filter = (request.args.get("shipping_line") or "").strip().upper()
    destination_filter = (request.args.get("destination") or "").strip().upper()
    type_filter = (request.args.get("evacuation_type") or "").strip().upper()

    rows = (
        db.session.query(Container, ContainerPosition, YardBay)
        .outerjoin(ContainerPosition, ContainerPosition.container_id == Container.id)
        .outerjoin(YardBay, YardBay.id == ContainerPosition.bay_id)
        .filter(
            Container.site_id == site_id,
            Container.is_in_yard == True,  # noqa: E712
            db.func.coalesce(Container.dispatch_status, "NORMAL") == "PARA_EVACUAR",
        )
        .order_by(Container.updated_at.desc())
        .all()
    )

    container_ids = [c.id for c, _, _ in rows]
    cls_by_container = _last_classification_by_container_ids(container_ids)

    items = []

    for c, pos, bay in rows:
        cls = cls_by_container.get(c.id)
        shipping_line = ((cls.get("shipping_line") if cls else "") or "").strip().upper()
        destination = (c.evacuation_destination or "").strip().upper()
        evacuation_type = (c.evacuation_type or "").strip().upper()

        if size_filter and (c.size or "").strip().upper() != size_filter:
            continue

        if shipping_line_filter and shipping_line != shipping_line_filter:
            continue

        if destination_filter and destination != destination_filter:
            continue

        if type_filter and evacuation_type != type_filter:
            continue

        items.append({
            "id": c.id,
            "code": c.code,
            "size": c.size,
            "shipping_line": shipping_line,
            "dispatch_status": c.dispatch_status or "NORMAL",
            "dispatch_marked_at": c.dispatch_marked_at,
            "evacuation_destination": c.evacuation_destination,
            "evacuation_type": c.evacuation_type,
            "evacuation_notes": c.evacuation_notes,
            "position": None if not pos else {
                "bay_code": bay.code if bay else None,
                "depth_row": pos.depth_row,
                "tier": pos.tier,
            },
        })

    size_options = sorted({(c.size or "").strip().upper() for c, _, _ in rows if c.size})
    shipping_line_options = sorted({
        ((cls_by_container.get(c.id) or {}).get("shipping_line") or "").strip().upper()
        for c, _, _ in rows
        if ((cls_by_container.get(c.id) or {}).get("shipping_line") or "").strip()
    })
    destination_options = sorted({
        (c.evacuation_destination or "").strip().upper()
        for c, _, _ in rows
        if (c.evacuation_destination or "").strip()
    })
    type_options = sorted({
        (c.evacuation_type or "").strip().upper()
        for c, _, _ in rows
        if (c.evacuation_type or "").strip()
    })

    return render_template(
        "inventory/evacuation_list.html",
        items=items,
        size_options=size_options,
        shipping_line_options=shipping_line_options,
        destination_options=destination_options,
        type_options=type_options,
        size_filter=size_filter,
        shipping_line_filter=shipping_line_filter,
        destination_filter=destination_filter,
        type_filter=type_filter,
    )

# =========================================================
# Carga masiva de contenedores
# =========================================================

BULK_REQUIRED_HEADERS = {
    "CONTENEDOR",
    "TAMAÑO",
    "NAVIERA",
    "ESTADO",
}

BULK_HEADERS = [
    "CONTENEDOR",
    "TAMAÑO",
    "NAVIERA",
    "ESTADO",
    "CLASIFICACION",
    "FECHA_INGRESO",
    "AÑO",
    "MAX_GROSS",
    "TARA",
    "NOTAS",
    "ORIGEN",
    "ESTIBA",
    "FILA",
    "NIVEL",
    "DESTINO_EVACUACION",
    "TIPO_EVACUACION",
    "OBS_EVACUACION",
]

BULK_STATUS_MAP = {
    "DISPONIBLE": "NORMAL",
    "NORMAL": "NORMAL",

    "NO_USAR": "NO_USAR",
    "NO USAR": "NO_USAR",
    "BLOQUEADO": "NO_USAR",

    "ASIGNADO": "PARA_DESPACHO",
    "PARA_DESPACHO": "PARA_DESPACHO",

    "EVACUAR": "PARA_EVACUAR",
    "PARA_EVACUAR": "PARA_EVACUAR",

    "ASIGNADO_EVACUAR": "EVACUAR_SOLICITADO",
    "EVACUAR_SOLICITADO": "EVACUAR_SOLICITADO",

    "DESPACHO_MONTADO": "DESPACHO_MONTADO",
    "EVACUACION_MONTADA": "EVACUACION_MONTADA",
}

BULK_VALID_SIZES = {
    "20ST",
    "20OT",
    "20RF",
    "20TQ",
    "40ST",
    "40HC",
    "40RF",
    "40OT",
    "45HC",
}

BULK_VALID_EVAC_TYPES = {
    "RT",
    "BARCO",
    "EVACUACION",
}

BULK_VALID_CLASSIFICATIONS = {
    "A+",
    "A-",
    "B+",
    "B-",
    "C",
    "A2",
    "B2",
    "CHATARRA",
}


def _bulk_clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _bulk_upper(value):
    return _bulk_clean(value).upper()


def _bulk_int(value):
    value = _bulk_clean(value)
    if not value:
        return None

    try:
        return int(float(value))
    except Exception:
        return None

def _bulk_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.replace(hour=0, minute=0, second=0, microsecond=0)

    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    raw = str(value).strip()

    if not raw:
        return None

    try:
        return datetime.strptime(raw, "%Y-%m-%d")
    except Exception:
        return "INVALID"

def _bulk_normalize_container_code(value):
    raw = _bulk_upper(value)
    raw = raw.replace(" ", "").replace("\t", "")

    if not raw:
        return ""

    if "-" in raw:
        return raw

    # Formato sin guiones: CSNU0012888 -> CSNU-001288-8
    if len(raw) == 11 and raw[:4].isalpha() and raw[4:].isdigit():
        return f"{raw[:4]}-{raw[4:10]}-{raw[10]}"

    return raw


def _bulk_headers_from_sheet(ws):
    headers = {}

    for idx, cell in enumerate(ws[1], start=1):
        name = _bulk_upper(cell.value)
        if name:
            headers[name] = idx

    return headers


def _bulk_get(row, headers, name):
    idx = headers.get(name)
    if not idx:
        return ""
    return _bulk_clean(row[idx - 1].value)


def _bulk_validate_position(
    *,
    site_id: int,
    container_size: str,
    bay_code: str,
    depth_row,
    tier,
):
    bay_code = _bulk_upper(bay_code)

    if not bay_code and not depth_row and not tier:
        return {
            "ok": True,
            "has_position": False,
            "bay": None,
            "depth_row": None,
            "tier": None,
        }

    if not bay_code or not depth_row or not tier:
        return {
            "ok": False,
            "message": "Si se indica ubicación, ESTIBA, FILA y NIVEL deben venir completos.",
        }

    bay = YardBay.query.filter_by(
        site_id=site_id,
        code=bay_code,
        is_active=True,
    ).first()

    if not bay:
        return {
            "ok": False,
            "message": f"La estiba {bay_code} no existe o está inactiva en este predio.",
        }

    try:
        depth_row = int(depth_row)
        tier = int(tier)
    except Exception:
        return {
            "ok": False,
            "message": "FILA y NIVEL deben ser numéricos.",
        }

    if depth_row < 1 or depth_row > int(bay.max_depth_rows or 0):
        return {
            "ok": False,
            "message": f"La fila {depth_row} está fuera del rango permitido para {bay.code}.",
        }

    if tier < 1 or tier > int(bay.max_tiers or 0):
        return {
            "ok": False,
            "message": f"El nivel {tier} está fuera del rango permitido para {bay.code}.",
        }

    size = _bulk_upper(container_size)
    bay_size = _bulk_upper(bay.container_size_type or "40")

    if bay_size == "20" and not size.startswith("20"):
        return {
            "ok": False,
            "message": f"La estiba {bay.code} solo acepta contenedores de 20 pies.",
        }

    if bay_size == "40" and not (size.startswith("40") or size.startswith("45")):
        return {
            "ok": False,
            "message": f"La estiba {bay.code} solo acepta contenedores de 40/45 pies.",
        }

    occupied = (
        db.session.query(Container, ContainerPosition)
        .join(ContainerPosition, ContainerPosition.container_id == Container.id)
        .filter(
            Container.site_id == site_id,
            Container.is_in_yard == True,  # noqa: E712
            ContainerPosition.bay_id == bay.id,
            ContainerPosition.depth_row == depth_row,
            ContainerPosition.tier == tier,
        )
        .first()
    )

    if occupied:
        c, _ = occupied
        return {
            "ok": False,
            "message": f"La posición {bay.code} F{depth_row:02d} N{tier} ya está ocupada por {c.code}.",
        }

    if tier > 1:
        support = (
            db.session.query(ContainerPosition)
            .join(Container, Container.id == ContainerPosition.container_id)
            .filter(
                Container.site_id == site_id,
                Container.is_in_yard == True,  # noqa: E712
                ContainerPosition.bay_id == bay.id,
                ContainerPosition.depth_row == depth_row,
                ContainerPosition.tier == tier - 1,
            )
            .first()
        )

        if not support:
            return {
                "ok": False,
                "message": f"No se puede colocar en N{tier} sin soporte debajo en N{tier - 1}.",
            }

    access_blocker = (
        db.session.query(Container, ContainerPosition)
        .join(ContainerPosition, ContainerPosition.container_id == Container.id)
        .filter(
            Container.site_id == site_id,
            Container.is_in_yard == True,  # noqa: E712
            ContainerPosition.bay_id == bay.id,
            ContainerPosition.depth_row > depth_row,
        )
        .first()
    )

    if access_blocker:
        c, p = access_blocker
        return {
            "ok": False,
            "message": f"La sidepick no puede acceder a F{depth_row:02d}; bloquea {c.code} en F{p.depth_row:02d} N{p.tier}.",
        }

    return {
        "ok": True,
        "has_position": True,
        "bay": bay,
        "depth_row": depth_row,
        "tier": tier,
    }


@inventory_bp.get("/inventory/bulk-upload/template")
@login_required
def inventory_bulk_upload_template():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "DATOS"

    ws.append(BULK_HEADERS)

    header_fill = PatternFill("solid", fgColor="0F3B63")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 18)

    example = [
        "CSNU-001288-8",
        "40HC",
        "ONE",
        "DISPONIBLE",
        "A+",
        "2026-06-29",
        2015,
        32500,
        3800,
        "Piso OK",
        "",      # ORIGEN: vacío, LIMON o CALDERA
        "A1E",
        1,
        1,
        "",
        "",
        "",
    ]

    ws.append(example)

    dv_size = DataValidation(
        type="list",
        formula1='"20ST,20OT,20RF,20TQ,40ST,40HC,40RF,40OT,45HC"',
        allow_blank=False,
    )

    dv_status = DataValidation(
        type="list",
        formula1='"DISPONIBLE,NO_USAR,ASIGNADO,EVACUAR,ASIGNADO_EVACUAR,DESPACHO_MONTADO,EVACUACION_MONTADA"',
        allow_blank=False,
    )

    dv_classification = DataValidation(
        type="list",
        formula1='"A+,A-,B+,B-,C,A2,B2,CHATARRA"',
        allow_blank=True,
    )

    dv_origin = DataValidation(
        type="list",
        formula1='"LIMON,CALDERA"',
        allow_blank=True,
    )

    dv_dest = DataValidation(
        type="list",
        formula1='"LIMON,CALDERA,OTRO"',
        allow_blank=True,
    )

    dv_type = DataValidation(
        type="list",
        formula1='"RT,BARCO,EVACUACION"',
        allow_blank=True,
    )

    ws.add_data_validation(dv_size)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_origin)
    ws.add_data_validation(dv_dest)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_classification)

    dv_size.add("B2:B5000")
    dv_status.add("D2:D5000")
    dv_origin.add("K2:K5000")
    dv_dest.add("O2:O5000")
    dv_type.add("P2:P5000")
    dv_classification.add("E2:E5000")

    ws2 = wb.create_sheet("INSTRUCCIONES")

    instructions = [
        ["CARGA MASIVA DE CONTENEDORES", ""],
        ["", ""],
        ["Campos obligatorios", "CONTENEDOR, TAMAÑO, NAVIERA, ESTADO"],
        ["Campos opcionales", "CLASIFICACION, FECHA_INGRESO, AÑO, MAX_GROSS, TARA, NOTAS, ORIGEN, ESTIBA, FILA, NIVEL, DESTINO_EVACUACION, TIPO_EVACUACION, OBS_EVACUACION"],
        ["", ""],
        ["Regla de ubicación", "Si ESTIBA, FILA y NIVEL vienen completos, se ubicará automáticamente."],
        ["Regla de ubicación", "Si falta alguno de los tres, quedará en patio pendiente de ubicar."],
        ["", ""],
        ["ESTADO EXCEL", "ESTADO SISTEMA"],
        ["DISPONIBLE", "NORMAL"],
        ["NO_USAR", "NO_USAR"],
        ["ASIGNADO", "PARA_DESPACHO"],
        ["EVACUAR", "PARA_EVACUAR"],
        ["ASIGNADO_EVACUAR", "EVACUAR_SOLICITADO"],
        ["DESPACHO_MONTADO", "DESPACHO_MONTADO"],
        ["EVACUACION_MONTADA", "EVACUACION_MONTADA"],
        ["", ""],
        ["Tamaños permitidos", "20ST, 20OT, 20RF, 20TQ, 40ST, 40HC, 40RF, 40OT, 45HC"],
        ["Tipos evacuación", "RT, BARCO, EVACUACION"],
        ["Destinos evacuación", "LIMON, CALDERA, OTRO"],
        ["Origen", "Puede venir vacío, LIMON o CALDERA"],
        ["Clasificación", "Valores permitidos: A+, A-, B+, B-, C, A2, B2, CHATARRA"],
        ["Fecha ingreso", "Formato obligatorio: YYYY-MM-DD. Ejemplo: 2026-06-29"],
    ]

    for row in instructions:
        ws2.append(row)

    for col in range(1, 3):
        ws2.column_dimensions[get_column_letter(col)].width = 38

    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A9"].font = Font(bold=True)
    ws2["B9"].font = Font(bold=True)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="plantilla_carga_masiva_contenedores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@inventory_bp.get("/inventory/bulk-upload")
@login_required
def inventory_bulk_upload_view():
    _ensure_active_site()

    return render_template(
        "inventory/bulk_upload.html",
        result=None,
        errors=[],
    )


@inventory_bp.post("/inventory/bulk-upload")
@login_required
def inventory_bulk_upload_post():
    site_id = _ensure_active_site()

    file = request.files.get("file")

    if not file or not file.filename:
        flash("Debe seleccionar un archivo Excel.", "warning")
        return redirect(url_for("inventory.inventory_bulk_upload_view"))

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("El archivo debe ser .xlsx o .xlsm.", "danger")
        return redirect(url_for("inventory.inventory_bulk_upload_view"))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception:
        flash("No se pudo leer el archivo Excel.", "danger")
        return redirect(url_for("inventory.inventory_bulk_upload_view"))

    ws = wb["DATOS"] if "DATOS" in wb.sheetnames else wb.active

    headers = _bulk_headers_from_sheet(ws)
    missing = sorted(BULK_REQUIRED_HEADERS - set(headers.keys()))

    if missing:
        return render_template(
            "inventory/bulk_upload.html",
            result=None,
            errors=[
                {
                    "row": 1,
                    "container": "—",
                    "message": f"Faltan columnas obligatorias: {', '.join(missing)}",
                }
            ],
        )

    errors = []
    parsed_rows = []
    seen_codes = set()

    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        raw_values = [cell.value for cell in row]

        if all(v is None or str(v).strip() == "" for v in raw_values):
            continue

        code = _bulk_normalize_container_code(_bulk_get(row, headers, "CONTENEDOR"))
        size = _bulk_upper(_bulk_get(row, headers, "TAMAÑO"))
        shipping_line = _bulk_upper(_bulk_get(row, headers, "NAVIERA"))
        status_excel = _bulk_upper(_bulk_get(row, headers, "ESTADO"))
        final_classification = _bulk_upper(_bulk_get(row, headers, "CLASIFICACION"))
        entry_date = _bulk_date(_bulk_get(row, headers, "FECHA_INGRESO"))

        year = _bulk_int(_bulk_get(row, headers, "AÑO"))
        max_gross_kg = _bulk_int(_bulk_get(row, headers, "MAX_GROSS"))
        tare_kg = _bulk_int(_bulk_get(row, headers, "TARA"))
        notes = _bulk_clean(_bulk_get(row, headers, "NOTAS"))
        origin = _bulk_upper(_bulk_get(row, headers, "ORIGEN"))

        bay_code = _bulk_upper(_bulk_get(row, headers, "ESTIBA"))
        depth_row = _bulk_int(_bulk_get(row, headers, "FILA"))
        tier = _bulk_int(_bulk_get(row, headers, "NIVEL"))

        evac_destination = _bulk_upper(_bulk_get(row, headers, "DESTINO_EVACUACION"))
        evac_type = _bulk_upper(_bulk_get(row, headers, "TIPO_EVACUACION"))
        evac_notes = _bulk_clean(_bulk_get(row, headers, "OBS_EVACUACION"))

        row_errors = []

        if not code:
            row_errors.append("CONTENEDOR es obligatorio.")

        if not size:
            row_errors.append("TAMAÑO es obligatorio.")
        elif size not in BULK_VALID_SIZES:
            row_errors.append(f"Tamaño inválido: {size}.")

        if not shipping_line:
            row_errors.append("NAVIERA es obligatoria.")

        if not status_excel:
            row_errors.append("ESTADO es obligatorio.")

        dispatch_status = BULK_STATUS_MAP.get(status_excel)

        if not dispatch_status:
            row_errors.append(f"Estado inválido: {status_excel}.")

        if final_classification and final_classification not in BULK_VALID_CLASSIFICATIONS:
            row_errors.append(
                f"CLASIFICACION inválida: {final_classification}. "
                "Valores permitidos: A+, A-, B+, B-, C, A2, B2, CHATARRA."
            )

        if entry_date == "INVALID":
            row_errors.append(
                "FECHA_INGRESO inválida. Use formato YYYY-MM-DD, ejemplo: 2026-06-29."
            )

        if code in seen_codes:
            row_errors.append(f"El contenedor {code} está duplicado dentro del Excel.")

        if code:
            seen_codes.add(code)

            exists = Container.query.filter_by(
                site_id=site_id,
                code=code,
            ).first()

            if exists:
                row_errors.append(f"El contenedor {code} ya existe en este predio.")

        if year is not None and (year < 1980 or year > 2100):
            row_errors.append(f"Año inválido: {year}.")

        if max_gross_kg is not None and max_gross_kg <= 0:
            row_errors.append("MAX_GROSS debe ser mayor a 0.")

        if tare_kg is not None and tare_kg <= 0:
            row_errors.append("TARA debe ser mayor a 0.")

        if origin and origin not in {"LIMON", "CALDERA"}:
            row_errors.append(f"ORIGEN inválido: {origin}. Debe ser LIMON, CALDERA o vacío.")

        if evac_type and evac_type not in BULK_VALID_EVAC_TYPES:
            row_errors.append(f"TIPO_EVACUACION inválido: {evac_type}.")

        is_mounted_status = dispatch_status in {
            "DESPACHO_MONTADO",
            "EVACUACION_MONTADA",
        }

        position_result = {
            "ok": True,
            "has_position": False,
            "bay": None,
            "depth_row": None,
            "tier": None,
        }

        if not is_mounted_status:
            position_result = _bulk_validate_position(
                site_id=site_id,
                container_size=size,
                bay_code=bay_code,
                depth_row=depth_row,
                tier=tier,
            )

            if not position_result.get("ok"):
                row_errors.append(position_result.get("message") or "Ubicación inválida.")

        if row_errors:
            for msg in row_errors:
                errors.append({
                    "row": row_number,
                    "container": code or "—",
                    "message": msg,
                })
            continue

        parsed_rows.append({
            "row_number": row_number,
            "code": code,
            "size": size,
            "shipping_line": shipping_line,
            "dispatch_status": dispatch_status,
            "year": year,
            "max_gross_kg": max_gross_kg,
            "tare_kg": tare_kg,
            "notes": notes,
            "origin": origin,
            "final_classification": final_classification,
            "entry_date": entry_date,
            "position": position_result,
            "evac_destination": evac_destination,
            "evac_type": evac_type,
            "evac_notes": evac_notes,
            "is_mounted_status": is_mounted_status,
        })

    if errors:
        db.session.rollback()

        return render_template(
            "inventory/bulk_upload.html",
            result={
                "ok": False,
                "created": 0,
                "validated": len(parsed_rows),
                "errors_count": len(errors),
            },
            errors=errors,
        )

    created_count = 0
    positioned_count = 0
    pending_location_count = 0
    mounted_count = 0
    BATCH_SIZE = 50

    try:
        for item in parsed_rows:
            position = item["position"]
            has_position = bool(position.get("has_position"))
            entry_at = item["entry_date"] or datetime.utcnow()

            status_notes = None

            if not has_position:
                status_notes = "PENDIENTE_UBICAR_EN_PATIO"

            c = Container(
                site_id=site_id,
                code=item["code"],
                size=item["size"],
                year=item["year"],
                status_notes=status_notes,
                gate_in_origin_port=item["origin"] or None,
                is_in_yard=True,
                dispatch_status=item["dispatch_status"],
            )

            if item["dispatch_status"] == "PARA_EVACUAR":
                c.dispatch_marked_at = entry_at
                c.dispatch_marked_by_user_id = current_user.id
                c.evacuation_destination = item["evac_destination"] or None
                c.evacuation_type = item["evac_type"] or None
                c.evacuation_notes = item["evac_notes"] or None

            if item["dispatch_status"] in {"DESPACHO_MONTADO", "EVACUACION_MONTADA"}:
                c.mounted_at = entry_at
                c.mounted_by_user_id = current_user.id

            db.session.add(c)
            db.session.flush([c])

            db.session.add(
                ContainerClassification(
                    site_id=site_id,
                    container_id=c.id,
                    classified_at=entry_at,
                    classified_by_user_id=current_user.id,
                    shipping_line=item["shipping_line"],
                    max_gross_kg=item["max_gross_kg"],
                    tare_kg=item["tare_kg"],
                    manufacture_year=item["year"],
                    needs_workshop=False,
                    final_classification=item["final_classification"] or None,
                    summary_text=item["notes"] or None,
                    notes=item["notes"] or None,
                )
            )

            db.session.add(
                Movement(
                    site_id=site_id,
                    container_id=c.id,
                    movement_type="GATE_IN",
                    occurred_at=entry_at,
                    bay_code=None,
                    depth_row=None,
                    tier=None,
                    created_by_user_id=current_user.id,
                    notes="BULK_IMPORT_GATE_IN",
                )
            )

            if item["is_mounted_status"]:
                mounted_count += 1

                db.session.add(
                    Movement(
                        site_id=site_id,
                        container_id=c.id,
                        movement_type="MOVE",
                        occurred_at=entry_at,
                        bay_code=None,
                        depth_row=None,
                        tier=None,
                        created_by_user_id=current_user.id,
                        notes=f"BULK_IMPORT_{item['dispatch_status']}_WITHOUT_POSITION",
                    )
                )

            elif has_position:
                bay = position["bay"]
                depth_row = position["depth_row"]
                tier = position["tier"]

                db.session.add(
                    ContainerPosition(
                        container_id=c.id,
                        bay_id=bay.id,
                        depth_row=depth_row,
                        tier=tier,
                        placed_at=entry_at,
                        placed_by_user_id=current_user.id,
                    )
                )

                db.session.add(
                    Movement(
                        site_id=site_id,
                        container_id=c.id,
                        movement_type="MOVE",
                        occurred_at=entry_at,
                        bay_code=bay.code,
                        depth_row=depth_row,
                        tier=tier,
                        created_by_user_id=current_user.id,
                        notes="BULK_IMPORT_PLACED_WITH_POSITION",
                    )
                )

                positioned_count += 1

            else:
                db.session.add(
                    Movement(
                        site_id=site_id,
                        container_id=c.id,
                        movement_type="MOVE",
                        occurred_at=entry_at,
                        bay_code=None,
                        depth_row=None,
                        tier=None,
                        created_by_user_id=current_user.id,
                        notes="BULK_IMPORT_PENDING_LOCATION",
                    )
                )

                pending_location_count += 1

            created_count += 1

            if created_count % BATCH_SIZE == 0:
                db.session.commit()

        db.session.commit()

    except Exception as exc:
        db.session.rollback()

        return render_template(
            "inventory/bulk_upload.html",
            result=None,
            errors=[
                {
                    "row": "—",
                    "container": "—",
                    "message": f"Error general importando archivo: {str(exc)}",
                }
            ],
        )

    flash(f"Carga masiva completada. {created_count} contenedores creados.", "success")

    return render_template(
        "inventory/bulk_upload.html",
        result={
            "ok": True,
            "created": created_count,
            "positioned": positioned_count,
            "pending_location": pending_location_count,
            "mounted": mounted_count,
            "errors_count": 0,
        },
        errors=[],
    )

@inventory_bp.post("/inventory/<int:container_id>/update-gate-in-origin")
@login_required
def update_gate_in_origin(container_id: int):
    c = Container.query.get_or_404(container_id)

    gate_in_origin_port = (
        request.form.get("gate_in_origin_port") or ""
    ).strip().upper()

    if gate_in_origin_port not in {"", "LIMON", "CALDERA"}:
        flash("Origen de ingreso inválido.", "danger")
        return redirect(url_for("inventory.inventory_detail", container_id=c.id))

    c.gate_in_origin_port = gate_in_origin_port or None
    c.updated_at = datetime.utcnow()

    db.session.add(c)

    audit_log(
        current_user.id,
        "CONTAINER_GATE_IN_ORIGIN_UPDATED",
        "container",
        c.id,
        {
            "container_code": c.code,
            "gate_in_origin_port": c.gate_in_origin_port,
        },
    )

    db.session.commit()

    flash("Origen de ingreso actualizado.", "success")
    return redirect(url_for("inventory.inventory_detail", container_id=c.id))

@inventory_bp.get("/inventory/evacuation-list/pdf")
@login_required
def evacuation_list_pdf():
    site_id = _ensure_active_site()

    size_filter = (request.args.get("size") or "").strip().upper()
    shipping_line_filter = (request.args.get("shipping_line") or "").strip().upper()
    destination_filter = (request.args.get("destination") or "").strip().upper()
    type_filter = (request.args.get("evacuation_type") or "").strip().upper()

    rows = (
        db.session.query(Container, ContainerPosition, YardBay)
        .outerjoin(ContainerPosition, ContainerPosition.container_id == Container.id)
        .outerjoin(YardBay, YardBay.id == ContainerPosition.bay_id)
        .filter(
            Container.site_id == site_id,
            Container.is_in_yard == True,  # noqa: E712
            db.func.coalesce(Container.dispatch_status, "NORMAL") == "PARA_EVACUAR",
        )
        .order_by(Container.updated_at.desc())
        .all()
    )

    container_ids = [c.id for c, _, _ in rows]
    cls_by_container = _last_classification_by_container_ids(container_ids)

    items = []

    for c, pos, bay in rows:
        cls = cls_by_container.get(c.id)
        shipping_line = ((cls.get("shipping_line") if cls else "") or "").strip().upper()
        destination = (c.evacuation_destination or "").strip().upper()
        evacuation_type = (c.evacuation_type or "").strip().upper()

        if size_filter and (c.size or "").strip().upper() != size_filter:
            continue

        if shipping_line_filter and shipping_line != shipping_line_filter:
            continue

        if destination_filter and destination != destination_filter:
            continue

        if type_filter and evacuation_type != type_filter:
            continue

        position_txt = "—"
        if pos:
            position_txt = f"{bay.code if bay else ''} F{int(pos.depth_row):02d} N{pos.tier}"

        marked_txt = ""
        if c.dispatch_marked_at:
            marked_txt = c.dispatch_marked_at.strftime("%d/%m/%Y %I:%M %p")

        items.append([
            c.code or "",
            c.size or "",
            shipping_line or "—",
            destination or "—",
            evacuation_type or "—",
            position_txt,
            marked_txt,
            c.evacuation_notes or "",
        ])

    bio = BytesIO()

    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(letter),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph("<b>Lista de Vacíos / Evacuación</b>", styles["Title"])
    elements.append(title)

    filters_txt = (
        f"Tamaño: {size_filter or 'Todos'} | "
        f"Naviera: {shipping_line_filter or 'Todas'} | "
        f"Destino: {destination_filter or 'Todos'} | "
        f"Tipo: {type_filter or 'Todos'}"
    )

    elements.append(Paragraph(filters_txt, styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [
        [
            "Contenedor",
            "Tamaño",
            "Naviera",
            "Destino",
            "Tipo",
            "Posición",
            "Marcado",
            "Notas",
        ]
    ]

    table_data.extend(items)

    if len(table_data) == 1:
        table_data.append(["—", "—", "—", "—", "—", "—", "—", "Sin registros"])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[90, 55, 70, 90, 55, 80, 105, 210],
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F3B63")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))

    elements.append(table)

    doc.build(elements)

    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="lista_vacios_evacuacion.pdf",
        mimetype="application/pdf",
    )


@inventory_bp.post("/inventory/<int:container_id>/mark-no-use")
@login_required
def mark_container_no_use(container_id: int):
    site_id = _ensure_active_site()

    c = Container.query.get_or_404(container_id)

    if c.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    if not c.is_in_yard:
        flash("Solo se pueden marcar como NO USAR contenedores que están en patio.", "warning")
        return redirect(url_for("inventory.inventory_index"))

    current_status = (c.dispatch_status or "NORMAL").strip().upper()

    if current_status != "NORMAL":
        flash("Solo se pueden marcar como NO USAR contenedores disponibles.", "warning")
        return redirect(url_for("inventory.inventory_index"))

    c.dispatch_status = "NO_USAR"
    c.dispatch_marked_at = datetime.utcnow()
    c.dispatch_marked_by_user_id = current_user.id

    db.session.commit()

    flash(f"Contenedor {c.code} marcado como NO USAR.", "success")
    return redirect(url_for("inventory.inventory_index"))


@inventory_bp.post("/inventory/<int:container_id>/unmark-no-use")
@login_required
def unmark_container_no_use(container_id: int):
    site_id = _ensure_active_site()

    c = Container.query.get_or_404(container_id)

    if c.site_id != site_id and getattr(current_user, "role", None) != "admin":
        abort(403)

    current_status = (c.dispatch_status or "NORMAL").strip().upper()

    if current_status != "NO_USAR":
        flash("Solo se puede habilitar un contenedor que está en estado NO USAR.", "warning")
        return redirect(url_for("inventory.inventory_index"))

    c.dispatch_status = "NORMAL"
    c.dispatch_marked_at = None
    c.dispatch_marked_by_user_id = None

    db.session.commit()

    flash(f"Contenedor {c.code} volvió a Disponible.", "success")
    return redirect(url_for("inventory.inventory_index"))